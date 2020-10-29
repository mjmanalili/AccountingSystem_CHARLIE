import sqlite3
from tkinter import *
from tkinter import messagebox
from tkinter.filedialog import askopenfilename
import tkinter.ttk as tk
import datetime
from openpyxl import Workbook
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import PatternFill
import openpyxl
from PIL import ImageTk, Image
import os
from random import choice
from num2words import num2words
from plyer import notification
# from ipaddress import IPv4Address
# from pyairmore.request import AirmoreSession
# from pyairmore.services.messaging import MessagingService
# from win32com import client
# import win32api

path = "C:/Users/Mj Manalili/Desktop/codingprojects/ASD/" #//Dbpsc-acctng-05/charlie database/
savepath = "C:/Users/Mj Manalili/Desktop/codingprojects/ASD/" #C:/Users/Mj Manalili/Desktop/codingprojects/ASD/"  #D:/charlie/
iconpath = "C:/Users/Mj Manalili/Desktop/codingprojects/ASD/iconpack/" #//Dbpsc-acctng-05/charlie database/
centers = ["CLIENT","ALL DEPTS.","ASD","AUDIT","BCD","DPD","FMD","GSAD","HRMDD","IT","LEGAL","MKTG","OP","TMSU"]
maintitle = "CHARLIE System"
geometry = "1095x700+200+0"
bcdgeometry = "1195x700+100+0"
version  = "2020.2.0"
mainbg = "gray20"
wc = "gray15"
fc = "white"
click = "lawn green"
buttonbg = "white"
totalbg = "white"
totalfg = "black"
codetitlebg = "white"
codetitlefg = "black"
dbpscbg = "royal blue"
dssibg = "maroon"
cancelbg = "#FDEA79"
cancelfg = "black"
cancelfont = ('calibri', 9)
fonts = ("calibri", 12)
boldfonts = ("calibri", 12, "bold")
fontreports = ("calibri", 10)
unreadfont = ("calibri", 10, "bold")
clearbuttonfont = ("calibri", 10)
iconfont = ("calibri", 8)
chatdatefont = ("calibri", 7)
underlinefont = ("calibri", 12, "underline")
fsbg = "F6F0F0"
cancelleddocument = '#ffb347'
offlinebg = "red2"
pad = 2
today = datetime.date.today()

class main:
    def __init__(self, master):
        self.master = master
        self.username = StringVar()
        self.password = StringVar()
        self.widgets()
        master.protocol("WM_DELETE_WINDOW", self.exit)
        
    def new_user(self, *args):
        user_entry = self.username.get()
        pass_entry1 = password_entry1.get()
        pass_entry2 = password_entry2.get()
        find_user = ("SELECT * FROM users WHERE username = ?")
        c.execute(find_user,[user_entry])
        results = c.fetchall()
        if results:
            for user_entry in results:
                messagebox.showerror("Error", "Username already taken!")
                break
        else:
            if pass_entry1 == pass_entry2:
                insert = "INSERT INTO users (username, password, type) VALUES (?,?,?)"
                c.execute(insert,[user_entry, pass_entry1, usertype_entry.get()])
                conn.commit()
                messagebox.showinfo("Success!", "Account Created Successfully!")
                username_entry.delete(0, END)
                password_entry1.delete(0, END)
                password_entry2.delete(0, END)
                usertype_entry.delete(0, END)
            else:
                messagebox.showerror("Error!", "Password entries not matched!")
                
    def login(self, *args):
        if username_entry["fg"] == codetitlefg and password_entry["fg"] == codetitlefg:
            user_entry = self.username.get()
            pass_entry = self.password.get()
            find_user = "SELECT * FROM users WHERE username = ? AND password = ?"
            c.execute(find_user,[user_entry, pass_entry])
            results = c.fetchall()
            if results:
                boylabo.destroy()
                log_frame.destroy()
                self.useraccess()
                self.company()
            else:
                messagebox.showerror("Error!", "Invalid username and password!")
        else:
            messagebox.showerror("Error!", "Invalid username and password!")
    
    def useraccess(self):
        global access, avatar
        user = self.username.get()
        identifier = "SELECT type, department, gender, avatar FROM users WHERE username = ?"
        c.execute(identifier, [user])
        access = c.fetchone()
        updater = "UPDATE users SET status = 'Online' WHERE username = ?"
        c.execute(updater, [user])
        conn.commit()
        avatar = ImageTk.PhotoImage(Image.open(iconpath + "avatars/" + access[2] + "/" + access[3] + ".png")) #path

    def exit(self, *args):
        try:
            updater = "UPDATE users SET status = 'Offline' WHERE username = ?"
            c.execute(updater, [self.username.get()])
            conn.commit()
        except:
            pass
        conn.close()
        self.master.destroy()

    def widgets(self):
        global login_frame, log_frame, username_entry, password_entry, login_button, change_password, themepicker, boylabo
        login_frame = LabelFrame(self.master, text = "CHARLIE System", font = fonts, fg = fc, bg = wc, padx = 10)
        login_frame.place(x = 385, y = 255)

        boylabo = Label(self.master, text = "CHARLIE System - ASD v" + version + " | Boy Labo Systems™", font = ("calibri", 7), bg = mainbg, fg = fc)
        boylabo.place(x = 385, y = 549)

        log_frame = Frame(login_frame, bg = wc)
        log_frame.grid(column = 0, row = 0)

        roundedlabel1 = Label(log_frame, bg = wc, image = roundedicon, width = 260)
        roundedlabel1.place(x = 0, y = 0)

        roundedlabel2 = Label(log_frame, bg = wc, image = roundedicon, width = 260)
        roundedlabel2.place(x = 0, y = 38)

        username_entry = Entry(log_frame, textvariable = self.username, font = fonts, fg = "grey", bd = 0)
        username_entry.grid(column = 1, row = 0, pady = 12, padx = 50)
        username_entry.insert(0, "Username")
        username_entry.bind("<Return>", self.login)
        username_entry.bind("<FocusIn>", self.usernamefocus)
        username_entry.bind("<FocusOut>", self.usernamefocus)

        password_entry = Entry(log_frame, textvariable = self.password, font = fonts, fg = "grey", bd = 0)
        password_entry.grid(column = 1, row = 1, pady = 5)
        password_entry.insert(0, "Password")
        password_entry.bind("<Return>", self.login)
        password_entry.bind("<FocusIn>", self.passwordfocus)
        password_entry.bind("<FocusOut>", self.passwordfocus)

        login_button = Button(log_frame, text = " Login", font = fonts, bd = 0, bg = wc, fg = fc, cursor = "hand2", width = 90, image = loginicon, compound = LEFT, command = self.login)
        login_button.grid(column = 1, row = 2, pady = 2)
        login_button.bind("<Return>", self.login)
        login_button.bind("<Enter>", lambda e: login_button.config(bg = fc, fg = wc))
        login_button.bind("<Leave>", lambda e: login_button.config(bg = wc, fg = fc))

        create_button = Button(log_frame, text = " Register", font = fonts, bd = 0, bg = wc, fg = fc, cursor = "hand2", width = 90, image = registericon, compound = LEFT, command = self.showadmin)
        create_button.grid(column = 1, row = 3, pady = 2)
        create_button.bind("<Return>", self.showadmin)
        create_button.bind("<Enter>", lambda e: create_button.config(bg = fc, fg = wc))
        create_button.bind("<Leave>", lambda e: create_button.config(bg = wc, fg = fc))

        change_password = Button(log_frame, text = " Change password", font = fonts, bd = 0, bg = wc, fg = fc, cursor = "hand2", width = 150, image = changeicon, compound = LEFT, command = self.showchange)
        change_password.grid(column = 1, row = 4, pady = 2)
        change_password.bind("<Return>", self.showchange)
        change_password.bind("<Enter>", lambda e: change_password.config(bg = fc, fg = wc))
        change_password.bind("<Leave>", lambda e: change_password.config(bg = wc, fg = fc))
        
        exit_button = Button(log_frame, text = "Exit", font = fonts, bd = 0, bg = wc, fg = fc, cursor = "hand2", width = 90, image = exiticon, compound = LEFT, command = self.exit)
        exit_button.grid(column = 1, row = 5, pady = 2)
        exit_button.bind("<Return>", self.exit)
        exit_button.bind("<Enter>", lambda e: exit_button.config(bg = fc, fg = wc))
        exit_button.bind("<Leave>", lambda e: exit_button.config(bg = wc, fg = fc))

        themepicker = tk.Combobox(log_frame, values = ["Day Light","Dark Knight","Cotton Candy","Ocean Deep","Luscious Grapes"], font = fonts, foreground = "grey")
        themepicker.grid(column = 1, row = 6, pady = 5)
        themepicker.bind("<<ComboboxSelected>>", self.colorchanger)
        themepicker.insert(0, "choose theme")

    def usernamefocus(self, *args):
        if username_entry.get() == "Username":
            username_entry.delete(0, END)
            username_entry.config(fg = codetitlefg)
        elif username_entry.get() == "":
            username_entry.config(fg = "grey")
            username_entry.insert(0, "Username")

    def passwordfocus(self, *args):
        if password_entry.get() == "Password":
            password_entry.delete(0, END)
            password_entry.config(fg = codetitlefg, show = "●")
        elif password_entry.get() == "":
            password_entry.config(fg = "grey", show = "")
            password_entry.insert(0, "Password")

    def colorchanger(self, *args):
        self.theme()
        boylabo.destroy()
        self.reg_exit()

    def theme(self):
        global mainbg, wc, fc, click, totalbg, totalfg, buttonbg, codetitlebg, codetitlefg, dbpscbg, dssibg
        if themepicker.get() == "Day Light":
            root.config(background = "#FEFADE")
            mainbg = "#FEFADE"
            wc = "#E8CCAE"
            fc = "black"
            click = "#F6CE4B"
            buttonbg = "white"
            totalbg = "white"
            totalfg = "black"
            codetitlebg = "white"
            codetitlefg = "black"
            dbpscbg = "#aec6cf"
            dssibg = "#ff6961"
        elif themepicker.get() == "Cotton Candy":
            root.config(background = "#FEC8D8")
            mainbg = "#FEC8D8"
            wc = "#FFDFD3"
            fc = "black"
            click = "#E0BBE4"
            buttonbg = "white"
            totalbg = "white"
            totalfg = "black"
            codetitlebg = "white"
            codetitlefg = "black"
            dbpscbg = "#82A0C2"
            dssibg = "#F7A992"
        elif themepicker.get() == "Ocean Deep":
            root.config(background = "#0D5C75")
            mainbg = "#0D5C75"
            wc = "#A5D1E1"
            fc = "black"
            click = "#FBBF7C"
            buttonbg = "white"
            totalbg = "white"
            totalfg = "black"
            codetitlebg = "white"
            codetitlefg = "black"
            dbpscbg = "#95C5C8"
            dssibg = "#F0C0AA"
        elif themepicker.get() == "Luscious Grapes":
            root.config(background = "#B45AC7")
            mainbg = "#B45AC7"
            wc = "#6F377A"
            fc = "white"
            click = "#F0BCFB"
            buttonbg = "white"
            totalbg = "white"
            totalfg = "black"
            codetitlebg = "white"
            codetitlefg = "black"
            dbpscbg = "#95C5C8"
            dssibg = "#F0C0AA"
        else:
            root.config(background = "gray20")
            mainbg = "gray20"
            wc = "gray15"
            fc = "white"
            click = "lawn green"
            buttonbg = "white"
            totalbg = "white"
            totalfg = "black"
            codetitlebg = "white"
            codetitlefg = "black"
            dbpscbg = "royal blue"
            dssibg = "maroon"

    def showchange(self, *args):
        log_frame.destroy()
        boylabo.destroy()

        change_frame = LabelFrame(login_frame, text = "Change Password", font = fonts, bg = wc, fg = fc)
        change_frame.grid(column = 0, row = 0)

        username = Label(change_frame, text = "Username", font = fonts, bg = wc, fg = fc)
        username.grid(column = 0, row = 0, padx = 10)

        password = Label(change_frame, text = "Current password", font = fonts, bg = wc, fg = fc)
        password.grid(column = 0, row = 1, padx = 10)

        password1 = Label(change_frame, text = "New password", font = fonts, bg = wc, fg = fc)
        password1.grid(column = 0, row = 2, padx = 10)

        password2 = Label(change_frame, text = "Confirm password", font = fonts, bg = wc, fg = fc)
        password2.grid(column = 0, row = 3, padx = 10)

        global username_entry, password_entry, password_entry1, password_entry2
        username_entry = Entry(change_frame, textvariable = self.username, font = fonts, bd = 3, relief = SUNKEN)
        username_entry.grid(column = 1, row = 0)
        username_entry.delete(0, END)

        password_entry = Entry(change_frame, bd = 3, font = fonts, relief = SUNKEN, show = "●")
        password_entry.grid(column = 1, row = 1)
        password_entry.delete(0, END)
 
        password_entry1 = Entry(change_frame, bd = 3, font = fonts, relief = SUNKEN, show = "●")
        password_entry1.grid(column = 1, row = 2)
        password_entry1.delete(0, END)

        password_entry2 = Entry(change_frame, bd = 3, font = fonts, relief = SUNKEN, show = "●")
        password_entry2.grid(column = 1, row = 3)
        password_entry2.delete(0, END)

        create_button = Button(change_frame, text = "Confirm", font = fonts, bg = buttonbg, width = 10, command = self.changepassword)
        create_button.grid(column = 1, row = 4, pady = 2)
        create_button.bind("<Return>", self.changepassword)

        exit_button = Button(change_frame, text = "Exit", font = fonts, bg = buttonbg, width = 10, command = self.reg_exit)
        exit_button.grid(column = 1, row = 5, pady = 2)
        exit_button.bind("<Return>", self.reg_exit)

    def changepassword(self, *args):
        changefinder = "SELECT * FROM users WHERE username = ? AND password = ?"
        c.execute(changefinder, [username_entry.get(), password_entry.get()])
        resultchange = c.fetchall()
        if resultchange:
            if password_entry1.get() == password_entry2.get():
                confirm = messagebox.askyesno("Change password", "Are you sure?")
                if confirm == True:
                    passchanger = "UPDATE users SET password = ? WHERE username = ? AND password = ?"
                    c.execute(passchanger, [password_entry1.get(),username_entry.get(), password_entry.get()])
                    conn.commit()
                    messagebox.showinfo("Success!", "Your password has been changed!")
                    username_entry.delete(0, END)
                    password_entry.delete(0, END)
                    password_entry1.delete(0, END)
                    password_entry2.delete(0, END)
                else:
                    pass
            else:
                messagebox.showerror("Error!", "Password entries not matched!")
        else:
            messagebox.showerror("Error!", "Invalid username/password!")
        
    def showregister(self, *args):
        admin.destroy()
        
        reg_frame = LabelFrame(login_frame, text = "Create Account", font = fonts, bg = wc, fg = fc)
        reg_frame.grid(column = 0, row = 0)
        
        username = Label(reg_frame, text = "Username", font = fonts, bg = wc, fg = fc)
        username.grid(column = 0, row = 0, padx = 10)

        password1 = Label(reg_frame, text = "Password", font = fonts, bg = wc, fg = fc)
        password1.grid(column = 0, row = 1, padx = 10)

        password2 = Label(reg_frame, text = "Confirm password", font = fonts, bg = wc, fg = fc)
        password2.grid(column = 0, row = 2, padx = 10)

        Type = Label(reg_frame, text = "Type", font = fonts, bg = wc, fg = fc)
        Type.grid(column = 0, row = 3, padx = 10)

        global username_entry, password_entry1, password_entry2
        username_entry = Entry(reg_frame, textvariable = self.username, font = fonts, bd = 3, relief = SUNKEN)
        username_entry.grid(column = 1, row = 0)
        username_entry.bind("<Return>", self.login)
        username_entry.focus()
        username_entry.delete(0, END)
        
        password_entry1 = Entry(reg_frame, bd = 3, font = fonts, relief = SUNKEN, show = "●")
        password_entry1.grid(column = 1, row = 1)
        password_entry1.bind("<Return>", self.login)
        password_entry1.delete(0, END)

        password_entry2 = Entry(reg_frame, bd = 3, font = fonts, relief = SUNKEN, show = "●")
        password_entry2.grid(column = 1, row = 2)
        password_entry2.bind("<Return>", self.login)
        password_entry2.delete(0, END)
    
        typelist = ["super", "regular", "viewer"]
        global usertype_entry
        usertype_entry = tk.Combobox(reg_frame, values = typelist, font = fonts, width = 10)
        usertype_entry.grid(column = 1, row = 3)

        create_button = Button(reg_frame, text = "Register", font = fonts, bg = buttonbg, width = 10, command = self.new_user)
        create_button.grid(column = 1, row = 4, pady = 2)
        create_button.bind("<Return>", self.new_user)

        global exit_button
        exit_button = Button(reg_frame, text = "Exit", font = fonts, bg = buttonbg, width = 10, command = self.reg_exit)
        exit_button.grid(column = 1, row = 5, pady = 2)
        exit_button.bind("<Return>", self.reg_exit)

    def showadmin(self, *args):
        log_frame.destroy()
        boylabo.destroy()

        global admin
        admin = Entry(login_frame, bd = 3, relief = SUNKEN, font = fonts, show = "●", fg = "red")
        admin.grid(column = 0, row = 0)
        admin.bind("<Return>", self.admin)
        admin.focus()

    def admin(self, *args):
        if admin.get() == "boylabo":
            self.showregister()
        else:
            messagebox.showwarning("Warning!", "Ooops! You are not allowed here.")
            self.reg_exit()
            
    def reg_exit(self, *args):
        login_frame.destroy()
        self.widgets()
        username_entry.delete(0, END)
        username_entry.insert(0, "Username")
        password_entry.delete(0, END)
        password_entry.insert(0, "Password")

    def company(self):
        global logout_button
        choose = Label(login_frame, text = "Hi, " + self.username.get() + "!" + "\n Choose Company", font = fonts, bg = wc, fg = fc)
        choose.grid(column = 0, row = 0, padx = 10)

        dbpsc_button = Button(login_frame, bd = 0, bg = dbpscbg, image = dbpscicon, cursor = "hand2", command = self.dbpscswitch)
        dbpsc_button.grid(column = 0, row = 1, pady = 5)
        dbpsc_button.bind("<Return>", self.dbpscswitch)
        dbpsc_button.bind("<Enter>", lambda e: dbpsc_button.config(bg = codetitlebg))
        dbpsc_button.bind("<Leave>", lambda e: dbpsc_button.config(bg = dbpscbg))

        dssi_button = Button(login_frame, bd = 0, bg = dssibg, image = dssiicon, cursor = "hand2", command = self.dssiswitch)
        dssi_button.grid(column = 0, row = 2, pady = 5)
        dssi_button.bind("<Return>", self.dssiswitch)
        dssi_button.bind("<Enter>", lambda e: dssi_button.config(bg = codetitlebg))
        dssi_button.bind("<Leave>", lambda e: dssi_button.config(bg = dssibg))

        logout_button = Button(login_frame, font = iconfont, bg = wc, fg = fc, bd = 0, image = logouticon, compound = LEFT, cursor = "hand2", command = self.logout)
        logout_button.grid(column = 0, row = 3, pady = 5)
        logout_button.bind("<Enter>", lambda e: logout_button.config(text = "Logout"))
        logout_button.bind("<Leave>", lambda e: logout_button.config(text = ""))

    def dbpscswitch(self, *args):
        global company
        company = "DBPSC"
        self.menu()
        if access[1] == "FMD":
            accounting_button.config(state = DISABLED)
            bcd_button.config(state = DISABLED)
            gsad_button.config(state = DISABLED)
            fmd_button.focus()
        else:
            pass

    def dssiswitch(self, *args):
        global company
        company = "DSSI"
        self.menu()
        if access[1] == "FMD":
            accounting_button.config(state = DISABLED)
            bcd_button.config(state = DISABLED)
            gsad_button.config(state = DISABLED)
            fmd_button.focus()
        else:
            pass
    
    def menu(self, *args):
        global greeting,menu_frame, accounting_button, fmd_button, bcd_button, gsad_button, logout_button, exit_button, book_label, avatar_button
        greeting = LabelFrame(self.master, text = "Hi, " + self.username.get() + "!" + "\n Welcome to CHARLIE System!", font = fonts, bg = wc, fg = fc)
        greeting.grid(column = 0, row = 0, ipadx = 20, sticky = NW)
        login_frame.destroy()

        menu_frame = LabelFrame(greeting, text = "Main Menu", font = fonts, bg = wc, fg = fc)
        menu_frame.grid(column = 0, row = 0)

        button_frame = Frame(menu_frame, bg = wc)
        button_frame.grid(column = 0, row = 5, pady = 10, sticky = N)

        accounting_button = Button(menu_frame, text = " Accounting             ", font = fonts, bd = 0, bg = wc, fg = fc, width = 240, image = accountingicon, compound = LEFT, cursor = "hand2", command = self.general_journal)
        accounting_button.grid(column = 0, row = 0, pady = 5)
        accounting_button.bind("<Return>", self.general_journal)
        accounting_button.bind("<Enter>", lambda e: accounting_button.config(bg = fc, fg = wc))
        accounting_button.bind("<Leave>", lambda e: accounting_button.config(bg = wc, fg = fc))

        fmd_button = Button(menu_frame, text = " Finance                  ", font = fonts, bd = 0, bg = wc, fg = fc, width = 240, image = financeicon, compound = LEFT, cursor = "hand2", command = self.showfmd)
        fmd_button.grid(column = 0, row = 1, pady = 5)
        fmd_button.bind("<Return>", self.showfmd)
        fmd_button.bind("<Enter>", lambda e: fmd_button.config(bg = fc, fg = wc))
        fmd_button.bind("<Leave>", lambda e: fmd_button.config(bg = wc, fg = fc))

        bcd_button = Button(menu_frame, text = " Billing & Collection", font = fonts, bd = 0, bg = wc, fg = fc, width = 240, image = bcdicon, compound = LEFT, cursor = "hand2", command = self.showbcd)
        bcd_button.grid(column = 0, row = 2, pady = 5)
        bcd_button.bind("<Return>", self.showbcd)
        bcd_button.bind("<Enter>", lambda e: bcd_button.config(bg = fc, fg = wc))
        bcd_button.bind("<Leave>", lambda e: bcd_button.config(bg = wc, fg = fc))

        gsad_button = Button(menu_frame, text = " GSAD                     ", font = fonts, bd = 0, bg = wc, fg = fc, width = 240, image = gsadicon, compound = LEFT, cursor = "hand2", command = self.showgsad)
        gsad_button.grid(column = 0, row = 3, pady = 5)
        gsad_button.bind("<Return>", self.showgsad)
        gsad_button.bind("<Enter>", lambda e: gsad_button.config(bg = fc, fg = wc))
        gsad_button.bind("<Leave>", lambda e: gsad_button.config(bg = wc, fg = fc))

        choosecompany_button= Button(menu_frame, text = " Switch Company   ", font = fonts, bd = 0, bg = wc, fg = fc, width = 240, image = companyicon, compound = LEFT, cursor = "hand2", command = self.choosecompany)
        choosecompany_button.grid(column = 0, row = 4, pady = 5)
        choosecompany_button.bind("<Return>", self.choosecompany)
        choosecompany_button.bind("<Enter>", lambda e: choosecompany_button.config(bg = fc, fg = wc))
        choosecompany_button.bind("<Leave>", lambda e: choosecompany_button.config(bg = wc, fg = fc))
        
        logout_button = Button(button_frame, font = iconfont, bg = wc, fg = fc, bd = 0, image = logouticon, compound = LEFT, cursor = "hand2", command = self.logout)
        logout_button.grid(column = 0, row = 0, pady = 5, padx = 10)
        logout_button.bind("<Return>", self.logout)
        logout_button.bind("<Enter>", lambda e: logout_button.config(text = "Logout"))
        logout_button.bind("<Leave>", lambda e: logout_button.config(text = ""))

        exit_button = Button(button_frame, font = iconfont, bg = wc, fg = fc, bd = 0, image = exiticon, compound = LEFT, cursor = "hand2", command = self.exit)
        exit_button.grid(column = 1, row = 0, pady = 5, padx = 10)
        exit_button.bind("<Return>", self.exit)
        exit_button.bind("<Enter>", lambda e: exit_button.config(text = "Exit"))
        exit_button.bind("<Leave>", lambda e: exit_button.config(text = ""))

        avatar_button = Button(self.master, font = cancelfont, bg = mainbg, fg = fc, bd = 0, image = avatar, compound = "left", cursor = "hand2", command = self.myprofile)
        avatar_button.place(x = 5, y = 615)
        avatar_button.bind("<Enter>", lambda e: avatar_button.config(text = " My Profile"))
        avatar_button.bind("<Leave>", lambda e: avatar_button.config(text = ""))

        book_label = Label(self.master, text = company + "   " + str(today.strftime('%m-%d-%Y')), font = fonts, anchor = W, width = 30, bg = wc, fg = fc)
        book_label.place(x = 5, y = 675)

        self.chatnotifier()

    def choosecompany(self, *args):
        greeting.destroy()
        try:
            boylabo.destroy()
        except:
            pass
        global login_frame
        login_frame = LabelFrame(self.master, text = "CHARLIE System", font = fonts, bg = wc, padx = 10, fg = fc)
        login_frame.place(x = 325, y = 250)
        self.company()
        
    def logout(self, *args):
        updater = "UPDATE users SET status = 'Offline' WHERE username = ?"
        c.execute(updater, [self.username.get()])
        conn.commit()
        for widgets in self.master.winfo_children():
            widgets.destroy()
        self.widgets()
        username_entry.delete(0, END)
        username_entry.insert(0, "Username")
        password_entry.delete(0, END)
        password_entry.insert(0, "Password")
        
    def back(self, *args):
        for widgets in self.master.winfo_children():
            widgets.destroy()
        self.menu()

    def general_journal(self, *args):
        menu_frame.destroy()
        global gj_frame
        gj_frame = LabelFrame(greeting, text = "Accounting", font = fonts, bg = wc, fg = fc)
        gj_frame.grid(column = 0, row = 1)

        button_frame = Frame(gj_frame, bg = wc)
        button_frame.grid(column = 0, row = 11, pady = 10, sticky = N)

        global viewrecord_button
        viewrecord_button = Button(gj_frame, text = " View/Update Record", font = fonts, bd = 0, bg = wc, fg = fc, cursor = "hand2", width = 240, image = updateicon, compound = LEFT, disabledforeground = fc, command = self.view_journal)
        viewrecord_button.grid(column = 0, row = 0, pady = 5)
        viewrecord_button.bind("<Return>", self.view_journal)
        viewrecord_button.bind("<Enter>", lambda e: viewrecord_button.config(bg = fc, fg = wc))
        viewrecord_button.bind("<Leave>", lambda e: viewrecord_button.config(bg = wc, fg = fc))

        global addrecord_buttond
        addrecord_buttond = Button(gj_frame, text = " Disbursements        ", font = fonts, bd = 0, bg = wc, fg = fc, cursor = "hand2", width = 240, image = plusicon, compound = LEFT, disabledforeground = fc, command = self.dtypeswitch)
        addrecord_buttond.grid(column = 0, row = 1, pady = 5)
        addrecord_buttond.bind("<Return>", self.dtypeswitch)
        addrecord_buttond.bind("<Enter>", lambda e: addrecord_buttond.config(bg = fc, fg = wc))
        addrecord_buttond.bind("<Leave>", lambda e: addrecord_buttond.config(bg = wc, fg = fc))

        global addrecord_buttonr
        addrecord_buttonr = Button(gj_frame, text = " Receipts                  ", font = fonts, bd = 0, bg = wc, fg = fc, cursor = "hand2", width = 240, image = plusicon, compound = LEFT, disabledforeground = fc, command = self.rtypeswitch)
        addrecord_buttonr.grid(column = 0, row = 2, pady = 5)
        addrecord_buttonr.bind("<Return>", self.rtypeswitch)
        addrecord_buttonr.bind("<Enter>", lambda e: addrecord_buttonr.config(bg = fc, fg = wc))
        addrecord_buttonr.bind("<Leave>", lambda e: addrecord_buttonr.config(bg = wc, fg = fc))

        global addrecord_buttons
        addrecord_buttons = Button(gj_frame, text = " Sales Journal           ", font = fonts, bd = 0, bg = wc, fg = fc, cursor = "hand2", width = 240, image = plusicon, compound = LEFT, disabledforeground = fc, command = self.stypeswitch)
        addrecord_buttons.grid(column = 0, row = 3, pady = 5)
        addrecord_buttons.bind("<Return>", self.stypeswitch)
        addrecord_buttons.bind("<Enter>", lambda e: addrecord_buttons.config(bg = fc, fg = wc))
        addrecord_buttons.bind("<Leave>", lambda e: addrecord_buttons.config(bg = wc, fg = fc))

        global addrecord_buttong
        addrecord_buttong = Button(gj_frame, text = " General Journal        ", font = fonts, bd = 0, bg = wc, fg = fc, cursor = "hand2", width = 240, image = plusicon, compound = LEFT, disabledforeground = fc, command = self.gtypeswitch)
        addrecord_buttong.grid(column = 0, row = 4, pady = 5)
        addrecord_buttong.bind("<Return>", self.gtypeswitch)
        addrecord_buttong.bind("<Enter>", lambda e: addrecord_buttong.config(bg = fc, fg = wc))
        addrecord_buttong.bind("<Leave>", lambda e: addrecord_buttong.config(bg = wc, fg = fc))

        global receivable_button
        receivable_button = Button(gj_frame, text = "Billing Importer", font = fonts, bd = 0, bg = wc, fg = fc, cursor = "hand2", width = 30, disabledforeground = fc, command = self.showbillimporter)
        # receivable_button.grid(column = 0, row = 5, pady = 5)
        receivable_button.bind("<Return>", self.showbillimporter)
        receivable_button.bind("<Enter>", lambda e: receivable_button.config(bg = fc, fg = wc))
        receivable_button.bind("<Leave>", lambda e: receivable_button.config(bg = wc, fg = fc))

        global payroll_button
        payroll_button = Button(gj_frame, text = " Payroll Importer       ", font = fonts, bd = 0, bg = wc, fg = fc, cursor = "hand2", width = 240, image = importericon, compound = LEFT, disabledforeground = fc, command = self.showpayroll)
        payroll_button.grid(column = 0, row = 6, pady = 5)
        payroll_button.bind("<Return>", self.showpayroll)
        payroll_button.bind("<Enter>", lambda e: payroll_button.config(bg = fc, fg = wc))
        payroll_button.bind("<Leave>", lambda e: payroll_button.config(bg = wc, fg = fc))

        global chart_button
        chart_button = Button(gj_frame, text = " Chart of Accounts    ", font = fonts, bd = 0, bg = wc, fg = fc, cursor = "hand2", width = 240, image = charticon, compound = LEFT, disabledforeground = fc, command = self.chart)
        chart_button.grid(column = 0, row = 7, pady = 5)
        chart_button.bind("<Return>", self.chart)
        chart_button.bind("<Enter>", lambda e: chart_button.config(bg = fc, fg = wc))
        chart_button.bind("<Leave>", lambda e: chart_button.config(bg = wc, fg = fc))

        global bir_button
        bir_button = Button(gj_frame, text = " BIR                           ", font = fonts, bd = 0, bg = wc, fg = fc, cursor = "hand2", width = 240, image = biricon, compound = LEFT, disabledforeground = fc, command = self.showbir)
        bir_button.grid(column = 0, row = 8, pady = 5)
        bir_button.bind("<Return>", self.showbir)
        bir_button.bind("<Enter>", lambda e: bir_button.config(bg = fc, fg = wc))
        bir_button.bind("<Leave>", lambda e: bir_button.config(bg = wc, fg = fc))

        global locker_button
        locker_button = Button(gj_frame, text = " Period Locker           ", font = fonts, bd = 0, bg = wc, fg = fc, cursor = "hand2", width = 240, image = lockicon, compound = LEFT, disabledforeground = fc, command = self.showlocker)
        locker_button.grid(column = 0, row = 9, pady = 5)
        locker_button.bind("<Return>", self.showlocker)
        locker_button.bind("<Enter>", lambda e: locker_button.config(bg = fc, fg = wc))
        locker_button.bind("<Leave>", lambda e: locker_button.config(bg = wc, fg = fc))

        global reports_button
        reports_button = Button(gj_frame, text = " Reports                    ", font = fonts, bd = 0, bg = wc, fg = fc, cursor = "hand2", width = 240, image = reporticon, compound = LEFT, disabledforeground = fc, command = self.showreports)
        reports_button.grid(column = 0, row = 10, pady = 5)
        reports_button.bind("<Return>", self.showreports)
        reports_button.bind("<Enter>", lambda e: reports_button.config(bg = fc, fg = wc))
        reports_button.bind("<Leave>", lambda e: reports_button.config(bg = wc, fg = fc))

        global back, logout_button, exit_button
        back = Button(button_frame, font = iconfont, bg = wc, fg = fc, bd = 0, image = homeicon, compound = LEFT, cursor = "hand2", command = self.back)
        back.grid(column = 0, row = 0, pady = 5, padx = 10)
        back.bind("<Return>", self.back)
        back.bind("<Enter>", lambda e: back.config(text = "Home"))
        back.bind("<Leave>", lambda e: back.config(text = ""))

        logout_button = Button(button_frame, font = iconfont, bg = wc, fg = fc, bd = 0, image = logouticon, compound = LEFT, cursor = "hand2", command = self.logout)
        logout_button.grid(column = 1, row = 0, pady = 5, padx = 10)
        logout_button.bind("<Return>", self.logout)
        logout_button.bind("<Enter>", lambda e: logout_button.config(text = "Logout"))
        logout_button.bind("<Leave>", lambda e: logout_button.config(text = ""))

        exit_button = Button(button_frame, font = iconfont, bg = wc, fg = fc, bd = 0, image = exiticon, compound = LEFT, cursor = "hand2", command = self.exit)
        exit_button.grid(column = 2, row = 0, pady = 5, padx = 10)
        exit_button.bind("<Return>", self.exit)
        exit_button.bind("<Enter>", lambda e: exit_button.config(text = "Exit"))
        exit_button.bind("<Leave>", lambda e: exit_button.config(text = ""))

    def menubuttons(self, status, *args):
        viewrecord_button.config(state = status)
        addrecord_buttond.config(state = status)
        addrecord_buttonr.config(state = status)
        addrecord_buttons.config(state = status)
        addrecord_buttong.config(state = status)
        receivable_button.config(state = status)
        payroll_button.config(state = status)
        reports_button.config(state = status)
        chart_button.config(state = status)
        bir_button.config(state = status)
        locker_button.config(state = status)
        back.config(state = status)
        logout_button.config(state = status)
        exit_button.config(state = status)

### MY PROFILE ###
    def myprofile(self):
        global top, avatar_logo
        top = Toplevel()
        top.title(maintitle)
        top.iconbitmap(iconpath + "icon.ico") #path
        top.geometry("400x400+600+200")
        top.config(background = mainbg)
        top.resizable(width = False, height = False)
        top.protocol("WM_DELETE_WINDOW", self.userstatusupdater("Online"))
        top.focus_force()
        self.countonlinenow()
        if len(chatcount) == 0:
            chattext = "Chat"
        else:
            chattext = f"Chat ({len(chatcount)})"

        dashboard_frame = Frame(top, bg = wc)
        dashboard_frame.grid(column = 0, row = 0, pady = pad, sticky = NW)

        avatar_logo = Button(dashboard_frame, bg = wc, bd = 0, image = avatar, cursor = "hand2", command = self.showavatar)
        avatar_logo.grid(column = 0, row = 0, padx = pad)

        to_do = Button(dashboard_frame, text = "To-do List", font = fontreports, bg = buttonbg, width = 13, cursor = "hand2", command = self.showtodo)
        to_do.grid(column = 1, row = 0, padx = pad)

        chat_mate = Button(dashboard_frame, text = chattext, font = fontreports, bg = buttonbg, width = 13, cursor = "hand2", command = self.showchat)
        chat_mate.grid(column = 2, row = 0, padx = pad)

        self.userstatusupdater("Active")
        active_now = Button(dashboard_frame, text = f"Online ({len(active)})", font = fontreports, bg = buttonbg, width = 13, cursor = "hand2", command = self.showactivenow)
        active_now.grid(column = 3, row = 0, padx = pad)
        
        top.mainloop()

    def userstatusupdater(self, status):
        update = "UPDATE users SET status = ? WHERE username = ?"
        c.execute(update, [status ,self.username.get()])
        conn.commit()

    def chatnotifier(self):
        select = "SELECT from_user, message, status FROM chat WHERE to_user = ? AND status = 'unread' GROUP BY from_user"
        c.execute(select, [self.username.get()])
        unread = c.fetchall()
        if unread:
            for i in unread:
                notification.notify(
                    title = f"New message from {i[0]}!",
                    message = f'"{i[1]}"',
                    app_icon = iconpath + "icon.ico",
                    timeout = 10
                )
    
    def readmessages(self, user):
        update = "UPDATE chat SET status = 'read' WHERE to_user = ? and from_user = ? AND status = 'unread'"
        if user != self.username.get():
            c.execute(update, [self.username.get(), user])
        else:
            c.execute(update, [user, self.username.get()])
        conn.commit()
    
    ### CHAT ###
    def showchat(self, *args):
        global profile_frame, messages_label, sub_frame, chat_frame
        try:
            profile_frame.destroy()
        except:
            pass
        profile_frame = Frame(top, bg = wc)
        profile_frame.grid(column = 0, row = 1, pady = pad, sticky = NW)

        sub_frame = Frame(profile_frame, bg = wc)
        sub_frame.grid(column = 0, row = 0, sticky = NW, pady = pad)

        chat_frame = Frame(profile_frame, bg = wc)
        chat_frame.grid(column = 0, row = 1, sticky = NW)

        messages_label = Label(sub_frame, text = "Messages", font = fontreports, relief = RIDGE, bg = wc, fg = fc, width = 56)
        messages_label.grid(column = 0, row = 0, pady = pad)

        self.showmessages(chat_frame)
    
    def showmessages(self, frame, *args):
        global user_frame
        select = "SELECT from_user, message, date, to_user, MAX(number), team, status FROM chat WHERE to_user = ? or from_user = ? GROUP BY team ORDER BY number DESC"
        c.execute(select, [self.username.get(), self.username.get()])
        result = c.fetchall()
        if result:
            self.imageresizer()
            self.itemscroller(frame, 1)
            canvas.config(width = 373, height = 290)
            scrollbar.config(width = 20)
            row = 0
            for i in result:
                user_frame = Frame(scrollable_frame, bg = wc)
                user_frame.grid(column = 0, row = row)

                getavatar = "SELECT avatar FROM users WHERE username = ?"
                if i[6] != "read" and i[0] != self.username.get():
                    chatfont = unreadfont
                else:
                    chatfont = fontreports
                if i[3] != self.username.get():
                    c.execute(getavatar, [i[3]])
                    get = c.fetchall()

                    user_avatar = Button(user_frame, text = f"  {i[3]}: {i[1]}", font = chatfont, width = 367, bg = wc, fg = fc, image = eval(get[0][0]), anchor = W, cursor = "hand2", compound = LEFT, command = lambda i=i[3]: self.showuserchat(i))
                    user_avatar.grid(column = 0, row = 0)
                else:
                    c.execute(getavatar, [i[0]])
                    get = c.fetchall()

                    user_avatar = Button(user_frame, text = f"  {i[0]}: {i[1]}", font = chatfont, width = 367, bg = wc, fg = fc, image = eval(get[0][0]), anchor = W, cursor = "hand2", compound = LEFT, command = lambda i=i[0]: self.showuserchat(i))
                    user_avatar.grid(column = 0, row = 0)
                row += 1
        else:
            pass

    def showuserchat(self, user, *args):
        global geticon, getmine, message_entry
        getavatar = "SELECT avatar, gender FROM users WHERE username = ?"
        c.execute(getavatar, [user])
        get = c.fetchall()
        geticon = ImageTk.PhotoImage(Image.open(f"{iconpath}avatars/{get[0][1]}/{get[0][0]}.png").resize((20,20), Image.ANTIALIAS))
        getmine = ImageTk.PhotoImage(Image.open(f"{iconpath}avatars/{access[2]}/{access[3]}.png").resize((20,20), Image.ANTIALIAS))

        messages_label.config(text = f" {user}", width = 355, image = eval(get[0][0]), compound = LEFT)

        home_button = Button(sub_frame, bg = wc, fg = fc, cursor = "hand2", image = previousicon, command = self.showchat)
        home_button.grid(column = 1, row = 0)

        scrollbox.destroy()
        self.itemscroller(chat_frame, 0)
        canvas.config(width = 373, height = 245)
        scrollbar.config(width = 20)
        self.fillmessages(user, geticon)
        self.readmessages(user)
        
        message_frame = Frame(profile_frame, bg = wc)
        message_frame.grid(column = 0, row = 2, sticky = W)

        message_entry = Entry(message_frame, font = ("calibri", 14), width = 36, insertbackground = fc, bg = wc, fg = "grey")
        message_entry.grid(column = 0, row = 0)
        message_entry.bind("<Return>", lambda e: self.sendmessage(user))
        message_entry.bind("<FocusIn>", lambda e: self.chatfocus(message_entry))
        message_entry.bind("<FocusOut>", lambda e: self.chatfocus(message_entry))
        message_entry.insert(0, "Aa")

        send_button = Button(message_frame, bg = wc, fg = fc, bd = 0, image = sendicon, cursor = "hand2", command = lambda: self.sendmessage(user))
        send_button.grid(column = 1, row = 0)
        send_button.bind("<Return>", lambda e: self.sendmessage(user))

    def sendmessage(self, user, *args):
        sortedteam = sorted((self.username.get(), user))
        team = f"{sortedteam[0]}{sortedteam[1]}"
        insert = "INSERT INTO chat (from_user, to_user, message, date, team, status) values (?,?,?,?,?,?)"
        c.execute(insert, [self.username.get(), user, message_entry.get(), today.strftime('%m-%d-%Y'), team, "unread"])
        conn.commit()
        self.updatemessages(user, geticon)
        message_entry.delete(0, END)

    def fillmessages(self, user, geticon):
        select = "SELECT message, date, from_user, number FROM chat WHERE to_user = ? AND from_user = ? UNION SELECT message, date, from_user, number FROM chat WHERE to_user = ? AND from_user = ? ORDER BY number"
        c.execute(select, [self.username.get(), user, user, self.username.get()])
        result = c.fetchall()
        for i in result:
            if i[2] != user:
                Label(scrollable_frame, text = f"{i[1]}              ", font = chatdatefont, bg = wc, fg = fc, width = 74, anchor = E).pack()
                Label(scrollable_frame, text = f"{i[0]}  ", font = fontreports, bg = wc, fg = fc, width = 370, image = getmine, compound = RIGHT, wraplength = 360, anchor = E, justify = RIGHT).pack()
            else:
                Label(scrollable_frame, text = f"              {i[1]}", font = chatdatefont, bg = wc, fg = fc, width = 74, anchor = W).pack()
                Label(scrollable_frame, text = f"  {i[0]}", font = fontreports, bg = wc, fg = fc, width = 370, image = geticon, compound = LEFT, wraplength = 360, anchor = W, justify = LEFT).pack()

    def updatemessages(self, user, geticon):
        select = "SELECT message, date, from_user, number FROM chat WHERE to_user = ? AND from_user = ? UNION SELECT message, date, from_user, number FROM chat WHERE to_user = ? AND from_user = ? ORDER BY number"
        c.execute(select, [self.username.get(), user, user, self.username.get()])
        result = c.fetchall()
        if result[-1][2] != user:
            Label(scrollable_frame, text = f"{result[-1][1]}              ", font = chatdatefont, bg = wc, fg = fc, width = 74, anchor = E).pack()
            Label(scrollable_frame, text = f"{result[-1][0]}  ", font = fontreports, bg = wc, fg = fc, width = 370, image = getmine, compound = RIGHT, wraplength = 360, anchor = E, justify = RIGHT).pack()
        else:
            Label(scrollable_frame, text = f"              {result[-1][1]}", font = chatdatefont, bg = wc, fg = fc, width = 74, anchor = W).pack()
            Label(scrollable_frame, text = f"  {result[-1][0]}", font = fontreports, bg = wc, fg = fc, width = 370, image = geticon, compound = LEFT, wraplength = 360, anchor = W, justify = LEFT).pack()
        # size = len(scrollable_frame.winfo_children())
        # canvas.yview_scroll(size, "pages")

    ### TO DO ###
    def showtodo(self, *args):
        global profile_frame, todo_entry, todo_list
        try:
            profile_frame.destroy()
        except:
            pass
        profile_frame = Frame(top, bg = wc)
        profile_frame.grid(column = 0, row = 1, pady = pad, sticky = NW)

        sub_frame = Frame(profile_frame, bg = wc)
        sub_frame.grid(column = 0, row = 0, sticky = NW)

        button_frame = Frame(profile_frame, bg = wc)
        button_frame.grid(column = 0, row = 1, sticky = NW)

        todo_list = tk.Treeview(sub_frame)
        todo_list["columns"] = ("Date","To-Do")
        todo_list.column("#0", width=0, minwidth=0, stretch = True)
        todo_list.column("Date", anchor = W, width = 75)
        todo_list.column("To-Do", anchor = W, width = 320)
        todo_list.heading("#0", text = "", anchor = W)
        todo_list.heading("Date", text = "Date", anchor = N)
        todo_list.heading("To-Do", text = "To-Do", anchor = N)

        select = "SELECT date, todo FROM todo WHERE user = ?"
        c.execute(select, [self.username.get()])
        result = c.fetchall()
        if result:
            for i in result:
                todo_list.insert("", "end", values = i)
        else:
            todo_list.insert("", "end", values = ("otoke", "otoke"))
        todo_list.pack()
        todo_list.bind("<ButtonRelease>", self.selecteditem)

        todo_entry = Entry(button_frame, font = fonts, fg = "grey", width = 49)
        todo_entry.grid(column = 0, row = 0, sticky = N)
        todo_entry.insert(0, "Aa")
        todo_entry.bind("<FocusIn>", lambda e: self.todofocus(todo_entry))
        todo_entry.bind("<FocusOut>", lambda e: self.todofocus(todo_entry))

        add_button = Button(button_frame, text = "Add", font = fonts, bg = wc, bd = 0, width = 95, image = buttonicon, compound = CENTER, cursor = "hand2", command = self.addtodo)
        add_button.grid(column = 0, row = 1, sticky = N)

        delete_button = Button(button_frame, text = "Delete", font = fonts, bg = wc, bd = 0, width = 95, image = buttonicon, compound = CENTER, cursor = "hand2", command = self.deletetodo)
        delete_button.grid(column = 0, row = 2, sticky = N)

    def addtodo(self, *args):
        add = "INSERT INTO todo (date, todo, user) values (?, ?, ?)"
        c.execute(add, [today.strftime("%m-%d-%Y"), todo_entry.get(), self.username.get()])
        try:
            conn.commit()
            profile_frame.destroy()
            self.showtodo()
        except:
            pass

    def deletetodo(self, *args):
        delete = "DELETE FROM todo WHERE user = ? AND todo = ?"
        try:
            c.execute(delete, [self.username.get(), currentitem])
            conn.commit()
            profile_frame.destroy()
            self.showtodo()
        except:
            pass

    def selecteditem(self, *args):
        global currentitem
        try:
            currentitem = todo_list.item(todo_list.focus(), "values")[1]
        except:
            pass

    def todofocus(self, entry, *args):
        if entry.get() == "Aa":
            entry.delete(0, END)
            entry.config(fg = codetitlefg)
        elif entry.get() == "":
            entry.config(fg = "grey")
            entry.insert(0, "Aa")

    def chatfocus(self, entry, *args):
        if entry.get() == "Aa":
            entry.delete(0, END)
            entry.config(fg = fc)
        elif entry.get() == "":
            entry.config(fg = "grey")
            entry.insert(0, "Aa")

    ### AVATAR ###
    def showavatar(self, *args):
        global idx, profile_frame
        try:
            profile_frame.destroy()
        except:
            pass
        profile_frame = Frame(top, bg = wc)
        profile_frame.grid(column = 0, row = 1, pady = pad, sticky = NW)

        self.avatarpicker()
        row = 0
        col = 0
        idx = []
        for i in avatarlist:
            avatarvar = IntVar()
            avatarlogo = Checkbutton(profile_frame, variable = avatarvar, bg = wc, image = i, cursor = "hand2", command = self.avatarchecker)
            avatarlogo.grid(column = col, row = row, padx = pad, pady = pad)
            idx.append(avatarvar)
            col += 1
            if access[2] == "male":
                if col == 3:
                    col = 0
                    row += 1
            else:
                if col == 4:
                    col = 0
                    row += 1

    def avatarchecker(self, *args):
        global chosenlogo
        try:
            chosenlogo.set(0)
        except:
            pass
        for i in idx:
            if i.get() == 1:
                chosenlogo = i
            i.set(0)
        try:
            chosenlogo.set(1)
        except:
            pass
        self.changeavatar()

    def changeavatar(self, *args):
        global avatar
        chosen = access[3][0] + str(idx.index(chosenlogo)+1)
        update = "UPDATE users SET avatar = ? WHERE username = ?"
        c.execute(update, [chosen, self.username.get()])
        conn.commit()
        top.destroy()
        avatar = ImageTk.PhotoImage(Image.open(iconpath + "avatars/" + access[2] + "/" + chosen + ".png")) #path
        avatar_button.config(image = avatar)
        self.myprofile()

    ### ACTIVE NOW ###
    def showactivenow(self):
        global profile_frame
        try:
            profile_frame.destroy()
        except:
            pass
        profile_frame = Frame(top, bg = wc)
        profile_frame.grid(column = 0, row = 1, pady = pad, sticky = NW)

        label_frame = Frame(profile_frame, bg = wc)
        label_frame.grid(column = 0, row = 0, sticky = NW)

        user_frame = Frame(profile_frame, bg = wc)
        user_frame.grid(column = 0, row = 1)

        name_label = Label(label_frame, text = "User", font = fontreports, bg = wc, fg = fc, width = 20, relief = RIDGE)
        name_label.grid(column = 0, row = 0)

        department_label = Label(label_frame, text = "Department", font = fontreports, bg = wc, fg = fc, width = 15, relief = RIDGE, anchor = N)
        department_label.grid(column = 1, row = 0)

        status_label = Label(label_frame, text = "Status", font = fontreports, bg = wc, fg = fc, width = 15, relief = RIDGE, anchor = N)
        status_label.grid(column = 2, row = 0)

        self.itemscroller(user_frame, 0)
        canvas.config(width = 368, height = 300)
        self.imageresizer()
        self.showalluserstatus(scrollable_frame)

    def showalluserstatus(self, frame):
        row = 0
        for i in active:
            user_frame = Frame(frame, bg = wc)
            user_frame.grid(column = 0, row = row)

            user_avatar = Button(user_frame, text = f"  {i[1]}", font = fontreports, width = 136, bg = wc, fg = fc, bd = 0, image = eval(i[0]), compound = LEFT, cursor = "hand2", anchor = W, command = lambda i=i[1]: self.tunneltochat(i))
            user_avatar.grid(column = 0, row = 0)

            user_department = Label(user_frame, text = i[2], font = fontreports, width = 15, bg = wc, fg = fc, anchor = N)
            user_department.grid(column = 1, row = 0)

            user_status = Label(user_frame, text = i[3], font = fontreports, width = 15, bg = fc, fg = wc, anchor = N)
            user_status.grid(column = 2, row = 0)
            row += 1

        for i in inactive:
            user_frame = Frame(frame, bg = wc)
            user_frame.grid(column = 0, row = row)

            user_avatar = Button(user_frame, text = f"  {i[1]}", font = fontreports, width = 136, bg = wc, fg = fc, bd = 0, image = eval(i[0]), compound = LEFT, cursor = "hand2", anchor = W, command = lambda i=i[1]: self.tunneltochat(i))
            user_avatar.grid(column = 0, row = 0)

            user_department = Label(user_frame, text = i[2], font = fontreports, width = 15, bg = wc, fg = fc, anchor = N)
            user_department.grid(column = 1, row = 0)

            user_status = Label(user_frame, text = i[3], font = fontreports, width = 15, bg = wc, fg = fc, anchor = N)
            user_status.grid(column = 2, row = 0)
            row += 1

    def tunneltochat(self, user, *args):
        profile_frame.destroy()
        self.showchat()
        self.showuserchat(user)

    def imageresizer(self):
        global B1,B2,B3,B4,B5,B6,B7,B8,B9,G1,G2,G3,G4,G5,G6,G7,G8,G9,G10,G11,G12,G13
        B1 = ImageTk.PhotoImage(Image.open(iconpath + "avatars/male/B1.png").resize((32,32), Image.ANTIALIAS))
        B2 = ImageTk.PhotoImage(Image.open(iconpath + "avatars/male/B2.png").resize((32,32), Image.ANTIALIAS))
        B3 = ImageTk.PhotoImage(Image.open(iconpath + "avatars/male/B3.png").resize((32,32), Image.ANTIALIAS))
        B4 = ImageTk.PhotoImage(Image.open(iconpath + "avatars/male/B4.png").resize((32,32), Image.ANTIALIAS))
        B5 = ImageTk.PhotoImage(Image.open(iconpath + "avatars/male/B5.png").resize((32,32), Image.ANTIALIAS))
        B6 = ImageTk.PhotoImage(Image.open(iconpath + "avatars/male/B6.png").resize((32,32), Image.ANTIALIAS))
        B7 = ImageTk.PhotoImage(Image.open(iconpath + "avatars/male/B7.png").resize((32,32), Image.ANTIALIAS))
        B8 = ImageTk.PhotoImage(Image.open(iconpath + "avatars/male/B8.png").resize((32,32), Image.ANTIALIAS))
        B9 = ImageTk.PhotoImage(Image.open(iconpath + "avatars/male/B9.png").resize((32,32), Image.ANTIALIAS))
        G1 = ImageTk.PhotoImage(Image.open(iconpath + "avatars/female/G1.png").resize((32,32), Image.ANTIALIAS))
        G2 = ImageTk.PhotoImage(Image.open(iconpath + "avatars/female/G2.png").resize((32,32), Image.ANTIALIAS))
        G3 = ImageTk.PhotoImage(Image.open(iconpath + "avatars/female/G3.png").resize((32,32), Image.ANTIALIAS))
        G4 = ImageTk.PhotoImage(Image.open(iconpath + "avatars/female/G4.png").resize((32,32), Image.ANTIALIAS))
        G5 = ImageTk.PhotoImage(Image.open(iconpath + "avatars/female/G5.png").resize((32,32), Image.ANTIALIAS))
        G6 = ImageTk.PhotoImage(Image.open(iconpath + "avatars/female/G6.png").resize((32,32), Image.ANTIALIAS))
        G7 = ImageTk.PhotoImage(Image.open(iconpath + "avatars/female/G7.png").resize((32,32), Image.ANTIALIAS))
        G8 = ImageTk.PhotoImage(Image.open(iconpath + "avatars/female/G8.png").resize((32,32), Image.ANTIALIAS))
        G9 = ImageTk.PhotoImage(Image.open(iconpath + "avatars/female/G9.png").resize((32,32), Image.ANTIALIAS))
        G10 = ImageTk.PhotoImage(Image.open(iconpath + "avatars/female/G10.png").resize((32,32), Image.ANTIALIAS))
        G11 = ImageTk.PhotoImage(Image.open(iconpath + "avatars/female/G11.png").resize((32,32), Image.ANTIALIAS))
        G12 = ImageTk.PhotoImage(Image.open(iconpath + "avatars/female/G12.png").resize((32,32), Image.ANTIALIAS))
        G13 = ImageTk.PhotoImage(Image.open(iconpath + "avatars/female/G13.png").resize((32,32), Image.ANTIALIAS))

    def avatarpicker(self):
        global avatarlist
        if access[2] == "male":
            B1 = ImageTk.PhotoImage(Image.open(iconpath + "avatars/male/B1.png"))
            B2 = ImageTk.PhotoImage(Image.open(iconpath + "avatars/male/B2.png"))
            B3 = ImageTk.PhotoImage(Image.open(iconpath + "avatars/male/B3.png"))
            B4 = ImageTk.PhotoImage(Image.open(iconpath + "avatars/male/B4.png"))
            B5 = ImageTk.PhotoImage(Image.open(iconpath + "avatars/male/B5.png"))
            B6 = ImageTk.PhotoImage(Image.open(iconpath + "avatars/male/B6.png"))
            B7 = ImageTk.PhotoImage(Image.open(iconpath + "avatars/male/B7.png"))
            B8 = ImageTk.PhotoImage(Image.open(iconpath + "avatars/male/B8.png"))
            B9 = ImageTk.PhotoImage(Image.open(iconpath + "avatars/male/B9.png"))
            avatarlist = [B1,B2,B3,B4,B5,B6,B7,B8,B9]
        else:
            G1 = ImageTk.PhotoImage(Image.open(iconpath + "avatars/female/G1.png"))
            G2 = ImageTk.PhotoImage(Image.open(iconpath + "avatars/female/G2.png"))
            G3 = ImageTk.PhotoImage(Image.open(iconpath + "avatars/female/G3.png"))
            G4 = ImageTk.PhotoImage(Image.open(iconpath + "avatars/female/G4.png"))
            G5 = ImageTk.PhotoImage(Image.open(iconpath + "avatars/female/G5.png"))
            G6 = ImageTk.PhotoImage(Image.open(iconpath + "avatars/female/G6.png"))
            G7 = ImageTk.PhotoImage(Image.open(iconpath + "avatars/female/G7.png"))
            G8 = ImageTk.PhotoImage(Image.open(iconpath + "avatars/female/G8.png"))
            G9 = ImageTk.PhotoImage(Image.open(iconpath + "avatars/female/G9.png"))
            G10 = ImageTk.PhotoImage(Image.open(iconpath + "avatars/female/G10.png"))
            G11 = ImageTk.PhotoImage(Image.open(iconpath + "avatars/female/G11.png"))
            G12 = ImageTk.PhotoImage(Image.open(iconpath + "avatars/female/G12.png"))
            G13 = ImageTk.PhotoImage(Image.open(iconpath + "avatars/female/G13.png"))
            avatarlist = [G1,G2,G3,G4,G5,G6,G7,G8,G9,G10,G11,G12,G13]

    def countonlinenow(self):
        global active, inactive, chatcount, allusers
        c.execute("SELECT avatar, username, department, status, gender FROM users")
        allusers = c.fetchall()
        active, inactive, chatcount = [], [], []
        for i in allusers:
            if i[3] == "Online" or i[3] == "Active":
                active.append(i)
            else:
                inactive.append(i)
        select = "SELECT status FROM chat WHERE status = 'unread' AND to_user = ?"
        c.execute(select, [self.username.get()])
        result = c.fetchall()
        if result:
            for i in result:
                if i[0] == "unread":
                    chatcount.append(i)

### GENERAL FUNCTIONS ###
    def uppercase(self, var, *args):
        converted = var.get().upper()
        var.delete(0, END)
        var.insert(0, converted)

    def amountvalidatormaster(self, amt, *args):
        amount = amt
        try:
            if amount.get() == "":
                amount.set(format(float(0), ',.2f'))
            else:
                try:
                    comma = format(float(amount.get()), ',.2f')
                    amount.set(comma)
                except:
                    amount.set(format(float(0), ',.2f'))
        except:
            amount.set(format(float(0), ',.2f'))

    def cancel(self, *args):
        if checkbox.get() == 1:
            cancelbox.config(bg = cancelbg, fg = cancelfg, font = cancelfont, text = 'Cancelled', )
        else:
            cancelbox.config(bg = wc, fg = fc, font = fonts, text = 'Cancel')

    def colorswitch(self, widget, back, *args):
        widget.config(bg = back, disabledforeground = wc)
        if widget["state"] == DISABLED:
            self.unbindaccounting()
        else:
            self.bindaccounting()
            widget.config(bg = wc, fg = fc, disabledforeground = fc)

    def colorswitchfmd(self, widget, back, *args):
        widget.config(bg = back, disabledforeground = wc)
        if widget["state"] == DISABLED:
            self.unbindfmd()
        else:
            self.bindfmd()
            widget.config(bg = wc, fg = fc, disabledforeground = fc)

    def unbindaccounting(self):
        viewrecord_button.unbind("<Enter>")
        viewrecord_button.unbind("<Leave>")
        addrecord_buttond.unbind("<Enter>")
        addrecord_buttond.unbind("<Leave>")
        addrecord_buttonr.unbind("<Enter>")
        addrecord_buttonr.unbind("<Leave>")
        addrecord_buttons.unbind("<Enter>")
        addrecord_buttons.unbind("<Leave>")
        addrecord_buttong.unbind("<Enter>")
        addrecord_buttong.unbind("<Leave>")
        receivable_button.unbind("<Enter>")
        receivable_button.unbind("<Leave>")
        payroll_button.unbind("<Enter>")
        payroll_button.unbind("<Leave>")
        reports_button.unbind("<Enter>")
        reports_button.unbind("<Leave>")
        chart_button.unbind("<Enter>")
        chart_button.unbind("<Leave>")
        bir_button.unbind("<Enter>")
        bir_button.unbind("<Leave>")
        locker_button.unbind("<Enter>")
        locker_button.unbind("<Leave>")
        back.unbind("<Enter>")
        back.unbind("<Leave>")
        logout_button.unbind("<Enter>")
        logout_button.unbind("<Leave>")
        exit_button.unbind("<Enter>")
        exit_button.unbind("<Leave>")

    def bindaccounting(self):
        viewrecord_button.bind("<Enter>", lambda e: viewrecord_button.config(bg = fc, fg = wc))
        viewrecord_button.bind("<Leave>", lambda e: viewrecord_button.config(bg = wc, fg = fc))
        addrecord_buttond.bind("<Enter>", lambda e: addrecord_buttond.config(bg = fc, fg = wc))
        addrecord_buttond.bind("<Leave>", lambda e: addrecord_buttond.config(bg = wc, fg = fc))
        addrecord_buttonr.bind("<Enter>", lambda e: addrecord_buttonr.config(bg = fc, fg = wc))
        addrecord_buttonr.bind("<Leave>", lambda e: addrecord_buttonr.config(bg = wc, fg = fc))
        addrecord_buttons.bind("<Enter>", lambda e: addrecord_buttons.config(bg = fc, fg = wc))
        addrecord_buttons.bind("<Leave>", lambda e: addrecord_buttons.config(bg = wc, fg = fc))
        addrecord_buttong.bind("<Enter>", lambda e: addrecord_buttong.config(bg = fc, fg = wc))
        addrecord_buttong.bind("<Leave>", lambda e: addrecord_buttong.config(bg = wc, fg = fc))
        receivable_button.bind("<Enter>", lambda e: receivable_button.config(bg = fc, fg = wc))
        receivable_button.bind("<Leave>", lambda e: receivable_button.config(bg = wc, fg = fc))
        payroll_button.bind("<Enter>", lambda e: payroll_button.config(bg = fc, fg = wc))
        payroll_button.bind("<Leave>", lambda e: payroll_button.config(bg = wc, fg = fc))
        reports_button.bind("<Enter>", lambda e: reports_button.config(bg = fc, fg = wc))
        reports_button.bind("<Leave>", lambda e: reports_button.config(bg = wc, fg = fc))
        chart_button.bind("<Enter>", lambda e: chart_button.config(bg = fc, fg = wc))
        chart_button.bind("<Leave>", lambda e: chart_button.config(bg = wc, fg = fc))
        bir_button.bind("<Enter>", lambda e: bir_button.config(bg = fc, fg = wc))
        bir_button.bind("<Leave>", lambda e: bir_button.config(bg = wc, fg = fc))
        locker_button.bind("<Enter>", lambda e: locker_button.config(bg = fc, fg = wc))
        locker_button.bind("<Leave>", lambda e: locker_button.config(bg = wc, fg = fc))
        back.bind("<Enter>", lambda e: back.config(text = "Home"))
        back.bind("<Leave>", lambda e: back.config(text = ""))
        logout_button.bind("<Enter>", lambda e: logout_button.config(text = "Logout"))
        logout_button.bind("<Leave>", lambda e: logout_button.config(text = ""))
        exit_button.bind("<Enter>", lambda e: exit_button.config(text = "Exit"))
        exit_button.bind("<Leave>", lambda e: exit_button.config(text = ""))

    def unbindfmd(self):
        disbursement_button.unbind("<Enter>")
        disbursement_button.unbind("<Leave>")
        viewer_button.unbind("<Enter>")
        viewer_button.unbind("<Leave>")
        finder_button.unbind("<Enter>")
        finder_button.unbind("<Leave>")
        sms_button.unbind("<Enter>")
        sms_button.unbind("<Leave>")
        reports_button.unbind("<Enter>")
        reports_button.unbind("<Leave>")
        back.unbind("<Enter>")
        back.unbind("<Leave>")
        logout_button.unbind("<Enter>")
        logout_button.unbind("<Leave>")
        exit_button.unbind("<Enter>")
        exit_button.unbind("<Leave>")

    def bindfmd(self):
        disbursement_button.bind("<Enter>", lambda e: disbursement_button.config(bg = fc, fg = wc))
        disbursement_button.bind("<Leave>", lambda e: disbursement_button.config(bg = wc, fg = fc))
        viewer_button.bind("<Enter>", lambda e: viewer_button.config(bg = fc, fg = wc))
        viewer_button.bind("<Leave>", lambda e: viewer_button.config(bg = wc, fg = fc))
        finder_button.bind("<Enter>", lambda e: finder_button.config(bg = fc, fg = wc))
        finder_button.bind("<Leave>", lambda e: finder_button.config(bg = wc, fg = fc))
        sms_button.bind("<Enter>", lambda e: sms_button.config(bg = fc, fg = wc))
        sms_button.bind("<Leave>", lambda e: sms_button.config(bg = wc, fg = fc))
        reports_button.bind("<Enter>", lambda e: reports_button.config(bg = fc, fg = wc))
        reports_button.bind("<Leave>", lambda e: reports_button.config(bg = wc, fg = fc))
        back.bind("<Enter>", lambda e: back.config(text = "Home"))
        back.bind("<Leave>", lambda e: back.config(text = ""))
        logout_button.bind("<Enter>", lambda e: logout_button.config(text = "Logout"))
        logout_button.bind("<Leave>", lambda e: logout_button.config(text = ""))
        exit_button.bind("<Enter>", lambda e: exit_button.config(text = "Exit"))
        exit_button.bind("<Leave>", lambda e: exit_button.config(text = ""))

    def validatenumber(self, num, *args):
        for i in num.get():
            if i.isdigit() == False:
                num.delete(0, END)
                break

    def itemscroller(self, master, rw):
        global scrollbox, scrollable_frame, canvas, scrollbar
        scrollbox = Frame(master)
        container = Frame(scrollbox)
        canvas = Canvas(container, bg = wc)
        scrollbar = Scrollbar(container, orient = "vertical", width = 25, command = canvas.yview)
        scrollable_frame = Frame(canvas)
        scrollable_frame.bind("<Configure>", lambda e: canvas.configure(scrollregion = canvas.bbox("all")))
        canvas.create_window((0,0), window = scrollable_frame, anchor = "nw")
        canvas.configure(yscrollcommand = scrollbar.set)

        scrollbox.grid(column = 0, row = rw, sticky = NW)
        container.pack()
        canvas.pack(side = "left", fill = "both", expand = True)
        canvas.config(height = 320, width = 760)
        scrollbar.pack(side = "right", fill = "y")

    def merchantlister(self, combo):
        global merchants
        merchants = []
        c.execute("SELECT name FROM merchant")
        result = c.fetchall()
        if result:
            for i in result:
                merchants.append(i[0])
            merchants.sort()
            combo.config(values = list(dict.fromkeys(merchants)))
        else:
            combo.config(values = "none")

### GSAD factory ###
    def showgsad(self, *args):
        global gsad_frame, purchase_button, merchant_button, view_button, back
        menu_frame.destroy()
        gsad_frame = LabelFrame(greeting, text = "General Services &\nAdministrative Department", font = fonts, bg = wc, fg = fc)
        gsad_frame.grid(column = 0, row = 1)

        view_button = Button(gsad_frame, text = "View/Edit P.O.", font = fonts, bg = buttonbg, width = 30, command = self.showviewpo)
        view_button.grid(column = 0, row = 0, pady = 5)
        view_button.bind("<Return>", self.showviewpo)

        purchase_button = Button(gsad_frame, text = "Purchase Order", font = fonts, bg = buttonbg, width = 30, command = lambda: self.showpurchaseorder(self.master))
        purchase_button.grid(column = 0, row = 1, pady = 5)
        purchase_button.bind("<Return>", lambda e: self.showpurchaseorder(self.master))

        merchant_button = Button(gsad_frame, text = "Manage Merchant", font = fonts, bg = buttonbg, width = 30, command = self.showmerchant)
        merchant_button.grid(column = 0, row = 2, pady = 5)
        merchant_button.bind("<Return>", self.showmerchant)

        back = Button(gsad_frame, text = "Main Menu", font = fonts, bg = buttonbg, width = 10, command = self.gsadback)
        back.grid(column = 0, row = 5, pady = 5)
        back.bind("<Return>", self.gsadback)

    def showviewpo(self, *args):
        self.gsadbuttons(DISABLED)
        self.colorswitch(view_button, click)
        global search_frame, result_frame, ponumberentry
        search_frame = LabelFrame(self.master, text = "\nView/Edit Purchase Order", font = fonts, bg = wc, fg = fc)
        search_frame.grid(column = 1, row = 0, sticky = NW)

        view_frame = Frame(search_frame, bg = wc)
        view_frame.grid(column = 0, row = 0)

        result_frame = Frame(search_frame, bg = wc)
        result_frame.grid(column = 0, row = 1)

        ponumberlabel = Label(view_frame, text = "P.O. #", font = fonts, width = 10, bg = wc, fg = fc, anchor = W)
        ponumberlabel.grid(column = 0, row = 0, sticky = W)

        ponumberentry = Entry(view_frame, font = fonts, width = 10)
        ponumberentry.grid(column = 1, row = 0, padx = pad)
        ponumberentry.bind("<KeyRelease>", lambda e: self.validatenumber(ponumberentry))
        ponumberentry.bind("<Return>", self.findpurchaseorder)

        ponumberbutton = Button(view_frame, text = "View", font = fonts, width = 10, command = self.findpurchaseorder)
        ponumberbutton.grid(column = 2, row = 0, padx = pad)
        ponumberbutton.bind("<Return>", self.findpurchaseorder)

        closeviewpo = Button(view_frame, text = "Close", font = fonts, width = 10, command = self.closeviewpo)
        closeviewpo.grid(column = 3, row = 0, padx = pad)
        closeviewpo.bind("<Return>", self.closeviewpo)

    def findpurchaseorder(self, *args):
        finder = "SELECT * FROM purchaseorders WHERE number = ? AND company = ?"
        c.execute(finder, [ponumberentry.get(), company])
        poresult = c.fetchall()
        if poresult:
            self.showpurchaseorder(result_frame)
        else:
            messagebox.showerror("View/Edit Purchase Order", "P.O. number not found!")

    def showpurchaseorder(self, master, *args):
        self.gsadbuttons(DISABLED)
        self.colorswitch(purchase_button, click)
        global po_frame, dateentry, dateent, ponumber, supcode, supname, supname_list, supadd, suptin, grossamtvar, vatamtvar, ewtamtvar, netamtvar, tax, particulars_entry, instruction_entry, addbutton, lessbutton
        po_frame = LabelFrame(master, text = "\nPurchase Order", font = fonts, bg = wc, fg = fc)
        po_frame.grid(column = 1, row = 0, sticky = NW)

        details_frame = Frame(po_frame, bg = wc)
        details_frame.grid(column = 0, row = 0, sticky = NW)

        items_frame = Frame(po_frame, bg = wc)
        items_frame.grid(column = 0, row = 1)

        buttons_frame = Frame(po_frame, bg = wc)
        buttons_frame.grid(column = 0, row = 2, stick = E, pady = pad)

        totals_frame = Frame(po_frame, bg = wc)
        totals_frame.place(x = 570, y = 0)
        ### labels ###
        datelabel = Label(details_frame, text = "Date", font = fonts, bg = wc, fg = fc, width = 10, anchor = W)
        datelabel.grid(column = 0, row = 0, pady = pad, sticky = W)

        ponumberlabel = Label(details_frame, text = "P.O. #", font = fonts, bg = wc, fg = fc, width = 10, anchor = W)
        ponumberlabel.place(x = 380, y = 0)

        supcodelabel = Label(details_frame, text = "Supplier Code", font = fonts, bg = wc, fg = fc, width = 13, anchor = W)
        supcodelabel.grid(column = 0, row = 1, pady = pad, sticky = W)

        supnamelabel = Label(details_frame, text = "Supplier Name", font = fonts, bg = wc, fg = fc, width = 13, anchor = W)
        supnamelabel.grid(column = 0, row = 2, pady = pad, sticky = W)

        supaddlabel = Label(details_frame, text = "Supplier Address", font = fonts, bg = wc, fg = fc, width = 13, anchor = W)
        supaddlabel.grid(column = 0, row = 3, pady = pad, sticky = W)

        suptinlabel = Label(details_frame, text = "Supplier TIN", font = fonts, bg = wc, fg = fc, width = 10, anchor = W)
        suptinlabel.grid(column = 0, row = 4, pady = pad, sticky = W)

        particularslabel = Label(details_frame, text = "Particulars", font = fonts, bg = wc, fg = fc, width = 10, anchor = W)
        particularslabel.grid(column = 0, row = 5, pady = pad, sticky = W)

        instructlabel = Label(details_frame, text = "Instructions", font = fonts, bg = wc, fg = fc, width = 10, anchor = W)
        instructlabel.grid(column = 0, row = 6, pady = pad, sticky = W)

        taxlabel = Label(details_frame, text = "Tax Code", font = fonts, bg = wc, fg = fc, width = 10, anchor = W)
        taxlabel.grid(column = 0, row = 7, pady = pad, sticky = W)

        ### entries ###
        dateent = StringVar()
        dateentry = Entry(details_frame, textvariable = dateent, font = fonts, width = 10)
        dateentry.grid(column = 1, row = 0, sticky = NW, padx = 15)
        dateentry.insert(0, today.strftime("%m-%d-%Y"))
        dateentry.bind("<FocusOut>", self.formatdate2)

        ponumber = StringVar()
        ponumberentry = Entry(details_frame, textvariable = ponumber, font = fonts, state = DISABLED, disabledforeground = codetitlefg, width = 10)
        ponumberentry.place(x = 460, y = 0)

        supcode = StringVar()
        supcode_list = Entry(details_frame, textvariable = supcode, font = fonts, state = DISABLED, disabledbackground = codetitlebg, disabledforeground = codetitlefg, width = 11)
        supcode_list.grid(column = 1, row = 1, sticky = NW, padx = 15)

        supname = StringVar()
        supname_list = tk.Combobox(details_frame, textvariable = supname, font = fonts, width = 50)
        supname_list.grid(column = 1, row = 2, sticky = NW, padx = 15)
        supname_list.bind("<<ComboboxSelected>>", self.merchantnameselected)
        supname_list.bind("<FocusOut>", self.merchantnameselected)

        supadd = StringVar()
        supadd_entry = Entry(details_frame, textvariable = supadd, font = fonts, state = DISABLED, disabledbackground = codetitlebg, disabledforeground = codetitlefg, width = 52)
        supadd_entry.grid(column = 1, row = 3, sticky = NW, padx = 15)

        suptin = StringVar()
        suptin_entry = Entry(details_frame, textvariable = suptin, font = fonts, state = DISABLED, disabledbackground = codetitlebg, disabledforeground = codetitlefg, width = 15)
        suptin_entry.grid(column = 1, row = 4, sticky = NW, padx = 15)

        particulars_entry = Entry(details_frame, font = fonts, width = 81)
        particulars_entry.grid(column = 1, row = 5, sticky = NW, padx = 15)

        instruction_entry = Entry(details_frame, font = fonts, width = 81)
        instruction_entry.grid(column = 1, row = 6, sticky = NW, padx = 15)

        tax = StringVar()
        tax_entry = Entry(details_frame, textvariable = tax, font = fonts, state = DISABLED, disabledbackground = codetitlebg, disabledforeground = codetitlefg, width = 15)
        tax_entry.grid(column = 1, row = 7, sticky = NW, padx = 15)

        addbutton = Button(details_frame, font = fonts, bg = wc,  width = 30, image = add, command = self.addpoitem)
        addbutton.grid(column = 0, row = 8, padx = pad)
        addbutton.bind("<Return>", self.addpoitem)

        lessbutton = Button(details_frame, font = fonts, bg = wc, width = 30, image = less, command = self.lesspoitem)
        lessbutton.grid(column = 1, row = 8)
        lessbutton.bind("<Return>", self.lesspoitem)
        ### totals ###
        grosslabel = Label(totals_frame, text = "Gross", font = fonts, width = 10, relief = RIDGE, bg = wc, fg = fc)
        grosslabel.grid(column = 2, row = 0)

        vatlabel = Label(totals_frame, text = "VAT", font = fonts, width = 10, relief = RIDGE, bg = wc, fg = fc)
        vatlabel.grid(column = 2, row = 1)

        ewtlabel = Label(totals_frame, text = "EWT", font = fonts, width = 10, relief = RIDGE, bg = wc, fg = fc)
        ewtlabel.grid(column = 2, row = 2)

        netlabel = Label(totals_frame, text = "Net", font = fonts, width = 10, relief = RIDGE, bg = wc, fg = fc)
        netlabel.grid(column = 2, row = 3)

        grossamtvar = StringVar()
        grossamt = Entry(totals_frame, textvariable = grossamtvar, font = fonts, width = 15, state = DISABLED, disabledbackground = totalbg, disabledforeground = totalfg, justify = RIGHT)
        grossamt.grid(column = 3, row = 0)
    
        vatamtvar = StringVar()
        vatamt = Entry(totals_frame, textvariable = vatamtvar, font = fonts, width = 15, state = DISABLED, disabledbackground = totalbg, disabledforeground = totalfg, justify = RIGHT)
        vatamt.grid(column = 3, row = 1)

        ewtamtvar = StringVar()
        ewtamt = Entry(totals_frame, textvariable = ewtamtvar, font = fonts, width = 15, state = DISABLED, disabledbackground = totalbg, disabledforeground = totalfg, justify = RIGHT)
        ewtamt.grid(column = 3, row = 2)

        netamtvar = StringVar()
        netamt = Entry(totals_frame, textvariable = netamtvar, font = fonts, width = 15, state = DISABLED, disabledbackground = totalbg, disabledforeground = totalfg, justify = RIGHT)
        netamt.grid(column = 3, row = 3)

        submitbutton = Button(buttons_frame, text = "Submit", font = fonts, bg = buttonbg, width = 10, command = self.submitpo)
        submitbutton.grid(column = 0, row = 0, padx = pad)
        submitbutton.bind("<Return>", self.submitpo)

        printbutton = Button(buttons_frame, text = "Print", font = fonts, bg = buttonbg, width = 10, command = None)
        printbutton.grid(column = 1, row = 0, padx = pad)
        printbutton.bind("<Return>", None)

        refreshbutton = Button(buttons_frame, text = "Refresh", font = fonts, bg = buttonbg, width = 10, command = self.refreshpo)
        refreshbutton.grid(column = 2, row = 0, padx = pad)
        refreshbutton.bind("<Return>", self.refreshpo)

        closebutton = Button(buttons_frame, text = "Close", font = fonts, bg = buttonbg, width = 10, command = self.closepo)
        closebutton.grid(column = 3, row = 0, padx = pad)
        closebutton.bind("<Return>", self.closepo)

        global items, allitemcodes, allitemdescs, allquantities, allunits, allcosts, alltotals, allcenters, checks
        items, allitemcodes, allitemdescs, allquantities, allunits, allcosts, alltotals, allcenters, checks = [], [], [], [], [], [], [], [], []
        self.merchantlister(supname_list)
        self.polabels(items_frame)
        self.itemscroller(items_frame, 1)
        self.poitems(scrollable_frame, 0)

    def merchantnameselected(self, *args):
        selected = "SELECT supcode, name, address, tin, code, description, unit, cost, tax FROM merchant WHERE name = ?"
        c.execute(selected, [supname.get()])
        result = c.fetchone()
        if result:
            supcode.set(str(result[0]).zfill(4))
            supname.set(result[1])
            supadd.set(result[2])
            suptin.set(result[3])
            tax.set(result[8])
            c.execute(selected, [supname.get()])
            result2 = c.fetchall()
            codes, descs = [], []
            for x in result2:
                codes.append(x[4])
                descs.append(x[5])
                codes.sort()
                descs.sort()
            for y in allitemcodes:
                try:
                    y.config(values = list(dict.fromkeys(codes)))
                except:
                    pass
            for z in allitemdescs:
                try:
                    z.config(values = list(dict.fromkeys(descs)))
                except:
                    pass

    def ponumbermaster(self):
        global newnumber
        c.execute("SELECT MAX(number) FROM purchaseorders")
        series = c.fetchone()
        if series[0] == None:
            newnumber = 1
            ponumber.set(str(newnumber).zfill(6))
        else:
            newnumber = series[0] + 1
            ponumber.set(str(newnumber).zfill(6))

    def submitpo(self, *args):
        ask = messagebox.askyesno("Submit Purchase Order", "Are you sure?")
        if ask == True:
            self.ponumbermaster()
            validitems = []
            for i in range(len(items)):
                try:
                    if float(alltotals[i].get().replace(",","")) != 0:
                        if tax.get().split("-")[0] == "WV":
                            vat = round(float(alltotals[i].get().replace(",",""))/1.12*.12, 2)
                            ewt = round(float(alltotals[i].get().replace(",",""))/1.12*int(tax.get().split("-")[1])/100, 2)
                        else:
                            vat = round(float(alltotals[i].get().replace(",","")), 2)
                            ewt = round(float(alltotals[i].get().replace(",",""))*int(tax.get().split("-")[1])/100, 2)
                        validitems.append([
                            company, dateentry.get(), newnumber, supcode.get(), particulars_entry.get(),
                            instruction_entry.get(), allitemcodes[i].get(), allitemdescs[i].get(), allquantities[i].get(), allunits[i].get(),
                            float(allcosts[i].get().replace(",","")), float(alltotals[i].get().replace(",","")), float(alltotals[i].get().replace(",","")), vat, ewt,
                            float(alltotals[i].get().replace(",",""))-ewt, allcenters[i].get(), self.username.get(), "idsantiago", dateentry.get().split("-")[0],
                            dateentry.get().split("-")[1], dateentry.get().split("-")[2], "VALID"
                        ])
                except Exception as e:
                    print(e)
            submit = """INSERT INTO purchaseorders (
                company, date, number, supcode, particulars,
                instructions, code, description, quantity, unit,
                cost, total, gross, vat, ewt,
                net, center, user, signatory, month,
                day, year, status)
                values (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)"""
            for i in validitems:
                c.execute(submit, i)
                conn.commit()
            messagebox.showinfo("Submit Purchase Order", "Purchase Order has been submitted!")
            self.podisabler()
    
    def podisabler(self):
        dateentry.config(state = DISABLED)
        supname_list.config(state = DISABLED)
        particulars_entry.config(state = DISABLED)
        instruction_entry.config(state = DISABLED)
        addbutton.config(state = DISABLED)
        lessbutton.config(state = DISABLED)
        for i in range(len(items)):
            allitemcodes[i].config(state = DISABLED)
            allitemdescs[i].config(state = DISABLED)
            allquantities[i].config(state = DISABLED)
            allunits[i].config(state = DISABLED)
            allcenters[i].config(state = DISABLED)

    def poitems(self, master, row):
        itemframe = Frame(master, bg = wc)
        itemframe.grid(column = 0, row = row, sticky = NW)
        items.append(itemframe)

        itemcode = tk.Combobox(itemframe, font = fonts, width = 8)
        itemcode.grid(column = 0, row = 1)
        itemcode.bind("<<ComboboxSelected>>", lambda e: self.poitemcodeselected(itemcode, itemdesc, quantity, unit, costamt, totalamt))
        itemcode.bind("<Return>", lambda e: self.poitemcodeselected(itemcode, itemdesc, quantity, unit, costamt, totalamt))
        allitemcodes.append(itemcode)

        itemdesc = tk.Combobox(itemframe, font = fonts, width = 25)
        itemdesc.grid(column = 1, row = 1)
        itemdesc.bind("<<ComboboxSelected>>", lambda e: self.poitemdescselected(itemcode, itemdesc, quantity, unit, costamt, totalamt))
        itemdesc.bind("<Return>", lambda e: self.poitemdescselected(itemcode, itemdesc, quantity, unit, costamt, totalamt))
        allitemdescs.append(itemdesc)

        quantity = Entry(itemframe, font = fonts, width = 3)
        quantity.grid(column = 2, row = 1)
        quantity.bind("<FocusOut>", lambda e: self.costmultiplier(quantity, costamt, totalamt))
        allquantities.append(quantity)

        unit = tk.Combobox(itemframe, font = fonts, width = 4)
        unit.grid(column = 3, row = 1)
        allunits.append(unit)

        costamt = StringVar()
        cost = Entry(itemframe, textvariable = costamt, font = fonts, state = DISABLED, disabledforeground = codetitlefg, disabledbackground = codetitlebg, width = 15, justify = RIGHT)
        cost.grid(column = 4, row = 1)
        allcosts.append(cost)

        totalamt = StringVar()
        total = Entry(itemframe, textvariable = totalamt, font = fonts, state = DISABLED, disabledforeground = codetitlefg, disabledbackground = codetitlebg, width = 15, justify = RIGHT)
        total.grid(column = 5, row = 1)
        alltotals.append(total)

        center = tk.Combobox(itemframe, values = centers, font = fonts, width = 8)
        center.grid(column = 6, row = 1)
        allcenters.append(center)

        checker = IntVar()
        checkbox = Checkbutton(itemframe, variable = checker, bd = 0, font = fonts, bg = wc)
        checkbox.grid(column = 7, row = 1, padx = 7)
        checks.append(checker)

    def poitemcodeselected(self, codevar, descvar, qtyvar, unitvar, costvar, total, *args):
        find = "SELECT code, description, unit, cost FROM merchant WHERE code = ? AND supcode = ?"
        c.execute(find, [codevar.get(), int(supcode.get())])
        result = c.fetchone()
        if result:
            descvar.delete(0, END)
            descvar.insert(0, result[1])
            qtyvar.delete(0, END)
            qtyvar.insert(0, 1)
            unitvar.delete(0, END)
            unitvar.insert(0, result[2])
            costvar.set(result[3])
            self.costmultiplier(qtyvar, costvar, total)

    def poitemdescselected(self, codevar, descvar, qtyvar, unitvar, costvar, total, *args):
        find = "SELECT code, description, unit, cost FROM merchant WHERE description = ? AND supcode = ?"
        c.execute(find, [descvar.get(), int(supcode.get())])
        result = c.fetchone()
        if result:
            codevar.delete(0, END)
            codevar.insert(0, result[1])
            qtyvar.delete(0, END)
            qtyvar.insert(0, 1)
            unitvar.delete(0, END)
            unitvar.insert(0, result[2])
            costvar.set(result[3])
            self.costmultiplier(qtyvar, costvar, total)

    def costmultiplier(self, qty, cost, total):
        try:
            totalcost = int(qty.get())*float(cost.get().replace(",",""))
            total.set(format(totalcost, ",.2f"))
        except:
            qty.delete(0, END)
            qty.insert(0, 1)
            try:
                total.set(format(float(cost.get().replace(",","")), ",.2f"))
            except:
                total.set(format(0, ",.2f"))
        finally:
            self.pototalupdater()

    def pototalupdater(self):
        grosssum, vatsum, ewtsum = [], [], []
        for i in alltotals:
            try:
                grosssum.append(float(i.get().replace(",","")))
                if tax.get().split("-")[0] == "WV":
                    vatsum.append(float(i.get().replace(",",""))/1.12*.12)
                    ewtsum.append(float(i.get().replace(",",""))/1.12*(int(tax.get().split("-")[1])/100))
                else:
                    vatsum.append(float(i.get().replace(",","")))
                    ewtsum.append(float(i.get().replace(",",""))*(int(tax.get().split("-")[1])/100))
            except:
                grosssum.append(0)
                vatsum.append(0)
                ewtsum.append(0)
        grossamtvar.set(format(sum(grosssum), ",.2f"))
        vatamtvar.set(format(sum(vatsum), ",.2f"))
        ewtamtvar.set(format(sum(ewtsum), ",.2f"))
        netamtvar.set(format(sum(grosssum)-sum(ewtsum), ",.2f"))

    def showmerchant(self, *args):
        self.gsadbuttons(DISABLED)
        self.colorswitch(merchant_button, click)
        global merchant_frame, viewmerchant_button, addmerchant_button, delmerchant_button
        merchant_frame = LabelFrame(self.master, text = "\nManage Merchant", font = fonts, bg = wc, fg = fc)
        merchant_frame.grid(column = 1, row = 0, sticky = NW)

        merchantmenu_frame = Frame(merchant_frame, bg = wc)
        merchantmenu_frame.grid(column = 0, row = 0, sticky = NW)

        viewmerchant_button = Button(merchantmenu_frame, text = "View/Edit Merchant", font = fonts, width = 15, command = self.showmerchantviewer)
        viewmerchant_button.grid(column = 0, row = 0, padx = pad)
        viewmerchant_button.bind("<Return>", self.showmerchantviewer)

        addmerchant_button = Button(merchantmenu_frame, text = "Add Merchant", font = fonts, width = 15, command = self.showmerchantadder)
        addmerchant_button.grid(column = 1, row = 0, padx = pad)
        addmerchant_button.bind("<Return>", self.showmerchantadder)

        delmerchant_button = Button(merchantmenu_frame, text = "Delete Merchant", font = fonts, width = 15, command = self.showmerchantdeleter)
        delmerchant_button.grid(column = 2, row = 0, padx = pad)
        delmerchant_button.bind("<Return>", self.showmerchantdeleter)

        closemerchant_button = Button(merchantmenu_frame, text = "Close", font = fonts, width = 10, command = self.closemerchant)
        closemerchant_button.grid(column = 3, row = 0, padx = pad)
        closemerchant_button.bind("<Return>", self.closemerchant)

    def showmerchantadder(self, *args):
        self.merchantbuttons(DISABLED)
        global merchantadder_frame, items_frame, items, codes, descs, units, costs, types, checks
        merchantadder_frame = Frame(merchant_frame, bg = wc)
        merchantadder_frame.grid(column = 0, row = 1, sticky = NW)

        details_frame = Frame(merchantadder_frame, bg = wc)
        details_frame.grid(column = 0, row = 0, sticky = NW, pady = pad)

        labels_frame  = Frame(merchantadder_frame, bg = wc)
        labels_frame.grid(column = 0, row = 1, sticky = NW, pady = pad)

        items_frame = Frame(merchantadder_frame, bg = wc)
        items_frame.grid(column = 0, row = 2, sticky = NW, pady = pad)

        buttons_frame = Frame(merchantadder_frame, bg = wc)
        buttons_frame.grid(column = 0, row = 3, sticky = NW, pady = pad)

        submitmerchant_button = Button(buttons_frame, text = "Save", font = fonts, width = 10, command = self.savemerchant)
        submitmerchant_button.grid(column = 0, row = 0, padx = pad)
        submitmerchant_button.bind("<Return>", self.savemerchant)

        closemerchantadder_button = Button(buttons_frame, text = "Close", font = fonts, width = 10, command = lambda: self.closemerchantmenu(merchantadder_frame))
        closemerchantadder_button.grid(column = 3, row = 0, padx = pad)
        closemerchantadder_button.bind("<Return>", lambda e: self.closemerchantmenu(merchantadder_frame))

        items, codes, descs, units, costs, types, checks = [], [], [], [], [], [], []
        self.showmerchantdetails(details_frame, 0)
        supcode_entry.config(state = DISABLED, disabledbackground = codetitlebg, disabledforeground = codetitlefg)
        self.pomerchantitemlabels(labels_frame, 0)
        self.itemscroller(items_frame, 0)
        self.showmerchantitems(scrollable_frame, 0)

        addbutton = Button(details_frame, font = fonts, bg = wc,  width = 30, image = add, command = self.additem)
        addbutton.grid(column = 0, row = 2, padx = pad)
        addbutton.bind("<Return>", self.additem)

        lessbutton = Button(details_frame, font = fonts, bg = wc, width = 30, image = less, command = self.lessitem)
        lessbutton.grid(column = 1, row = 2)
        lessbutton.bind("<Return>", self.lessitem)

    def showmerchantviewer(self, *args):
        self.merchantbuttons(DISABLED)
        global merchantviewer_frame, merchantviewersub_frame, merchant_list
        merchantviewer_frame = Frame(merchant_frame, bg = wc)
        merchantviewer_frame.grid(column = 0, row = 1, pady = pad, sticky = NW)

        merchantviewersub_frame = Frame(merchantviewer_frame, bg = wc)
        merchantviewersub_frame.grid(column = 0, row = 0, sticky = NW)

        merchant_list = tk.Combobox(merchantviewersub_frame, font = fonts, width = 50)
        merchant_list.grid(column = 0, row = 0)
        merchant_list.bind("<<ComboboxSelected>>", self.viewmerchant)

        closemerchantviewer_button = Button(merchantviewersub_frame, text = "Close", font = fonts, width = 10, command = lambda: self.closemerchantmenu(merchantviewer_frame))
        closemerchantviewer_button.grid(column = 2, row = 0, padx = pad)
        closemerchantviewer_button.bind("<Return>", lambda e: self.closemerchantmenu(merchantviewer_frame))

        self.merchantlister(merchant_list)

    def showmerchantdeleter(self, *args):
        self.showmerchantviewer()
        merchant_list.unbind("<<ComboboxSelected>>")

        deletemerchant_button = Button(merchantviewersub_frame, text = "Delete", font = fonts, width = 10, command = self.deletemerchant)
        deletemerchant_button.grid(column = 1, row = 0, padx = pad)
        deletemerchant_button.bind("<Return>", self.deletemerchant)

    def viewmerchant(self, *args):
        global items, codes, descs, units, costs, types, checks
        try:
            viewmerchant_frame.destroy()
        except:
            pass
        finally:
            viewmerchant_frame = Frame(merchantviewer_frame, bg = wc)
            viewmerchant_frame.grid(column = 0, row = 1, pady = pad, sticky = NW)

            viewbutton_frame = Frame(merchantviewer_frame, bg = wc)
            viewbutton_frame.grid(column = 0, row = 2, pady = pad, sticky = NW)

        items, codes, descs, units, costs, types, checks = [], [], [], [], [], [], []
        self.showmerchantdetails(viewmerchant_frame, 0)
        finder = "SELECT * FROM merchant WHERE name = ?"
        c.execute(finder, [merchant_list.get()])
        result = c.fetchall()
        supcode.set(str(result[0][0]).zfill(4))
        supcode_entry.config(state = DISABLED)
        supname_entry.delete(0, END)
        supname_entry.insert(0, result[0][1])
        supadd_entry.delete(0, END)
        supadd_entry.insert(0, result[0][2])
        suptin_entry.delete(0, END)
        suptin_entry.insert(0, result[0][3])
        supcon_entry.delete(0, END)
        supcon_entry.insert(0, result[0][9])
        tax_list.delete(0, END)
        tax_list.insert(0, result[0][4])

        addbutton = Button(merchantdetails_frame, font = fonts, bg = wc,  width = 30, image = add, command = self.additem)
        addbutton.grid(column = 0, row = 6, padx = pad)
        addbutton.bind("<Return>", self.additem)

        lessbutton = Button(merchantdetails_frame, font = fonts, bg = wc, width = 30, image = less, command = self.lessitem)
        lessbutton.grid(column = 1, row = 6)
        lessbutton.bind("<Return>", self.lessitem)

        self.pomerchantitemlabels(viewmerchant_frame, 1)
        self.itemscroller(viewmerchant_frame, 2)
        self.showmerchantitems(scrollable_frame, 0)
        for i in range(len(result)):
            self.additem()
            codes[i].insert(0, result[i][5])
            descs[i].insert(0, result[i][6])
            units[i].insert(0, result[i][7])
            costs[i].set(result[i][8])
            types[i].insert(0, result[i][9])

        updatemerchant_button = Button(viewbutton_frame, text = "Update", font = fonts, width = 10, command = self.updatemerchant)
        updatemerchant_button.grid(column = 0, row = 0, padx = pad)
        updatemerchant_button.bind("<Return>", self.updatemerchant)

    def updatemerchant(self, *args):
        if supname_entry.get() == "" or supadd_entry.get() == "" or suptin_entry.get() == "" or tax_list.get() not in atc:
            messagebox.showerror("Update Merchant", "Please fill-out all required details!")
        else:
            ask = messagebox.askyesno("Update Merchant", "Are you sure?")
            if ask == True:
                suppliercode = int(supcode.get())
                deleter = "DELETE FROM merchant WHERE supcode = ?"
                c.execute(deleter, [suppliercode])
                conn.commit()
                blankitems = []
                for i in range(len(costs)):
                    try:
                        if float(costs[i].get().replace(",","")) == 0:
                            blankitems.append(costs[i])
                    except:
                        pass
                if len(blankitems) == 0:
                    validitems = []
                    for i in range(len(costs)):
                        try:
                            if float(costs[i].get().replace(",","")) > 0:
                                validitems.append([
                                    suppliercode, supname_entry.get(), supadd_entry.get(), suptin_entry.get(), codes[i].get(),
                                    descs[i].get(), units[i].get(), float(costs[i].get().replace(",","")), types[i].get(), supcon_entry.get(), self.username.get(), tax_list.get()
                                    ])
                        except:
                            pass
                    if len(validitems) != 0:
                        saver = """INSERT INTO merchant (
                            supcode, name, address, tin, code,
                            description, unit, cost, type, contact, user, tax
                        ) values (?,?,?,?,?,?,?,?,?,?,?,?)"""
                        for i in validitems:
                            c.execute(saver, i)
                            conn.commit()
                        messagebox.showinfo("Update Merchant", "Merchant saved successfully!")
                    else:
                        messagebox.showerror("Update Merchant", "Error in valid items!")
                else:
                    messagebox.showerror("Update Merchant", "Don't leave an item cost blank!")

    def savemerchant(self, *args):
        if supname_entry.get() == "" or supadd_entry.get() == "" or suptin_entry.get() == "" or tax_list.get() not in atc:
            messagebox.showerror("Add Merchant", "Please fill-out all required details!")
        else:
            ask = messagebox.askyesno("Add Merchant", "Are you sure?")
            if ask == True:
                blankitems = []
                for i in range(len(costs)):
                    try:
                        if float(costs[i].get().replace(",","")) == 0:
                            blankitems.append(costs[i])
                    except:
                        pass
                if len(blankitems) == 0:
                    self.merchantnumbermaster()
                    validitems = []
                    for i in range(len(costs)):
                        try:
                            if float(costs[i].get().replace(",","")) > 0:
                                validitems.append([
                                    int(supcode.get()), supname_entry.get(), supadd_entry.get(), suptin_entry.get(), codes[i].get(),
                                    descs[i].get(), units[i].get(), float(costs[i].get().replace(",","")), types[i].get(), supcon_entry.get(), self.username.get(), tax_list.get()
                                    ])
                        except:
                            pass
                    if len(validitems) != 0:
                        saver = """INSERT INTO merchant (
                            supcode, name, address, tin, code,
                            description, unit, cost, type, contact, user, tax
                        ) values (?,?,?,?,?,?,?,?,?,?,?,?)"""
                        for i in validitems:
                            c.execute(saver, i)
                            conn.commit()
                        messagebox.showinfo("Add Merchant", "Merchant saved successfully!")
                    else:
                        messagebox.showerror("Add Merchant", "Error in valid items!")
                else:
                    messagebox.showerror("Add Merchant", "Don't leave an item cost blank!")

    def deletemerchant(self, *args):
        ask = messagebox.askyesno("Delete Merchant", "Are you sure?")
        if ask == True:
            deleter = "DELETE FROM merchant WHERE name = ?"
            c.execute(deleter, [merchant_list.get()])
            conn.commit()
            messagebox.showinfo("Delete Merchant", "Merchant deleted successfully!")
            self.merchantlister(merchant_list)

    def merchantnumbermaster(self):
        global newnumber
        c.execute("SELECT MAX(supcode) FROM merchant")
        series = c.fetchall()
        if series[0][0] == None:
            newnumber = 1
            supcode.set(str(newnumber).zfill(4))
        else:
            newnumber = series[0][0] + 1
            supcode.set(str(newnumber).zfill(4))

    def polabels(self, master):
        polabels_frame = Frame(master, bg = wc)
        polabels_frame.grid(column = 0, row = 0, sticky = NW)

        codelabel = Label(polabels_frame, text = "Code", font = fonts, width = 10, relief = RIDGE, bg = wc, fg = fc)
        codelabel.grid(column = 0, row = 0)

        itemlabel = Label(polabels_frame, text = "Item", font = fonts, width = 27, relief = RIDGE, bg = wc, fg = fc)
        itemlabel.grid(column = 1, row = 0)

        quantitylabel = Label(polabels_frame, text = "Qty", font = fonts, width = 3, relief = RIDGE, bg = wc, fg = fc)
        quantitylabel.grid(column = 2, row = 0)

        unitlabel = Label(polabels_frame, text = "U.O.M", font = fonts, width = 6, relief = RIDGE, bg = wc, fg = fc)
        unitlabel.grid(column = 3, row = 0)

        costlabel = Label(polabels_frame, text = "Cost", font = fonts, width = 15, relief = RIDGE, bg = wc, fg = fc)
        costlabel.grid(column = 4, row = 0)

        totallabel = Label(polabels_frame, text = "Total", font = fonts, width = 15, relief = RIDGE, bg = wc, fg = fc)
        totallabel.grid(column = 5, row = 0)

        centerlabel = Label(polabels_frame, text = "Cost Center", font = fonts, width = 10, relief = RIDGE, bg = wc, fg = fc)
        centerlabel.grid(column = 6, row = 0)

        addlabel = Label(polabels_frame, text = " □ ", font = fonts, width = 4, relief = RIDGE, bg = wc, fg = fc)
        addlabel.grid(column = 7, row = 0)

    def pomerchantitemlabels(self, master, rw):
        polabels_frame = Frame(master, bg = wc)
        polabels_frame.grid(column = 0, row = rw, sticky = NW)

        codelabel = Label(polabels_frame, text = "Code", font = fonts, width = 10, relief = RIDGE, bg = wc, fg = fc)
        codelabel.grid(column = 0, row = 0)

        itemlabel = Label(polabels_frame, text = "Item", font = fonts, width = 36, relief = RIDGE, bg = wc, fg = fc)
        itemlabel.grid(column = 1, row = 0)

        unitlabel = Label(polabels_frame, text = "U.O.M", font = fonts, width = 10, relief = RIDGE, bg = wc, fg = fc)
        unitlabel.grid(column = 2, row = 0)

        costlabel = Label(polabels_frame, text = "Cost", font = fonts, width = 15, relief = RIDGE, bg = wc, fg = fc)
        costlabel.grid(column = 3, row = 0)

        typelabel = Label(polabels_frame, text = "Type", font = fonts, width = 15, relief = RIDGE, bg = wc, fg = fc)
        typelabel.grid(column = 4, row = 0)

        addlabel = Label(polabels_frame, text = " □ ", font = fonts, width = 4, relief = RIDGE, bg = wc, fg = fc)
        addlabel.grid(column = 5, row = 0)

    def showmerchantdetails(self, frame, row, *args):
        global merchantdetails_frame, supcode_entry, supcode, supname_entry, supadd_entry, suptin_entry, supcon_entry, tax_list, atc
        merchantdetails_frame = Frame(frame, bg = wc)
        merchantdetails_frame.grid(column = 0, row = row, sticky = NW)

        supcodelabel = Label(merchantdetails_frame, text = "Code", font = fonts, bg = wc, fg = fc, width = 13, anchor = W)
        supcodelabel.grid(column = 0, row = 0, pady = pad, sticky = W)

        supnamelabel = Label(merchantdetails_frame, text = "Name", font = fonts, bg = wc, fg = fc, width = 13, anchor = W)
        supnamelabel.grid(column = 0, row = 1, pady = pad, sticky = W)

        supaddlabel = Label(merchantdetails_frame, text = "Address", font = fonts, bg = wc, fg = fc, width = 13, anchor = W)
        supaddlabel.grid(column = 0, row = 2, pady = pad, sticky = W)

        suptinlabel = Label(merchantdetails_frame, text = "TIN", font = fonts, bg = wc, fg = fc, width = 10, anchor = W)
        suptinlabel.grid(column = 0, row = 3, pady = pad, sticky = W)

        supconlabel = Label(merchantdetails_frame, text = "Contact", font = fonts, bg = wc, fg = fc, width = 10, anchor = W)
        supconlabel.grid(column = 0, row = 4, pady = pad, sticky = W)

        taxlabel = Label(merchantdetails_frame, text = "Tax Code", font = fonts, bg = wc, fg = fc, width = 10, anchor = W)
        taxlabel.grid(column = 0, row = 5, pady = pad, sticky = W)

        supcode = StringVar()
        supcode_entry = Entry(merchantdetails_frame, textvariable = supcode, font = fonts, width = 13)
        supcode_entry.grid(column = 1, row = 0, sticky = NW, padx = 15)

        supname_entry = Entry(merchantdetails_frame, font = fonts, width = 52)
        supname_entry.grid(column = 1, row = 1, sticky = NW, padx = 15)
        supname_entry.bind("<FocusOut>", lambda e: self.uppercase(supname_entry))

        supadd_entry = Entry(merchantdetails_frame, font = fonts, width = 52)
        supadd_entry.grid(column = 1, row = 2, sticky = NW, padx = 15)
        supadd_entry.bind("<FocusOut>", lambda e: self.uppercase(supadd_entry))

        suptin_entry = Entry(merchantdetails_frame, font = fonts, width = 15)
        suptin_entry.grid(column = 1, row = 3, sticky = NW, padx = 15)

        supcon_entry = Entry(merchantdetails_frame, font = fonts, width = 20)
        supcon_entry.grid(column = 1, row = 4, sticky = NW, padx = 15)

        atc = ["WV-01","WV-02","NV-01","NV-02","NV-00"]
        tax_list = tk.Combobox(merchantdetails_frame, values = atc, font = fonts, width = 13)
        tax_list.grid(column = 1, row = 5, sticky = NW, padx = 15)

    def additem(self, *args):
        itemrow = len(items) + 1
        self.showmerchantitems(scrollable_frame, itemrow)

    def lessitem(self, *args):
        for i in checks:
            if i.get() == 1:
                items[checks.index(i)].destroy()

    def addpoitem(self, *args):
        itemrow = len(items) + 1
        self.poitems(scrollable_frame, itemrow)
        self.merchantnameselected()

    def lesspoitem(self, *args):
        for i in checks:
            if i.get() == 1:
                items[checks.index(i)].destroy()
        self.pototalupdater()

    def showmerchantitems(self, frame, row, *args):
        itemframe = Frame(frame, bg = wc)
        itemframe.grid(column = 0, row = row, sticky = NW)
        items.append(itemframe)

        codeentry = Entry(itemframe, font = fonts, width = 10)
        codeentry.grid(column = 0, row = 0, padx = 1)
        codeentry.bind("<FocusOut>", lambda e: self.uppercase(codeentry))
        codes.append(codeentry)

        descentry = Entry(itemframe, font = fonts, width = 36)
        descentry.grid(column = 1, row = 0, padx = 1)
        descentry.bind("<FocusOut>", lambda e: self.uppercase(descentry))
        descs.append(descentry)

        unitentry = tk.Combobox(itemframe, values = ["piece","box","lot","gallon","liter","kilogram","yard","meter","ream"], width = 8, font = fonts)
        unitentry.grid(column = 2, row = 0)
        units.append(unitentry)

        cost = StringVar()
        costentry = Entry(itemframe, textvariable = cost, font = fonts, width = 15, justify = RIGHT)
        costentry.grid(column = 3, row = 0)
        costentry.bind("<FocusOut>", lambda e: self.amountvalidatormaster(cost))
        costs.append(cost)

        typeentry = tk.Combobox(itemframe, values = ["office","computer","security","janitorial","various forms"], width = 13, font = fonts)
        typeentry.grid(column = 4, row = 0)
        types.append(typeentry)

        checker = IntVar()
        checkbox = Checkbutton(itemframe, variable = checker, bd = 0, font = fonts, bg = wc)
        checkbox.grid(column = 5, row = 0, padx = 7)
        checks.append(checker)

    def closemerchant(self, *args):
        merchant_frame.destroy()
        self.colorswitch(merchant_button, buttonbg)
        self.gsadbuttons(NORMAL)

    def closemerchantmenu(self, menu, *args):
        menu.destroy()
        self.merchantbuttons(NORMAL)
        self.colorswitch(merchant_button, click)

    def closeviewpo(self, *args):
        search_frame.destroy()
        self.gsadbuttons(NORMAL)
        self.colorswitch(view_button, buttonbg)

    def closepo(self, *args):
        po_frame.destroy()
        self.colorswitch(purchase_button, buttonbg)
        self.gsadbuttons(NORMAL)

    def refreshpo(self, *args):
        po_frame.destroy()
        self.showpurchaseorder(self.master)

    def gsadbuttons(self, status):
        purchase_button.config(state = status)
        merchant_button.config(state = status)
        view_button.config(state = status)
        back.config(state = status)

    def merchantbuttons(self, status):
        viewmerchant_button.config(state = status)
        addmerchant_button.config(state = status)
        delmerchant_button.config(state = status)

    def gsadback(self, *args):
        gsad_frame.destroy()
        greeting.destroy()
        self.menu()
        if access[1] == "GSAD":
            accounting_button.config(state = DISABLED)
            fmd_button.config(state = DISABLED)
            bcd_button.config(state = DISABLED)

### BCD factory ###
    def showbcd(self, *args):
        root.resizable(width = True, height = False)
        root.geometry(bcdgeometry)
        root.resizable(width = False, height = False)
        global bcd_frame, view_button, billing_button, collection_button, reports_button, back
        menu_frame.destroy()
        bcd_frame = LabelFrame(greeting, text = "Billing & Collection", font = fonts, bg = wc, fg = fc)
        bcd_frame.grid(column = 0, row = 1)

        view_button = Button(bcd_frame, text = "View/Update Record", font = fonts, bg = buttonbg, width = 30, command = self.showviewbill)
        view_button.grid(column = 0, row = 0, pady = 5)
        view_button.bind("<Return>", self.showviewbill)

        billing_button = Button(bcd_frame, text = "Billing", font = fonts, bg = buttonbg, width = 30, command = self.showbilling)
        billing_button.grid(column = 0, row = 1, pady = 5)
        billing_button.bind("<Return>", self.showbilling)

        collection_button = Button(bcd_frame, text = "Collection", font = fonts, bg = buttonbg, width = 30, command = self.showcollection)
        collection_button.grid(column = 0, row = 2, pady = 5)
        collection_button.bind("<Return>", self.showcollection)

        reports_button = Button(bcd_frame, text = "Reports", font = fonts, bg = buttonbg, width = 30, command = self.showbcdreports)
        reports_button.grid(column = 0, row = 3, pady = 5)
        reports_button.bind("<Return>", self.showbcdreports)

        back = Button(bcd_frame, text = "Main Menu", font = fonts, bg = buttonbg, width = 10, command = self.bcdback)
        back.grid(column = 0, row = 5, pady = 5)
        back.bind("<Return>", self.bcdback)

    def showviewbill(self, *args):
        self.colorswitch(view_button, click)
        self.bcdbuttons(DISABLED)
        global viewbillframe, findsoaentry
        viewbillframe = LabelFrame(self.master, text = "\nView/Update Record", font = fonts, bg = wc, fg = fc)
        viewbillframe.grid(column = 1, row = 0, sticky = NW)

        findbillframe = Frame(viewbillframe, bg = wc)
        findbillframe.grid(column = 0, row = 0, sticky = NW)

        findsoalabel = Label(findbillframe, text = "SOA No.", font = fonts, bg = wc, fg = fc, width = 10, anchor = W)
        findsoalabel.grid(column = 0, row = 0, padx = pad)

        findsoaentry = Entry(findbillframe, font = fonts, width = 10)
        findsoaentry.grid(column = 1, row = 0, padx = pad)
        findsoaentry.bind("<Key>", lambda e: self.validatenumber(findsoaentry))
        findsoaentry.bind("<Return>", self.findbill)
        findsoaentry.focus()

        findbillbutton = Button(findbillframe, text = "View", font = fonts, bg = buttonbg, width = 10, command = self.findbill)
        findbillbutton.grid(column = 2, row = 0, pady = pad)
        findbillbutton.bind("<Return>", self.findbill)

        closeviewbillbutton = Button(findbillframe, text = "Close", font = fonts, bg = buttonbg, width = 10, command = self.closeviewbill)
        closeviewbillbutton.grid(column = 3, row = 0, pady = pad)
        closeviewbillbutton.bind("<Return>", self.closeviewbill)

    def showbilling(self, *args):
        self.colorswitch(billing_button, click)
        self.bcdbuttons(DISABLED)
        global billingframe, submitbutton, closebutton, dateentry, soaentry, periodentry, remarksentry, itemsframe, entryframe, cancelbox, checkbox, staffcategory, clientcategory, cliententry1, cliententry2, categoryentry1, categoryentry2, grossamt, grossamtvar, vatamt, vatamtvar, ewtamt, ewtamtvar, fvatamt, fvatamtvar, netamt, netamtvar
        if company == "DBPSC":
            staffcategory = ["Clerical", "Janitorial"]
        else:
            staffcategory = ["Security"]
        billingframe = LabelFrame(self.master, text = "\nBilling", font = fonts, bg = wc, fg = fc)
        billingframe.grid(column = 1, row = 0, sticky = NW)

        billingsubframe = Frame(billingframe, bg = wc)
        billingsubframe.grid(column = 0, row = 0, sticky = NW)

        labelsframe = Frame(billingframe, bg = wc)
        labelsframe.grid(column = 0, row = 1, sticky = NW)

        itemsframe = Frame(billingframe, bg = wc)
        itemsframe.grid(column = 0, row = 2, sticky = NW)

        entryframe = Frame(billingframe, bg = wc)
        entryframe.grid(column = 0, row = 3, sticky = NW)

        buttonframe = Frame(billingframe, bg = wc)
        buttonframe.place(x = 780, y = 230)

        datelabel = Label(billingsubframe, text = "Date", font = fonts, bg = wc, fg = fc, width = 10, anchor = W)
        datelabel.grid(column = 0, row = 0, pady = pad)

        clientlabel = Label(billingsubframe, text = "Client", font = fonts, bg = wc, fg = fc, width = 10, anchor = W)
        clientlabel.grid(column = 0, row = 1, pady = pad)

        subclientlabel = Label(billingsubframe, text = "Sub-Group", font = fonts, bg = wc, fg = fc, width = 10, anchor = W)
        subclientlabel.grid(column = 0, row = 2, pady = pad)

        categorylabel1 = Label(billingsubframe, text = "Category1", font = fonts, bg = wc, fg = fc, width = 10, anchor = W)
        categorylabel1.grid(column = 0, row = 3, pady = pad)

        categorylabel2 = Label(billingsubframe, text = "Category2", font = fonts, bg = wc, fg = fc, width = 10, anchor = W)
        categorylabel2.grid(column = 0, row = 4, pady = pad)

        soalabel = Label(billingsubframe, text = "SOA No.", font = fonts, bg = wc, fg = fc, width = 10, anchor = W)
        soalabel.grid(column = 0, row = 5, pady = pad)

        periodlabel = Label(billingsubframe, text = "Period", font = fonts, bg = wc, fg = fc, width = 10, anchor = W)
        periodlabel.grid(column = 0, row = 6, pady = pad)

        remarkslabel = Label(billingsubframe, text = "Particulars", font = fonts, bg = wc, fg = fc, width = 10, anchor = W)
        remarkslabel.grid(column = 0, row = 7, pady = pad)

        dateentry = Entry(billingsubframe, font = fonts, width = 10)
        dateentry.grid(column = 1, row = 0, sticky = NW, padx = 15)
        dateentry.insert(0, datetime.datetime.strftime(today, "%m-%d-%Y"))
        dateentry.bind("<FocusOut>", self.formatdate2)

        cliententry1 = tk.Combobox(billingsubframe, font = fonts, width = 67)
        cliententry1.grid(column = 1, row = 1, sticky = NW, padx = 15)
        cliententry1.bind("<<ComboboxSelected>>", self.listclients2)
        cliententry1.bind("<<FocusOut>>", self.listclients2)

        cliententry2 = tk.Combobox(billingsubframe, font = fonts, width = 67)
        cliententry2.grid(column = 1, row = 2, sticky = NW, padx = 15)

        categoryentry1 = tk.Combobox(billingsubframe, values = staffcategory, font = fonts, width = 30)
        categoryentry1.grid(column = 1, row = 3, sticky = NW, padx = 15)
        if company == "DSSI":
            categoryentry1.insert(0, staffcategory[0])

        clientcategory = ["Private", "Government"]
        categoryentry2 = tk.Combobox(billingsubframe, values = clientcategory, font = fonts, width = 30)
        categoryentry2.grid(column = 1, row = 4, sticky = NW, padx = 15)
        categoryentry2.bind("<<ComboboxSelected>>", self.updatebilltotal)

        soaentry = Entry(billingsubframe, font = fonts, width = 10)
        soaentry.grid(column = 1, row = 5, sticky = NW, padx = 15)
        soaentry.bind("<Key>", lambda e: self.validatenumber(soaentry))

        periodentry = Entry(billingsubframe, font = fonts, width = 32)
        periodentry.grid(column = 1, row = 6, sticky = NW, padx = 15)

        remarksentry = Entry(billingsubframe, font = fonts, width = 69)
        remarksentry.grid(column = 1, row = 7, sticky = NW, padx = 15)

        grosslabel = Label(billingsubframe, text = "Gross", font = fonts, width = 10, relief = RIDGE, bg = wc, fg = fc)
        grosslabel.grid(column = 2, row = 0)

        vatlabel = Label(billingsubframe, text = "VAT", font = fonts, width = 10, relief = RIDGE, bg = wc, fg = fc)
        vatlabel.grid(column = 2, row = 1)

        ewtlabel = Label(billingsubframe, text = "EWT", font = fonts, width = 10, relief = RIDGE, bg = wc, fg = fc)
        ewtlabel.grid(column = 2, row = 2)

        fvatlabel = Label(billingsubframe, text = "FVAT", font = fonts, width = 10, relief = RIDGE, bg = wc, fg = fc)
        fvatlabel.grid(column = 2, row = 3)

        netlabel = Label(billingsubframe, text = "Net", font = fonts, width = 10, relief = RIDGE, bg = wc, fg = fc)
        netlabel.grid(column = 2, row = 4)

        grossamtvar = StringVar()
        grossamt = Entry(billingsubframe, textvariable = grossamtvar, font = fonts, width = 15, state = DISABLED, disabledbackground = totalbg, disabledforeground = totalfg, justify = RIGHT)
        grossamt.grid(column = 3, row = 0)
    
        vatamtvar = StringVar()
        vatamt = Entry(billingsubframe, textvariable = vatamtvar, font = fonts, width = 15, state = DISABLED, disabledbackground = totalbg, disabledforeground = totalfg, justify = RIGHT)
        vatamt.grid(column = 3, row = 1)

        ewtamtvar = StringVar()
        ewtamt = Entry(billingsubframe, textvariable = ewtamtvar, font = fonts, width = 15, state = DISABLED, disabledbackground = totalbg, disabledforeground = totalfg, justify = RIGHT)
        ewtamt.grid(column = 3, row = 2)

        fvatamtvar = StringVar()
        fvatamt = Entry(billingsubframe, textvariable = fvatamtvar, font = fonts, width = 15, state = DISABLED, disabledbackground = totalbg, disabledforeground = totalfg, justify = RIGHT)
        fvatamt.grid(column = 3, row = 3)

        netamtvar = StringVar()
        netamt = Entry(billingsubframe, textvariable = netamtvar, font = fonts, width = 15, state = DISABLED, disabledbackground = totalbg, disabledforeground = totalfg, justify = RIGHT)
        netamt.grid(column = 3, row = 4)

        checkbox = IntVar()
        cancelbox = Checkbutton(billingsubframe, text = 'Cancel', variable = checkbox, bg = wc, fg = fc, font = fonts, command = self.cancel)
        cancelbox.bind('<Button-1>', self.cancel)

        submitbutton = Button(buttonframe, text = "Submit", font = fonts, bg = buttonbg, width = 10, command = self.checkbillingentries)
        submitbutton.grid(column = 0, row = 0, pady = pad)
        submitbutton.bind("<Return>", self.checkbillingentries)

        printbutton = Button(buttonframe, text = "Print", font = fonts, bg = buttonbg, width = 10, command = None)
        printbutton.grid(column = 0, row = 1, pady = pad)
        printbutton.bind("<Return>", None)

        closebutton = Button(buttonframe, text = "Close", font = fonts, bg = buttonbg, width = 10, command = self.closebilling)
        closebutton.grid(column = 0, row = 5, pady = pad)
        closebutton.bind("<Return>", self.closebilling)

        self.listclients1()
        self.showitemlabels(labelsframe)
        self.showbillingitems(itemsframe)

    def showitemlabels(self, master):
        descriptionlabel = Label(master, text = "Description", font = fonts, width = 60, relief = RIDGE, bg = wc, fg = fc)
        descriptionlabel.grid(column = 0, row = 0)

        amountlabel = Label(master, text = "Amount", font = fonts, width = 15, relief = RIDGE, bg = wc, fg = fc)
        amountlabel.grid(column = 1, row = 0)

        taxlabel = Label(master, text = "Tax Code", font = fonts, width = 12, relief = RIDGE, bg = wc, fg = fc)
        taxlabel.grid(column = 2, row = 0)

        clearlabel = Label(master, text = "<<", font = fonts, width = 5, relief = RIDGE, bg = wc, fg = fc)
        clearlabel.grid(column = 3, row = 0)

    def showbillingitems(self, master):
        global allbillamt, allbill, alltax, alldesc
        atc = ["WV-02","NV-02","NV-00","Zero-Rated"]
        alldesc, allbillamt, allbill, alltax = [], [], [], []
        ### line1 ###
        descentry1 = Entry(master, font = fonts, width = 60)
        descentry1.grid(column = 0, row = 0)

        billamt1 = StringVar()
        billamount1 = Entry(master, textvariable = billamt1, font = fonts, width = 15, justify = RIGHT)
        billamount1.grid(column = 1, row = 0, padx = 1)
        billamount1.bind("<FocusOut>", lambda e: self.validateamount(billamt1))

        billtaxlist1 = tk.Combobox(master, values = atc, font = fonts, width = 10)
        billtaxlist1.grid(column = 2, row = 0)
        billtaxlist1.bind("<<ComboboxSelected>>", self.updatebilltotal)

        billclear1 = Button(master, text = "clear", font = clearbuttonfont, width = 5, command = lambda: self.clearbillitem(descentry1, billamount1, billtaxlist1))
        billclear1.grid(column = 3, row = 0, padx = 1)
        billclear1.bind("<Return>", lambda e: self.clearbillitem(descentry1, billamount1, billtaxlist1))

        alldesc.append(descentry1)
        allbillamt.append(billamt1)
        allbill.append(billamount1)
        alltax.append(billtaxlist1)
        ### line2 ###
        descentry2 = Entry(master, font = fonts, width = 60)
        descentry2.grid(column = 0, row = 1)

        billamt2 = StringVar()
        billamount2 = Entry(master, textvariable = billamt2, font = fonts, width = 15, justify = RIGHT)
        billamount2.grid(column = 1, row = 1, padx = 1)
        billamount2.bind("<FocusOut>", lambda e: self.validateamount(billamt2))

        billtaxlist2 = tk.Combobox(master, values = atc, font = fonts, width = 10)
        billtaxlist2.grid(column = 2, row = 1)
        billtaxlist2.bind("<<ComboboxSelected>>", self.updatebilltotal)

        billclear2 = Button(master, text = "clear", font = clearbuttonfont, width = 5, command = lambda: self.clearbillitem(descentry2, billamount2, billtaxlist2))
        billclear2.grid(column = 3, row = 1, padx = 1)
        billclear2.bind("<Return>", lambda e: self.clearbillitem(descentry2, billamount2, billtaxlist2))

        alldesc.append(descentry2)
        allbillamt.append(billamt2)
        allbill.append(billamount2)
        alltax.append(billtaxlist2)
        ### line3 ###
        descentry3 = Entry(master, font = fonts, width = 60)
        descentry3.grid(column = 0, row = 2)

        billamt3 = StringVar()
        billamount3 = Entry(master, textvariable = billamt3, font = fonts, width = 15, justify = RIGHT)
        billamount3.grid(column = 1, row = 2, padx = 1)
        billamount3.bind("<FocusOut>", lambda e: self.validateamount(billamt3))

        billtaxlist3 = tk.Combobox(master, values = atc, font = fonts, width = 10)
        billtaxlist3.grid(column = 2, row = 2)
        billtaxlist3.bind("<<ComboboxSelected>>", self.updatebilltotal)

        billclear3 = Button(master, text = "clear", font = clearbuttonfont, width = 5, command = lambda: self.clearbillitem(descentry3, billamount3, billtaxlist3))
        billclear3.grid(column = 3, row = 2, padx = 1)
        billclear3.bind("<Return>", lambda e: self.clearbillitem(descentry3, billamount3, billtaxlist3))

        alldesc.append(descentry3)
        allbillamt.append(billamt3)
        allbill.append(billamount3)
        alltax.append(billtaxlist3)
        ### line4 ###
        descentry4 = Entry(master, font = fonts, width = 60)
        descentry4.grid(column = 0, row = 3)

        billamt4 = StringVar()
        billamount4 = Entry(master, textvariable = billamt4, font = fonts, width = 15, justify = RIGHT)
        billamount4.grid(column = 1, row = 3, padx = 1)
        billamount4.bind("<FocusOut>", lambda e: self.validateamount(billamt4))

        billtaxlist4 = tk.Combobox(master, values = atc, font = fonts, width = 10)
        billtaxlist4.grid(column = 2, row = 3)
        billtaxlist4.bind("<<ComboboxSelected>>", self.updatebilltotal)

        billclear4 = Button(master, text = "clear", font = clearbuttonfont, width = 5, command = lambda: self.clearbillitem(descentry4, billamount4, billtaxlist4))
        billclear4.grid(column = 3, row = 3, padx = 1)
        billclear4.bind("<Return>", lambda e: self.clearbillitem(descentry4, billamount4, billtaxlist4))

        alldesc.append(descentry4)
        allbillamt.append(billamt4)
        allbill.append(billamount4)
        alltax.append(billtaxlist4)
        ### line5 ###
        descentry5 = Entry(master, font = fonts, width = 60)
        descentry5.grid(column = 0, row = 4)

        billamt5 = StringVar()
        billamount5 = Entry(master, textvariable = billamt5, font = fonts, width = 15, justify = RIGHT)
        billamount5.grid(column = 1, row = 4, padx = 1)
        billamount5.bind("<FocusOut>", lambda e: self.validateamount(billamt5))

        billtaxlist5 = tk.Combobox(master, values = atc, font = fonts, width = 10)
        billtaxlist5.grid(column = 2, row = 4)
        billtaxlist5.bind("<<ComboboxSelected>>", self.updatebilltotal)

        billclear5 = Button(master, text = "clear", font = clearbuttonfont, width = 5, command = lambda: self.clearbillitem(descentry5, billamount5, billtaxlist5))
        billclear5.grid(column = 3, row = 4, padx = 1)
        billclear5.bind("<Return>", lambda e: self.clearbillitem(descentry5, billamount5, billtaxlist5))

        alldesc.append(descentry5)
        allbillamt.append(billamt5)
        allbill.append(billamount5)
        alltax.append(billtaxlist5)
        ### line6 ###
        descentry6 = Entry(master, font = fonts, width = 60)
        descentry6.grid(column = 0, row = 5)

        billamt6 = StringVar()
        billamount6 = Entry(master, textvariable = billamt6, font = fonts, width = 15, justify = RIGHT)
        billamount6.grid(column = 1, row = 5, padx = 1)
        billamount6.bind("<FocusOut>", lambda e: self.validateamount(billamt6))

        billtaxlist6 = tk.Combobox(master, values = atc, font = fonts, width = 10)
        billtaxlist6.grid(column = 2, row = 5)
        billtaxlist6.bind("<<ComboboxSelected>>", self.updatebilltotal)

        billclear6 = Button(master, text = "clear", font = clearbuttonfont, width = 5, command = lambda: self.clearbillitem(descentry6, billamount6, billtaxlist6))
        billclear6.grid(column = 3, row = 5, padx = 1)
        billclear6.bind("<Return>", lambda e: self.clearbillitem(descentry6, billamount6, billtaxlist6))

        alldesc.append(descentry6)
        allbillamt.append(billamt6)
        allbill.append(billamount6)
        alltax.append(billtaxlist6)
        ### line7 ###
        descentry7 = Entry(master, font = fonts, width = 60)
        descentry7.grid(column = 0, row = 6)

        billamt7 = StringVar()
        billamount7 = Entry(master, textvariable = billamt7, font = fonts, width = 15, justify = RIGHT)
        billamount7.grid(column = 1, row = 6, padx = 1)
        billamount7.bind("<FocusOut>", lambda e: self.validateamount(billamt7))

        billtaxlist7 = tk.Combobox(master, values = atc, font = fonts, width = 10)
        billtaxlist7.grid(column = 2, row = 6)
        billtaxlist7.bind("<<ComboboxSelected>>", self.updatebilltotal)

        billclear7 = Button(master, text = "clear", font = clearbuttonfont, width = 5, command = lambda: self.clearbillitem(descentry7, billamount7, billtaxlist7))
        billclear7.grid(column = 3, row = 6, padx = 1)
        billclear7.bind("<Return>", lambda e: self.clearbillitem(descentry7, billamount7, billtaxlist7))

        alldesc.append(descentry7)
        allbillamt.append(billamt7)
        allbill.append(billamount7)
        alltax.append(billtaxlist7)

    def listclients1(self):
        select = "SELECT name FROM bcdclients WHERE company = ?"
        c.execute(select, [company])
        result = c.fetchall()
        clientlist = []
        for i in result:
            clientlist.append(i[0])
        clientlist.sort()
        cliententry1.config(values = list(dict.fromkeys(clientlist)))

    def listclients2(self, *args):
        select = "SELECT subgroup FROM bcdclients WHERE name = ? AND company = ?"
        c.execute(select, [cliententry1.get(), company])
        result = c.fetchall()
        sublist = []
        if result:
            for i in result:
                sublist.append(i[0])
            sublist.sort()
            cliententry2.config(values = list(dict.fromkeys(sublist)))
            cliententry2.delete(0, END)

        selectcat = "SELECT category FROM bcdclients WHERE name = ? AND company = ?"
        c.execute(selectcat, [cliententry1.get(), company])
        result = c.fetchall()
        if result:
            categoryentry2.delete(0, END)
            categoryentry2.insert(0, result[0])
        else:
            pass

    def showcollection(self, *args):
        collectionframe = LabelFrame(self.master, text = "\nCollection", font = fonts, bg = wc, fg = fc)
        collectionframe.grid(column = 1, row = 0, sticky = NW)

    def showbcdreports(self, *args):
        pass

    def checkbillingentries(self, *args):
        if len(dateentry.get()) == 10 and cliententry1.get() != "" and categoryentry1.get() in staffcategory and categoryentry2.get() in clientcategory and soaentry.get() != "" and periodentry.get() != "" and float(netamt.get().replace(",","")) != 0:
            decide = messagebox.askyesno("Submission", "Are you sure?")
            if decide == True:
                self.submitbilling()
        else:
            messagebox.showerror("Submission", "Please fill-out all required details!")

    def checkbillingentries2(self, *args):
        if len(dateentry.get()) == 10 and cliententry1.get() != "" and categoryentry1.get() in staffcategory and categoryentry2.get() in clientcategory and soaentry.get() != "" and periodentry.get() != "" and float(netamt.get().replace(",","")) != 0:
            decide = messagebox.askyesno("Update", "Are you sure?")
            if decide == True:
                self.updatebill()
        else:
            messagebox.showerror("Update", "Please fill-out all required details!")
        
    def submitbilling(self):
        self.showentry()
        if checkbox.get() == 1:
            status = "CANCELLED"
        else:
            status = "VALID"
        validbills = []
        for i in range(7):
            try:
                if float(allbill[i].get().replace(",","")) > 0:
                    validbills.append([
                        company, dateentry.get(), soaentry.get(), cliententry1.get(), cliententry2.get(),
                        categoryentry1.get(), categoryentry2.get(), periodentry.get(), remarksentry.get(), alldesc[i].get(),
                        float(allbill[i].get().replace(",","")), allvat[i], allewt[i], allfvat[i], allnet[i],
                        alltax[i].get(), accountcode[0], status, self.username.get(), today.strftime('%m-%d-%Y'),
                        today.strftime('%m-%d-%Y'), dateentry.get().split("-")[0], dateentry.get().split("-")[1], dateentry.get().split("-")[2]
                    ])
            except:
                pass
        submit = """INSERT INTO receivables (
            company, date, soa, client, subgroup,
            category1, category2, period, remarks, description,
            gross, vat, ewt, fvat, net,
            tax, account, status, user, created,
            modified, month, day, year)
            values (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
            """
        for i in validbills:
            c.execute(submit, i)
            conn.commit()
        messagebox.showinfo("Submission", "Billing has been saved!")

    def printbilling(self, *args):
        pass

    def findbill(self, *args):
        global findbillresult
        finder = "SELECT * FROM receivables WHERE soa LIKE ? AND company = ?"
        c.execute(finder, [findsoaentry.get(), company])
        findbillresult = c.fetchall()
        if findbillresult:
            self.showbilling()
            self.colorswitch(billing_button, buttonbg)
            cancelbox.grid(column = 2, row = 5)
            submitbutton.config(text = "Update", command = self.checkbillingentries2)
            submitbutton.bind("<Return>", self.checkbillingentries2)
            closebutton.config(command = lambda: billingframe.destroy())
            closebutton.bind("<Return>", lambda e: billingframe.destroy())
            dateentry.delete(0, END)
            dateentry.insert(0, findbillresult[0][1])
            cliententry1.insert(0, findbillresult[0][3])
            cliententry2.insert(0, findbillresult[0][4])
            categoryentry1.insert(0, findbillresult[0][5])
            categoryentry2.insert(0, findbillresult[0][6])
            soaentry.insert(0, findbillresult[0][2])
            periodentry.insert(0, findbillresult[0][7])
            remarksentry.insert(0, findbillresult[0][8])
            for i in range(len(findbillresult)):
                alldesc[i].insert(0, findbillresult[i][9])
                allbill[i].insert(0, findbillresult[i][10])
                alltax[i].insert(0, findbillresult[i][15])
                self.validateamount(allbillamt[i])
            if findbillresult[0][17] == "CANCELLED":
                checkbox.set(1)
                self.cancel()

    def updatebill(self, *args):
        soa = findbillresult[0][2]
        cre = findbillresult[0][19]
        use = findbillresult[0][18]
        canceller = "DELETE FROM receivables WHERE soa = ? AND company = ?"
        c.execute(canceller, [soa, company])
        conn.commit()
        self.showentry()
        if checkbox.get() == 1:
            status = "CANCELLED"
        else:
            status = "VALID"
        validbills = []
        for i in range(7):
            try:
                if float(allbill[i].get().replace(",","")) > 0:
                    validbills.append([
                        company, dateentry.get(), soa, cliententry1.get(), cliententry2.get(),
                        categoryentry1.get(), categoryentry2.get(), periodentry.get(), remarksentry.get(), alldesc[i].get(),
                        float(allbill[i].get().replace(",","")), allvat[i], allewt[i], allfvat[i], allnet[i],
                        alltax[i].get(), accountcode[0], status, use, cre,
                        today.strftime('%m-%d-%Y'), dateentry.get().split("-")[0], dateentry.get().split("-")[1], dateentry.get().split("-")[2]
                    ])
            except:
                pass
        submit = """INSERT INTO receivables (
            company, date, soa, client, subgroup,
            category1, category2, period, remarks, description,
            gross, vat, ewt, fvat, net,
            tax, account, status, user, created,
            modified, month, day, year)
            values (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
            """
        for i in validbills:
            c.execute(submit, i)
            conn.commit()
        messagebox.showinfo("Update", "Record has been updated!")

    def formatdate2(self, *args):
        global date_input
        date_input = dateentry.get()
        try:
            if len(date_input) == 10:
                try:
                    date_format = datetime.datetime.strptime(date_input, '%m-%d-%Y')
                except:
                    date_format = datetime.datetime.strptime(date_input, '%m/%d/%Y')
                finally:
                    dateentry.delete(0, END)
                    dateentry.insert(0, date_format.strftime('%m-%d-%Y'))
            else:
                messagebox.showerror("Date Checker", "Date entered is invalid!")
                dateentry.delete(0, END)
        except:
            dateentry.delete(0, END)

    def validateamount(self, amt, *args):
        amount = amt
        try:
            if amountentry.get() == "":
                amount.set(format(float(0), ',.2f'))
            else:
                try:
                    comma = format(float(amount.get()), ',.2f')
                    amount.set(comma)
                except:
                    amount.set(format(float(0), ',.2f'))
        except:
            amount.set(format(float(0), ',.2f'))
        self.updatebilltotal()

    def updatebilltotal(self, *args):
        global allgross, allvat, allewt, allfvat, allnet
        allgross, allvat, allewt, allfvat, allnet = [], [], [], [], []
        for i in range(7):
            try:
                allgross.append(float(allbill[i].get().replace(",","")))
                if categoryentry2.get() == "Private":
                    if alltax[i].get() == "WV-02":
                        allvat.append(float(allbill[i].get().replace(",",""))*.12)
                        allewt.append(float(allbill[i].get().replace(",",""))*.02)
                        allfvat.append(0)
                        allnet.append((float(allbill[i].get().replace(",","")))+(float(allbill[i].get().replace(",",""))*.12)-(float(allbill[i].get().replace(",",""))*.02))
                    elif alltax[i].get() == "NV-02":
                        allvat.append(0)
                        allewt.append(float(allbill[i].get().replace(",",""))*.02)
                        allfvat.append(0)
                        allnet.append((float(allbill[i].get().replace(",","")))-(float(allbill[i].get().replace(",",""))*.02))
                    elif alltax[i].get() == "NV-00":
                        allvat.append(0)
                        allewt.append(0)
                        allfvat.append(0)
                        allnet.append(float(allbill[i].get().replace(",","")))
                    elif alltax[i].get() == "Zero-Rated":
                        allvat.append(0)
                        allewt.append(0)
                        allfvat.append(0)
                        allnet.append(float(allbill[i].get().replace(",","")))
                    else:
                        allvat.append(0)
                        allewt.append(0)
                        allfvat.append(0)
                        allnet.append(float(allbill[i].get().replace(",","")))
                elif categoryentry2.get() == "Government":
                    if alltax[i].get() == "WV-02":
                        allvat.append(float(allbill[i].get().replace(",",""))*.12)
                        allewt.append(float(allbill[i].get().replace(",",""))*.02)
                        allfvat.append(float(allbill[i].get().replace(",",""))*.05)
                        allnet.append((float(allbill[i].get().replace(",","")))+(float(allbill[i].get().replace(",",""))*.12)-(float(allbill[i].get().replace(",",""))*.02)-(float(allbill[i].get().replace(",",""))*.05))
                    elif alltax[i].get() == "NV-02":
                        allvat.append(0)
                        allewt.append(float(allbill[i].get().replace(",",""))*.02)
                        allfvat.append(0)
                        allnet.append((float(allbill[i].get().replace(",","")))-(float(allbill[i].get().replace(",",""))*.02))
                    elif alltax[i].get() == "NV-00":
                        allvat.append(0)
                        allewt.append(0)
                        allfvat.append(0)
                        allnet.append(float(allbill[i].get().replace(",","")))
                    elif alltax[i].get() == "Zero-Rated":
                        allvat.append(0)
                        allewt.append(0)
                        allfvat.append(0)
                        allnet.append(float(allbill[i].get().replace(",","")))
                    else:
                        allvat.append(0)
                        allewt.append(0)
                        allfvat.append(0)
                        allnet.append(float(allbill[i].get().replace(",","")))
                else:
                    allvat.append(0)
                    allewt.append(0)
                    allfvat.append(0)
                    allnet.append(float(allbill[i].get().replace(",","")))
            except:
                allvat.append(0)
                allewt.append(0)
                allfvat.append(0)
                allnet.append(0)
        grossamtvar.set(format(sum(allgross), ",.2f"))
        vatamtvar.set(format(sum(allvat), ",.2f"))
        ewtamtvar.set(format(sum(allewt), ",.2f"))
        fvatamtvar.set(format(sum(allfvat), ",.2f"))
        netamtvar.set(format(sum(allnet), ",.2f"))

    def showentry(self):
        global accountcode
        select = "SELECT account FROM bcdclients WHERE name = ? AND company = ?"
        c.execute(select, [cliententry1.get(), company])
        result = c.fetchone()
        if result:
            if company == "DBPSC":
                chartselect = "SELECT code FROM chart WHERE code = ?"
            else:
                chartselect = "SELECT code FROM dssichart WHERE code = ?"
            c.execute(chartselect, [result[0]])
            accountcode = c.fetchone()
        else:
            pass
        
    def clearbillitem(self, desc, amt, tx, *args):
        desc.delete(0, END)
        amt.delete(0, END)
        tx.delete(0, END)
        self.updatebilltotal()

    def closebilling(self, *args):
        billingframe.destroy()
        self.bcdbuttons(NORMAL)
        self.colorswitch(billing_button, buttonbg)

    def closeviewbill(self, *args):
        viewbillframe.destroy()
        self.bcdbuttons(NORMAL)
        self.colorswitch(view_button, buttonbg)

    def bcdbuttons(self, status):
        view_button.config(state = status)
        billing_button.config(state = status)
        collection_button.config(state = status)
        reports_button.config(state = status)
        back.config(state = status)

    def bcdback(self, *args):
        bcd_frame.destroy()
        greeting.destroy()
        root.resizable(width = True, height = False)
        root.geometry(geometry)
        root.resizable(width = False, height = False)
        self.menu()
        if access[1] == "BCD":
            accounting_button.config(state = DISABLED)
            fmd_button.config(state = DISABLED)

### FMD factory ###
    def showfmd(self, *args):
        global atc, typelister, dmbanks, banks, modes, dmparticulars1, dmparticulars2, loiparticulars1, loiparticulars2
        atc = ['NV-00','NV-01','NV-02','NV-05','NV-10','NV-15','WV-01','WV-02','WV-05','V-12']
        typelister = ['13th MP','Abuloy','Allowances','Association dues','Bonus','CA','CA contractuals','Cashcard reversals','Communications','Consultancy fee','Deathclaims','Dividends','Donations','Financial assistance','FT','Gas allowance','Government','Gratuity pay','Hazard pay','HDMF SAI','Honorarium','Insurance','IT','Liquidation','Loan payments','Marketing expenses','Maxicare','Monetization','Others','Overtime pay','Paternity claims','PCF replenishment','Per diem','Pest Control','Photocopy','Quitclaims','Refund','Reimbursements','Rent','Repairs','Replacements','Retainer fee','Retirement pay','Salary','Seminars','Separation pay','Solo parent leave','SSS claims','Supplies','Tax refund','Triple H','Uniforms','Utilities','Vehicles']
        dmbanks = ['BDO Ara1 - 3970069089','BDO Ara2 - 3970093672','BDO - 418000153','BDO Ara5 - 3970069097','BDO Ara3 - 3970072187','DBP SA - 00-5-11281-405-8','DBP CA - 00-0-00832-405-0','DBP DAVAO - 915010136080','DBP DSSI - 00-0-06231-405-9']
        banks = ['BDO (469)','RCBC','BDO Ara1 (089)','BDO Ara2 (672)','BDO Stocks','BDO DSSI','LBP Buendia','LBP DSSI','DBP CA','DBP SA','DBP DSSI','BPI','BOC Petron','PNB Diliman','Metrobank','CA Fund']
        modes = ['Cash','Ck#','E-Banking','E-Transfer','MC']
        dmparticulars1 = 'This is to authorize RCBC Bank to debit our DBPSC Savings Account No. 9001735822 with RCBC Makati Avenue Branch the amount of '
        dmparticulars2 = 'Please credit said amount to his/her RCBC Makati Ave. Account No. '
        loiparticulars1 = 'This is to authorize DBP Bank to debit our DBPSC Security Service, Inc. with Current Account no. 00-0-06231-405-9 in the amount of '
        loiparticulars2 = 'Please credit said amount to his/her DBP Account No. '

        root.resizable(width = True, height = False)
        root.geometry(bcdgeometry)
        root.resizable(width = False, height = False)
        global fmd_frame, disbursement_button, viewer_button, finder_button, sms_button, bir_button, reports_button, back, logout_button, exit_button
        menu_frame.destroy()
        fmd_frame = LabelFrame(greeting, text = "Finance", font = fonts, bg = wc, fg = fc)
        fmd_frame.grid(column = 0, row = 1)

        button_frame = Frame(fmd_frame, bg = wc)
        button_frame.grid(column = 0, row = 11, pady = 10, sticky = N)

        disbursement_button = Button(fmd_frame, text = " Disbursements", font = fonts, bg = wc, fg = fc, bd = 0, cursor = "hand2", width = 240, image = disbursementicon, compound = LEFT, disabledforeground = fc, command = self.showdisbursements)
        disbursement_button.grid(column = 0, row = 0, pady = 5)
        disbursement_button.bind("<Return>", self.showdisbursements)
        disbursement_button.bind("<Enter>", lambda e: disbursement_button.config(bg = fc, fg = wc))
        disbursement_button.bind("<Leave>", lambda e: disbursement_button.config(bg = wc, fg = fc))

        viewer_button = Button(fmd_frame, text = " View/Update   ", font = fonts, bg = wc, fg = fc, bd = 0, cursor = "hand2", width = 240, image = compilericon, compound = LEFT, disabledforeground = fc, command = self.showcompiler)
        viewer_button.grid(column = 0, row = 1, pady = 5)
        viewer_button.bind("<Return>", self.showcompiler)
        viewer_button.bind("<Enter>", lambda e: viewer_button.config(bg = fc, fg = wc))
        viewer_button.bind("<Leave>", lambda e: viewer_button.config(bg = wc, fg = fc))

        # claims_button = Button(fmd_frame, text = "Quitclaims", font = fonts, bg = buttonbg, width = 30, command = self.showquitclaims)
        # claims_button.grid(column = 0, row = 2, pady = 5)
        # claims_button.bind("<Return>", self.showquitclaims)

        finder_button = Button(fmd_frame, text = " Finder              ", font = fonts, bg = wc, fg = fc, bd = 0, cursor = "hand2", width = 240, image = findicon, compound = LEFT, disabledforeground = fc, command = self.showfmdfinder)
        finder_button.grid(column = 0, row = 2, pady = 5)
        finder_button.bind("<Return>", self.showfmdfinder)
        finder_button.bind("<Enter>", lambda e: finder_button.config(bg = fc, fg = wc))
        finder_button.bind("<Leave>", lambda e: finder_button.config(bg = wc, fg = fc))

        sms_button = Button(fmd_frame, text = " Text Master (BETA)      ", font = fonts, bg = wc, fg = fc, bd = 0, cursor = "hand2", width = 240, image = texticon, compound = LEFT, disabledforeground = fc, command = self.showtextmaster, anchor = E)
        sms_button.grid(column = 0, row = 3, pady = 5)
        sms_button.bind("<Return>", self.showtextmaster)
        sms_button.bind("<Enter>", lambda e: sms_button.config(bg = fc, fg = wc))
        sms_button.bind("<Leave>", lambda e: sms_button.config(bg = wc, fg = fc))

        reports_button = Button(fmd_frame, text = " Reports          ", font = fonts, bg = wc, fg = fc, bd = 0, cursor = "hand2", width = 240, image = reporticon, compound = LEFT, disabledforeground = fc, command = self.showfmdreports)
        reports_button.grid(column = 0, row = 4, pady = 5)
        reports_button.bind("<Return>", self.showfmdreports)
        reports_button.bind("<Enter>", lambda e: reports_button.config(bg = fc, fg = wc))
        reports_button.bind("<Leave>", lambda e: reports_button.config(bg = wc, fg = fc))

        back = Button(button_frame, font = iconfont, bg = wc, fg = fc, bd = 0, cursor = "hand2", image = homeicon, compound = LEFT, command = self.fmdback)
        back.grid(column = 0, row = 0, pady = 5, padx = 10)
        back.bind("<Return>", self.fmdback)
        back.bind("<Enter>", lambda e: back.config(text = "Home"))
        back.bind("<Leave>", lambda e: back.config(text = ""))

        logout_button = Button(button_frame, font = iconfont, bg = wc, fg = fc, bd = 0, cursor = "hand2", image = logouticon, compound = LEFT, command = self.logout)
        logout_button.grid(column = 1, row = 0, pady = 5, padx = 10)
        logout_button.bind("<Return>", self.logout)
        logout_button.bind("<Enter>", lambda e: logout_button.config(text = "Logout"))
        logout_button.bind("<Leave>", lambda e: logout_button.config(text = ""))

        exit_button = Button(button_frame, font = iconfont, bg = wc, fg = fc, bd = 0, cursor = "hand2", image = exiticon, compound = LEFT, command = self.exit)
        exit_button.grid(column = 2, row = 0, pady = 5, padx = 10)
        exit_button.bind("<Return>", self.exit)
        exit_button.bind("<Enter>", lambda e: exit_button.config(text = "Exit"))
        exit_button.bind("<Leave>", lambda e: exit_button.config(text = ""))

    ### quitclaims ###
    def showquitclaims(self, *args):
        global claimsframe
        self.fmdmenubuttons(DISABLED)
        self.colorswitch(claims_button, click)
        claimsframe = LabelFrame(self.master, text = '\nQuitclaims', font = fonts, bg = wc, fg = fc)
        claimsframe.grid(column = 1, row = 0, sticky = NW)

        claims_subframe = Frame(claimsframe, bg = wc)
        claims_subframe.grid(column = 0, row = 0, sticky = NW, pady = pad)

        button_frame = Frame(claimsframe, bg = wc)
        button_frame.grid(column = 0, row = 1, sticky = E, pady = pad)

        close_button = Button(button_frame, text = "Close", font = fonts, width = 10, command = lambda: self.closefmdmenu(claimsframe, claims_button))
        close_button.grid(column = 2, row = 0, padx = pad)
        close_button.bind("<Return>", lambda e: self.closefmdmenu(claimsframe, claims_button))

    ### text master ###
    def showtextmaster(self, *args):
        global smsframe, smssubframe, listframe, buttonframe, statusframe, smsoption, importbutton, sendbutton
        self.fmdmenubuttons(DISABLED)
        self.colorswitchfmd(sms_button, fc)
        smsframe = LabelFrame(self.master, text = '\nText Master (BETA)', font = fonts, bg = wc, fg = fc)
        smsframe.grid(column = 1, row = 0, sticky = NW)

        smssubframe = Frame(smsframe, bg = wc)
        smssubframe.grid(column = 0, row = 0, sticky = NW)

        smsoption = Frame(smsframe, bg = wc)
        smsoption.grid(column = 0, row = 1, sticky = NW)

        buttonframe = Frame(smsframe, bg = wc)
        buttonframe.grid(column = 0, row = 2, sticky = E)

        listframe = Frame(smsframe, bg = wc)
        listframe.grid(column = 0, row = 3, sticky = NW)

        statusframe = Frame(smsframe, bg = wc)
        statusframe.grid(column = 0, row = 4, sticky = E)

        text_button = Button(smssubframe, text = "Quick Send", font = fonts, bg = wc, bd = 0, width = 95,image = buttonicon, compound = CENTER, cursor = "hand2", command = self.showquicksend)
        text_button.grid(column = 0, row = 0, padx = pad, pady = pad)
        text_button.bind("<Return>", self.showquicksend)

        textmanual_button = Button(smssubframe, text = "Manual Send", font = fonts, bg = wc, bd = 0, width = 95,image = buttonicon, compound = CENTER, cursor = "hand2", command = self.showquicksendmanual)
        textmanual_button.grid(column = 1, row = 0, padx = pad, pady = pad)
        textmanual_button.bind("<Return>", self.showquicksendmanual)

        sent_button = Button(smssubframe, text = "Sent Items", font = fonts, bg = wc, bd = 0, width = 95,image = buttonicon, compound = CENTER, cursor = "hand2", command = self.showsentitems)
        sent_button.grid(column = 3, row = 0, padx = pad, pady = pad)
        sent_button.bind("<Return>", self.showsentitems)

        payreg_button = Button(smssubframe, text = "PayRegister", font = fonts, bg = wc, bd = 0, width = 95,image = buttonicon, compound = CENTER, cursor = "hand2", command = self.showpayrollregister)
        payreg_button.grid(column = 4, row = 0, padx = pad, pady = pad)
        payreg_button.bind("<Return>", self.showpayrollregister)

        phonebook_button = Button(smssubframe, text = "Phone Book", font = fonts, bg = wc, bd = 0, width = 95,image = buttonicon, compound = CENTER, cursor = "hand2", state = DISABLED, command = self.showphonebook)
        phonebook_button.grid(column = 5, row = 0, padx = pad, pady = pad)
        phonebook_button.bind("<Return>", self.showphonebook)

        sendbutton = Button(buttonframe, text = "Send", font = fonts, bg = wc, bd = 0, width = 95, image = buttonicon, compound = CENTER, cursor = "hand2", command = self.sendexcelfile)
        sendbutton.bind("<Return>", self.sendexcelfile)

        importbutton = Button(buttonframe, text = "Import", font = fonts, bg = wc, bd = 0, width = 95, image = buttonicon, compound = CENTER, cursor = "hand2", command = self.importexcelfile)
        importbutton.bind("<Return>", self.importexcelfile)

        closebutton = Button(buttonframe, text = "Close", font = fonts, bg = wc, bd = 0, width = 95, image = buttonicon, compound = CENTER, cursor = "hand2", command = lambda: self.closefmdmenu(smsframe, sms_button))
        closebutton.grid(column = 2, row = 0, padx = pad, pady = pad)
        closebutton.bind("<Return>", lambda e: self.closefmdmenu(smsframe, sms_button))

    def showpayrollregister(self, *args):
        global listframe, excelvar
        optionframe = LabelFrame(smsoption, text = "PayRegister", font = fonts, bg = wc, fg = fc)
        optionframe.grid(column = 0, row = 0, sticky = NW)

        importlistframe = Frame(optionframe, bg = wc)
        importlistframe.grid(column = 0, row = 0, sticky = NW, pady = pad)

        listframe = Frame(optionframe, bg = wc)
        listframe.grid(column = 0, row = 1, sticky = NW, pady = pad)

        importlabel = Label(importlistframe, text = "Excel File", font = fonts, width = 15, bg = wc, fg = fc)
        importlabel.grid(column = 0, row = 0, padx = pad, pady = pad)

        excelvar = StringVar()
        excelfile = Entry(importlistframe, textvariable = excelvar, font = fonts, width = 60)
        excelfile.grid(column = 1, row = 0, padx = pad)

        import_button = Button(importlistframe, text = "...", font = fonts, bg = wc, bd = 0, width = 95, image = buttonicon, compound = CENTER, cursor = "hand2", command = self.openexcelfile)
        import_button.grid(column = 2, row = 0, padx = pad)
        import_button.bind("<Return>", self.openexcelfile)

        importbutton.config(text = "Convert", command = self.convertpayreg)
        importbutton.bind("<Return>", self.convertpayreg)

    def convertpayreg(self, *args):
        global payreg
        wb = openpyxl.load_workbook(filename)
        st = wb.active
        wb2 = openpyxl.Workbook()
        st2 = wb2.active
        st2.append(["sender","particulars","receiver","number","reference no.","amount"])
        payreg = []
        try:
            for i in range(st.max_row):
                if len(str(st["B" + str(i+1)].value).split(",")) == 2:
                    particulars = "SALARY FROM DBPSC"
                    receiver = st["B" + str(i+1)].value
                    number = self.trimphonenumber(st["K" + str(i+1)].value)
                    amount = st["I" + str(i+1)].value
                    payreg.append(["", particulars, receiver, number, "", amount])
                    st2.append(["", particulars, receiver, number, "", amount])
            wb2.save(savepath + "converted.xlsx")
            os.startfile(savepath + "converted.xlsx", "open")
        except Exception as e:
            print(e)
            messagebox.showerror("Convert", "Error converting excel file!")

    def trimphonenumber(self, num):
        old = str(num).replace("/","").replace("-","").strip()
        if len(old) != 11:
            if old[0:2] == "00":
                new = old[1:12]
            elif old[0:2] == "09":
                new = old[0:11]
            else:
                new = old
        else:
            new = old
        return new

    def showquicksendmanual(self, *args):
        self.smsbuttons(DISABLED)
        global optionframe, detailsframe, android_connection, connection, sendmanualbutton, sender, number, receiver, particulars, reference, referencevar, principal, principalvar
        optionframe = LabelFrame(smsoption, text = "Manual Send", font = fonts, bg = wc, fg = fc)
        optionframe.grid(column = 0, row = 0, sticky = NW)

        androidconnectframe = Frame(optionframe, bg = wc)
        androidconnectframe.grid(column = 0, row = 0, sticky = NW, pady = pad)

        detailsframe = Frame(optionframe, bg = wc)
        detailsframe.grid(column = 0, row = 1, sticky = NW, pady = pad)

        # android connection
        connection = StringVar()
        android_connection = Label(androidconnectframe, textvariable = connection, font = boldfonts, width = 15, bg = offlinebg, fg = fc)
        android_connection.grid(column = 0, row = 0, padx = pad, pady = pad)
        connection.set("Offline")

        get_connection = Button(androidconnectframe, text = "Connect", font = fonts, bg = wc, bd = 0, width = 95, image = buttonicon, compound = CENTER, cursor = "hand2", command = self.adbconnect)
        get_connection.grid(column = 2, row = 0, padx = pad)
        get_connection.bind("<Return>", self.adbconnect)
        # message body
        senderlabel = Label(detailsframe, text = "Sender", font = fonts, bg = wc, fg = fc, width = 10, anchor = W)
        senderlabel.grid(column = 0, row = 0, padx = pad)

        numberlabel = Label(detailsframe, text = "Number", font = fonts, bg = wc, fg = fc, width = 10, anchor = W)
        numberlabel.grid(column = 0, row = 1, padx = pad)

        receiverlabel = Label(detailsframe, text = "Receiver", font = fonts, bg = wc, fg = fc, width = 10, anchor = W)
        receiverlabel.grid(column = 0, row = 2, padx = pad)

        particularslabel = Label(detailsframe, text = "Particulars", font = fonts, bg = wc, fg = fc, width = 10, anchor = W)
        particularslabel.grid(column = 0, row = 3, padx = pad)

        referencelabel = Label(detailsframe, text = "KPTN", font = fonts, bg = wc, fg = fc, width = 10, anchor = W)
        referencelabel.grid(column = 0, row = 4, padx = pad)

        amountlabel = Label(detailsframe, text = "Amount", font = fonts, bg = wc, fg = fc, width = 10, anchor = W)
        amountlabel.grid(column = 0, row = 5, padx = pad)

        senders = ["HANZEL R. CLEMENTE", "JHAY L. BANGALAN", "JULIUS BRYAN M. PEÑAFLORIDA"]
        sender = tk.Combobox(detailsframe, values = senders, font = fonts, width = 30)
        sender.grid(column = 1, row = 0, padx = pad, sticky = W)

        number = Entry(detailsframe, font = fonts, width = 15)
        number.grid(column = 1, row = 1, padx = pad, sticky = W)
        number.bind("<FocusOut>", lambda e: self.checkphonenumber(number))

        receiver = Entry(detailsframe, font = fonts, width = 32)
        receiver.grid(column = 1, row = 2, padx = pad, sticky = W)

        particulars = Entry(detailsframe, font = fonts, width = 32)
        particulars.grid(column = 1, row = 3, padx = pad, sticky = W)

        referencevar = StringVar()
        reference = Entry(detailsframe, textvariable = referencevar, font = fonts, width = 32)
        reference.grid(column = 1, row = 4, padx = pad, sticky = W)
        reference.bind("<FocusOut>", lambda e: self.formatreference(referencevar))

        principalvar = StringVar()
        principal = Entry(detailsframe, textvariable = principalvar, font = fonts, width = 15, justify = RIGHT)
        principal.grid(column = 1, row = 5, padx = pad, sticky = W)
        principal.bind("<FocusOut>", lambda e: self.amountvalidatormaster(principalvar))

        sendmanualbutton = Button(buttonframe, text = "Send", font = fonts, bg = wc, bd = 0, width = 95, image = buttonicon, compound = CENTER, cursor = "hand2", state = DISABLED, command = self.sendtextmanual)
        sendmanualbutton.grid(column = 0, row = 0, padx = pad)
        sendmanualbutton.bind("<Return>", self.sendtextmanual)

        refreshbutton = Button(buttonframe, text = "Refresh", font = fonts, bg = wc, bd = 0, width = 95, image = buttonicon, compound = CENTER, cursor = "hand2", command = self.refreshsendmanual)
        refreshbutton.grid(column = 1, row = 0, padx = pad)
        refreshbutton.bind("<Return>", self.refreshsendmanual)

    def sendtextmanual(self, *args):
        self.adbconnect()
        now = str(today.strftime('%m-%d-%Y'))
        tab = "cmd /c adb shell input keyevent 22"
        send = "cmd /c adb shell input keyevent 66"
        sendmanual = "INSERT INTO messaging (sender, particulars, receiver, number, reference, amount, date, status, user) VALUES (?,?,?,?,?,?,?,?,?)"
        if connection.get() == "Online":
            try:
                message = f"{particulars.get()}, Sender: {sender.get()}, KPTN: {reference.get()}, Amount: P{principalvar.get()}"
                # smsService.send_message(i[3], f"{i[1]}, Sender: {i[0]}, Receiver: {i[2]}, Reference #: {i[4]}, Amount: P{i[5]}")
                os.system(f"cmd /c adb shell am start -a android.intent.action.SENDTO -d sms:{number.get()} --es sms_body '{message}' --ez exit_on_sent true")
                os.system(tab)
                os.system(send)
                c.execute(sendmanual, [sender.get(), particulars.get(), receiver.get(), number.get(), reference.get(), float(principalvar.get().replace(",","")), now, "sent", self.username.get()])
                conn.commit()
                messagebox.showinfo("Send", "Message sent!")
                sendmanualbutton.config(state = DISABLED)
                for widget in detailsframe.winfo_children():
                    widget.config(state = DISABLED)
            except:
                c.execute(sendmanual, [sender.get(), particulars.get(), receiver.get(), number.get(), reference.get(), float(principalvar.get().replace(",","")), now, "failed", self.username.get()])
                conn.commit()
                messagebox.showinfo("Send", "Message sending failed!")

    def showquicksend(self, *args):
        self.smsbuttons(DISABLED)
        global android_connection, connection, excelvar
        optionframe = LabelFrame(smsoption, text = "Quick Send", font = fonts, bg = wc, fg = fc)
        optionframe.grid(column = 0, row = 0, sticky = NW)

        androidconnectframe = Frame(optionframe, bg = wc)
        androidconnectframe.grid(column = 0, row = 0, sticky = NW)

        importlistframe = Frame(optionframe, bg = wc)
        importlistframe.grid(column = 0, row = 1, sticky = NW)

        connection = StringVar()
        android_connection = Label(androidconnectframe, textvariable = connection, font = boldfonts, width = 15, bg = offlinebg, fg = fc)
        android_connection.grid(column = 0, row = 0, padx = pad, pady = pad)
        connection.set("Offline")

        # this code block is for emergency texting - Airmore
        # ipaddress_entry = Entry(androidconnectframe, font = fonts, fg = "grey", width = 15, justify = CENTER)
        # ipaddress_entry.grid(column = 1, row = 0, padx = pad)
        # ipaddress_entry.bind("<Return>", self.connectairmore)
        # ipaddress_entry.bind("<FocusIn>", lambda e: self.ipaddressfocus(ipaddress_entry))
        # ipaddress_entry.bind("<FocusOut>", lambda e: self.ipaddressfocus(ipaddress_entry))
        # ipaddress_entry.insert(0, "IP Address...")

        get_connection = Button(androidconnectframe, text = "Connect", font = fonts, bg = wc, bd = 0, width = 95, image = buttonicon, compound = CENTER, cursor = "hand2", command = self.adbconnect)
        get_connection.grid(column = 2, row = 0, padx = pad)
        get_connection.bind("<Return>", self.adbconnect)

        importlabel = Label(importlistframe, text = "Excel File", font = fonts, width = 15, bg = wc, fg = fc)
        importlabel.grid(column = 0, row = 0, padx = pad, pady = pad)

        excelvar = StringVar()
        excelfile = Entry(importlistframe, textvariable = excelvar, font = fonts, width = 60)
        excelfile.grid(column = 1, row = 0, padx = pad)

        importbutton = Button(importlistframe, text = "...", font = fonts, bg = wc, bd = 0, width = 95, image = buttonicon, compound = CENTER, cursor = "hand2", command = self.openexcelfile)
        importbutton.grid(column = 2, row = 0, padx = pad)
        importbutton.bind("<Return>", self.openexcelfile)

        templatebutton = Button(importlistframe, text = "Get template", font = fonts, bg = wc, fg = fc, bd = 0, width = 10, cursor = "hand2", command = self.showtemplate)
        templatebutton.grid(column = 3, row = 0, padx = pad)
        templatebutton.bind("<Return>", self.showtemplate)
        templatebutton.bind("<Enter>", lambda e: templatebutton.config(font = underlinefont))
        templatebutton.bind("<Leave>", lambda e: templatebutton.config(font = fonts))

        helpbutton = Button(importlistframe, text = "Help", font = fonts, bg = wc, fg = fc, bd = 0, width = 10, cursor = "hand2", command = self.showhelp)
        helpbutton.grid(column = 4, row = 0, padx = pad)
        helpbutton.bind("<Return>", self.showhelp)
        helpbutton.bind("<Enter>", lambda e: helpbutton.config(font = underlinefont))
        helpbutton.bind("<Leave>", lambda e: helpbutton.config(font = fonts))

    def showhelp(self, *args):
        messagebox.showinfo("Text Master - CHARLIE System", """
        ###### HOW TO USE TEXT MASTER - Four (4) Easy Steps! #######\n\n
        1. Connect your android device to your PC using USB cable.\n
        2. On your device, go to "Settings" > "Developer Options" then enable "USB Debugging".\n
        3. On Text Master, click "Connect" to establish connection between Android and PC.\n
        4. Import your excel file then click "Send".
        """)

        # "Text Master - CHARLIE System", """
        # ###### HOW TO USE TEXT MASTER IN THE OFFICE - Seven (7) Easy Steps! #######\n\n
        # Assuming:\n
        # ---You have a laptop and two (2) android devices (namely "Android-1" and "Android-2").\n
        # ---You have already downloaded and installed "Airmore" application from Google Playstore on Android-1 (primary device for texting).\n
        # ---You have already encoded your ToSendList in an excel file.\n
        # ---You have the latest stand-alone Text Master - CHARLIE System application on your laptop (get a copy from MJ if you don't have one)\n\n
        # 1. On Android-2, turn on Wifi Hotspot. (No Mobile Data needed)\n
        # 2. Connect both Android-1 and laptop to the Wifi Hotspot established by Android-2.\n
        # 3. On Android-1, open Airmore application on your device, go to "More" and under the "Other tools" section click "Hotspot" (signal tower symbol), then remember the IP address (e.g. "192.168.00.00" - do not include colon and last 4 digits ":0000").\n
        # 4. On Text Master Application, enter the IP address and click "Connect", then wait until it goes "Online".\n
        # 5. On Text Master Application, "Import" your excel file ("ToSendList.xlsx") then click "Send".\n
        # 6. On your device, you will be prompted to accept the connection. Click "Accept".\n\n
        # 7. Your ToSendList will now be sent automatically and a pop-up message on your laptop will appear if it is done.\n
        # """

    def showtemplate(self, *args):
        wb = openpyxl.Workbook()
        st = wb.active
        st.append(["sender","particulars","receiver","number","reference no.","amount"])
        wb.save(savepath + "template.xlsx")
        os.startfile(savepath + "template.xlsx", "open")

    def sendexcelfile(self, *args):
        self.adbconnect()
        tab = "cmd /c adb shell input keyevent 22"
        send = "cmd /c adb shell input keyevent 66"
        sender = "INSERT INTO messaging (sender, particulars, receiver, number, reference, amount, date, status, user) VALUES (?,?,?,?,?,?,?,?,?)"
        sent = []
        failed = []
        now = str(today.strftime('%m-%d-%Y'))
        if connection.get() == "Online":
            ask = messagebox.askyesno("Send", "Are you sure?")
            if ask == True:
                row = 2
                count = 0
                st["G1"] = "status"
                st["H1"] = "date sent"
                for i in sendlist:
                    if i[3] != "invalid":
                        try:
                            message = f"{i[1]}, Sender: {i[0]}, KPTN: {i[4]}, Amount: P{i[5]}"
                            # smsService.send_message(i[3], f"{i[1]}, Sender: {i[0]}, Receiver: {i[2]}, Reference #: {i[4]}, Amount: P{i[5]}")
                            os.system(f"cmd /c adb shell am start -a android.intent.action.SENDTO -d sms:{i[3]} --es sms_body '{message}' --ez exit_on_sent true")
                            os.system(tab)
                            os.system(send)
                            st["G" + str(row)] = "sent"
                            st["H" + str(row)] = now
                            sent.append(1)
                            c.execute(sender, [i[0], i[1], i[2], i[3], i[4], float(i[5].replace(",","")), now, "sent", self.username.get()])
                            conn.commit()
                        except:
                            st["G" + str(row)] = "failed"
                            st["H" + str(row)] = now
                            failed.append(1)
                            c.execute(sender, [i[0], i[1], i[2], i[3], i[4], float(i[5].replace(",","")), now, "failed", self.username.get()])
                            conn.commit()
                    else:
                        st["G" + str(row)] = "failed"
                        st["H" + str(row)] = now
                        failed.append(1)
                        c.execute(sender, [i[0], i[1], i[2], i[3], i[4], float(i[5].replace(",","")), now, "failed", self.username.get()])
                        conn.commit()
                    row += 1
                    count += 1
                    print(round((count/len(sendlist))*100, 0))
                    progress["value"] = round((count/len(sendlist))*100, 0)
                    progress.update_idletasks()
                messagebox.showinfo("Send", f"Sending completed!\nSent: {len(sent)}\nFailed: {len(failed)}\nTotal Records: {len(sent)+len(failed)}")
                wb.save(filename)
                os.startfile(filename, "open")
        else:
            messagebox.showerror("Send", "Sending failed due to lost connection!")

    def importexcelfile(self, *args):
        global sendlist, wb, st
        wb = openpyxl.load_workbook(filename)
        st = wb.active
        self.itemlabels()
        self.itemscroller(listframe, 1)
        sendlist = []
        try:
            for i in range(st.max_row-1):
                sender = st["A" + str(i+2)].value
                particulars = st["B" + str(i+2)].value
                receiver = st["C" + str(i+2)].value
                number = st["D" + str(i+2)].value
                reference = st["E" + str(i+2)].value
                amount = st["F" + str(i+2)].value
                if len(self.validatephonenumber(number)) == 11:
                    sendlist.append([sender.upper(), particulars, receiver.upper(), self.validatephonenumber(number), self.quickformatreference(reference), self.amountformatmaster(amount)])
                    itembg = codetitlebg
                else:
                    sendlist.append([sender.upper(), particulars, receiver.upper(), "invalid", self.quickformatreference(reference), self.amountformatmaster(amount)])
                    itembg = cancelleddocument
                item = Frame(scrollable_frame, bg = wc)
                item.grid(column = 0, row = i)

                sender_entry = Entry(item, font = fonts, bg = itembg, width = 15, relief = RIDGE, bd = 1)
                sender_entry.grid(column = 0, row = 0)
                sender_entry.insert(0, sender.upper())

                particulars_entry = Entry(item, font = fonts, bg = itembg, width = 15, relief = RIDGE, bd = 1)
                particulars_entry.grid(column = 1, row = 0)
                particulars_entry.insert(0, particulars)

                receiver_entry = Entry(item, font = fonts, bg = itembg, width = 26, relief = RIDGE, bd = 1)
                receiver_entry.grid(column = 2, row = 0)
                receiver_entry.insert(0, receiver.upper())

                number_entry = Entry(item, font = fonts, bg = itembg, width = 11, relief = RIDGE, bd = 1)
                number_entry.grid(column = 3, row = 0)
                number_entry.insert(0, self.validatephonenumber(number))

                reference_entry = Entry(item, font = fonts, bg = itembg, width = 23, relief = RIDGE, bd = 1, justify = RIGHT)
                reference_entry.grid(column = 4, row = 0)
                reference_entry.insert(0, self.quickformatreference(reference))

                amount_entry = Entry(item, font = fonts, bg = itembg, width = 15, relief = RIDGE, bd = 1, justify = RIGHT)
                amount_entry.grid(column = 5, row = 0)
                amount_entry.insert(0, self.amountformatmaster(amount))

            sendbutton.grid(column = 0, row = 0, padx = pad, pady = pad)
            self.showrecords(sendlist)
        except:
            messagebox.showerror("Import", "Please use proper template!")

    def showrecords(self, count):
        global progress
        records_frame = Frame(statusframe, bg = wc)
        records_frame.grid(column = 0, row = 0, sticky = E, padx = 10)

        total_label = Label(records_frame, text = "Total Records", font = fonts, width = 15, bg = wc, fg = fc)
        total_label.grid(column = 0, row = 0, sticky = E)

        count_label = Label(records_frame, text = str(len(count)), font = fonts, width = 10, bg = wc, fg = fc, relief = RIDGE)
        count_label.grid(column = 1, row = 0, sticky = E, padx = pad)

        progress = tk.Progressbar(statusframe, orient = HORIZONTAL, length = 885, mode = "determinate")
        progress.grid(column = 0, row = 1, padx = 10, sticky = W)

    def itemlabels(self):
        global senderlabel
        labelframe = Frame(listframe, bg = wc)
        labelframe.grid(column = 0, row = 0, sticky = NW)

        senderlabel = Label(labelframe, text = "Sender", font = fonts, bg = wc, fg = fc, width = 15, relief = RIDGE)
        senderlabel.grid(column = 1, row = 0)

        particularslabel = Label(labelframe, text = "Particulars", font = fonts, bg = wc, fg = fc, width = 15, relief = RIDGE)
        particularslabel.grid(column = 2, row = 0)

        receiverlabel = Label(labelframe, text = "Receiver", font = fonts, bg = wc, fg = fc, width = 25, relief = RIDGE)
        receiverlabel.grid(column = 3, row = 0)

        numberlabel = Label(labelframe, text = "Number", font = fonts, bg = wc, fg = fc, width = 11, relief = RIDGE)
        numberlabel.grid(column = 4, row = 0)

        referencelabel = Label(labelframe, text = "Reference no.", font = fonts, bg = wc, fg = fc, width = 23, relief = RIDGE)
        referencelabel.grid(column = 5, row = 0)

        amountlabel = Label(labelframe, text = "Amount", font = fonts, bg = wc, fg = fc, width = 15, relief = RIDGE)
        amountlabel.grid(column = 6, row = 0)

    def openexcelfile(self, *args):
        global filename
        filename = askopenfilename(filetypes = (("Excel files", "*.xlsx"),("All files", "*.*")))
        if filename:
            excelvar.set(filename)
            importbutton.grid(column = 1, row = 0, padx = pad, pady = pad)
        else:
            excelvar.set("")
            importbutton.grid_forget()

    def showsentitems(self, *args):
        self.smsbuttons(DISABLED)
        global listframe
        optionframe = LabelFrame(smsoption, text = "Sent Items", font = fonts, bg = wc, fg = fc)
        optionframe.grid(column = 0, row = 0, sticky = NW)

        subframe = Frame(optionframe, bg = wc)
        subframe.grid(column = 0, row = 0, sticky = NW)

        listframe =  Frame(optionframe, bg = wc)
        listframe.grid(column = 0, row = 1, sticky = NW)

        scrollframe = Frame(optionframe, bg = wc)
        scrollframe.grid(column = 0, row = 2, sticky = NW)

        self.itemlabels()
        senderlabel.config(text = "Date Sent")
        self.itemscroller(scrollframe, 0)
        c.execute("SELECT date, particulars, receiver, number, reference, amount, status FROM messaging ORDER BY receiver")
        result = c.fetchall()
        row = 0
        for i in result:
            if i[6] == "sent":
                itembg = codetitlebg
            else:
                itembg = cancelleddocument
            item = Frame(scrollable_frame, bg = wc)
            item.grid(column = 0, row = row)

            date_entry = Entry(item, font = fonts, bg = itembg, width = 15, relief = RIDGE, bd = 1)
            date_entry.grid(column = 0, row = 0)
            date_entry.insert(0, i[0])

            particulars_entry = Entry(item, font = fonts, bg = itembg, width = 15, relief = RIDGE, bd = 1)
            particulars_entry.grid(column = 1, row = 0)
            particulars_entry.insert(0, i[1])

            receiver_entry = Entry(item, font = fonts, bg = itembg, width = 26, relief = RIDGE, bd = 1)
            receiver_entry.grid(column = 2, row = 0)
            receiver_entry.insert(0, i[2].upper())

            number_entry = Entry(item, font = fonts, bg = itembg, width = 11, relief = RIDGE, bd = 1)
            number_entry.grid(column = 3, row = 0)
            number_entry.insert(0, self.validatephonenumber(i[3]))

            reference_entry = Entry(item, font = fonts, bg = itembg, width = 23, relief = RIDGE, bd = 1, justify = RIGHT)
            reference_entry.grid(column = 4, row = 0)
            reference_entry.insert(0, i[4])

            amount_entry = Entry(item, font = fonts, bg = itembg, width = 15, relief = RIDGE, bd = 1, justify = RIGHT)
            amount_entry.grid(column = 5, row = 0)
            amount_entry.insert(0, self.amountformatmaster(i[5]))

            row += 1

    def showphonebook(self, *args):
        self.smsbuttons(DISABLED)

    def smsbuttons(self, status):
        for widget in smssubframe.winfo_children():
            widget.config(state = status)

    def connectairmore(self, *args): ### option 2 ###
        global androidIP, androidSession
        if ipaddress_entry.get() != "":
            try:
                androidIP = IPv4Address(ipaddress_entry.get())
                androidSession = AirmoreSession(androidIP)
                if androidSession.is_server_running == True:
                    messagebox.showinfo("Android Connection", "Your android is now connected!")
                    connection.set("Online")
                    android_connection.config(bg = click)
                else:
                    messagebox.showerror("Android Connection", "Your android is not connected! Please open your Airmore application.")
                    connection.set("Offline")
                    android_connection.config(bg = offlinebg)
            except:
                messagebox.showerror("Android Connection", "IP address is not valid")
                connection.set("Offline")
                android_connection.config(bg = offlinebg)
                ipaddress_entry.delete(0, END)
                ipaddress_entry.insert(0, "IP Address...")
        else:
            messagebox.showerror("Android Connection", "Open Airmore application on your android and get the IP address!")
            connection.set("Offline")
            android_connection.config(bg = offlinebg)

    def adbconnect(self, *args): ### option 1 ###
        if os.system(f"cmd /c adb shell am start") == 1:
        # if os.system("cmd /c adb shell") == 1:
            connection.set("Offline")
            android_connection.config(bg = offlinebg)
        else:
            connection.set("Online")
            android_connection.config(bg = click)

    def ipaddressfocus(self, ip):
        if ip.get() == "IP Address...":
            ip.delete(0, END)
            ip.config(fg = codetitlefg)
        elif ip.get() == "":
            ip.config(fg = "grey")
            ip.insert(0, "IP Address...")

    def itemscroller(self, master, rw):
        global scrollbox, scrollable_frame, canvas, scrollbar
        scrollbox = Frame(master)
        container = Frame(scrollbox)
        canvas = Canvas(container, bg = wc)
        scrollbar = Scrollbar(container, orient = "vertical", width = 25, command = canvas.yview)
        scrollable_frame = Frame(canvas)
        scrollable_frame.bind("<Configure>", lambda e: canvas.configure(scrollregion = canvas.bbox("all")))
        canvas.create_window((0,0), window = scrollable_frame, anchor = "nw")
        canvas.configure(yscrollcommand = scrollbar.set)

        scrollbox.grid(column = 0, row = rw, sticky = NW)
        container.pack()
        canvas.pack(side = "left", fill = "both", expand = True)
        canvas.config(height = 295, width = 870)
        scrollbar.pack(side = "right", fill = "y")

    def amountformatmaster(self, amt):
        amount = amt
        try:
            if amount == 0:
                return str(format(float(0), ',.2f'))
            else:
                try:
                    comma = format(float(amount), ',.2f')
                    return str(comma)
                except:
                    return str(format(float(0), ',.2f'))
        except:
            return str(format(float(0), ',.2f'))

    def validatephonenumber(self, num):
        try:
            if num > 0:
                return "0" + str(num)
        except:
            return num

    def checkphonenumber(self, num):
        if len(num.get()) != 11:
            num.delete(0, END)
            sendmanualbutton.config(state = DISABLED)
        else:
            sendmanualbutton.config(state = NORMAL)
            self.searchphonebook(num.get())

    def searchphonebook(self, num):
        find = "SELECT receiver FROM messaging WHERE number = ?"
        c.execute(find, [num])
        result = c.fetchone()
        if result:
            receiver.delete(0, END)
            receiver.insert(0, result[0])
        else:
            receiver.delete(0, END)

    def refreshsendmanual(self, *args):
        optionframe.destroy()
        self.showquicksendmanual()

    def formatreference(self, ref, *args):
        kptn = ref.get().replace(" ","")
        if len(kptn) == 18:
            ref.set(f"{kptn[0:4]} {kptn[4:7]} {kptn[7:11]} {kptn[11:14]} {kptn[14:18]}")
        else:
            ref.set("")

    def quickformatreference(self, ref):
        kptn = str(ref).replace(" ","")
        if len(kptn) == 18:
            return f"{kptn[0:4]} {kptn[4:7]} {kptn[7:11]} {kptn[11:14]} {kptn[14:18]}"
        else:
            return kptn

    ### compiler ###
    def showcompiler(self, *args):
        global compilerframe, compilersubframe, numberentry, buttonframe, checkbox, cancelbox, detailsbutton, printbutton
        self.fmdmenubuttons(DISABLED)
        self.colorswitchfmd(viewer_button, fc)
        compilerframe = LabelFrame(self.master, text = '\nView/Update', font = fonts, bg = wc, fg = fc)
        compilerframe.grid(column = 1, row = 0, sticky = NW)

        compilersubframe = Frame(compilerframe, bg = wc)
        compilersubframe.grid(column = 0, row = 0, sticky = NW)

        numberlabel = Label(compilersubframe, text = 'Number', font = fonts, bg = wc, fg = fc)
        numberlabel.grid(column = 0, row = 0, sticky = W, pady = pad)

        numberentry = Entry(compilersubframe, font = fonts, width = 10)
        numberentry.grid(column = 1, row = 0, padx = 16, sticky = W)
        numberentry.bind('<Return>', self.fmdfinder)

        findbutton = Button(compilersubframe, text = 'View', font = fonts, bg = wc, bd = 0, width = 95, image = buttonicon, compound = CENTER, cursor = "hand2", command = self.fmdfinder)
        findbutton.grid(column = 2, row = 0, padx = 10, sticky = NW)
        findbutton.bind('<Return>', self.fmdfinder)

        checkbox = IntVar() 
        cancelbox = Checkbutton(compilersubframe, text = 'Cancel', variable = checkbox, bg = wc, fg = fc, font = fonts, command = self.cancel)
        cancelbox.bind('<Button-1>', self.cancel)

        buttonframe = Frame(compilerframe, bg = wc)
        buttonframe.grid(column = 0, row = 3, sticky = E)

        printbutton = Button(buttonframe, text = 'Print', font = fonts, bg = wc, bd = 0, width = 95, image = buttonicon, compound = CENTER, cursor = "hand2", command = self.printdoc)
        printbutton.bind('<Return>', self.printdoc)

        detailsbutton = Button(buttonframe, text = 'Details', font = fonts, bg = wc, bd = 0, width = 95, image = buttonicon, compound = CENTER, cursor = "hand2", command = self.adddetails)
        detailsbutton.bind('<Return>', self.adddetails)

        closebutton = Button(buttonframe, text = 'Close', font = fonts, bg = wc, bd = 0, width = 95, image = buttonicon, compound = CENTER, cursor = "hand2", command = lambda: self.closefmdmenu(compilerframe, viewer_button))
        closebutton.grid(column = 3, row = 0, padx = pad, pady = pad)
        closebutton.bind('<Return>', lambda e: self.closefmdmenu(compilerframe, viewer_button))

        numberentry.focus()

    def fmdfinder(self, *args):
        global result
        finder = 'SELECT * FROM disbursements WHERE number = ? AND company = ?'
        c.execute(finder, [numberentry.get(), company])
        result = c.fetchall()
        if result:
            try:
                detailsframe.destroy()
                linecontent.destroy()
                contentframe.destroy()
            except:
                pass
            finally:
                try:
                    updatebutton.grid_forget()
                    printbutton.grid_forget()
                    cancelbox.grid_forget()
                except:
                    pass
                self.showdetails()
                doc.set(result[0][2])
                com.set(result[0][0])
                dat.set(result[0][1])
                typ.set(result[0][4])
                par.set(result[0][6].split('_')[0])
                pre.set(result[0][21])
                if len(result[0][6].split('_')) == 2:
                    sup.set(result[0][5].split('_')[0])
                else:
                    sup.set(result[0][5])
                self.showitems()
                if result[0][20] == 'VALID':
                    checkbox.set(0)
                else:
                    checkbox.set(1)
                self.cancel()
        else:
            messagebox.showerror('Compiler Protocol', 'Document Number is not in records!')
            try:
                detailsframe.destroy()
                lineframes.destroy()
                contentframe.destroy()
            except:
                pass
            try:
                updatebutton.grid_forget()
                printbutton.grid_forget()
                detailsbutton.grid_forget()
                cancelbox.grid_forget()
            except:
                pass

    def showdetails(self):
        global updatebutton, detailsframe, preparedlabel, payeelabel, particularslabel, documententry, companyentry, dateentry, typeentry, supplierlist, particularsentry, preparedentry, signatoryentry1, signatoryentry2, pre, doc, com, dat, typ, sup, par
        detailsframe = Frame(compilerframe, bg = wc)
        detailsframe.grid(column = 0, row = 1, sticky = NW)
        
        documentlabel = Label(detailsframe, text = 'Document', font = fonts, bg = wc, fg = fc)
        documentlabel.grid(column = 0, row = 0, sticky = W, pady = pad)

        companylabel = Label(detailsframe, text = 'Company', font = fonts, bg = wc, fg = fc)
        companylabel.grid(column = 0, row = 1, sticky = W, pady = pad)

        datelabel = Label(detailsframe, text = 'Date', font = fonts, bg = wc, fg = fc)
        datelabel.grid(column = 0, row = 2, sticky = W, pady = pad)
        
        typelabel = Label(detailsframe, text = 'Type', font = fonts, bg = wc, fg = fc)
        typelabel.grid(column = 0, row = 3, sticky = W, pady = pad)

        payeelabel = Label(detailsframe, text = 'Payee', font = fonts, bg = wc, fg = fc)
        payeelabel.grid(column = 2, row = 0, sticky = W, pady = pad)

        particularslabel = Label(detailsframe, text = 'Particulars', font = fonts, bg = wc, fg = fc)
        particularslabel.grid(column = 2, row = 1, sticky = W, pady = pad)

        preparedlabel = Label(detailsframe, text = 'Prepared by', font = fonts, bg = wc, fg = fc)
        preparedlabel.grid(column = 2, row = 2, sticky = W, pady = pad)
        
        doc = StringVar()
        documententry = Entry(detailsframe, textvariable = doc, font = fonts, width = 10)
        documententry.grid(column = 1, row = 0, sticky = W)

        com = StringVar()
        companyentry = tk.Combobox(detailsframe, textvariable = com, values = ['DBPSC','DSSI'], font = fonts, width = 8)
        companyentry.grid(column = 1, row = 1, sticky = W)
        companyentry.set(company)
        companyentry.config(state = DISABLED)

        dat = StringVar()
        dateentry = Entry(detailsframe, textvariable = dat, font = fonts, width = 10)
        dateentry.grid(column = 1, row = 2, sticky = W)

        typ = StringVar()
        typeentry = tk.Combobox(detailsframe, textvariable = typ, values = typelister, font = fonts, width = 15)
        typeentry.grid(column = 1, row = 3, sticky = W)
        typeentry.bind('<<ComboboxSelected>>', self.chosendmtxn)
        typeentry.bind('<FocusOut>', self.intypelist)

        sup = StringVar()
        supplierlist = tk.Combobox(detailsframe, textvariable = sup, font = fonts, width = 58)
        supplierlist.grid(column = 3, row = 0, sticky = W)

        par = StringVar()
        particularsentry = tk.Combobox(detailsframe, textvariable = par, font = fonts, width = 58)
        particularsentry.grid(column = 3, row = 1, sticky = W)

        pre = StringVar()
        preparedentry = Entry(detailsframe, textvariable = pre, font = fonts, width = 60, state = DISABLED)
        preparedentry.grid(column = 3, row = 2, sticky = W)

        signatoryentry1 = tk.Combobox(detailsframe, font = fonts, width = 59)
        signatoryentry2 = tk.Combobox(detailsframe, font = fonts, width = 59)

        updatebutton = Button(buttonframe, text = 'Update', font = fonts, bg = wc, bd = 0, width = 95, image = buttonicon, compound = CENTER, cursor = "hand2", command = self.updatecompiler)
        updatebutton.bind('<Return>', self.updatecompiler)

        # button2307 = Button(buttonframe, text = '2307', font = fonts, width = 10, command = self.generate2307)
        # button2307.bind('<Return>', self.generate2307)

        self.merchantlister(supplierlist)

    def showitems(self):
        global lineframes, linecontent, amountentry, amount, assignmententry, accountentry, assignmentlabel, accountlabel, amountlabel, payeeentry, bankentry, destinationentry
        if result[0][2] == 'General Voucher':
            lineframes = Frame(compilerframe, bg = wc)
            lineframes.grid(column = 0, row = 2, sticky = NW)

            linecontent = Frame(lineframes, bg = wc)
            linecontent.grid(column = 0, row = 0, sticky = NW)

            signatoryentry1.insert(0, result[0][7])
            signatoryentry2.insert(0, result[0][8])

            self.showlabels(linecontent, 0, 0)
            self.showlines(linecontent, 0, 1)
            self.showtotals(linecontent, 0, 2)
            if len(result) == 1:
                allitems[0].insert(0, result[0][9])
                allbanks[0].insert(0, result[0][10])
                allmodes[0].insert(0, result[0][11])
                allamounts[0].set(result[0][12])
                alltaxes[0].delete(0, END)
                alltaxes[0].insert(0, result[0][13])
                self.gvamountvalidator(allamounts[0])
                self.computegvtax()
            else:
                for i in range(len(result)):
                    allitems[i].insert(0, result[i][9])
                    allbanks[i].insert(0, result[i][10])
                    allmodes[i].insert(0, result[i][11])
                    allamounts[i].set(result[i][12])
                    alltaxes[i].delete(0, END)
                    alltaxes[i].insert(0, result[i][13])
                    self.gvamountvalidator(allamounts[i])
            self.computegvtax()
            detailsbutton.grid(column = 2, row = 0, padx = pad)
        else:
            amount = StringVar()
            amountentry = Entry(detailsframe, textvariable = amount, font = fonts, width = 15, justify = RIGHT)
            try:
                detailsbutton.grid_forget()
            except:
                pass
            if typ.get() != 'FT':
                try:
                    lineframes.destroy()
                except:
                    pass
                lineframes = Frame(compilerframe, bg = wc)
                lineframes.grid(column = 0, row = 2, sticky = NW)

                self.showdmcontent()
                payeeentry = supplierlist

                assignmentlabel = Label(detailsframe, text = 'Assignment', font = fonts, bg = wc, fg = fc)
                assignmentlabel.grid(column = 2, row = 2, sticky = W, pady = pad)

                accountlabel = Label(detailsframe, text = 'Account No.', font = fonts, bg = wc, fg = fc)
                accountlabel.grid(column = 2, row = 3, sticky = W, pady = pad)

                amountlabel = Label(detailsframe, text = 'Amount', font = fonts, bg = wc, fg = fc)
                amountlabel.grid(column = 2, row = 4, sticky = W, pady = pad)

                assignmententry = Entry(detailsframe, font = fonts, width = 60)
                assignmententry.grid(column = 3, row = 2, sticky = W)
                assignmententry.bind('<FocusOut>', self.dmassembly)
                
                accountentry = Entry(detailsframe, font = fonts, width = 60)
                accountentry.grid(column = 3, row = 3, sticky = W)
                accountentry.bind('<FocusOut>', self.dmassembly)
                try:
                    if len(result[0][6].split('_')) == 2:
                        assignmententry.insert(0, result[0][5].split('_')[1])
                        accountentry.insert(0, result[0][6].split('_')[1])
                        sup.set(result[0][5].split('_')[0])
                    else:
                        assignmententry.insert(0, result[0][5])
                        sup.set('')
                except:
                    pass
                finally:
                    preparedlabel.grid(column = 2, row = 5, sticky = W)
                    preparedentry.grid(column = 3, row = 5, sticky = W)

                    amountentry.bind('<FocusOut>', lambda e: self.amountvalidatormaster(amount))
                    amountentry.grid(column = 3, row = 4, sticky = W)
                    amount.set(result[0][14])
                    self.amountvalidatormaster(amount)
                    self.dmassembly()
            else:
                try:
                    lineframes.destroy()
                    payeelabel.destroy()
                    particularslabel.destroy()
                    particularsentry.destroy()
                except:
                    pass
                finally:
                    try:
                        supplierlist.destroy()
                    except:
                        pass
                    banklabel = Label(detailsframe, text = 'Bank', font = fonts, bg = wc, fg = fc)
                    banklabel.grid(column = 0, row = 4, sticky = W, pady = pad)

                    amountlabel = Label(detailsframe, text = 'Amount', font = fonts, bg = wc, fg = fc)
                    amountlabel.grid(column = 2, row = 0, sticky = W, pady = pad)

                    destinationlabel = Label(detailsframe, text = 'Destination', font = fonts, bg = wc, fg = fc)
                    destinationlabel.grid(column = 2, row = 1, sticky = W, pady = pad)

                    bankentry = tk.Combobox(detailsframe, values = dmbanks, font = fonts, width = 15)
                    bankentry.grid(column = 1, row = 4, sticky = W)
                    bankentry.bind('<<ComboboxSelected>>', self.destbank)
                    bankentry.insert(0, result[0][10])

                    destinationentry = tk.Combobox(detailsframe, font = fonts, width = 58   )
                    destinationentry.grid(column = 3, row = 1, sticky = W)
                    destinationentry.insert(0, result[0][6].split('to ')[1])

                    preparedlabel.grid(column = 2, row = 2, sticky = W)
                    preparedentry.grid(column = 3, row = 2, sticky = W)

                    amountentry.bind('<FocusOut>', lambda e: self.amountvalidatormaster(amount))
                    amountentry.grid(column = 3, row = 0, sticky = W)
                    amount.set(result[0][12])
                    self.amountvalidatormaster(amount)

        cancelbox.grid(column = 3, row = 0, sticky = NW, padx = 20)
        printbutton.grid(column = 1, row = 0, padx = pad)
        updatebutton.grid(column = 0, row = 0, padx = pad)

    def updatecompiler(self, *args):
        global sig1, sig2
        deleter = 'DELETE FROM disbursements WHERE number = ?'
        if documententry.get() == 'General Voucher':
            sig1 = result[0][7]
            sig2 = result[0][8]
        else:
            pass
        c.execute(deleter, [result[0][3]])
        conn.commit()
        self.updatedoc()

    def updatedoc(self, *args):
        if documententry.get() == 'General Voucher':
            if numberentry.get() == '':
                messagebox.showerror('Posting Protocol','General Voucher number is missing!')
            elif dateentry.get() == '':
                messagebox.showerror('Posting Protocol','Date is missing!')
            elif typeentry.get() == '':
                messagebox.showerror('Posting Protocol','Type is missing!')
            elif supplierlist.get() == '':
                messagebox.showerror('Posting Protocol','Payee is missing!')
            elif float(total.get().replace(',','')) == 0:
                messagebox.showerror('Posting Protocol','Please check amounts!')
            else:
                ask = messagebox.askyesno('Update Document', 'Are you sure?')
                if ask == True:
                    if checkbox.get() == 1:
                        status = 'CANCELLED'
                    else:
                        status = 'VALID'
                    validitems = []
                    for i in range(5):
                        try:
                            if float(allnets[i].get().replace(',','')) != 0 and allnets[i].get() != '':
                                validitems.append([
                                    com.get(), dat.get(), doc.get(), numberentry.get(), typ.get(),
                                    sup.get(), par.get(), sig1, sig2, allitems[i].get(),
                                    allbanks[i].get(), allmodes[i].get(), float(allamounts[i].get().replace(',','')), alltaxes[i].get(), float(allnets[i].get().replace(',','')),
                                    round(allewt[i],2), round(allvat[i],2), dateentry.get().split('-')[0], dateentry.get().split('-')[1], dateentry.get().split('-')[2],
                                    status, pre.get()
                                ])
                        except:
                            pass
                    inserter = """INSERT INTO disbursements (
                        company, date, document, number, type,
                        payee, particulars, signatory1, signatory2, item,
                        bank, mode, gross, tax, net,
                        ewt, vat, month, day, year,
                        status, user) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
                        """
                    for v in validitems:
                        c.execute(inserter, v)
                        conn.commit()
                        self.disabler()
                    messagebox.showinfo('Updating Protocol', 'Disbursement record has been updated!')
        else:
            if dateentry.get() == '':
                messagebox.showerror('Updating Protocol','Date is missing!')
            elif typeentry.get() == '':
                messagebox.showerror('Posting Protocol','Type is missing!')
            elif float(amount.get().replace(',','')) == 0 or amount.get() == '':
                messagebox.showerror('Updating Protocol','Please check amounts!')
            else:
                ask = messagebox.askyesno('Update Document', 'Are you sure?')
                if ask == True:
                    if checkbox.get() == 1:
                        status = 'CANCELLED'
                    else:
                        status = 'VALID'
                    defaultmode = 'DM'
                    defaulttax = 'NV-00'
                    if typeentry.get() == 'FT' or typeentry.get() == 'Fund transfer' or typeentry.get() == 'Fund Transfer' or typeentry.get() == 'FUND TRANSFER' or typeentry.get() == 'ft':
                        inserter = """INSERT INTO disbursements (
                            company, date, document, number, type,
                            payee, particulars, bank, mode, gross,
                            tax, net, month, day, year,
                            status, user) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
                            """
                        defaultpayee = 'DBPSC'
                        defaultparticulars = 'Fund transfer to '+ destinationentry.get()
                        values = [
                            com.get(), dat.get(), doc.get(), numberentry.get(), typ.get(),
                            defaultpayee,defaultparticulars,bankentry.get().split(' - ')[0],defaultmode,float(amount.get().replace(',','')),
                            defaulttax,float(amount.get().replace(',','')),dateentry.get().split('-')[0],dateentry.get().split('-')[1],dateentry.get().split('-')[2],
                            status, pre.get()
                            ]
                    else:
                        inserter = """INSERT INTO disbursements (
                            company, date, document, number, type,
                            payee, particulars, bank, mode, gross,
                            tax, net, month, day, year,
                            status, user) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
                            """
                        defaultbank = 'RCBC'
                        if accountentry.get() != '':
                            fullparticulars = particularsentry.get() + '_' + accountentry.get()
                            payeeassign = payeeentry.get()
                        else:
                            fullparticulars = particularsentry.get()
                            payeeassign = assignmententry.get()
                        values = [
                            com.get(), dat.get(), doc.get(), numberentry.get(), typ.get(),
                            payeeassign,fullparticulars,defaultbank,defaultmode,float(amount.get().replace(',','')),
                            defaulttax,float(amount.get().replace(',','')),dateentry.get().split('-')[0],dateentry.get().split('-')[1],dateentry.get().split('-')[2],
                            status,pre.get()
                            ]
                    c.execute(inserter, values)
                    conn.commit()
                    self.disabler()
                    messagebox.showinfo('Updating Protocol', 'Disbursement record has been updated!')
                    updatebutton.config(state = DISABLED)
                else:
                    pass

    def adddetails(self, *args):
        global top, lineframe, numb, comp, dater, savebutton, updatebutton, printbutton, tinput, totalamount
        top = Toplevel()
        top.title(maintitle)
        top.iconbitmap(iconpath + "icon.ico")
        top.geometry('960x590+300+0')
        top.config(bg = wc)
        top.resizable(height = False, width = False)
        top.grab_set()

        detailssubframe = Frame(top, bg = wc)
        detailssubframe.grid(column = 0, row = 0, sticky = NW, pady = pad)

        labelframe = Frame(top, bg = wc)
        labelframe.grid(column = 0, row = 1, sticky = NW, pady = pad)

        lineframe = Frame(top, bg = wc)
        lineframe.grid(column = 0, row = 2, sticky = NW)

        buttonframe = Frame(top, bg = wc)
        buttonframe.grid(column = 0, row = 3, sticky = E, pady = pad)

        comlabel = Label(detailssubframe, text = 'Company', font = fonts, width = 14, bg = wc, fg = fc)
        comlabel.grid(column = 0, row = 0, sticky = NW)

        datlabel = Label(detailssubframe, text = 'Date', font = fonts, width = 14, bg = wc, fg = fc)
        datlabel.grid(column = 0, row = 1, sticky = NW)

        numlabel = Label(detailssubframe, text = 'Number', font = fonts, width = 14, bg = wc, fg = fc)
        numlabel.grid(column = 0, row = 2, sticky = NW)

        comp = StringVar()
        comentry = Entry(detailssubframe, textvariable = comp, font = fonts, state = DISABLED, disabledbackground = codetitlebg, disabledforeground = codetitlefg, width = 14)
        comentry.grid(column = 1, row = 0, sticky = E)
        comp.set(companyentry.get())

        dater = StringVar()
        datentry = Entry(detailssubframe, textvariable = dater, font = fonts, state = DISABLED, disabledbackground = codetitlebg, disabledforeground = codetitlefg, width = 14)
        datentry.grid(column = 1, row = 1, sticky = E)
        dater.set(dateentry.get())

        numb = StringVar()
        numentry = Entry(detailssubframe, textvariable = numb, font = fonts, state = DISABLED, disabledbackground = codetitlebg, disabledforeground = codetitlefg, width = 14)
        numentry.grid(column = 1, row = 2, sticky = E)
        numb.set(numberentry.get())

        addbutton = Button(detailssubframe, font = fonts, bg = wc,  width = 30, image = add, cursor = "hand2", command = self.addcheque)
        addbutton.grid(column = 1, row = 8, pady = pad)
        addbutton.bind("<Return>", self.addcheque)

        lessbutton = Button(detailssubframe, font = fonts, bg = wc, width = 30, image = less, cursor = "hand2", command = self.lesscheque)
        lessbutton.grid(column = 2, row = 8, pady = pad)
        lessbutton.bind("<Return>", self.lesscheque)

        tinput = StringVar()
        totalamount = Entry(top, textvariable = tinput, font = fonts, justify = RIGHT, state = DISABLED, disabledbackground = codetitlebg, disabledforeground = codetitlefg, width = 15)
        totalamount.place(x = 555, y = 98)

        paylabel = Label(labelframe, text = 'Payee', font = fonts, width = 40, bg = wc, fg = fc, relief = RIDGE)
        paylabel.grid(column = 0, row = 0)

        banlabel = Label(labelframe, text = 'Bank', font = fonts, width = 10, bg = wc, fg = fc, relief = RIDGE)
        banlabel.grid(column = 1, row = 0)

        modlabel = Label(labelframe, text = 'Check #', font = fonts, width = 17, bg = wc, fg = fc, relief = RIDGE)
        modlabel.grid(column = 2, row = 0)

        amtlabel = Label(labelframe, text = 'Amount', font = fonts, width = 15, bg = wc, fg = fc, relief = RIDGE)
        amtlabel.grid(column = 3, row = 0)
        
        stalabel = Label(labelframe, text = 'Status', font = fonts, width = 10, bg = wc, fg = fc, relief = RIDGE)
        stalabel.grid(column = 4, row = 0)

        remlabel = Label(labelframe, text = 'Remarks', font = fonts, width = 15, bg = wc, fg = fc, relief = RIDGE)
        remlabel.grid(column = 5, row = 0)

        boxlabel = Label(labelframe, text = " □ ", font = fonts, width = 4, bg = wc, fg = fc, relief = RIDGE)
        boxlabel.grid(column = 6, row = 0)

        global chequeframes, chequeitems, chequebanks, chequemodes, chequeamounts, chequestatus, chequeremarks, chequeboxes
        chequeframes, chequeitems, chequebanks, chequemodes, chequeamounts, chequestatus, chequeremarks, chequeboxes = [], [], [], [], [], [], [] ,[]
        self.itemscroller(lineframe, 0)
        canvas.config(height = 400, width = 930)
        self.chequeitem(scrollable_frame, 0)
        
        savebutton = Button(buttonframe, text = 'Save', font = fonts, bg = wc, bd = 0, width = 95, image = buttonicon, compound = CENTER, cursor = "hand2", command = self.savedetails)
        savebutton.grid(column = 0, row = 0, padx = pad)
        savebutton.bind('<Return>', self.savedetails)

        updatebutton = Button(buttonframe, text = 'Update', font = fonts, bg = wc, bd = 0, width = 95, image = buttonicon, compound = CENTER, cursor = "hand2", state = DISABLED, command = self.updatedetails)
        updatebutton.grid(column = 1, row = 0, padx = pad)
        updatebutton.bind('<Return>', self.updatedetails)

        printbutton = Button(buttonframe, text = 'Export', font = fonts, bg = wc, bd = 0, width = 95, image = buttonicon, compound = CENTER, cursor = "hand2", state = DISABLED, command = self.printdetails)
        printbutton.grid(column = 2, row = 0, padx = pad)
        printbutton.bind('<Return>', self.printdetails)

        closebutton = Button(buttonframe, text = 'Close', font = fonts, bg = wc, bd = 0, width = 95, image = buttonicon, compound = CENTER, cursor = "hand2", command = self.closedetails)
        closebutton.grid(column = 3, row = 0, padx = pad)
        closebutton.bind('<Return>', self.closedetails)

        self.updatechecker()
        self.totalupdater()

    def savedetails(self, *args):
        ask = messagebox.askyesno('Save Details', 'Are you sure?')
        if ask == True:
            validdetails = []
            for i in range(len(chequeframes)):
                try:
                    if float(chequeamounts[i].get().replace(",","")) > 0 and chequeitems[i].get() != '' and chequebanks[i].get() != '' and chequemodes[i].get() != '' and chequestatus[i].get() != '':
                        validdetails.append([comp.get(),dater.get(),numb.get(),chequeitems[i].get(),chequebanks[i].get(),chequemodes[i].get(),float(chequeamounts[i].get().replace(",","")),chequestatus[i].get(),chequeremarks[i].get()])
                except:
                    pass
            saver = """INSERT INTO details (
                company, date, number, payee, bank, mode, amount, status, remarks)
                VALUES (?,?,?,?,?,?,?,?,?)
                """
            if len(validdetails) != 0:
                for v in validdetails:
                    c.execute(saver, v)
                    conn.commit()
                messagebox.showinfo('Save Details', 'Details have been saved!')
                savebutton.config(state = DISABLED)
                updatebutton.config(state = NORMAL)
            else:
                messagebox.showerror('Save Details', 'Please check details!')
        else:
            pass

    def printdetails(self, *args):
        wb = openpyxl.load_workbook(path + "report.xlsx") #path
        st = wb.active
        st.append([comp.get()])
        st.append(["GV Details"])
        st.append([f"{numb.get()}_{dater.get()}"])
        st.append(["Payee","Bank","Mode","Amount","Status","Remarks"])
        for i in range(len(chequeframes)):
            if chequeamounts[i].get() != "":
                st.append([
                    chequeitems[i].get(),chequebanks[i].get(),chequemodes[i].get(),
                    float(chequeamounts[i].get().replace(",","")),chequestatus[i].get(),chequeremarks[i].get()
                    ])
        wb.save(savepath + "details.xlsx") #path
        os.startfile(savepath + "details.xlsx", "open") #path

    def chequeitem(self, master, rw):
        item = Frame(master, bg = wc)
        item.grid(column = 0, row = rw)
        chequeframes.append(item)

        itementry = Entry(item, font = fonts, width = 40)
        itementry.grid(column = 0, row = 0, padx = 1)
        itementry.bind("<FocusOut>", lambda e: self.uppercase(itementry))
        chequeitems.append(itementry)

        bankentry = tk.Combobox(item, values = banks, font = fonts, width = 8)
        bankentry.grid(column = 1, row = 0)
        chequebanks.append(bankentry)

        modeentry = Entry(item, font = fonts, width = 17)
        modeentry.grid(column = 2, row = 0)
        chequemodes.append(modeentry)

        chequeamt = StringVar()
        amountentry = Entry(item, textvariable = chequeamt, font = fonts, width = 15, justify = RIGHT)
        amountentry.grid(column = 3, row = 0)
        amountentry.bind('<FocusOut>', lambda e: self.chequetotalmaster(chequeamt))
        chequeamounts.append(chequeamt)
        
        statusentry = tk.Combobox(item, values = ['Valid','Cancelled','Replaced'], font = fonts, width = 8)
        statusentry.grid(column = 4, row = 0)
        chequestatus.append(statusentry)

        remarksentry = Entry(item, font = fonts, width = 15)
        remarksentry.grid(column = 5, row = 0, padx = 1)
        chequeremarks.append(remarksentry)

        checker = IntVar()
        checkbox = Checkbutton(item, variable = checker, bd = 0, font = fonts, bg = wc)
        checkbox.grid(column = 6, row = 0, padx = 10)
        chequeboxes.append(checker)

    def addcheque(self, *args):
        itemrow = len(chequeitems) + 1
        self.chequeitem(scrollable_frame, itemrow)

    def lesscheque(self, *args):
        for i in chequeboxes:
            if i.get() == 1:
                chequeframes[chequeboxes.index(i)].destroy()

    def totalupdater(self, *args):
        totalcatcher = []
        for i in chequeamounts:
            try:
                if float(i.get().replace(",","")) > 0:
                    totalcatcher.append(float(i.get().replace(",","")))
            except:
                pass
        tinput.set(format(sum(totalcatcher), ',.2f'))

    def updatedetails(self, *args):
        ask = messagebox.askyesno('Update Details', 'Are you sure?')
        if ask == True:
            deleter = 'DELETE FROM details WHERE company = ? AND number = ? AND date = ?'
            c.execute(deleter, [comp.get(),numb.get(),dater.get()])
            conn.commit()
            validdetails = []
            for i in range(len(chequeframes)):
                try:
                    if float(chequeamounts[i].get().replace(",","")) > 0 and chequeitems[i].get() != '' and chequebanks[i].get() != '' and chequemodes[i].get() != '' and chequestatus[i].get() != '':
                        validdetails.append([comp.get(),dater.get(),numb.get(),chequeitems[i].get(),chequebanks[i].get(),chequemodes[i].get(),float(chequeamounts[i].get().replace(",","")),chequestatus[i].get(),chequeremarks[i].get()])
                except:
                    pass
            saver = """INSERT INTO details (
                company, date, number, payee, bank, mode, amount, status, remarks)
                VALUES (?,?,?,?,?,?,?,?,?)
                """
            if len(validdetails) != 0:
                for v in validdetails:
                    c.execute(saver, v)
                    conn.commit()
                messagebox.showinfo('Update Details', 'Details have been updated!')
            else:
                messagebox.showerror('Update Details', 'Please check details!')

    def updatechecker(self):
        checker = 'SELECT * FROM details WHERE number = ?'
        c.execute(checker, [numb.get()])
        result = c.fetchall()
        if result:
            updatebutton.config(state = NORMAL)
            printbutton.config(state = NORMAL)
            savebutton.config(state = DISABLED)
            i = 0
            for x in result:
                chequeitems[i].insert(0, x[3])
                chequebanks[i].insert(0, x[4])
                chequemodes[i].insert(0, x[5])
                chequeamounts[i].set(x[6])
                chequestatus[i].insert(0, x[7])
                chequeremarks[i].insert(0, x[8])
                self.chequetotalmaster(chequeamounts[i])
                i += 1
                self.addcheque()
        else:
            updatebutton.config(state = DISABLED)
            savebutton.config(state = NORMAL)

    def closedetails(self, *args):
        top.grab_release()
        top.destroy()

    def chequetotalmaster(self, amt, *args):
        self.amountvalidatormaster(amt)
        chequetotal = []
        for i in chequeamounts:
            try:
                chequetotal.append(float(i.get().replace(",","")))
            except:
                chequetotal.append(0)
        tinput.set(format(sum(chequetotal), ",.2f"))

    ### disbursements ###
    def showdisbursements(self, *args):
        global disbursementsframe, disbursementssubframe, lineframes, buttonframe, documententry, checkbox, cancelbox, submitbutton, refreshbutton, printbutton, closebutton
        self.fmdmenubuttons(DISABLED)
        self.colorswitchfmd(disbursement_button, fc)
        disbursementsframe = LabelFrame(self.master, text = '\nDisbursements', font = fonts, bg = wc, fg = fc)
        disbursementsframe.grid(column = 1, row = 0, sticky = NW)

        documentframe = Frame(disbursementsframe, bg = wc)
        documentframe.grid(column = 0, row = 0, sticky = NW, pady = pad)

        disbursementssubframe = Frame(disbursementsframe, bg = wc)
        disbursementssubframe.grid(column = 0, row = 1, sticky = NW, pady = pad)

        lineframes = Frame(disbursementsframe, bg = wc)
        lineframes.grid(column = 0, row = 2, sticky = NW, pady = pad)

        buttonframe = Frame(disbursementsframe, bg = wc)
        buttonframe.grid(column = 0, row = 3, sticky = E, pady = pad)

        documentlabel = Label(documentframe, text = 'Document', font = fonts, bg = wc, fg = fc)
        documentlabel.grid(column = 0, row = 0, sticky = W, padx = pad)

        documententry = tk.Combobox(documentframe, values = ['General Voucher','Debit Memo','DM - DBP Davao','Letter of Instruction'], font = fonts, width = 15)
        documententry.grid(column = 1, row = 0, padx = pad, sticky = W)
        documententry.bind('<<ComboboxSelected>>', self.showdocument)

        checkbox = IntVar() 
        cancelbox = Checkbutton(documentframe, text = 'Cancel', variable = checkbox, bg = wc, fg = fc, font = fonts, command = self.cancel)
        cancelbox.bind('<Button-1>', self.cancel)

        submitbutton = Button(buttonframe, text = 'Submit', font = fonts, bg = wc, bd = 0, width = 95, image = buttonicon, compound = CENTER, cursor = "hand2", command = self.submitdoc)
        submitbutton.bind('<Return>', self.submitdoc)
        
        printbutton = Button(buttonframe, text = 'Print', font = fonts, bg = wc, bd = 0, width = 95, image = buttonicon, compound = CENTER, cursor = "hand2", command = self.printdoc)
        printbutton.bind('<Return>', self.printdoc)

        refreshbutton = Button(buttonframe, text = 'Refresh', font = fonts, bg = wc, bd = 0, width = 95, image = buttonicon, compound = CENTER, cursor = "hand2", command = self.fmdrefresh)
        refreshbutton.bind('<Return>', self.fmdrefresh)
        
        closebutton = Button(buttonframe, text = 'Close', font = fonts, bg = wc, bd = 0, width = 95, image = buttonicon, compound = CENTER, cursor = "hand2", command = lambda: self.closefmdmenu(disbursementsframe, disbursement_button))
        closebutton.bind('<Return>', lambda e: self.closefmdmenu(disbursementsframe, disbursement_button))
        closebutton.grid(column = 2, row = 0, padx = pad)

        # button2307 = Button(buttonframe, text = '2307', font = fonts, bg = buttonbg, width = 10, command = showgenerate)
        # button2307.bind('<Return>', showgenerate)
    
    def showdocument(self, *args):
        global companyentry, dateentry, number, numberentry, typeentry, supplierlist, particularsentry, payeeentry, signatoryentry1, signatoryentry2, amount, amountentry, assignmententry, accountentry, bankentry, banklabel, payeelabel, assignmentlabel, particularslabel, accountlabel
        if documententry.get() == 'General Voucher':
            companylabel = Label(disbursementssubframe, text = 'Company', font = fonts, bg = wc, fg = fc)
            companylabel.grid(column = 0, row = 0, sticky = W, pady = pad)

            datelabel = Label(disbursementssubframe, text = 'Date', font = fonts, bg = wc, fg = fc)
            datelabel.grid(column = 0, row = 1, sticky = W, pady = pad)

            numberlabel = Label(disbursementssubframe, text = 'Number', font = fonts, bg = wc, fg = fc)
            numberlabel.grid(column = 0, row = 2, sticky = W, pady = pad)
            
            typelabel = Label(disbursementssubframe, text = 'Type', font = fonts, bg = wc, fg = fc)
            typelabel.grid(column = 0, row = 3, sticky = W, pady = pad)

            payeelabel = Label(disbursementssubframe, text = 'Payee', font = fonts, bg = wc, fg = fc)
            payeelabel.grid(column = 2, row = 0, sticky = W, pady = pad)

            particularslabel = Label(disbursementssubframe, text = 'Particulars', font = fonts, bg = wc, fg = fc)
            particularslabel.grid(column = 2, row = 1, sticky = W, pady = pad)

            signatorylabel1 = Label(disbursementssubframe, text = 'Signatory1', font = fonts, bg = wc, fg = fc)
            signatorylabel1.grid(column = 2, row = 2, sticky = W, pady = pad)

            signatorylabel2 = Label(disbursementssubframe, text = 'Signatory2', font = fonts, bg = wc, fg = fc)
            signatorylabel2.grid(column = 2, row = 3, sticky = W, pady = pad)

            companyentry = tk.Combobox(disbursementssubframe, values = ['DBPSC','DSSI'], font = fonts, width = 8)
            companyentry.grid(column = 1, row = 0, sticky = W)
            companyentry.insert(0, company)
            companyentry.config(state = DISABLED)

            dateentry = Entry(disbursementssubframe, font = fonts, width = 10)
            dateentry.grid(column = 1, row = 1, sticky = W)
            dateentry.bind('<FocusOut>', self.formatdate2)
            dateentry.insert(0, today.strftime('%m-%d-%Y'))

            numberentry = Entry(disbursementssubframe, font = fonts, width = 10)
            numberentry.grid(column = 1, row = 2, sticky = W)
            numberentry.bind('<FocusOut>', lambda e: self.checkdoc(numberentry))

            typeentry = tk.Combobox(disbursementssubframe, values = typelister, font = fonts, width = 15)
            typeentry.grid(column = 1, row = 3, sticky = W)
            typeentry.bind('<<ComboboxSelected>>', self.chosendmtxn)
            typeentry.bind('<FocusOut>', self.intypelist)

            supplierlist = tk.Combobox(disbursementssubframe, font = fonts, width = 59)
            supplierlist.grid(column = 3, row = 0, sticky = W)
            supplierlist.bind('<FocusOut>', lambda e: self.uppercase(supplierlist))

            particularsentry = tk.Combobox(disbursementssubframe, font = fonts, width = 59)
            particularsentry.grid(column = 3, row = 1, sticky = W)

            signatoryentry1 = tk.Combobox(disbursementssubframe, font = fonts, width = 59)
            signatoryentry1.grid(column = 3, row = 2, sticky = W)

            signatoryentry2 = tk.Combobox(disbursementssubframe, font = fonts, width = 59)
            signatoryentry2.grid(column = 3, row = 3, sticky = W)

            self.showlabels(lineframes, 0, 1)
            self.showlines(lineframes, 0, 2)
            self.showtotals(lineframes, 0, 3)
            self.merchantlister(supplierlist)
            self.fillsignature2()
            cancelbox.grid(column = 3, row = 0, padx = 75)

        else:
            companylabel = Label(disbursementssubframe, text = 'Company', font = fonts, bg = wc, fg = fc)
            companylabel.grid(column = 0, row = 0, sticky = W, pady = pad)

            datelabel = Label(disbursementssubframe, text = 'Date', font = fonts, bg = wc, fg = fc)
            datelabel.grid(column = 0, row = 1, sticky = W, pady = pad)

            numberlabel = Label(disbursementssubframe, text = 'Number', font = fonts, bg = wc, fg = fc)
            numberlabel.grid(column = 0, row = 2, sticky = W, pady = pad)
            
            typelabel = Label(disbursementssubframe, text = 'Type', font = fonts, bg = wc, fg = fc)
            typelabel.grid(column = 0, row = 3, sticky = W, pady = pad)

            banklabel = Label(disbursementssubframe, text = 'Bank', font = fonts, bg = wc, fg = fc)

            amountlabel = Label(disbursementssubframe, text = 'Amount', font = fonts, bg = wc, fg = fc)
            amountlabel.grid(column = 2, row = 0, sticky = W, pady = pad)

            payeelabel = Label(disbursementssubframe, text = 'Payee', font = fonts, bg = wc, fg = fc)
            payeelabel.grid(column = 2, row = 1, sticky = W, pady = pad)

            assignmentlabel = Label(disbursementssubframe, text = 'Assignment', font = fonts, bg = wc, fg = fc)
            assignmentlabel.grid(column = 2, row = 2, sticky = W, pady = pad)

            particularslabel = Label(disbursementssubframe, text = 'Particulars', font = fonts, bg = wc, fg = fc)
            particularslabel.grid(column = 2, row = 3, sticky = W, pady = pad)

            accountlabel = Label(disbursementssubframe, text = 'Account No.', font = fonts, bg = wc, fg = fc)
            accountlabel.grid(column = 2, row = 4, sticky = W, pady = pad)

            companyentry = tk.Combobox(disbursementssubframe, values = ['DBPSC','DSSI'], font = fonts, width = 8)
            companyentry.grid(column = 1, row = 0, sticky = W)
            companyentry.insert(0, company)
            companyentry.config(state = DISABLED)

            dateentry = Entry(disbursementssubframe, font = fonts, width = 10)
            dateentry.grid(column = 1, row = 1, sticky = W)
            dateentry.bind('<FocusOut>', self.formatdate2)
            dateentry.insert(0, today.strftime('%m-%d-%Y'))

            number = StringVar()
            numberentry = Entry(disbursementssubframe, textvariable = number, font = fonts, width = 10)
            numberentry.grid(column = 1, row = 2, sticky = W)
            numberentry.bind("<FocusOut>", lambda e: self.checkdoc(numberentry))

            typeentry = tk.Combobox(disbursementssubframe, values = typelister, font = fonts, width = 15)
            typeentry.grid(column = 1, row = 3, sticky = W)
            typeentry.bind('<<ComboboxSelected>>', self.chosendmtxn)
            typeentry.bind('<FocusOut>', self.intypelist)

            bankentry = tk.Combobox(disbursementssubframe, values = dmbanks, font = fonts, width = 20)
            bankentry.bind('<<ComboboxSelected>>', self.destbank)
            bankentry.bind('<FocusOut>', lambda e: self.inbanklist(bankentry, dmbanks))

            amount = StringVar()
            amountentry = Entry(disbursementssubframe, textvariable = amount, font = fonts, width = 15, justify = RIGHT)
            amountentry.grid(column = 3, row = 0, sticky = W)
            amountentry.bind('<FocusOut>', lambda e: self.amountvalidatormaster(amount))

            payeeentry = Entry(disbursementssubframe, font = fonts, width = 61)
            payeeentry.grid(column = 3, row = 1, sticky = W)
            payeeentry.bind('<FocusOut>', self.dmassembly)

            assignmententry = Entry(disbursementssubframe, font = fonts, width = 61)
            assignmententry.grid(column = 3, row = 2, sticky = W)
            assignmententry.bind('<FocusOut>', self.dmassembly)

            particularsentry = tk.Combobox(disbursementssubframe, font = fonts, width = 59)
            particularsentry.grid(column = 3, row = 3, sticky = W)
            particularsentry.bind('<FocusOut>', self.dmassembly)

            accountentry = Entry(disbursementssubframe, font = fonts, width = 61)
            accountentry.grid(column = 3, row = 4, sticky = W)
            accountentry.bind('<FocusOut>', self.dmassembly)

            self.showdmcontent()
        documententry.config(state = DISABLED)
        submitbutton.grid(column = 0, row = 0, padx = pad)
        printbutton.grid(column = 1, row = 0, padx = pad)
        refreshbutton.grid(column = 2, row = 0, padx = pad)
        closebutton.grid(column = 3, row = 0, padx = pad)

    def checkdoc(self, num, *args):
        if documententry.get() == 'General Voucher':
            for i in num.get():
                if i.isdigit() == False:
                    num.delete(0, END)
                    break
        else:
            if len(num.get().split('-')) != 2:
                num.delete(0, END)
        if num.get() != "":
            find = "SELECT number FROM disbursements WHERE number = ? AND document = ? AND company = ?"
            c.execute(find, [num.get(), documententry.get(), companyentry.get()])
            result = c.fetchall()
            if result:
                messagebox.showerror("Number Checker", "Document number is already used!")
                num.delete(0, END)

    def submitdoc(self, *args):
        if documententry.get() == 'General Voucher':
            if numberentry.get() == '':
                messagebox.showerror('Posting Protocol','General Voucher number is missing!')
            elif dateentry.get() == '':
                messagebox.showerror('Posting Protocol','Date is missing!')
            elif typeentry.get() == '':
                messagebox.showerror('Posting Protocol','Type is missing!')
            elif supplierlist.get() == '':
                messagebox.showerror('Posting Protocol','Payee is missing!')
            elif float(total.get().replace(',','')) == 0:
                messagebox.showerror('Posting Protocol','Please check amounts!')
            else:
                ask = messagebox.askyesno('Submit Document', 'Are you sure?')
                if ask == True:
                    if checkbox.get() == 1:
                        status = 'CANCELLED'
                    else:
                        status = 'VALID'
                    validitems = []
                    for i in range(5):
                        try:
                            if float(allnets[i].get().replace(',','')) != 0 and allnets[i].get() != '':
                                self.computetax(alltaxes[i],allamounts[i])
                                validitems.append([
                                    companyentry.get(),dateentry.get(),documententry.get(),numberentry.get(),typeentry.get(),
                                    supplierlist.get(),particularsentry.get(),signatoryentry1.get(),signatoryentry2.get(),allitems[i].get(),
                                    allbanks[i].get(),allmodes[i].get(),float(allamounts[i].get().replace(',','')),alltaxes[i].get(),float(allnets[i].get().replace(',','')),
                                    round(allewt[i],2), round(allvat[i],2),dateentry.get().split('-')[0],dateentry.get().split('-')[1],dateentry.get().split('-')[2],
                                    status, self.username.get()
                                ])
                        except:
                            pass
                    inserter = """INSERT INTO disbursements (
                        company, date, document, number, type,
                        payee, particulars, signatory1, signatory2, item,
                        bank, mode, gross, tax, net,
                        ewt, vat, month, day, year,
                        status, user) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
                        """
                    for v in validitems:
                        c.execute(inserter, v)
                        conn.commit()
                    self.disabler()
                    submitbutton.config(state = DISABLED)
                    messagebox.showinfo('Posting Protocol', 'Disbursement record has been added!')
        else:
            if dateentry.get() == '':
                messagebox.showerror('Posting Protocol','Date is missing!')
            elif typeentry.get() == '':
                messagebox.showerror('Posting Protocol','Type is missing!')
            elif float(amount.get().replace(',','')) == 0 or amount.get() == '':
                messagebox.showerror('Posting Protocol','Please check amounts!')
            else:
                ask = messagebox.askyesno('Submit Document', 'Are you sure?')
                if ask == True:
                    if checkbox.get() == 1:
                        status = 'CANCELLED'
                    else:
                        status = 'VALID'
                    # self.dmnumbermaster()
                    if documententry.get() == 'Letter of Instruction':
                        doctype = 'LOI'
                    else:
                        doctype = 'DM'
                    defaultmode = doctype
                    defaulttax = 'NV-00'
                    if typeentry.get() == 'FT' or typeentry.get() == 'Fund transfer' or typeentry.get() == 'Fund Transfer' or typeentry.get() == 'FUND TRANSFER' or typeentry.get() == 'ft':
                        inserter = """INSERT INTO disbursements (
                            company, date, document, number, type,
                            payee, particulars, bank, mode, gross,
                            tax, net, month, day, year,
                            status, user) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
                            """
                        defaultpayee = 'DBPSC'
                        defaultparticulars = 'Fund transfer to '+destinationentry.get()
                        values = [
                            companyentry.get(),dateentry.get(),documententry.get(),numberentry.get(),typeentry.get(),
                            defaultpayee,defaultparticulars,bankentry.get(),defaultmode,float(amount.get().replace(',','')),
                            defaulttax,float(amount.get().replace(',','')),dateentry.get().split('-')[0],dateentry.get().split('-')[1],dateentry.get().split('-')[2],
                            status,self.username.get()
                            ]
                    else:
                        inserter = """INSERT INTO disbursements (
                            company, date, document, number, type,
                            payee, particulars, bank, mode, gross,
                            tax, net, month, day, year,
                            status, user) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
                            """
                        if documententry.get() == 'DM - DBP Davao':
                            defaultbank = 'DBP Davao'
                        else:
                            defaultbank = 'RCBC'
                        if accountentry.get() != '':
                            fullparticulars = particularsentry.get() + '_' + accountentry.get()
                            payeeassign = payeeentry.get() + '_' +assignmententry.get()
                        else:
                            fullparticulars = particularsentry.get()
                            payeeassign = assignmententry.get()
                        values = [
                            companyentry.get(),dateentry.get(),documententry.get(),numberentry.get(),typeentry.get(),
                            payeeassign,fullparticulars,defaultbank,defaultmode,float(amount.get().replace(',','')),
                            defaulttax,float(amount.get().replace(',','')),dateentry.get().split('-')[0],dateentry.get().split('-')[1],dateentry.get().split('-')[2],
                            status,self.username.get()
                            ]
                    c.execute(inserter, values)
                    conn.commit()
                    self.disabler()
                    submitbutton.config(state = DISABLED)
                    messagebox.showinfo('Posting Protocol', 'Disbursement record has been added!')

    def dmnumbermaster(self):
        global newnumber
        sequence = "SELECT MAX(number) FROM disbursements WHERE company = ? AND document = ? AND year = ?"
        c.execute(sequence, [companyentry.get(),documententry.get(),dateentry.get().split('-')[2]])
        result = c.fetchone()[0]
        if result:
            plusone = str((int(result.split('-')[1]))+1)
            newnumber = dateentry.get().split('-')[2]+'-'+plusone.zfill(4)
        else:
            newnumber = dateentry.get().split('-')[2]+'-'+'0001'
        number.set(newnumber)

    def computetax(self, taxvar, amtent):
        global vatamount, ewtamount
        percent = float(taxvar.get().split('-')[1])/100
        if amtent.get() != '':
            try:
                base = float(amtent.get().replace(',',''))
                if taxvar.get().split('-')[0] == 'WV':
                    vatamount = (base/1.12)*.12
                    ewtamount = (base/1.12)*percent
                elif taxvar.get().split('-')[0] == 'NV':
                    vatamount = 0
                    ewtamount = base*percent
                elif taxvar.get().split('-')[0] == 'V':
                    vatamount = (base/1.12)*percent
                    ewtamount = 0
            except:
                base = 0
                vatamount = 0
                ewtamount = 0
        else:
            base = 0
            vatamount = 0
            ewtamount = 0

    def disabler(self):
        companyentry.config(state = DISABLED)
        dateentry.config(state = DISABLED)
        numberentry.config(state = DISABLED)
        if documententry.get() == 'General Voucher':
            documententry.config(state = DISABLED)
            typeentry.config(state = DISABLED)  
            supplierlist.config(state = DISABLED)
            particularsentry.config(state = DISABLED)
            signatoryentry1.config(state = DISABLED)
            signatoryentry2.config(state = DISABLED)
            cancelbox.config(state = DISABLED)
            for line in alllines:
                for widget in line.winfo_children():
                    widget.config(state = DISABLED)
        else:
            if typeentry.get() == 'FT':
                documententry.config(state = DISABLED)
                typeentry.config(state = DISABLED)
                bankentry.config(state = DISABLED)
                amountentry.config(state = DISABLED)
                destinationentry.config(state = DISABLED)
                cancelbox.config(state = DISABLED)
            else:
                documententry.config(state = DISABLED)
                typeentry.config(state = DISABLED)
                amountentry.config(state = DISABLED) 
                payeeentry.config(state = DISABLED) 
                assignmententry.config(state = DISABLED) 
                particularsentry.config(state = DISABLED) 
                accountentry.config(state = DISABLED) 
                cancelbox.config(state = DISABLED)

    def printdoc(self, *args):
        if documententry.get() == 'General Voucher':
            if numberentry.get() == '':
                messagebox.showerror('Printing Protocol','General Voucher number is missing!')
            elif dateentry.get() == '':
                messagebox.showerror('Printing Protocol','Date is missing!')
            elif supplierlist.get() == '':
                messagebox.showerror('Printing Protocol','Payee is missing!')
            elif float(total.get().replace(',','')) == 0:
                messagebox.showerror('Printing Protocol','Please check amounts!')
            else:
                validnets = []
                validgross = []
                validitems = []
                validbanks = []
                validmodes = []
                validvat = []
                validewt = []
                for i in range(5):
                    try:
                        if float(allnets[i].get().replace(',','')) != 0:
                            validitems.append(allitems[i].get())
                            validbanks.append(allbanks[i].get())
                            validmodes.append(allmodes[i].get())
                            validgross.append(float(allamounts[i].get().replace(',','')))
                            validnets.append(float(allnets[i].get().replace(',','')))
                            validvat.append(allvat[i])
                            validewt.append(allewt[i])
                    except:
                        pass
                wb = openpyxl.load_workbook(path + 'GV.xlsx') #path
                st = wb.active
                st['B9'] = supplierlist.get().upper()
                st['E6'] = str(datetime.datetime.strptime(dateentry.get(), '%m-%d-%Y').strftime('%b %d %Y')).upper()
                st['A12'] = particularsentry.get()
                st['A15'] = 'per attached supporting papers - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -P'
                if len(validnets) == 1 and sum(validewt) == 0:
                    st['A21'] = validitems[0]
                    st['B21'] = validbanks[0]+' '+validmodes[0]
                    st['D21'] = validgross[0]
                    st.cell(column = 4, row = 21).border = Border(bottom = Side(style='double'))
                elif len(validnets) > 1 and sum(validewt) == 0:
                    rcbctotal, nonrcbctotal = [], []
                    for i in range(len(validnets)):
                        if validbanks[i] == "RCBC":
                            rcbctotal.append(validnets[i])
                        else:
                            nonrcbctotal.append(validnets[i])

                    if len(rcbctotal) > 1 and len(nonrcbctotal) >= 1:
                        for i in range(len(validnets)):
                            if validbanks[i] == "RCBC":
                                st['A'+str(20-len(validnets)+i)] = validitems[i]
                                st['B'+str(20-len(validnets)+i)] = validbanks[i]+' '+validmodes[i]
                                st['D'+str(20-len(validnets)+i)] = validgross[i]
                        st.cell(column = 4, row = 20-len(validnets)+i-1).border = Border(bottom = Side(style='thin'))
                        st['A'+str(20-len(validnets)+i)] = "RCBC Cash - Total"
                        st['D'+str(20-len(validnets)+i)] = sum(rcbctotal)

                        for i in range(len(validnets)):
                            if validbanks[i] != "RCBC":
                                st['A'+str(20-len(validnets)+len(nonrcbctotal)+i)] = validitems[i]
                                st['B'+str(20-len(validnets)+len(nonrcbctotal)+i)] = validbanks[i]+' '+validmodes[i]
                                st['D'+str(20-len(validnets)+len(nonrcbctotal)+i)] = validgross[i]
                        st.cell(column = 4, row = 20-len(validnets)+len(nonrcbctotal)+i).border = Border(bottom = Side(style='thin'))
                        st['D'+str(20-len(validnets)+len(nonrcbctotal)+i+1)] = sum(validnets)
                        st['A'+str(20-len(validnets)+len(nonrcbctotal)+i+1)] = "Total"
                        st.cell(column = 4, row = 20-len(validnets)+len(nonrcbctotal)+i+1).border = Border(bottom = Side(style='double'))

                    elif len(rcbctotal) > 1 and len(nonrcbctotal) == 0:
                        for i in range(len(validnets)):
                            st['A'+str(20-len(validnets)+i)] = validitems[i]
                            st['B'+str(20-len(validnets)+i)] = validbanks[i]+' '+validmodes[i]
                            st['D'+str(20-len(validnets)+i)] = validgross[i]
                        st.cell(column = 4, row = 20-len(validnets)+i).border = Border(bottom = Side(style='thin'))
                        st['A'+str(20-len(validnets)+i+1)] = "RCBC Cash - Total"
                        st['D'+str(20-len(validnets)+i+1)] = sum(validnets)
                        st.cell(column = 4, row = 20-len(validnets)+i+1).border = Border(bottom = Side(style='double'))

                    elif len(rcbctotal) == 0 and len(nonrcbctotal) > 1:
                        for i in range(len(validnets)):
                            st['A'+str(20-len(validnets)+i)] = validitems[i]
                            st['B'+str(20-len(validnets)+i)] = validbanks[i]+' '+validmodes[i]
                            st['D'+str(20-len(validnets)+i)] = validgross[i]
                        st.cell(column = 4, row = 20-len(validnets)+i).border = Border(bottom = Side(style='thin'))
                        st['D'+str(20-len(validnets)+i+1)] = sum(validnets)
                        st.cell(column = 4, row = 20-len(validnets)+i+1).border = Border(bottom = Side(style='double'))

                elif len(validnets) == 1 and sum(validewt) > 0:
                    st['A21'] = validitems[0]
                    st['B21'] = validbanks[0]+' '+validmodes[0]
                    st['D21'] = validgross[0]
                    st['B22'] = 'Less: EWT '+ str(percent*100)+'%'
                    st['D22'] = sum(validewt)
                    st.cell(column = 4, row = 22).border = Border(bottom = Side(style='thin'))
                    st['D23'] = sum(validgross)-sum(validewt)
                    st.cell(column = 4, row = 23).border = Border(bottom = Side(style='double'))
                elif len(validnets) > 1 and sum(validewt) > 0:
                    for i in range(len(validnets)):
                        st['A'+str(22-len(validnets)+i)] = validitems[i]
                        st['B'+str(22-len(validnets)+i)] = validbanks[i]+' '+validmodes[i]
                        st['D'+str(22-len(validnets)+i)] = validgross[i]
                    st.cell(column = 4, row = 22-len(validnets)+i).border = Border(bottom = Side(style='thin'))
                    st['D' + str(22-len(validnets)+i+1)] = sum(validnets)
                    st['D' + str(22-len(validnets)+i+2)] = sum(validewt)
                    st.cell(column = 4, row = 22-len(validnets)+i+2).border = Border(bottom = Side(style='thin'))
                    st['D' + str(22-len(validnets)+i+3)] = sum(validgross)-sum(validewt)
                    st.cell(column = 4, row = 22-len(validnets)+i+3).border = Border(bottom = Side(style='double'))
                st['E15'] = sum(validnets)
                self.numbertowords(totalnet)
                st['A25'] = "*** " + words + " ***"
                st['A31'] = signatoryentry1.get().upper()
                sigfind = "SELECT position FROM signatory WHERE name = ?"
                c.execute(sigfind, [signatoryentry1.get().upper()])
                sig1result = c.fetchone()
                if sig1result:
                    st['A32'] = sig1result[0]
                else:
                    st['A32'] = signatoryentry1.get().upper()
                st['D31'] = signatoryentry2.get().upper()
                c.execute(sigfind, [signatoryentry2.get().upper()])
                sig2result = c.fetchone()
                if sig2result:
                    st['D32'] = sig2result[0]
                else:
                    st['D32'] = signatoryentry2.get().upper()
                st['A40'] = self.username.get().upper()
                wb.save(savepath + 'test.xlsx') #path
                os.startfile(savepath + 'test.xlsx', 'open') #path
        else:
            if typeentry.get() != 'FT':
                if dateentry.get() == '':
                    messagebox.showerror('Printing Protocol','Date is missing!')
                elif numberentry.get() == '':
                    messagebox.showerror('Printing Protocol','Debit Memo number is missing!')
                elif float(amount.get().replace(',','')) == 0 or amount.get() == None:
                    messagebox.showerror('Printing Protocol','Please check amounts!')
                else:
                    if documententry.get() == 'Letter of Instruction':
                        doctype = 'LOI '
                        wb = openpyxl.load_workbook(path + 'DMDSSI.xlsx') #path
                    elif documententry.get() == 'DM - DBP Davao':
                        doctype = 'DM '
                        wb = openpyxl.load_workbook(path + 'DMDAVAO.xlsx') #path
                    else:
                        doctype = 'DM '
                        wb = openpyxl.load_workbook(path + 'DM.xlsx') #path
                    st = wb.active
                    st['H1'] = doctype + numberentry.get()
                    st['G3'] = datetime.datetime.strptime(dateentry.get(), '%m-%d-%Y').strftime('%B %d, %Y')
                    st['A16'] = content.get()
                    wb.save(savepath + 'test.xlsx') #path
                    os.startfile(savepath + 'test.xlsx', 'open') #path
            else:
                self.numbertowords(amount)
                if bankentry.get().split(' ')[0] == 'BDO':
                    wb = openpyxl.load_workbook(path + 'DMFTBDO.xlsx') #path
                    st = wb.active
                    intro = 'May we request that you debit today, ' + datetime.datetime.strptime(dateentry.get(), '%m-%d-%Y').strftime('%B %d, %Y') + ' our Savings Account No.'
                    creditto = ' and credit same amount to our Savings Account No.'
                    outro = ' maintained at your branch.'
                    try:
                        origin = bankentry.get().split(' - ')[1]
                    except:
                        origin = bankentry.get()
                    destiny = destinationentry.get().split(' - ')[1]
                    st['H1'] = 'DM ' + numberentry.get()
                    st['F3'] = datetime.datetime.strptime(dateentry.get(), '%m-%d-%Y').strftime('%B %d, %Y')
                    st['A15'] = intro + origin + ' for ' + words + ' (P' + amount.get() + ')' + creditto + destiny + outro
                elif bankentry.get().split(' ')[0] == 'DBP':
                    if bankentry.get().split(' ')[1] == 'SA':
                        accounttype = 'Savings Account No. '
                        creditto = ' and credit same amount to our Checking Account No.'
                    elif bankentry.get().split(' ')[1] == 'CA':
                        accounttype = 'Checking Account No. '
                        creditto = ' and credit same amount to our Savings Account No.'
                    else:
                        accounttype = 'Savings Account No. '
                        creditto = ' and credit same amount to our Savings Account No.'
                    wb = openpyxl.load_workbook(path + 'DMFTDBP.xlsx') #path
                    st = wb.active
                    intro = 'May we request that you debit today, ' + datetime.datetime.strptime(dateentry.get(), '%m-%d-%Y').strftime('%B %d, %Y') + ' our '
                    outro = ' maintained at your branch.'
                    try:
                        origin = bankentry.get().split(' - ')[1]
                    except:
                        origin = bankentry.get()
                    destiny = destinationentry.get().split(' - ')[1]
                    if documententry.get() == 'Letter of Instruction':
                        doctype = 'LOI '
                    else:
                        doctype = 'DM '
                    st['H1'] = doctype + numberentry.get()
                    st['F3'] = datetime.datetime.strptime(dateentry.get(), '%m-%d-%Y').strftime('%B %d, %Y')
                    st['A15'] = intro + accounttype + origin + ' for ' + words + ' (P' + amount.get() + ')' + creditto + destiny + outro
                else:
                    wb = openpyxl.load_workbook(path + 'DMFTBDO.xlsx') #path
                    st = wb.active
                wb.save(savepath + 'test.xlsx') #path
                os.startfile(savepath + 'test.xlsx', 'open') #path

    def numbertowords(self, amounttoconvert):
        global words
        try:
            if float(amounttoconvert.get().split('.')[1]) > 0:
                if documententry.get() == 'General Voucher':
                    decimal = ' AND ' + str(int(amounttoconvert.get().split('.')[1])) +'/100'
                else:
                    decimal = ' AND ' + str(int(amounttoconvert.get().split('.')[1])) +'/100'
            else:
                decimal = ' ONLY '
            converted = num2words(amounttoconvert.get().replace(',','').split('.')[0]).upper()
            words =  converted.replace(" AND","") + ' PESOS' + decimal
        except:
            pass

    def fmdrefresh(self, *args):
        if documententry.get() == 'General Voucher':
            disbursementsframe.destroy()
            self.showdisbursements()
            documententry.insert(0, 'General Voucher')
            self.showdocument()
        else:
            disbursementsframe.destroy()
            self.showdisbursements()
            documententry.insert(0, 'Debit Memo')
            self.showdocument()

    def chosendmtxn(self, *args):
        chosen = typeentry.get()
        if chosen == '13th MP':
            particularsentry.delete(0, END)
            particularsentry.insert(0, 'Payment of 13th month pay for the year 202X')
        elif chosen == 'Abuloy':
            particularsentry.delete(0, END)
            particularsentry.insert(0, 'Advance payment of abuloy')
        elif chosen == 'CA':
            particularsentry.delete(0, END)
            particularsentry.insert(0, 'Payment of cash advance no. XXXX')
        elif chosen == 'CA contractuals':
            particularsentry.delete(0, END)
            particularsentry.insert(0, 'Payment of cash advance')
        elif chosen == 'Cashcard reversals':
            particularsentry.delete(0, END)
            particularsentry.insert(0, 'Reversed balance of RCBC mywallet card')
        elif chosen == 'Deathclaims':
            particularsentry.delete(0, END)
            particularsentry.insert(0, 'Full payment of deathclaims')
        elif chosen == 'Financial assistance':
            particularsentry.delete(0, END)
            particularsentry.insert(0, 'Payment of financial assistance')
        elif chosen == 'CA contractuals':
            particularsentry.delete(0, END)
            particularsentry.insert(0, 'Payment of cash advance')
        elif chosen == 'Gas allowance':
            particularsentry.delete(0, END)
            particularsentry.insert(0, 'Payment of gasoline allowance for thep period')
        elif chosen == 'Gratuity pay':
            particularsentry.delete(0, END)
            particularsentry.insert(0, 'Payment of gratuity pay for the year 202X')
        elif chosen == 'Hazard pay':
            particularsentry.delete(0, END)
            particularsentry.insert(0, 'Payment of hazard pay')
        elif chosen == 'HDMF SAI':
            particularsentry.delete(0, END)
            particularsentry.insert(0, 'Payment of semi-annual incentive (SAI) for the year 202X')
        elif chosen == 'Monetization':
            particularsentry.delete(0, END)
            particularsentry.insert(0, 'Payment of leave monetization for the year 202X')
        elif chosen == 'Overtime pay':
            particularsentry.delete(0, END)
            particularsentry.insert(0, 'Payment of overtime pay for the period')
        elif chosen == 'Paternity claims':
            particularsentry.delete(0, END)
            particularsentry.insert(0, 'Payment of paternity claims')
        elif chosen == 'PCF replenishment':
            particularsentry.delete(0, END)
            particularsentry.insert(0, 'Replenishment of petty cash fund for the period')
        elif chosen == 'Quitclaims':
            particularsentry.delete(0, END)
            particularsentry.insert(0, 'Payment of quitclaims')
        elif chosen == 'Refund':
            particularsentry.delete(0, END)
            particularsentry.insert(0, 'Payment of tax refund for the year 202X')
        elif chosen == 'Retirement pay':
            particularsentry.delete(0, END)
            particularsentry.insert(0, 'Payment of retirement pay')
        elif chosen == 'Salary':
            particularsentry.delete(0, END)
            particularsentry.insert(0, 'Payment of salary')
        elif chosen == 'Separation pay':
            particularsentry.delete(0, END)
            particularsentry.insert(0, 'Payment of separation pay')
        elif chosen == 'Solo parent leave':
            particularsentry.delete(0, END)
            particularsentry.insert(0, 'Payment of solo parent leave')
        elif chosen == 'SSS claims':
            particularsentry.delete(0, END)
            particularsentry.insert(0, 'Advance payment of sss sickness & maternity claims')
        elif chosen == 'FT' or chosen == 'Fund transfer' or chosen == 'Fund Transfer' or chosen == 'FUND TRANSFER' or chosen == 'ft' or chosen == 'fund transfer' or chosen == 'fundtransfer':
            if documententry.get() == 'General Voucher':
                particularsentry.delete(0, END)
                particularsentry.insert(0, 'Fund Transfer from ')
            else:
                contentframe.destroy()
                typeentry.config(width = 20, state = DISABLED)
                payeelabel.destroy()
                payeeentry.destroy()
                assignmentlabel.destroy()
                assignmententry.destroy()
                particularslabel.destroy()
                particularsentry.destroy()
                accountlabel.destroy()
                accountentry.destroy()
                banklabel.grid(column = 0, row = 4, sticky = W, pady = pad)
                bankentry.grid(column = 1, row = 4, sticky = W)
                amount.set(0)
                amountentry.bind('<FocusOut>', None)
                self.showft()
        else:
            particularsentry.delete(0, END)

    def showlabels(self, master, cl, rw):
        labelframe = Frame(master, bg = wc)
        labelframe.grid(column = cl, row = rw, sticky = NW)

        itemlabel = Label(labelframe, text = 'Description', font = fonts, width = 35, relief = RIDGE, bg = wc, fg = fc)
        itemlabel.grid(column = 0, row = 0)

        banklabel = Label(labelframe, text = 'Bank', font = fonts, width = 12, relief = RIDGE, bg = wc, fg = fc)
        banklabel.grid(column = 1, row = 0)

        modelabel = Label(labelframe, text = 'Mode', font = fonts, width = 19, relief = RIDGE, bg = wc, fg = fc)
        modelabel.grid(column = 2, row = 0)

        grosslabel = Label(labelframe, text = 'Gross', font = fonts, width = 15, relief = RIDGE, bg = wc, fg = fc)
        grosslabel.grid(column = 3, row = 0)

        taxlabel = Label(labelframe, text = 'Tax', font = fonts, width = 11, relief = RIDGE, bg = wc, fg = fc)
        taxlabel.grid(column = 4, row = 0)

        netlabel = Label(labelframe, text = 'Net', font = fonts, width = 15, relief = RIDGE, bg = wc, fg = fc)
        netlabel.grid(column = 5, row = 0)

    def showlines(self, master, cl, rw):
        global linegroup, alllines, allitems, allbanks, allmodes, allamounts, alltaxes, allnets
        alllines, allitems, allbanks, allmodes, allamounts, alltaxes, allnets = [], [], [], [], [], [], []

        linegroup = Frame(master, bg = wc)
        linegroup.grid(column = cl, row = rw, sticky = W)

        for i in range(5):
            self.gvline(linegroup, i)

    def showtotals(self, master, cl, rw):
        global totalnet, totalnet, total, vatnet, vat, ewtnet, ewt
        totalframe = Frame(master, bg = wc)
        totalframe.grid(column = cl, row = rw, sticky = NW)

        blanklabel = Label(totalframe, text = '', font = fonts, bg = wc, fg = fc)
        blanklabel.grid(column = 0, row = 0, ipadx = 256)

        vat = StringVar()
        vatnet = Entry(totalframe, textvariable = vat, font = fonts, width = 15, bg = wc, fg = fc, disabledbackground = wc, disabledforeground = fc, justify = RIGHT)
        vatnet.grid(column = 1, row = 0)
        vatnet.config(state = DISABLED)
        vat.set('VAT: ' + format(float(0), ',.2f'))

        ewt = StringVar()
        ewtnet = Entry(totalframe, textvariable = ewt, font = fonts, width = 15, bg = wc, fg = fc, disabledbackground = wc, disabledforeground = fc, justify = RIGHT)
        ewtnet.grid(column = 2, row = 0)
        ewtnet.config(state = DISABLED)
        ewt.set('EWT: ' + format(float(0), ',.2f'))

        total = StringVar()
        totalnet = Entry(totalframe, textvariable = total, font = fonts, width = 15, bg = wc, fg = fc, disabledbackground = wc, disabledforeground = fc, justify = RIGHT)
        totalnet.grid(column = 3, row = 0)
        totalnet.config(state = DISABLED)
        total.set(format(float(0), ',.2f'))

    def gvline(self, master, rw):
        lineframe = Frame(master, bg = wc)
        lineframe.grid(column = 0, row = rw, padx = 1)
        alllines.append(lineframe)

        itementry = Entry(lineframe, font = fonts, width = 35)
        itementry.grid(column = 0, row = 1)
        allitems.append(itementry)

        bankentry = tk.Combobox(lineframe, values = banks, font = fonts, width = 10)
        bankentry.grid(column = 1, row = 1)
        allbanks.append(bankentry)

        modeentry = tk.Combobox(lineframe, values = modes, font = fonts, width = 17)
        modeentry.grid(column = 2, row = 1)
        allmodes.append(modeentry)

        amount = StringVar()
        amountentry = Entry(lineframe, textvariable = amount, font = fonts, width = 15, justify = RIGHT)
        amountentry.grid(column = 3, row = 1)
        amountentry.bind('<FocusOut>', lambda e: self.gvamountvalidator(amount))
        amount.set(0)
        allamounts.append(amount)
        
        taxentry = tk.Combobox(lineframe, values = atc, font = fonts, width = 9)
        taxentry.grid(column = 4, row = 1)
        taxentry.bind('<<ComboboxSelected>>', self.computegvtax)
        taxentry.bind('<FocusOut>', self.computegvtax)
        taxentry.insert(0, 'NV-00')
        alltaxes.append(taxentry)

        net = StringVar()
        netentry = Entry(lineframe, textvariable = net, state = DISABLED, disabledbackground = codetitlebg, disabledforeground = codetitlefg, font = fonts, width = 15, justify = RIGHT)
        netentry.grid(column = 5, row = 1)
        allnets.append(net)

    def gvamountvalidator(self, amt, *args):
        self.amountvalidatormaster(amt)
        self.computegvtax()

    def computegvtax(self, *args):
        global allvat, allewt, percent
        allvat, allewt = [], []
        for i in range(5):
            if alltaxes[i].get() in atc and float(allamounts[i].get().replace(",","")) > 0:
                percent = float(alltaxes[i].get().split('-')[1])/100
                base = float(allamounts[i].get().replace(',',''))
                if alltaxes[i].get().split('-')[0] == 'WV':
                    vatamount = (base/1.12)*.12
                    ewtamount = (base/1.12)*percent
                elif alltaxes[i].get().split('-')[0] == 'NV':
                    vatamount = 0
                    ewtamount = base*percent
                elif alltaxes[i].get().split('-')[0] == 'V':
                    vatamount = (base/1.12)*percent
                    ewtamount = 0
                else:
                    vatamount = 0
                    ewtamount = 0
            else:
                vatamount = 0
                ewtamount = 0
            allvat.append(vatamount)
            allewt.append(ewtamount)
            allnets[i].set(format(float(allamounts[i].get().replace(",",""))-ewtamount, ",.2f"))
        self.updategvtotal()

    def updategvtotal(self):
        allnetamounts = []
        for i in allnets:
            allnetamounts.append(float(i.get().replace(",","")))
        total.set(format(sum(allnetamounts), ',.2f'))
        vat.set('VAT: ' + format(sum(allvat), ',.2f'))
        ewt.set('EWT: ' + format(sum(allewt), ',.2f'))
        # showbutton2307()

    def fillsignature2(self):
        global siglister
        c.execute('SELECT * FROM signatory')
        result = c.fetchall()
        siglister = []
        for i in result:
            siglister.append(i[1])
        siglister.sort()
        if result:
            signatoryentry1.config(values = siglister)
            signatoryentry2.config(values = siglister)
        else:
            signatoryentry1.config(values = ['no records'])
            signatoryentry2.config(values = ['no records'])

    def inbanklist(self, widget, banklist, *args):
        if widget.get() in banklist:
            pass
        else:
            widget.delete(0, END)

    def destbank(self, *args):
        destbank = []
        if bankentry.get() in dmbanks[0:4]:
            destbank = dmbanks[0:4]
            destbank.pop(destbank.index(bankentry.get()))
        elif bankentry.get() in dmbanks[4:7]:
            destbank = dmbanks[4:7]
            destbank.pop(destbank.index(bankentry.get()))
        else:
            destbank = []
        destinationentry.config(values = destbank)

    def dmassembly(self, *args):
        global payeeprefix, particulars
        self.numbertowords(amount)
        if documententry.get() == 'Letter of Instruction':
            dmintro = loiparticulars1
            dmoutro = loiparticulars2
        elif documententry.get() == 'DM - DBP Davao':
            dmintro = "Kindly debit today, " + datetime.datetime.strptime(dateentry.get(), '%m-%d-%Y').strftime('%B %d, %Y') + ", our Demand Deposit Account No. 0-01031-915-6 for "
        else:
            dmintro = dmparticulars1
            dmoutro = dmparticulars2
        if documententry.get() != 'DM - DBP Davao':
            if accountentry.get() != '':
                payeeprefix = 'payable to '
                particulars = payeeentry.get().upper() + ', assigned at ' + assignmententry.get().upper() + ', representing ' + particularsentry.get() + '. ' + dmoutro + accountentry.get() + '.'
            else:
                payeeprefix = 'representing '
                particulars = particularsentry.get() + ' to assigned staff at ' + assignmententry.get().upper() + ', details per attached payroll listing.'
        else:
            payeeprefix = payeeentry.get().upper()
            particulars = " with account no. " + accountentry.get() + ", representing " + particularsentry.get()
        try:
            if documententry.get() != 'DM - DBP Davao':
                content.set(dmintro + words + ' (P ' + amount.get() + ') ' + payeeprefix + particulars)
            else:
                content.set(dmintro + words + ' (P ' + amount.get() + ') and credit same amount to ' + payeeprefix + particulars)
        except:
            content.set(dmintro + "input amount" + ' (P ' + amount.get() + ') ' + "payable to/representing " + "input particulars and assignment")
        self.uppercase(payeeentry)
        self.uppercase(assignmententry)

    def showdmcontent(self):
        global content, contentframe
        contentframe = LabelFrame(lineframes, text = 'Content', font = fonts, bg = wc, fg = fc)
        contentframe.grid(column = 0, row = 0)

        content = StringVar()
        contentlabel = Label(contentframe, textvariable = content, font = fonts, bg = '#FDEA79', fg = codetitlefg, wraplength = 890, justify = LEFT)
        contentlabel.grid(column = 0, row = 0)
    
    def intypelist(self, *args):
        if typeentry.get() in typelister:
            self.chosendmtxn()
        else:
            typeentry.delete(0, END)

    def showft(self):
        global destinationentry
        destinationlabel = Label(disbursementssubframe, text = 'Destination', font = fonts, bg = wc, fg = fc)
        destinationlabel.grid(column = 2, row = 1, sticky = W, pady = pad)

        destinationentry = tk.Combobox(disbursementssubframe, font = fonts, width = 59)
        destinationentry.grid(column = 3, row = 1, sticky = W)

    ### finder ###
    def showfmdfinder(self, *args):
        self.fmdmenubuttons(DISABLED)
        self.colorswitchfmd(finder_button, fc)
        global fmdreportsframe, find_list, find_entry, fmdviewerframe, exportbutton
        fmdreportsframe = LabelFrame(self.master, text = '\nFinder', font = fonts, bg = wc, fg = fc)
        fmdreportsframe.grid(column = 1, row = 0, sticky = NW)

        fmdfindersubframe = Frame(fmdreportsframe, bg = wc)
        fmdfindersubframe.grid(column = 0, row = 0, sticky = NW)

        fmdviewerframe = Frame(fmdreportsframe, bg = wc)
        fmdviewerframe.grid(column = 0, row = 2, sticky = N)

        fmdbuttonsframe = Frame(fmdreportsframe, bg = wc)
        fmdbuttonsframe.grid(column = 0, row = 3, pady = pad, sticky = N)

        findby_label = Label(fmdfindersubframe, text = "Find by", font = fonts, bg = wc, fg = fc)
        findby_label.grid(column = 0, row = 0, padx = pad)

        findtypes = ["Payee","Cheque no."]
        find_list = tk.Combobox(fmdfindersubframe, values = findtypes, font = fonts, width = 10)
        find_list.grid(column = 1, row = 0, padx = pad)

        find_entry = Entry(fmdfindersubframe, font = fonts, width = 30)
        find_entry.grid(column = 2, row = 0, padx = pad, pady = pad)
        find_entry.bind("<Return>", self.findviewer)

        viewbutton = Button(fmdfindersubframe, text = "Find", font = fonts, bg = wc, bd = 0, width = 95, image = buttonicon, compound = CENTER, cursor = "hand2", command = self.findviewer)
        viewbutton.grid(column = 3, row = 0, pady = pad, padx = pad)
        viewbutton.bind("<Return>", self.findviewer)

        closefmdbutton = Button(fmdfindersubframe, text = "Close", font = fonts, bg = wc, bd = 0, width = 95, image = buttonicon, compound = CENTER, cursor = "hand2", command = self.closefmdfinder)
        closefmdbutton.grid(column = 4, row = 0, pady = pad, padx = pad)
        closefmdbutton.bind("<Return>", self.closefmdfinder)

        closebutton = Button(fmdbuttonsframe, text = "Close", font = fonts, bg = wc, bd = 0, width = 95, image = buttonicon, compound = CENTER, cursor = "hand2", command = self.findcloser)   
        closebutton.bind("<Return>", self.findcloser)

    def findviewer(self, *args):
        selectedlist = find_list.get()
        findentered = find_entry.get()
        if findentered != "":
            if selectedlist == "Payee":
                find = "SELECT company, date, number, payee, particulars, bank, mode, net, status FROM disbursements WHERE company = ? AND payee LIKE ?"
            elif selectedlist == "Cheque no.":
                find = "SELECT company, date, number, payee, particulars, bank, mode, net, status FROM disbursements WHERE company = ? AND mode LIKE ?"
            c.execute(find, [company, f"%{findentered}%"])
            result = c.fetchall()
            if result:
                self.showdisblabels(fmdviewerframe, 0, 1)
                self.scrollreports(fmdviewerframe, result, 0, 2)
            else:
                messagebox.showerror("FMD Finder", "No match found!")
                selectedlist = find_list.get()
                findentered = find_entry.get()
                fmdreportsframe.destroy()
                self.showfmdfinder()
                find_list.insert(0, selectedlist)
                find_entry.insert(0, findentered)
                find_entry.focus()
        else:
            messagebox.showerror("FMD Finder", "What are you looking for?")

    ### reports ###       
    def showfmdreports(self, *args):
        self.fmdmenubuttons(DISABLED)
        self.colorswitchfmd(reports_button, fc)
        global fmdreportsframe, fmdreportssubframe, fmdbuttonsframe, fmdviewerframe, report_list, closebutton, viewbutton, exportbutton 
        fmdreportsframe = LabelFrame(self.master, text = '\nReports', font = fonts, bg = wc, fg = fc)
        fmdreportsframe.grid(column = 1, row = 0, sticky = NW)

        fmdreportssubframe = Frame(fmdreportsframe, bg = wc)
        fmdreportssubframe.grid(column = 0, row = 0, sticky = NW)

        fmdviewerframe = Frame(fmdreportsframe, bg = wc)
        fmdviewerframe.grid(column = 0, row = 2, sticky = N)

        fmdbuttonsframe = Frame(fmdreportsframe, bg = wc)
        fmdbuttonsframe.grid(column = 0, row = 3, pady = pad, sticky = N)

        reporttypes = ["Disbursement report","Unaccounted GV/DM","ISO report"]
        report_list = tk.Combobox(fmdreportssubframe, values = reporttypes, font = fonts, width = 30)
        report_list.grid(column = 0, row = 0, padx = pad)
        report_list.bind("<<ComboboxSelected>>", self.showselected)

        closefmdbutton = Button(fmdreportssubframe, text = "Close", font = fonts, bg = wc, bd = 0, width = 95, image = buttonicon, compound = CENTER, cursor = "hand2", command = self.closefmdreport)
        closefmdbutton.grid(column = 1, row = 0, pady = pad, padx = pad)
        closefmdbutton.bind("<Return>", self.closefmdreport)

        viewbutton = Button(fmdbuttonsframe, text = "View", font = fonts, bg = wc, bd = 0, width = 95, image = buttonicon, compound = CENTER, cursor = "hand2", command = self.fmdreportviewer)
        viewbutton.bind("<Return>", self.fmdreportviewer)

        exportbutton = Button(fmdbuttonsframe, text = "Export", font = fonts, bg = wc, bd = 0, width = 95, image = buttonicon, compound = CENTER, cursor = "hand2", state = DISABLED, command = self.extractreport)
        exportbutton.bind("<Return>", self.extractreport)

        closebutton = Button(fmdbuttonsframe, text = "Close", font = fonts, bg = wc, bd = 0, width = 95, image = buttonicon, compound = CENTER, cursor = "hand2", command = self.closedisbreport)   
        closebutton.bind("<Return>", self.closedisbreport)

    def fmdreportviewer(self, *args):
        global result
        if report_list.get() == "Disbursement report":
            if monthlist.get() != "ALL" and daylist.get() != "ALL" and typelist.get() != "ALL":
                finder = "SELECT company, date, number, payee, particulars, bank, mode, net, status FROM disbursements WHERE company = ? AND year = ? AND month = ? AND day = ? AND type = ? ORDER BY number"
                c.execute(finder, [companyentry.get(), yearlist.get(), monthlist.get(), daylist.get(), typelist.get()])
            elif monthlist.get() == "ALL" and daylist.get() != "ALL" and typelist.get() != "ALL":
                finder = "SELECT company, date, number, payee, particulars, bank, mode, net, status FROM disbursements WHERE company = ? AND year = ? AND type = ? ORDER BY number"
                c.execute(finder, [companyentry.get(), yearlist.get(), typelist.get()])
            elif monthlist.get() == "ALL" and daylist.get() != "ALL" and typelist.get() == "ALL":
                finder = "SELECT company, date, number, payee, particulars, bank, mode, net, status FROM disbursements WHERE company = ? AND year = ? ORDER BY number"
                c.execute(finder, [companyentry.get(), yearlist.get()])
            elif monthlist.get() != "ALL" and daylist.get() != "ALL" and typelist.get() == "ALL":
                finder = "SELECT company, date, number, payee, particulars, bank, mode, net, status FROM disbursements WHERE company = ? AND year = ? AND month = ? AND day = ? ORDER BY number"
                c.execute(finder, [companyentry.get(), yearlist.get(), monthlist.get(), daylist.get()])
            elif monthlist.get() != "ALL" and daylist.get() == "ALL" and typelist.get() == "ALL":
                finder = "SELECT company, date, number, payee, particulars, bank, mode, net, status FROM disbursements WHERE company = ? AND year = ? AND month = ? ORDER BY number"
                c.execute(finder, [companyentry.get(), yearlist.get(), monthlist.get()])
            elif monthlist.get() != "ALL" and daylist.get() == "ALL" and typelist.get() != "ALL":
                finder = "SELECT company, date, number, payee, particulars, bank, mode, net, status FROM disbursements WHERE company = ? AND year = ? AND month = ? AND type = ? ORDER BY number"
                c.execute(finder, [companyentry.get(), yearlist.get(), monthlist.get(), typelist.get()])
            try:
                labelframe.destroy()
                scrollbox.destroy()
            except:
                pass
            result = c.fetchall()
            if result:
                exportbutton.config(state = NORMAL)
                self.showdisblabels(fmdviewerframe, 0, 1)
                self.scrollreports(fmdviewerframe, result, 0, 2)
            else:
                messagebox.showerror("FMD Reports", "No records found!")
                exportbutton.config(state = DISABLED)

    def scrollreports(self, master, result, cl, rw):
        global scrollable_frame, scrollbox
        scrollbox = Frame(master)
        container = Frame(scrollbox)
        canvas = Canvas(container, bg = wc)
        scrollbar = Scrollbar(container, orient = "vertical", width = 20, command = canvas.yview)
        scrollable_frame = Frame(canvas)
        scrollable_frame.bind("<Configure>", lambda e: canvas.configure(scrollregion = canvas.bbox("all")))
        canvas.create_window((0,0), window = scrollable_frame, anchor = NW)
        canvas.configure(yscrollcommand = scrollbar.set)

        total = []
        row = 0
        for v in result:
            if v[8] == "VALID":
                validparticulars = v[4].split("_")[0]
                validnet = format(v[7], ",.2f")
                fortotal = v[7]
                validbg = None
            else:
                validparticulars = f"{v[8]} {v[4].split('_')[0]}"
                validnet = format(0, ",.2f")
                fortotal = 0
                validbg = cancelleddocument
            com = Entry(scrollable_frame, width = 5, font = fontreports, bg = validbg, relief = RIDGE)
            com.grid(column = 0, row = row)
            com.insert(0, v[0])

            dat = Entry(scrollable_frame, width = 10, font = fontreports, bg = validbg, relief = RIDGE)
            dat.grid(column = 1, row = row)
            dat.insert(0, v[1])

            num = Entry(scrollable_frame, width = 9, font = fontreports, bg = validbg, relief = RIDGE)
            num.grid(column = 2, row = row)
            num.insert(0, v[2])

            pay = Entry(scrollable_frame, width = 20, font = fontreports, bg = validbg, relief = RIDGE)
            pay.grid(column = 3, row = row)
            pay.insert(0, v[3].split("_")[0])

            par = Entry(scrollable_frame, width = 41, font = fontreports, bg = validbg, relief = RIDGE)
            par.grid(column = 4, row = row)
            par.insert(0, validparticulars)

            ban = Entry(scrollable_frame, width = 10, font = fontreports, bg = validbg, relief = RIDGE)
            ban.grid(column = 5, row = row)
            ban.insert(0, v[5])

            mod = Entry(scrollable_frame, width = 10, font = fontreports, bg = validbg, relief = RIDGE)
            mod.grid(column = 6, row = row)
            mod.insert(0, v[6])

            net = Entry(scrollable_frame, width = 15, font = fontreports, bg = validbg, justify = RIGHT, relief = RIDGE)
            net.grid(column = 7, row = row)
            net.insert(0, validnet)
            total.append(fortotal)
            row += 1

        net = Entry(scrollable_frame, width = 15, font = fontreports, justify = RIGHT, relief = RIDGE)
        net.grid(column = 7, row = row)
        net.insert(0, format(sum(total), ",.2f"))

        scrollbox.grid(column = cl, row = rw, sticky = NW)
        container.pack()
        canvas.pack(side = LEFT, fill = "both", expand = True)
        canvas.config(height = 300, width = 872)
        scrollbar.pack(side = RIGHT, fill = "y")

    def extractreport(self, *args):
        if companyentry.get() == 'DBPSC':
            displaycompany = 'DBP SERVICE CORPORATION'
        else:
            displaycompany = 'DBPSC SECURITY SERVICE, INC.'
        wb = openpyxl.load_workbook(path + 'reports.xlsx') #path
        st = wb.active
        st['A1'] = displaycompany
        try:
            dailydate = datetime.datetime.strptime(f'{monthlist.get()}-{daylist.get()}-{yearlist.get()}', '%m-%d-%Y').strftime('%B %d, %Y')
        except:
            dailydate = f'{monthlist.get()}-{yearlist.get()}'
        st['B5'] = dailydate
        total = []
        rw = 8
        for i in result:
            if i[8] == 'VALID':
                validamount = i[7]
                validlabel = ''
            else:
                validamount = 0
                validlabel = '[CANCELLED]'
            st['A'+str(rw)] = i[2]
            st['B'+str(rw)] = i[3].split(' -- ')[0]
            st['C'+str(rw)] = validlabel + i[4].split(' -- ')[0]
            st['D'+str(rw)] = i[5]
            st['E'+str(rw)] = i[6]
            st['F'+str(rw)] = validamount
            total.append(validamount)
            rw += 1
        st['E'+str(rw)] = 'Total'
        st['F'+str(rw)] = sum(total)
        st['A'+str(rw+2)] = 'Prepared by:'
        st['A'+str(rw+4)] = 'JHAY L. BANGALAN'
        st['A'+str(rw+5)] = 'Disbursement Assistant - FMD'
        wb.save(savepath + 'test.xlsx')
        os.startfile(savepath + 'test.xlsx', 'open')

    def showdisblabels(self, master, cl, rw):
        global labelframe
        labelframe = Frame(master, bg = wc)
        labelframe.grid(column = cl, row = rw, sticky = NW, pady = pad)

        comlabel = Label(labelframe, text = 'Book', font = fontreports, width = 5, bg = wc, fg = fc, relief = RIDGE)
        comlabel.grid(column = 0, row = 0)

        datlabel = Label(labelframe, text = 'Date', font = fontreports, width = 10, bg = wc, fg = fc, relief = RIDGE)
        datlabel.grid(column = 1, row = 0)

        numlabel = Label(labelframe, text = 'Number', font = fontreports, width = 9, bg = wc, fg = fc, relief = RIDGE)
        numlabel.grid(column = 2, row = 0)

        paylabel = Label(labelframe, text = 'Payee', font = fontreports, width = 20, bg = wc, fg = fc, relief = RIDGE)
        paylabel.grid(column = 3, row = 0)

        parlabel = Label(labelframe, text = 'Particulars', font = fontreports, width = 40, bg = wc, fg = fc, relief = RIDGE)
        parlabel.grid(column = 4, row = 0)

        banlabel = Label(labelframe, text = 'Bank', font = fontreports, width = 10, bg = wc, fg = fc, relief = RIDGE)
        banlabel.grid(column = 5, row = 0)

        modlabel = Label(labelframe, text = 'Mode', font = fontreports, width = 9, bg = wc, fg = fc, relief = RIDGE)
        modlabel.grid(column = 6, row = 0)

        netlabel = Label(labelframe, text = 'Total', font = fontreports, width = 15, bg = wc, fg = fc, relief = RIDGE)
        netlabel.grid(column = 7, row = 0)

    def showselected(self, *args):
        if report_list.get() == "Disbursement report":
            self.showdailyreport()
        elif report_list.get() == "Unaccounted GV/DM":
            report_list.config(state = DISABLED)
            self.showunforwardedreport(fmdviewerframe)
            print_button.grid_forget()
            close_button.grid_forget()
            viewbutton.config(state = DISABLED)
            # viewbutton.config(command = self.viewunforwarded)
            # viewbutton.bind("<Return>", self.viewunforwarded)
            exportbutton.config(command = self.printunforwarded, state = NORMAL)
            exportbutton.bind("<Return>", self.printunforwarded)
            self.fmdyearlister2()
        elif report_list.get() == "ISO report":
            self.showisoreport(fmdviewerframe)
            self.fmdyearlister2()
            viewbutton.config(command = self.viewisoreport)
            viewbutton.bind("<Return>", self.viewisoreport)
            exportbutton.config(command = self.printisoreport, state = DISABLED)
            exportbutton.bind("<Return>", self.printisoreport)

        viewbutton.grid(column = 0, row = 0, pady = pad, padx = pad)
        exportbutton.grid(column = 1, row = 0, pady = pad, padx = pad)
        closebutton.grid(column = 2, row = 0, pady = pad, padx = pad)

    def showisoreport(self, master, *args):
        report_list.config(state = DISABLED)
        global isoreportframe, year_entry, month_list
        isoreportframe = LabelFrame(master, text = "ISO Report", font = fonts, bg = wc, fg = fc)
        isoreportframe.grid(column = 0, row = 0)

        year_label = Label(isoreportframe, text = "Year", font = fonts, width = 5, bg = wc, fg = fc)
        year_label.grid(column = 0, row = 0, padx = 2)

        month_label = Label(isoreportframe, text = "Month", font = fonts, width = 8, bg = wc, fg = fc)
        month_label.grid(column = 1, row = 0, padx = 2)

        year_entry = tk.Combobox(isoreportframe, font = fonts, width = 5)
        year_entry.grid(column = 0, row = 1, padx = 2)
        year_entry.bind("<<ComboboxSelected>>", self.fmdmonthlister2)

        month_list = tk.Combobox(isoreportframe, font = fonts, width = 8)
        month_list.grid(column = 1, row = 1, padx = 2)

    def viewisoreport(self, *args):
        global isolabels, result, nets
        find = "SELECT type, SUM(net) FROM disbursements WHERE company = ? AND month = ? AND year = ? AND status = 'VALID' GROUP BY type"
        c.execute(find, [company, month_list.get(), year_entry.get()])
        result = c.fetchall()
        if result:
            try:
                isolabels.destroy()
                scrollbox.destroy()
            except:
                pass
            isolabels = Frame(fmdviewerframe, bg = wc)
            isolabels.grid(column = 0, row = 1, sticky = NW, pady = pad)

            typelabel = Label(isolabels, text = "Type", font = fonts, bg = wc, fg = fc, relief = RIDGE, width = 40, anchor = CENTER)
            typelabel.grid(column = 0, row = 0)
        
            netlabel = Label(isolabels, text = "Total", font = fonts, bg = wc, fg = fc, relief = RIDGE, width = 15, anchor = CENTER)
            netlabel.grid(column = 1, row = 0)

            exportbutton.config(state = NORMAL)
            self.itemscroller(fmdviewerframe, 2)
            canvas.config(width = 450)
            nets = []
            row = 0
            for i in result:
                typevar = StringVar()
                typeentry = Entry(scrollable_frame, textvariable = typevar, font = fonts, width = 40, state = DISABLED, disabledbackground = codetitlebg, disabledforeground = codetitlefg)
                typeentry.grid(column = 0, row = row)
                typevar.set(i[0])

                netentry = Entry(scrollable_frame, font = fonts, width = 15, justify = RIGHT)
                netentry.grid(column = 1, row = row)
                netentry.insert(0, format(i[1], ",.2f"))
                nets.append(i[1])
                row += 1
            totalentry = Entry(scrollable_frame, font = fonts, width = 15, justify = RIGHT)
            totalentry.grid(column = 1, row = row)
            totalentry.insert(0, format(sum(nets), ",.2f"))
        else:
            try:
                exportbutton.config(state = DISABLED)
                isolabels.destroy()
                scrollbox.destroy()
            except:
                pass

    def printisoreport(self, *args):
        wb = openpyxl.load_workbook(path + "report.xlsx") #path
        st = wb.active
        st.append([company])
        st.append(["ISO Report"])
        st.append([f"for the month of {month_list.get()}-{year_entry.get()}"])
        st.append(["Type","Total"])
        for i in result:
            st.append([i[0],i[1]])
        st.append(["",sum(nets)])
        wb.save(savepath + "isoreport.xlsx") #path
        os.startfile(savepath + "isoreport.xlsx", "open") #path

    def showdailyreport(self):
        report_list.config(state = DISABLED)
        global companyentry, yearlist, monthlist, daylist, typelist, disbreportframe
        disbreportframe = LabelFrame(fmdviewerframe, text = "Disbursement Report", font = fonts, bg = wc, fg = fc)
        disbreportframe.grid(column = 0, row = 0, sticky = NW)

        companylabel = Label(disbreportframe, text = 'Company', font = fonts, bg = wc, fg = fc)
        companylabel.grid(column = 0, row = 0, pady = pad, sticky = NW)

        yearlabel = Label(disbreportframe, text = 'Year', font = fonts, bg = wc, fg = fc)
        yearlabel.grid(column = 0, row = 1, pady = pad, sticky = NW)

        monthlabel = Label(disbreportframe, text = 'Month', font = fonts, bg = wc, fg = fc)
        monthlabel.grid(column = 0, row = 2, pady = pad, sticky = NW)

        daylabel = Label(disbreportframe, text = 'Day', font = fonts, bg = wc, fg = fc)
        daylabel.grid(column = 0, row = 3, pady = pad, sticky = NW)

        typelabel = Label(disbreportframe, text = 'Type', font = fonts, bg = wc, fg = fc)
        typelabel.grid(column = 0, row = 4, pady = pad, sticky = NW)

        companyentry = tk.Combobox(disbreportframe, values = ['DBPSC','DSSI'], font = fonts, width = 15)
        companyentry.grid(column = 1, row = 0, padx = pad, sticky = E)
        companyentry.insert(0, company)
        companyentry.config(state = DISABLED)
        companyentry.bind("<<ComboboxSelected>>", self.fmdyearlister)
        companyentry.bind("<FocusOut>", self.fmdyearlister)

        yearlist = tk.Combobox(disbreportframe, font = fonts, width = 15)
        yearlist.grid(column = 1, row = 1, padx = pad, sticky = E)
        yearlist.bind("<<ComboboxSelected>>", self.fmdmonthlister)
        yearlist.bind("<FocusOut>", self.fmdmonthlister)

        monthlist = tk.Combobox(disbreportframe, font = fonts, width = 15)
        monthlist.grid(column = 1, row = 2, padx = pad, sticky = E)
        monthlist.bind("<<ComboboxSelected>>", self.fmddaylister)
        monthlist.bind("<FocusOut>", self.fmddaylister)

        daylist = tk.Combobox(disbreportframe, font = fonts, width = 15)
        daylist.grid(column = 1, row = 3, padx = pad, sticky = E)
        daylist.bind("<<ComboboxSelected>>", self.fmdtypelister)
        daylist.bind("<FocusOut>", self.fmdtypelister)

        typelist = tk.Combobox(disbreportframe, font = fonts, width = 15)
        typelist.grid(column = 1, row = 4, padx = pad, sticky = E)

        self.fmdyearlister()
        
    def fmdyearlister(self, *args):
        fmdyears = []
        finder = "SELECT year FROM disbursements WHERE company = ?"
        c.execute(finder, [companyentry.get()])
        result = c.fetchall()
        if result:
            yearlist.delete(0, END)
            monthlist.delete(0, END)
            daylist.delete(0, END)
            typelist.delete(0, END)
            for y in result:
                fmdyears.append(y[0])
            fmdyears.sort()
            yearlist.config(values = list(dict.fromkeys(fmdyears)))
        else:
            yearlist.config(values = "none")
        self.fmdmonthlister()
        self.fmddaylister()
        self.fmdtypelister()

    def fmdmonthlister(self, *args):
        fmdmonths = []
        finder = "SELECT month FROM disbursements WHERE company = ? AND year = ?"
        c.execute(finder, [companyentry.get(), yearlist.get()])
        result = c.fetchall()
        if result:
            monthlist.delete(0, END)
            daylist.delete(0, END)
            typelist.delete(0, END)
            fmdmonths.append("ALL")
            for y in result:
                fmdmonths.append(y[0])
            fmdmonths.sort()
            monthlist.config(values = list(dict.fromkeys(fmdmonths)))
        else:
            monthlist.config(values = "none")
        self.fmddaylister()
        self.fmdtypelister()

    def fmddaylister(self, *args):
        if monthlist.get() == "ALL":
            daylist.delete(0, END)
            daylist.config(state = DISABLED)
        else:
            daylist.config(state = NORMAL)
            fmddays = []
            finder = "SELECT day FROM disbursements WHERE company = ? AND year = ? AND month = ?"
            c.execute(finder, [companyentry.get(), yearlist.get(), monthlist.get()])
            result = c.fetchall()
            if result:
                daylist.delete(0, END)
                typelist.delete(0, END)
                fmddays.append("ALL")
                for y in result:
                    fmddays.append(y[0])
                fmddays.sort()
                daylist.config(values = list(dict.fromkeys(fmddays)))
            else:
                daylist.config(values = "none")
        self.fmdtypelister()

    def fmdtypelister(self, *args):
        fmdtypes = []
        if daylist.get() == "ALL":
            finder = "SELECT type FROM disbursements WHERE company = ? AND year = ? AND month = ?"
            c.execute(finder, [companyentry.get(), yearlist.get(), monthlist.get()])
        elif monthlist.get() == "ALL":
            finder = "SELECT type FROM disbursements WHERE company = ? AND year = ?"
            c.execute(finder, [companyentry.get(), yearlist.get()])
        else:
            finder = "SELECT type FROM disbursements WHERE company = ? AND year = ? AND month = ? AND day = ?"
            c.execute(finder, [companyentry.get(), yearlist.get(), monthlist.get(), daylist.get()])
        result = c.fetchall()
        if result:
            typelist.delete(0, END)
            typelist.insert(0, "ALL")
            fmdtypes.append("ALL")
            for y in result:
                fmdtypes.append(y[0])
            fmdtypes.sort()
            typelist.config(values = list(dict.fromkeys(fmdtypes)))
        else:
            typelist.config(values = "none")

    def fmdback(self, *args):
        fmd_frame.destroy()
        greeting.destroy()
        root.resizable(width = True, height = False)
        root.geometry(geometry)
        root.resizable(width = False, height = False)
        self.menu()
        if access[1] == "FMD":
            accounting_button.config(state = DISABLED)
            bcd_button.config(state = DISABLED)
            gsad_button.config(state = DISABLED)

    def fmdmenubuttons(self, status, *args):
        disbursement_button.config(state = status)
        finder_button.config(state = status)
        sms_button.config(state = status)
        reports_button.config(state = status)
        # claims_button.config(state = status)
        viewer_button.config(state = status)
        back.config(state = status)
        logout_button.config(state = status)
        exit_button.config(state = status)

    def findcloser(self, *args):
        fmdreportsframe.destroy()
        self.showfmdfinder()
        report_list.config(state = NORMAL)

    def closefmdreport(self, *args):
        fmdreportsframe.destroy()
        self.fmdmenubuttons(NORMAL)
        self.colorswitchfmd(reports_button, fc)

    def closedisbreport(self, *args):
        fmdreportsframe.destroy()
        self.showfmdreports()
        report_list.config(state = NORMAL)

    def closefmdviewer(self, *args):
        fmdviewerframe.destroy()
        self.fmdmenubuttons(NORMAL)
        self.colorswitchfmd(viewer_button, fc)

    def closefmdfinder(self, *args):
        fmdreportsframe.destroy()
        self.fmdmenubuttons(NORMAL)
        self.colorswitchfmd(finder_button, fc)

    def closefmdmenu(self, frame, button, *args):
        frame.destroy()
        self.fmdmenubuttons(NORMAL)
        self.colorswitchfmd(button, fc)

### BILLING factory ###
    def showbillimporter(self, *args):
        global client_type
        self.menubuttons(DISABLED)
        self.colorswitch(receivable_button, fc)
        global newtype, billing_frame, book
        newtype = "R"
        book = "SJ"
        billing_frame = LabelFrame(self.master, text = "\nBilling Importer", font = fonts, bg = wc, fg = fc)
        billing_frame.grid(column = 1, row = 0, sticky = NW)

        import_frame = LabelFrame(billing_frame, text = "Import", font = fonts, bg = wc, fg = fc)
        import_frame.grid(column = 0, row = 0, sticky = NW)

        global paydate_type
        paydate = os.listdir(savepath + "payroll_files") #path
        paydate_type = tk.Combobox(import_frame, font = fonts, values = paydate, width = 9)
        paydate_type.grid(column = 0, row = 0, sticky = W, padx = 3)
        paydate_type.insert(0, "paydate")
        paydate_type.bind("<<ComboboxSelected>>", self.clientdir)
        paydate_type.bind("<FocusOut>", self.clientdir)
        
        global client_name
        client_name = tk.Combobox(import_frame, font = fonts, width = 18)
        client_name.grid(column = 1, row = 0, padx = 3)
        client_name.insert(0, "client")

        import_button = Button(import_frame, text = "Import", font = fonts, bg = buttonbg, width = 10, command = self.importbilling)
        import_button.grid(column = 2, row = 0, padx = 8)
        import_button.bind("<Return>", self.importbilling)
        
        global post_button
        post_button = Button(import_frame, text = "Post", font = fonts, bg = buttonbg, width = 10, command = self.debitcreditchecker)
        post_button.grid(column = 4, row = 0, padx = 8)
        post_button.bind("<Return>", self.debitcreditchecker)
        post_button.config(state = DISABLED)

        global print_button
        print_button = Button(import_frame, text = "Print", font = fonts, bg = buttonbg, width = 10, command = self.printer)
        print_button.grid(column = 5, row = 0, padx = 8)
        print_button.bind("<Return>", self.printer)
        print_button.config(state = DISABLED)
        
        details_back = Button(import_frame, text = "Close", font = fonts, bg = buttonbg, width = 10, command = self.closebillimporter)
        details_back.grid(column = 6, row = 0, padx = 8)
        details_back.bind("<Return>", self.closebillimporter)

    def importbilling(self, *args):
        try:
            details_frame.destroy()
            details.destroy()
            entries_frame.destroy()
            status_frame.destroy()
            scrollbox.desroy()
        except:
            pass
        finally:
            global addlist, getclient
            addlist = []
            getclient = str(client_name.get())   
            addlist.append(getclient)   
            self.showpayrolldetails(billing_frame)  
            self.billingmaster(addlist)
            self.labels(billing_frame, 2)
            self.showalllines(billing_frame, NORMAL)
            self.billarranger()
            self.billlister()
            code_entry1.focus()

    def closebillimporter(self, *args):
        billing_frame.destroy()
        self.menubuttons(NORMAL)
        self.colorswitch(receivable_button, fc)

    def billingmaster(self, addlist):
        global pclient, pperiod, clerical, janitorial, security, maintenance, technical, billamount
        for payroll in addlist:
            wb = openpyxl.load_workbook(savepath + "payroll_files/" + paydate_type.get() + "/" + payroll) #path
            s1 = wb.active
            pclient = s1['A5'].value.split('NAME: ')[1]
            pperiod = s1['A6'].value.split('Period : ')[1]
            clerical = s1['T39'].value
            janitorial = s1['T41'].value
            security = s1['T43'].value
            maintenance = s1['T45'].value
            technical = s1['T47'].value
            billamount = s1['T50'].value
            wb.close()

    def billarranger(self):
        global receivable, clericalincome, janitorialincome, securityincome, jfcincome
        receivable = billamount
        clericalincome = clerical
        janitorialincome = sum([janitorial, maintenance, technical])
        securityincome = security
        jfcincome = sum([clerical, janitorial, maintenance, technical])

    def billlister(self):
        rate = 0.12
        vat = 1.12
        output = 2119
        clericalcode = 5101
        janitorialcode = 5102
        jfccode = 5105
        securitycode = 5207
        self.allvars()
        jj = ['self.journalcode_val1()','self.journalcode_val2()','self.journalcode_val3()','self.journalcode_val4()']
        client_entry.insert(0, pclient)
        particulars_entry.insert(0, "BILLING FOR THE PERIOD " + pperiod)
        findaccount = "SELECT account, tax FROM bcdclients WHERE name = ? AND company = ? LIMIT 1"
        c.execute(findaccount, [pclient, company])
        accountresult = c.fetchone()
        if accountresult:
            codelist[0].insert(0, accountresult[0])
            eval(jj[0])
            drs[0].insert(0, format(billamount, ",.2f"))
            crs[0].insert(0, 0)
            if pclient.split(" ")[0] == "BURGER" or pclient.split(" ")[0] == "MANG" or pclient.split(" ")[0] == "JOLLIBEE":
                codelist[1].insert(0, output)
                eval(jj[1])
                drs[1].insert(0, 0)
                crs[1].insert(0, format(billamount/vat*rate, ",.2f"))

                codelist[2].insert(0, jfccode)
                eval(jj[2])
                drs[2].insert(0, 0)
                crs[2].insert(0, format(jfcincome/vat, ",.2f"))
            else:
                if accountresult[1] == "V":
                    codelist[1].insert(0, output)
                    eval(jj[1])
                    drs[1].insert(0, 0)
                    crs[1].insert(0, format(billamount/vat*rate, ",.2f"))
                    if clericalincome > 0:
                        codelist[2].insert(0, clericalcode)
                        eval(jj[2])
                        drs[2].insert(0, 0)
                        crs[2].insert(0, format(clericalincome/vat, ",.2f"))
                    if janitorialincome > 0:
                        codelist[2].insert(0, janitorialcode)
                        eval(jj[2])
                        drs[2].insert(0, 0)
                        crs[2].insert(0, format(janitorialincome/vat, ",.2f"))
                    if securityincome > 0:
                        codelist[2].insert(0, securitycode)
                        eval(jj[2])
                        drs[2].insert(0, 0)
                        crs[2].insert(0, format(securityincome/vat, ",.2f"))
                else:
                    row = 1
                    if clericalincome > 0:
                        codelist[row].insert(0, clericalcode)
                        eval(jj[row])
                        drs[row].insert(0, 0)
                        crs[row].insert(0, format(clericalincome, ",.2f"))
                        row += 1
                    if janitorialincome > 0:
                        codelist[row].insert(0, janitorialcode)
                        eval(jj[row])
                        drs[row].insert(0, 0)
                        crs[row].insert(0, format(janitorialincome, ",.2f"))
                        row += 1
                    if securityincome > 0:
                        codelist[row].insert(0, securitycode)
                        eval(jj[row])
                        drs[row].insert(0, 0)
                        crs[row].insert(0, format(securityincome, ",.2f"))
        else:
            pass

### BIR factory ###
    def showbir(self, *args):
        self.menubuttons(DISABLED)
        self.colorswitch(bir_button, fc)
        self.showbirframe(self.master)
        
    def showbirframe(self, master, *args):
        global bir_frame, closebir_button, bir_list
        bir_frame = LabelFrame(master, text = "\nBIR", font = fonts, bg = wc, fg = fc)
        bir_frame.grid(column = 1, row = 0, ipadx = 22, sticky = NW)

        bir_subframe = Frame(bir_frame, bg = wc)
        bir_subframe.grid(column = 0, row = 0, sticky = NW)

        forms = ["Expanded Withholding Tax","2307 Generator", "Supplier Options"]
        bir_list = tk.Combobox(bir_subframe, font = fonts, values = forms, width = 20)
        bir_list.grid(column = 0, row = 0, padx = 10)
        bir_list.bind("<<ComboboxSelected>>", self.selectedbiroption)

        closebir_button = Button(bir_subframe, text = "Close", font = fonts, bg = wc, bd = 0, width = 95, image = buttonicon, compound = CENTER, cursor = "hand2", command = self.closebir)
        closebir_button.grid(column = 1, row = 0, pady = 5, sticky = N)
        closebir_button.bind("<Return>", self.closebir)

    def selectedbiroption(self, *args):
        if bir_list.get() == "2307 Generator":
            self.generator(bir_frame, 1)
        elif bir_list.get() == "Supplier Options":
            self.supplieroptions()
        elif bir_list.get() == "Expanded Withholding Tax":
            self.showewt()
        bir_list.config(state = DISABLED)

    def showewt(self, *args):
        global ewt_frame, addewt_button, updewt_button, ewtreports_button
        ewt_frame = LabelFrame(bir_frame, text = "Expanded Withholding Tax", font = fonts, bg = wc, fg = fc)
        ewt_frame.grid(column = 0, row = 1, sticky = NW)

        ewtsub_frame = Frame(ewt_frame, bg = wc)
        ewtsub_frame.grid(column = 0, row = 0, pady = pad, sticky = NW)

        addewt_button = Button(ewtsub_frame, text = "Add", font = fonts, bg = wc, bd = 0, width = 95, image = buttonicon, compound = CENTER, cursor = "hand2", command = self.showaddewt)
        addewt_button.grid(column = 0, row = 0, padx = pad)
        addewt_button.bind("<Return>", self.showaddewt)

        updewt_button = Button(ewtsub_frame, text = "Update", font = fonts, bg = wc, bd = 0, width = 95, image = buttonicon, compound = CENTER, cursor = "hand2", command = self.showupdewt)
        updewt_button.grid(column = 1, row = 0, padx = pad)
        updewt_button.bind("<Return>", self.showupdewt)

        ewtreports_button = Button(ewtsub_frame, text = "Reports", font = fonts, bg = wc, bd = 0, width = 95, image = buttonicon, compound = CENTER, cursor = "hand2", command = self.showewtreport)
        ewtreports_button.grid(column = 2, row = 0, padx = pad)
        ewtreports_button.bind("<Return>", self.showewtreport)

        closeewtoptions_button = Button(ewtsub_frame, text = "Close", font = fonts, bg = wc, bd = 0, width = 95, image = buttonicon, compound = CENTER, cursor = "hand2", command = lambda: self.closemanagesup(ewt_frame))
        closeewtoptions_button.grid(column = 3, row = 0, padx = pad)
        closeewtoptions_button.bind("<Return>", lambda e: self.closemanagesup(ewt_frame))

    def showaddewt(self, *args):
        self.ewtmenubuttons(DISABLED)
        global addewt_frame
        addewt_frame = LabelFrame(ewt_frame, text = "Add EWT Record", font = fonts, bg = wc, fg = fc)
        addewt_frame.grid(column = 0, row = 1, sticky = NW)

        addewt_subframe = Frame(addewt_frame, bg = wc)
        addewt_subframe.grid(column = 0, row = 0, sticky = NW)

        button_frame = Frame(addewt_frame, bg = wc)
        button_frame.grid(column = 0, row = 1, sticky = E)

        self.ewttemplate(addewt_subframe, 0)
        self.merchantlister(supplier_list)

        submit_button = Button(button_frame, text = "Submit", font = fonts, bg = wc, bd = 0, width = 95, image = buttonicon, compound = CENTER, cursor = "hand2", command = self.addewt)
        submit_button.grid(column = 0, row = 0, pady = pad, padx = pad)
        submit_button.bind("<Return>", self.addewt)

        generate_button = Button(button_frame, text = "Generate", font = fonts, bg = wc, bd = 0, width = 95, image = buttonicon, compound = CENTER, cursor = "hand2", command = self.generate)
        generate_button.grid(column = 1, row = 0, pady = pad, padx = pad)
        generate_button.bind("<Return>", self.generate)

        closeaddewt_button = Button(button_frame, text = "Close", font = fonts, bg = wc, bd = 0, width = 95, image = buttonicon, compound = CENTER, cursor = "hand2", command = lambda: self.closeewtoption(addewt_frame))
        closeaddewt_button.grid(column = 2, row = 0, pady = pad, padx = pad)
        closeaddewt_button.bind("<Return>", lambda e: self.closeewtoption(addewt_frame))

    def addewt(self, *args):
        if supplier_list.get() not in merchants or period_list.get() not in periods or sign_list.get() not in signs or float(ewtamt.get().replace(",","")) == 0 or number_entry.get() == "":
            messagebox.showerror("Add EWT Record", "Please fill-out all required details!")
        else:
            ask = messagebox.askyesno("Add EWT Record", "Are you sure?")
            if ask == True:
                addrecord = """INSERT INTO ewt (
                    company, number, date, name, period,
                    atc, expense, ewt, remarks, user,
                    signature, month, day, year)
                    values (?,?,?,?,?,?,?,?,?,?,?,?,?,?)"""
                c.execute(addrecord, [
                    company, number_entry.get(), today.strftime("%m-%d-%Y"), supplier_list.get(), period_list.get(),
                    atc_list.get(), float(exvat.get().replace(",","")), float(ewtamt.get().replace(",","")), remarks_entry.get(), self.username.get(),
                    sign_list.get(), today.strftime("%m-%d-%Y").split("-")[0], today.strftime("%m-%d-%Y").split("-")[1], today.strftime("%m-%d-%Y").split("-")[2]
                ])
                conn.commit()
                messagebox.showinfo("Add EWT Record", "A new EWT record has been saved!")
                supplier_list.config(state = DISABLED)
                period_list.config(state = DISABLED)
                atc_list.config(state = DISABLED)
                sign_list.config(state = DISABLED)
                amount_entry.config(state = DISABLED)
                ewt_entry.config(state = DISABLED)
                number_entry.config(state = DISABLED)
                remarks_entry.config(state = DISABLED)

    def computeewt(self, amt, ex, ewt, *args):
        rates = [.15,.05,.01,.02,.1,.1,.05,.1,.01,.02]
        try:
            if vat_list.get() == "VAT":
                gross = float(amt.get().replace(",",""))/1.12
                ex.set(gross)
            else:
                gross = float(amt.get().replace(",",""))
                ex.set(gross)
            ewtresult = gross*rates[atc.index(atc_list.get())]
            ewt.set(ewtresult)
        except:
            amt.set(format(0, ",.2f"))
            ex.set(0)
            ewt.set(0)
        self.amountvalidatormaster(ex)
        self.amountvalidatormaster(ewt)

    def showupdewt(self, *args):
        self.ewtmenubuttons(DISABLED)
        global ewtupdate_frame, view_frame, supplier_records, gv_records, boxtype, update_button, delete_button, generate_button
        ewtupdate_frame = LabelFrame(ewt_frame, text = "View/Update EWT Record", font = fonts, bg = wc, fg = fc)
        ewtupdate_frame.grid(column = 0, row = 1, sticky = NW)

        sub_frame = Frame(ewtupdate_frame, bg = wc)
        sub_frame.grid(column = 0, row = 0, pady = pad, sticky = NW)

        button_frame = Frame(ewtupdate_frame, bg = wc)
        button_frame.grid(column = 0, row = 2, pady = pad, sticky = E)

        supplier_label = Label(sub_frame, text = "Supplier", font = fonts, width = 10, bg = wc, fg = fc, anchor = W)
        supplier_label.grid(column = 0, row = 0, padx = pad)

        number_label = Label(sub_frame, text = "GV no.", font = fonts, width = 10, bg = wc, fg = fc, anchor = W)
        number_label.grid(column = 0, row = 1, padx = pad)

        supplier_records = tk.Combobox(sub_frame, font = fonts, width = 60)
        supplier_records.grid(column = 1, row = 0, pady = pad, sticky = W)
        supplier_records.bind("<<ComboboxSelected>>", self.fillgvrecord)

        gv_records = tk.Combobox(sub_frame, font = fonts, width = 10)
        gv_records.grid(column = 1, row = 1, pady = pad, sticky = W)
        gv_records.bind("<Return>", self.viewewtrecord)

        view_button = Button(sub_frame, text = "View", font = fonts, bg = wc, bd = 0, width = 95, image = buttonicon, compound = CENTER, cursor = "hand2", command = self.viewewtrecord)
        view_button.grid(column = 1, row = 2, pady = pad, sticky = W)
        view_button.bind("<Return>", self.viewewtrecord)

        update_button = Button(button_frame, text = "Update", font = fonts, bg = wc, bd = 0, width = 95, image = buttonicon, compound = CENTER, cursor = "hand2", command = self.updateewtrecord)
        update_button.bind("<Return>", self.updateewtrecord)

        delete_button = Button(button_frame, text = "Delete", font = fonts, bg = wc, bd = 0, width = 95, image = buttonicon, compound = CENTER, cursor = "hand2", command = self.deleteewtrecord)
        delete_button.bind("<Return>", self.deleteewtrecord)

        generate_button = Button(button_frame, text = "Generate", font = fonts, bg = wc, bd = 0, width = 95, image = buttonicon, compound = CENTER, cursor = "hand2", command = self.generate)
        generate_button.bind("<Return>", self.generate)

        closeupdewt_button = Button(button_frame, text = "Close", font = fonts, bg = wc, bd = 0, width = 95, image = buttonicon, compound = CENTER, cursor = "hand2", command = lambda: self.closeewtoption(ewtupdate_frame))
        closeupdewt_button.grid(column = 3, row = 0, padx = pad)
        closeupdewt_button.bind("<Return>", lambda e: self.closeewtoption(ewtupdate_frame))

        boxtype = "supplier"
        self.clientbox()
        self.fillewtrecords()

    def updateewtrecord(self, *args):
        if supplier_list.get() not in merchants or period_list.get() not in periods or sign_list.get() not in signs or float(ewtamt.get().replace(",","")) == 0 or number_entry.get() == "":
            messagebox.showerror("Add EWT Record", "Please fill-out all required details!")
        else:
            ask = messagebox.askyesno("Update Record", "Are you sure?")
            if ask == True:
                update = "UPDATE ewt SET date = ?, name = ?, period = ?, atc = ?, signature = ?, expense = ?, ewt = ?, number = ?, remarks = ?, month = ?, day = ?, year = ?, user = ? WHERE number = ? AND company = ?"
                c.execute(update, [
                    today.strftime("%m-%d-%Y"),supplier_list.get(), period_list.get(), atc_list.get(),
                    sign_list.get(), exvat.get(), ewtamt.get(), remarks_entry.get(),
                    today.strftime("%m-%d-%Y").split("-")[0], today.strftime("%m-%d-%Y").split("-")[1], today.strftime("%m-%d-%Y").split("-")[2], self.username.get(),
                    result[0], company
                    ])
                conn.commit()
                messagebox.showinfo("Update Record", "EWT record has been updated!")
                supplier_list.config(state = DISABLED)
                period_list.config(state = DISABLED)
                atc_list.config(state = DISABLED)
                sign_list.config(state = DISABLED)
                amount_entry.config(state = DISABLED)
                ewt_entry.config(state = DISABLED)
                number_entry.config(state = DISABLED)
                remarks_entry.config(state = DISABLED)

    def deleteewtrecord(self, *args):
        if supplier_list.get() not in merchants or period_list.get() not in periods or sign_list.get() not in signs or float(ewtamt.get().replace(",","")) == 0 or number_entry.get() == "":
            messagebox.showerror("Add EWT Record", "Please fill-out all required details!")
        else:
            ask = messagebox.askyesno("Delete Record", "Are you sure?")
            if ask == True:
                delete = "DELETE FROM ewt WHERE number = ? AND company = ?"
                c.execute(delete, [result[0]])
                conn.commit()
                messagebox.showinfo("Delete Record","EWT record has been deleted!")
                supplier_list.delete(0, END)
                period_list.delete(0, END)
                atc_list.delete(0, END)
                sign_list.delete(0, END)
                exvat.set(0)
                ewtamt.set(0)
                number_entry.delete(0, END)
                remarks_entry.delete(0, END)

    def viewewtrecord(self, *args):
        global result, view_frame
        find = "SELECT number, date, name, period, atc, expense, ewt, remarks, signature FROM ewt WHERE name = ? AND number = ? AND company = ?"
        c.execute(find, [supplier_records.get(), gv_records.get(), company])
        result = c.fetchone()
        if result:
            try:
                view_frame.destroy()
                update_button.grid_forget()
                delete_button.grid_forget()
                generate_button.grid_forget()
            except:
                pass
            finally:
                view_frame = Frame(ewtupdate_frame, bg = wc)
                view_frame.grid(column = 0, row = 1, pady = pad, sticky = NW)
                self.ewttemplate(view_frame, 0)
                self.merchantlister(supplier_list)
                supplier_list.insert(0, result[2])
                period_list.insert(0, result[3])
                atc_list.insert(0, result[4])
                sign_list.insert(0, result[8])
                exvat.set(format(result[5], ",.2f"))
                ewtamt.set(format(result[6], ",.2f"))
                number_entry.insert(0, result[0])
                remarks_entry.insert(0, result[7])

                update_button.grid(column = 0, row = 0, padx = pad)
                delete_button.grid(column = 1, row = 0, padx = pad)
                generate_button.grid(column = 2, row = 0, padx = pad)
        else:
            messagebox.showerror("View Record", "Record not found!")
            try:
                view_frame.destroy()
                update_button.grid_forget()
                delete_button.grid_forget()
                generate_button.grid_forget()
            except:
                pass

    def fillgvrecord(self, *args):
        gv_records.delete(0, END)
        find = "SELECT number FROM ewt WHERE name = ? AND company = ?"
        c.execute(find, [supplier_records.get(), company])
        result = c.fetchall()
        gvs = []
        if result:
            for i in result:
                gvs.append(i[0])
            gv_records.config(values = sorted(gvs))

    def fillewtrecords(self):
        find = "SELECT name FROM ewt WHERE company = ?"
        c.execute(find, [company])
        result = c.fetchall()
        ewtrecords = []
        if result:
            for i in result:
                ewtrecords.append(i[0])
            supplier_records.config(values = list(dict.fromkeys(sorted(ewtrecords))))

    def showewtreport(self, *args):
        self.ewtmenubuttons(DISABLED)
        global ewtreport_frame, year_list, month_list, print_button
        ewtreport_frame = LabelFrame(ewt_frame, text = "Reports", font = fonts, bg = wc, fg = fc)
        ewtreport_frame.grid(column = 0, row = 1, sticky = NW)

        sub_frame = Frame(ewtreport_frame, bg = wc)
        sub_frame.grid(column = 0, row = 0, pady = pad, sticky = NW)

        button_frame = Frame(ewtreport_frame, bg = wc)
        button_frame.grid(column = 0, row = 2, pady = pad, sticky = E)

        year_label = Label(sub_frame, text = "Year", font = fonts, width = 10, bg = wc, fg = fc, anchor = W)
        year_label.grid(column = 0, row = 0, pady = pad)

        month_label = Label(sub_frame, text = "Month", font = fonts, width = 10, bg = wc, fg = fc, anchor = W)
        month_label.grid(column = 0, row = 1, pady = pad)

        year_list = tk.Combobox(sub_frame, font = fonts, width = 10)
        year_list.grid(column = 1, row = 0)
        year_list.bind("<<ComboboxSelected>>", self.ewtmonthlister)

        month_list = tk.Combobox(sub_frame, font = fonts, width = 10)
        month_list.grid(column = 1, row = 1)

        view_button = Button(button_frame, text = "View", font = fonts, bg = wc, bd = 0, width = 95, image = buttonicon, compound = CENTER, cursor = "hand2", command = self.viewewtreport)
        view_button.grid(column = 0, row = 0, padx = pad)
        view_button.bind("<Return>", self.viewewtreport)

        print_button = Button(button_frame, text = "Export", font = fonts, bg = wc, bd = 0, width = 95, image = buttonicon, compound = CENTER, cursor = "hand2", command = self.printewtreport)
        print_button.bind("<Return>", self.printewtreport)

        close_button = Button(button_frame, text = "Close", font = fonts, bg = wc, bd = 0, width = 95, image = buttonicon, compound = CENTER, cursor = "hand2", command = lambda: self.closeewtoption(ewtreport_frame))
        close_button.grid(column = 2, row = 0, padx = pad)
        close_button.bind("<Return>", lambda e: self.closeewtoption(ewtreport_frame))

        self.ewtyearlister(year_list)

    def printewtreport(self, *args):
        wb = openpyxl.load_workbook(path + "report.xlsx") #path
        st = wb.active
        st.append([company])
        st.append(["Expanded Withholding Tax"])
        st.append([year_list.get() + "-" + month_list.get()])
        st.append(["company","number","date","tin","name","address","period","atc","expense","ewt","remarks"])
        for i in result:
            st.append([
                i[0], i[1], i[2], i[10], i[3], i[9],
                i[4], i[5], i[6], i[7], i[8]
            ])
        st.append(["","","","","","","","",sum(allexpenses),sum(allewt)])
        wb.save(savepath + "ewtreport.xlsx")
        os.startfile(savepath + "ewtreport.xlsx", "open")

    def ewtyearlister(self, year):
        find = "SELECT year FROM ewt WHERE company = ?"
        c.execute(find, [company])
        result = c.fetchall()
        years = []
        if result:
            for i in result:
                years.append(i[0])
            year.config(values = list(dict.fromkeys(sorted(years))))

    def ewtmonthlister(self, *args):
        month_list.delete(0, END)
        find = "SELECT month FROM ewt WHERE year = ? AND company = ?"
        c.execute(find, [year_list.get(), company])
        result = c.fetchall()
        months = []
        if result:
            for i in result:
                months.append(i[0])
            month_list.config(values = list(dict.fromkeys(sorted(months))))

    def ewtlabels(self, master, rw):
        global labelframe
        labelframe = Frame(master, bg = wc)
        labelframe.grid(column = 0, row = rw, sticky = NW)

        booklabel = Label(labelframe, text = 'Book', font = fontreports, width = 6, relief = RIDGE, bg = wc, fg = fc)
        booklabel.grid(column = 0, row = 0)

        numberlabel = Label(labelframe, text = 'GV no.', font = fontreports, width = 7, relief = RIDGE, bg = wc, fg = fc)
        numberlabel.grid(column = 1, row = 0)

        datelabel = Label(labelframe, text = 'Date', font = fontreports, width = 9, relief = RIDGE, bg = wc, fg = fc)
        datelabel.grid(column = 2, row = 0)

        merchantlabel = Label(labelframe, text = 'Supplier', font = fontreports, width = 24, relief = RIDGE, bg = wc, fg = fc)
        merchantlabel.grid(column = 3, row = 0)

        periodlabel = Label(labelframe, text = 'Period', font = fontreports, width = 5, relief = RIDGE, bg = wc, fg = fc)
        periodlabel.grid(column = 4, row = 0)

        atclabel = Label(labelframe, text = 'ATC', font = fontreports, width = 6, relief = RIDGE, bg = wc, fg = fc)
        atclabel.grid(column = 5, row = 0)

        expenselabel = Label(labelframe, text = 'Expense', font = fontreports, width = 15, relief = RIDGE, bg = wc, fg = fc)
        expenselabel.grid(column = 6, row = 0)

        ewtlabel = Label(labelframe, text = 'EWT', font = fontreports, width = 14, relief = RIDGE, bg = wc, fg = fc)
        ewtlabel.grid(column = 7, row = 0)

        remarkslabel = Label(labelframe, text = 'Remarks', font = fontreports, width = 15, relief = RIDGE, bg = wc, fg = fc)
        remarkslabel.grid(column = 8, row = 0)

    def viewewtreport(self, *args):
        global result, allexpenses, allewt, view_frame
        find = "SELECT company, number, date, ewt.name, period, atc, expense, ewt, remarks, merchant.address, merchant.tin FROM ewt INNER JOIN merchant ON ewt.name = merchant.name WHERE year = ? AND month = ? AND company = ? ORDER by date"
        c.execute(find, [year_list.get(), month_list.get(), company])
        result = c.fetchall()
        if result:
            view_frame = Frame(ewtreport_frame, bg = wc)
            view_frame.grid(column = 0, row = 1, pady = pad, sticky = NW)
            self.ewtlabels(view_frame, 0)
            self.itemscroller(view_frame, 1)
            print_button.grid(column = 1, row = 0, padx = pad)
            rw = 0
            allexpenses, allewt = [], []
            for i in result:
                itemframe = Frame(scrollable_frame, bg = wc)
                itemframe.grid(column = 0, row = rw, sticky = W)

                bookentry = Entry(itemframe, font = fontreports, width = 6, relief = RIDGE)
                bookentry.grid(column = 0, row = 0)
                bookentry.insert(0, i[0])
                
                numberentry = Entry(itemframe, font = fontreports, width = 7, relief = RIDGE)
                numberentry.grid(column = 1, row = 0)
                numberentry.insert(0, i[1])

                dateentry = Entry(itemframe, font = fontreports, width = 10, relief = RIDGE)
                dateentry.grid(column = 2, row = 0)
                dateentry.insert(0, i[2])

                merchantentry = Entry(itemframe, font = fontreports, width = 24, relief = RIDGE)
                merchantentry.grid(column = 3, row = 0)
                merchantentry.insert(0, i[3])

                periodentry = Entry(itemframe, font = fontreports, width = 5, relief = RIDGE)
                periodentry.grid(column = 4, row = 0)
                periodentry.insert(0, i[4])

                atcentry = Entry(itemframe, font = fontreports, width = 6, relief = RIDGE)
                atcentry.grid(column = 5, row = 0)
                atcentry.insert(0, i[5])

                expenseentry = Entry(itemframe, font = fontreports, width = 15, relief = RIDGE, justify = RIGHT)
                expenseentry.grid(column = 6, row = 0)
                expenseentry.insert(0, format(i[6], ",.2f"))
                allexpenses.append(i[6])

                ewtentry = Entry(itemframe, font = fontreports, width = 15, relief = RIDGE, justify = RIGHT)
                ewtentry.grid(column = 7, row = 0)
                ewtentry.insert(0, format(i[7], ",.2f"))
                allewt.append(i[7])

                remarksentry = Entry(itemframe, font = fontreports, width = 15, relief = RIDGE)
                remarksentry.grid(column = 8, row = 0)
                remarksentry.insert(0, i[8])

                rw += 1

            expenseentry = Entry(itemframe, font = fontreports, width = 15, relief = RIDGE, justify = RIGHT)
            expenseentry.grid(column = 6, row = rw)
            expenseentry.insert(0, format(sum(allexpenses), ",.2f"))

            ewtentry = Entry(itemframe, font = fontreports, width = 15, relief = RIDGE, justify = RIGHT)
            ewtentry.grid(column = 7, row = rw)
            ewtentry.insert(0, format(sum(allewt), ",.2f"))
        else:
            try:
                labelframe.destroy()
                view_frame.destroy()
                print_button.grid_forget()
            except:
                pass

    def generator(self, master, rw):
        global boxtype, button_frame, generator_frame, generator_subframe, generate_button, closegenerator_button
        boxtype = "supplier"
        self.clientbox()
        generator_frame = LabelFrame(master, text = "2307 Generator", font = fonts, bg = wc, fg = fc)
        generator_frame.grid(column = 0, row = 1, sticky = NW)

        generator_subframe = Frame(generator_frame, bg = wc)
        generator_subframe.grid(column = 0, row = 0, sticky = NW)

        button_frame = Frame(generator_frame, bg = wc)
        button_frame.grid(column = 0, row = 1, sticky = E)

        self.ewttemplate(generator_subframe, 0)

        generate_button = Button(button_frame, text = "Generate", font = fonts, bg = wc, bd = 0, width = 95, image = buttonicon, compound = CENTER, cursor = "hand2", command = self.generate)
        generate_button.grid(column = 1, row = 0, pady = pad, padx = pad)
        generate_button.bind("<Return>", self.generate)

        closegenerator_button = Button(button_frame, text = "Close", font = fonts, bg = wc, bd = 0, width = 95, image = buttonicon, compound = CENTER, cursor = "hand2", command = lambda: self.closemanagesup(generator_frame))
        closegenerator_button.grid(column = 2, row = 0, pady = pad, padx = pad)
        closegenerator_button.bind("<Return>", lambda e: self.closemanagesup(generator_frame))

    def ewttemplate(self, master, rw):
        global grossvat, exvat, exvat_entry, amount_entry, supplier_list, periods, period_list, atc, atc_list, signs, sign_list, vat_list, number_entry, ewtamt, ewt_entry, remarks_entry
        supplier_label = Label(master, text = "Supplier", font = fonts, width = 10, bg = wc, fg = fc, anchor = W)
        supplier_label.grid(column = 0, row = 0, pady = pad)

        period_label = Label(master, text = "Period", font = fonts, width = 10, bg = wc, fg = fc, anchor = W)
        period_label.grid(column = 0, row = 1, pady = pad)

        atc_label = Label(master, text = "ATC", font = fonts, width = 10, bg = wc, fg = fc, anchor = W)
        atc_label.grid(column = 0, row = 2, pady = pad)

        sign_label = Label(master, text = "Signatory", font = fonts, width = 10, bg = wc, fg = fc, anchor = W)
        sign_label.grid(column = 0, row = 3, pady = pad)

        vat_label = Label(master, text = "Type", font = fonts, width = 10, bg = wc, fg = fc, anchor = W)
        vat_label.grid(column = 0, row = 4, pady = pad)

        gross_label = Label(master, text = "Gross", font = fonts, width = 10, bg = wc, fg = fc, anchor = W)
        gross_label.grid(column = 0, row = 5, pady = pad)

        amount_label = Label(master, text = "Expense", font = fonts, width = 10, bg = wc, fg = fc, anchor = W)
        amount_label.grid(column = 0, row = 6, pady = pad)

        ewt_label = Label(master, text = "EWT", font = fonts, width = 10, bg = wc, fg = fc, anchor = W)
        ewt_label.grid(column = 0, row = 7, pady = pad)

        number_label = Label(master, text = "GV No.", font = fonts, width = 10, bg = wc, fg = fc, anchor = W)
        number_label.grid(column = 0, row = 8, pady = pad)

        remarks_label = Label(master, text = "Remarks", font = fonts, width = 10, bg = wc, fg = fc, anchor = W)
        remarks_label.grid(column = 0, row = 9, pady = pad)

        supplier_list = tk.Combobox(master, font = fonts, width = 60)
        supplier_list.grid(column = 1, row = 0, padx = pad, sticky = W)
        supplier_list.focus()

        periods = ["JAN","FEB","MAR","APR","MAY","JUN","JUL","AUG","SEP","OCT","NOV","DEC"]
        period_list = tk.Combobox(master, values = periods, font = fonts, width = 17)
        period_list.grid(column = 1, row = 1, padx = pad, sticky = W)

        atc = ["WC 140","WC 100","WC 158","WC 160","WI 050","WI 091","WI 100","WI 141","WI 158","WI 160"]
        atc_list = tk.Combobox(master, values = atc, font = fonts, width = 17)
        atc_list.grid(column = 1, row = 2, padx = pad, sticky = W)
        atc_list.bind("<<ComboboxSelected>>", lambda e: self.computeewt(grossvat, exvat, ewtamt))

        signs = ["MBS","CSA","AHE"]
        sign_list = tk.Combobox(master, values = signs, font = fonts, width = 17)
        sign_list.grid(column = 1, row = 3, padx = pad, sticky = W)

        vat_list = tk.Combobox(master, values = ["VAT","NONVAT"], font = fonts, width = 17)
        vat_list.grid(column = 1, row = 4, padx = pad, sticky = W)
        vat_list.bind("<<ComboboxSelected>>", lambda e: self.computeewt(grossvat, exvat, ewtamt))
        
        grossvat = StringVar()
        amount_entry = Entry(master, textvariable = grossvat, font = fonts, width = 15, justify = RIGHT)
        amount_entry.grid(column = 1, row = 5, padx = pad, sticky = W)
        amount_entry.bind("<FocusOut>", lambda e: self.computeewt(grossvat, exvat, ewtamt))

        exvat = StringVar()
        exvat_entry = Entry(master, textvariable = exvat, font = fonts, state = DISABLED, disabledbackground = codetitlebg, disabledforeground = codetitlefg, width = 15, justify = RIGHT)
        exvat_entry.grid(column = 1, row = 6, padx = pad, sticky = W)

        ewtamt = StringVar()
        ewt_entry = Entry(master, textvariable = ewtamt, state = DISABLED, disabledbackground = codetitlebg, disabledforeground = codetitlefg, font = fonts, width = 15, justify = RIGHT)
        ewt_entry.grid(column = 1, row = 7, padx = pad, sticky = W)

        number_entry = Entry(master, font = fonts, width = 15)
        number_entry.grid(column = 1, row = 8, padx = pad, sticky = W)

        remarks_entry = Entry(master, font = fonts, width = 15)
        remarks_entry.grid(column = 1, row = 9, padx = pad, sticky = W)

    ### supplier options
    def supplieroptions(self):
        global supplier_frame
        supplier_frame = LabelFrame(bir_frame, text = "Supplier Options", font = fonts, bg = wc, fg = fc)
        supplier_frame.grid(column = 0, row = 1, sticky = NW)

        addsup_button = Button(supplier_frame, text = "Add", font = fonts, bg = wc, bd = 0, width = 95, image = buttonicon, compound = CENTER, cursor = "hand2", command = self.showaddsupplier)
        addsup_button.grid(column = 0, row = 0, padx = pad)
        addsup_button.bind("<Return>", self.showaddsupplier)

        updsup_button = Button(supplier_frame, text = "Update", font = fonts, bg = wc, bd = 0, width = 95, image = buttonicon, compound = CENTER, cursor = "hand2", command = self.showupdatesupplier)
        updsup_button.grid(column = 1, row = 0, padx = pad)
        updsup_button.bind("<Return>", self.showupdatesupplier)

        closesupoptions_button = Button(supplier_frame, text = "Close", font = fonts, bg = wc, bd = 0, width = 95, image = buttonicon, compound = CENTER, cursor = "hand2", command = lambda: self.closemanagesup(supplier_frame))
        closesupoptions_button.grid(column = 2, row = 0, padx = pad)
        closesupoptions_button.bind("<Return>", lambda e: self.closemanagesup(supplier_frame))

    def updatefiller(self, *args):
        self.lookup(supplier_list.get())
        name_entry.config(state = NORMAL)
        tin_entry.config(state = NORMAL)
        address_entry.config(state = NORMAL)
        if result:
            name_entry.delete(0, END)
            tin_entry.delete(0, END)
            address_entry.delete(0, END)
            name_entry.insert(0, result[0][0])
            tin_entry.insert(0, result[0][1])
            address_entry.insert(0, result[0][2])

    def updatesupplier(self, *args):
        goupdate = messagebox.askyesno("Update Protocol", "Are you sure?")
        if goupdate == True:
            updater = "UPDATE merchant SET name = ?, tin = ?, address = ? WHERE name = ?"
            c.execute(updater,[name_entry.get(),tin_entry.get(),address_entry.get(),result[0][0]])
            conn.commit()
            messagebox.showinfo("Update Protocol","Supplier info has been updated!")
            self.merchantlister(supplier_list)
            name_entry.delete(0, END)
            tin_entry.delete(0, END)
            address_entry.delete(0, END)
            name_entry.config(state = DISABLED)
            tin_entry.config(state = DISABLED)
            address_entry.config(state = DISABLED)   
        else:
            pass

    def showaddsupplier(self, *args):
        supplier_frame.destroy()
        global mansup_subframe, name_entry, tin_entry, address_entry, name, supcode
        mansup_subframe = LabelFrame(bir_frame, text = "Add Supplier", font = fonts, bg = wc, fg = fc)
        mansup_subframe.grid(column = 0, row = 1, sticky = NW)
        
        supcode = StringVar()
        supcode_entry = Entry(mansup_subframe, textvariable = supcode, font = fonts, width = 13)

        name_label = Label(mansup_subframe, text = "Name", font = fonts, width = 20, bg = wc, fg = fc)
        name_label.grid(column = 0, row = 0, padx = 2)

        tin_label = Label(mansup_subframe, text = "TIN", font = fonts, width = 10, bg = wc, fg = fc)
        tin_label.grid(column = 1, row = 0, padx = 2)

        address_label = Label(mansup_subframe, text = "Address", font = fonts, width = 20, bg = wc, fg = fc)
        address_label.grid(column = 2, row = 0, padx = 2)

        name = StringVar()
        name_entry = Entry(mansup_subframe, textvariable = name, font = fonts, relief = SUNKEN, width = 25, bd = 3, justify = LEFT)
        name_entry.grid(column = 0, row = 1, padx = 2)
        name_entry.bind("<FocusOut>", lambda e: self.uppercase(name_entry))
        name_entry.focus()
        
        tin_entry = Entry(mansup_subframe, relief = SUNKEN, font = fonts, width = 15, bd = 3, justify = LEFT)
        tin_entry.grid(column = 1, row = 1, padx = 2)

        address_entry = Entry(mansup_subframe, relief = SUNKEN, font = fonts, width = 25, bd = 3, justify = LEFT)
        address_entry.grid(column = 2, row = 1, padx = 2)

        regsup_button= Button(mansup_subframe, text = "Add", font = fonts, bg = wc, bd = 0, width = 95, image = buttonicon, compound = CENTER, cursor = "hand2", command = self.tindigitchecker)
        regsup_button.grid(column = 3, row = 1, padx = 2)
        regsup_button.bind("<Return>", self.tindigitchecker)

        closeaddsup_button = Button(mansup_subframe, text = "Close", font = fonts, bg = wc, bd = 0, width = 95, image = buttonicon, compound = CENTER, cursor = "hand2", command = lambda: self.closemanagesup(mansup_subframe))
        closeaddsup_button.grid(column = 4, row = 1, padx = 2)
        closeaddsup_button.bind("<Return>", lambda e: self.closemanagesup(mansup_subframe))

    def showupdatesupplier(self, *args):
        supplier_frame.destroy()
        global mansup_subframe, name_entry, tin_entry, address_entry, name, supplier_list
        mansup_subframe = LabelFrame(bir_frame, text = "Update Supplier", font = fonts, bg = wc, fg = fc)
        mansup_subframe.grid(column = 0, row = 1, sticky = NW)

        supplier_list = tk.Combobox(mansup_subframe, font = fonts, width = 25)
        supplier_list.grid(column = 0, row = 0, padx = 2)
        supplier_list.bind("<<ComboboxSelected>>", self.updatefiller)
        supplier_list.set("Choose client")
        self.merchantlister(supplier_list)

        name_label = Label(mansup_subframe, text = "Supplier", font = fonts, width = 21, bg = wc, fg = fc)
        name_label.grid(column = 0, row = 1, padx = 2)

        tin_label = Label(mansup_subframe, text = "TIN", font = fonts, width = 10, bg = wc, fg = fc)
        tin_label.grid(column = 1, row = 1, padx = 2)

        address_label = Label(mansup_subframe, text = "Address", font = fonts, width = 20, bg = wc, fg = fc)
        address_label.grid(column = 2, row = 1, padx = 2)

        name = StringVar()
        name_entry = Entry(mansup_subframe, textvariable = name, font = fonts, relief = SUNKEN, width = 25, bd = 3, justify = LEFT)
        name_entry.grid(column = 0, row = 2, padx = 2)
        name_entry.bind("<FocusOut>", lambda e: self.uppercase(name_entry))
        name_entry.config(state = DISABLED)

        tin_entry = Entry(mansup_subframe, relief = SUNKEN, font = fonts, width = 15, bd = 3, justify = LEFT)
        tin_entry.grid(column = 1, row = 2, padx = 2)
        tin_entry.bind("<FocusOut>", self.tindigitchecker)
        tin_entry.config(state = DISABLED)
        
        address_entry = Entry(mansup_subframe, relief = SUNKEN, font = fonts, width = 25, bd = 3, justify = LEFT)
        address_entry.grid(column = 2, row = 2, padx = 2)
        address_entry.config(state = DISABLED)

        updsup_button= Button(mansup_subframe, text = "Update", font = fonts, bg = wc, bd = 0, width = 95, image = buttonicon, compound = CENTER, cursor = "hand2", command = self.updatesupplier)
        updsup_button.grid(column = 3, row = 2, padx = 2)
        updsup_button.bind("<Return>", self.updatesupplier)

        closeupdsup_button = Button(mansup_subframe, text = "Close", font = fonts, bg = wc, bd = 0, width = 95, image = buttonicon, compound = CENTER, cursor = "hand2", command = lambda: self.closemanagesup(mansup_subframe))
        closeupdsup_button.grid(column = 4, row = 2, padx = 2)
        closeupdsup_button.bind("<Return>", lambda e: self.closemanagesup(mansup_subframe))

    def tindigitchecker(self, *args):
        tinlist = list(tin_entry.get().replace("-",""))
        if len(tin_entry.get()) == 15:
            if len(tin_entry.get().replace("-","")) == 12:
                nondigit = []
                for i in range(12):
                    if tinlist[i].isdigit():
                        pass
                    else:
                        nondigit.append(tinlist[i])
                        break
                if len(nondigit) == 0:
                    self.tinduplichecker()
            else:
                messagebox.showerror("Registration Protocol", "TIN entered is invalid!")
                tin_entry.delete(0, END)
        else:
            messagebox.showerror("Registration Protocol", "TIN entered is invalid!")
            tin_entry.delete(0, END)
                             
    def tinduplichecker(self, *args):
        if tin_entry.get() != "":
            checker = "SELECT name FROM merchant where tin = ?"
            c.execute(checker, [tin_entry.get()])
            checked = c.fetchone()
            if checked:
                messagebox.showerror("Registration Protocol", "TIN " + tin_entry.get() + " is already registered to " + checked[0] + "!")
            else:
                self.registersupplier()
        else:
            messagebox.showerror("Register Supplier", "Please fill-out all the required details!")
       
    def registersupplier(self, *args):
        goregister = messagebox.askyesno("Registration Protocol", "Are you sure?")
        if goregister == True:
            self.merchantnumbermaster()
            registrar = "INSERT INTO merchant(supcode,name,tin,address) VALUES (?,?,?,?)"
            c.execute(registrar,[newnumber,name_entry.get(),tin_entry.get(),address_entry.get()])
            conn.commit()
            messagebox.showinfo("Registration Protocol","Supplier has been added to the database!")
            name_entry.delete(0, END)
            tin_entry.delete(0, END)
            address_entry.delete(0, END)
        else:
            pass

    def lookup(self, sup):
        global result
        lookup = "SELECT name, tin, address FROM merchant WHERE name = ?"
        c.execute(lookup, [sup])
        result = c.fetchall()

    def generate(self, *args):
        supplier = supplier_list.get()
        period = period_list.get()
        atc = atc_list.get()
        sign = sign_list.get()
        if supplier in merchants:
            self.lookup(supplier)
            if result:
                wb = openpyxl.load_workbook(path + "generator.xlsx") #path
                sheet = wb.active
                sheet['N14'] = result[0][1].split('-')[0]
                sheet['R14'] = result[0][1].split('-')[1]
                sheet['V14'] = result[0][1].split('-')[2]
                sheet['Z14'] = result[0][1].split('-')[3]
                sheet['B17'] = result[0][0]
                sheet['B20'] = result[0][2]
                sheet['AR12'] = period
                sheet['L38'] = atc
                sheet['AD38'] = float(exvat_entry.get().replace(",",""))
                sheet['AR39'] = company
                sheet['AR42'] = sign
                sheet['A73'] = result[0][1]
                wb.save(savepath + '2307.xlsx') #path
                os.startfile(savepath + '2307.xlsx', 'open') #path
            else:
                messagebox.showerror("BIR Protocol", "Supplier is not in database!")
        else:
            messagebox.showerror("BIR Protocol", "Please fill-out all the required details!")

    def closebir(self, *args):
        bir_frame.destroy()
        self.menubuttons(NORMAL)
        self.colorswitch(bir_button, fc)

    def closemanagesup(self, frame, *args):
        try:
            mansup_subframe.destroy()
        except:
            pass
        frame.destroy()
        bir_list.config(state = NORMAL)
        bir_list.delete(0, END)

    def closeewtoption(self, frame, *args):
        frame.destroy()
        self.ewtmenubuttons(NORMAL)
    
    def ewtmenubuttons(self, status):
        addewt_button.config(state = status)
        updewt_button.config(state = status)
        ewtreports_button.config(state = status)
        
### locker factory ###
    def showlocker(self, *args):
        self.menubuttons(DISABLED)
        self.colorswitch(locker_button, fc)
        global locker_frame, year_entry, show_button, update_button, close_button
        locker_frame = LabelFrame(self.master, text = "\nPeriod Locker", font = fonts, bg = wc, fg = fc)
        locker_frame.grid(column = 1, row = 0, ipadx = 22, sticky = N)

        year_entry = tk.Combobox(locker_frame, font = fonts, width = 8)
        year_entry.grid(column = 0, row = 0, padx = 5, pady = pad)
        self.lockeryearlister()
        year_entry.insert(0, 2020)
        
        show_button = Button(locker_frame, text = "Show", font = fonts, bg = wc, bd = 0, width = 95, image = buttonicon, compound = CENTER, cursor = "hand2", command = self.showstats)
        show_button.grid(column = 1, row = 0, padx = pad, pady = pad)
        show_button.bind("<Return>", self.showstats)

        update_button = Button(locker_frame, text = "Update", font = fonts, bg = wc, bd = 0, width = 95, image = buttonicon, compound = CENTER, cursor = "hand2", command = self.checkstats)
        update_button.bind("<Return>", self.checkstats)

        close_button = Button(locker_frame, text = "Close", font = fonts, bg = wc, bd = 0, width = 95, image = buttonicon, compound = CENTER, cursor = "hand2", command = self.closelocker)
        close_button.grid(column = 2, row = 0, padx = pad, pady = pad)
        close_button.bind("<Return>", self.closelocker)
        if access[0] != "super":
            update_button.config(state = DISABLED)

    def showstats(self, *args):
        global yearentry, monthentry, statusentry, yearlist, monthlist, statlist, statframe
        periodfinder = "SELECT * FROM locker WHERE company = ? AND year = ?"
        c.execute(periodfinder, [company, year_entry.get()])
        periods = c.fetchall()
        if periods:
            try:
                statframe.destroy()
            except:
                pass
            statframe = Frame(locker_frame, bg = wc)
            statframe.grid(column = 0, row = 3, pady = pad)

            year_label = Label(statframe, text = "Year", font = fonts, relief = RIDGE, width = 10, bg = codetitlebg, fg = codetitlefg)
            year_label.grid(column = 0, row = 0)
            month_label = Label(statframe, text = "Month", font = fonts, relief = RIDGE, width = 10, bg = codetitlebg, fg = codetitlefg)
            month_label.grid(column = 1, row = 0)
            status_label = Label(statframe, text = "Status", font = fonts, relief = RIDGE, width = 12, bg = codetitlebg, fg = codetitlefg)
            status_label.grid(column = 2, row = 0)

            addrow = 1
            yearlist, monthlist, statlist = [], [], []
            for p in periods:
                yearentry = Label(statframe, text = p[0], font = fonts, bd = 1, width = 10, bg = totalbg, fg = totalfg)
                yearentry.grid(column = 0, row = addrow)
                monthentry = Label(statframe, text = p[1], font = fonts, bd = 1, width = 10, bg = totalbg, fg = totalfg)
                monthentry.grid(column = 1, row = addrow)
                statusentry = tk.Combobox(statframe, values = ["open","lock"], font = fonts, width = 10)
                statusentry.grid(column = 2, row = addrow)
                statusentry.insert(0, p[2])
                yearlist.append(yearentry)
                monthlist.append(monthentry)
                statlist.append(statusentry)
                addrow += 1

            show_button.grid(column = 0, row = 1)
            update_button.grid(column = 0, row = 4, padx = pad, pady = pad)
            close_button.grid(column = 0, row = 5)
        else:
            try:
                statframe.destroy()
                show_button.grid(column = 1, row = 0)
                update_button.grid_forget()
                close_button.grid(column = 2, row = 0)
            except:
                pass

    def checkstats(self, *args):
        staterror = []
        for i in statlist:
            if i.get() == "open" or i.get() == "lock" or i.get() == "locked":
                pass
            else:
                staterror.append(i)
        if len(staterror) == 0:
            self.statusupdater()
        else:
            messagebox.showerror("Period Locker", "Please check inputs!")

    def statusupdater(self, *args):
        updater = "UPDATE locker SET status = ? WHERE company = ? AND year = ? AND month = ?"
        validstats = []
        for i in range(12):
            if statlist[i].get() == "lock":
                validstats.append([statlist[i].get()+"ed", company, yearlist[i].cget("text"), monthlist[i].cget("text")])
            else:
                validstats.append([statlist[i].get(), company, yearlist[i].cget("text"), monthlist[i].cget("text")])
        if len(validstats) == 0:
            messagebox.showerror("Period Locker", "Please contact MJ ASAP!")
        else:
            for v in validstats:
                c.execute(updater,v)
                conn.commit()
            messagebox.showinfo("Period Locker", "Locker Table has been updated!")
            statframe.destroy()
            self.showstats()

    def lockeryearlister(self):
        select = "SELECT year FROM locker WHERE company = ?"
        c.execute(select, [company])
        result = c.fetchall()
        years = []
        for i in result:
            years.append(i[0])
        year_entry.config(values = list(dict.fromkeys(sorted(years))))

    def closelocker(self, *args):
        locker_frame.destroy()
        self.menubuttons(NORMAL)
        self.colorswitch(locker_button, fc)

    def lockchecker(self):
        global datelockresult
        if date_entry.get() != "":
            datelockcheck = "SELECT * FROM locker WHERE year = ? AND month = ? AND company = ?"
            c.execute(datelockcheck, [date_entry.get().split("-")[2],date_entry.get().split("-")[0],company])
            datelockresult = c.fetchone()[2]
        else:
            messagebox.showerror("Date Checker", "Please enter a valid date!")

### reports factory ###
    def showreports(self, *args):
        self.menubuttons(DISABLED)
        self.colorswitch(reports_button, fc)
        global reports_frame, report_list
        reports_frame = LabelFrame(self.master, text = "\nReports", font = fonts, bg = wc, fg = fc)
        reports_frame.grid(column = 1, row = 0, ipadx = 22, sticky = NW)

        report_options = Frame(reports_frame, bg = wc)
        report_options.grid(column = 0, row = 0, sticky = NW)

        reports = ["End-of-month","Account Analysis","Trial Balance","Finder","Bookkeeping","GL Excel","Financial Statements","Unaccounted GV/DM","Advanced SSS Claims"]
        report_list = tk.Combobox(report_options, values = reports, font = fonts, width = 20)
        report_list.grid(column = 0, row = 0, padx = 10)
        report_list.bind("<<ComboboxSelected>>", self.showmonthend)

        closereport_button = Button(report_options, text = "Close", font = fonts, bg = wc, bd = 0, width = 95, image = buttonicon, compound = CENTER, cursor = "hand2", command = self.closereports)
        closereport_button.grid(column = 1, row = 0, pady = 5, sticky = N)
        closereport_button.bind("<Return>", self.closereports)

    def disablereports(self, widget1, status):
        widget1.config(state = status)

    def showmonthend(self, *args):
        global fs_frame, excel_frame, monthend_frame, code_entry, title_label, year_entry, printreport_button, analysis_frame, scroll_frame, total_frame, column_list, finder_frame, finder_entry, month_list, book_list, month_list, book_list, book_frame, tb_frame, printtb_button, fs_list, print_button
        self.disablereports(report_list,DISABLED)
        if report_list.get() == "End-of-month":
            monthend_frame = LabelFrame(reports_frame, text = "End-of-Month", font = fonts, bg = wc, fg = fc)
            monthend_frame.grid(column = 0, row = 1, sticky = NW)

            viewreport_frame = Frame(monthend_frame, bg = wc)
            viewreport_frame.grid(column = 0, row = 0, sticky = NW)
            
            scroll_frame = Frame(monthend_frame, bg = wc)
            scroll_frame.grid(column = 0, row = 1, sticky = NW)

            total_frame = Frame(monthend_frame, bg = wc)
            total_frame.grid(column = 0, row = 2, sticky = NW)

            year_label = Label(viewreport_frame, text = "Year", font = fonts, width = 10, bg = wc, fg = fc)
            year_label.grid(column = 0, row = 0, padx = 5)

            month_label = Label(viewreport_frame, text = "Month", font = fonts, width = 10, bg = wc, fg = fc)
            month_label.grid(column = 1, row = 0, padx = 5)

            year_entry = tk.Combobox(viewreport_frame, font = fonts, width = 5)
            year_entry.grid(column = 0, row = 1)
            year_entry.bind("<<ComboboxSelected>>", self.yearlister)

            month_list = tk.Combobox(viewreport_frame, font = fonts, width = 5)
            month_list.grid(column = 1, row = 1)

            extract_button = Button(viewreport_frame, text = "View", font = fonts, bg = wc, bd = 0, width = 95, image = buttonicon, compound = CENTER, cursor = "hand2", command = self.extracteom)
            extract_button.grid(column = 2, row = 1, padx = 5)
            extract_button.bind("<Return>", self.extracteom)

            printreport_button = Button(viewreport_frame, text = "Export", font = fonts, bg = wc, bd = 0, width = 95, image = buttonicon, compound = CENTER, cursor = "hand2", command = self.reportprinter)
            printreport_button.grid(column = 3, row = 1, padx = 5)
            printreport_button.bind("<Return>", self.reportprinter)
            printreport_button.config(state = DISABLED)

            closeeom_button = Button(viewreport_frame, text = "Close", font = fonts, bg = wc, bd = 0, width = 95, image = buttonicon, compound = CENTER, cursor = "hand2", command = self.closeeom)
            closeeom_button.grid(column = 4, row = 1, padx = 5)
            closeeom_button.bind("<Return>", self.closeeom)

        elif report_list.get() == "Account Analysis":
            analysis_frame = LabelFrame(reports_frame, text = "Account Analysis", font = fonts, bg = wc, fg = fc)
            analysis_frame.grid(column = 0, row = 1, sticky = NW)

            viewreport_frame = Frame(analysis_frame, bg = wc)
            viewreport_frame.grid(column = 0, row = 0, sticky = NW)

            scroll_frame = Frame(analysis_frame, bg = wc)
            scroll_frame.grid(column = 0, row = 1, sticky = NW)

            total_frame = Frame(analysis_frame, bg = wc)
            total_frame.grid(column = 0, row = 2, sticky = NW)

            code_label = Label(viewreport_frame, text = "Code", font = fonts, width = 5, bg = wc, fg = fc)
            code_label.grid(column = 0, row = 0, padx = 2)

            title_label = Label(viewreport_frame, text = "Title", font = fonts, width = 25, bg = wc, fg = fc)
            title_label.grid(column = 1, row = 0, padx = 2)

            year_label = Label(viewreport_frame, text = "Year", font = fonts, width = 6, bg = wc, fg = fc)
            year_label.grid(column = 2, row = 0, padx = 2)

            month_label = Label(viewreport_frame, text = "Month", font = fonts, width = 4, bg = wc, fg = fc)
            month_label.grid(column = 3, row = 0, padx = 2)

            code_entry = Entry(viewreport_frame, font = fonts, width = 5, relief = SUNKEN, bd = 2, justify = "center")
            code_entry.grid(column = 0, row = 1, padx = pad)
            code_entry.bind("<FocusOut>", self.journalcode_val)

            title_label = Label(viewreport_frame, text = "", font = fonts, relief = SUNKEN, bd = 2, width = 25, anchor = W, bg = codetitlebg, fg = codetitlefg)
            title_label.grid(column = 1, row = 1, padx = pad)

            year_entry = tk.Combobox(viewreport_frame, font = fonts, width = 5)
            year_entry.grid(column = 2, row = 1, padx = pad)
            year_entry.bind("<<ComboboxSelected>>", self.yearlister)

            month_list = tk.Combobox(viewreport_frame, font = fonts, width = 5)
            month_list.grid(column = 3, row = 1, padx = pad)

            extract_button = Button(viewreport_frame, text = "View", font = fonts, bg = wc, bd = 0, width = 95, image = buttonicon, compound = CENTER, cursor = "hand2", command = self.accountanalysis)
            extract_button.grid(column = 4, row = 1, padx = 2)
            extract_button.bind("<Return>", self.accountanalysis)

            printreport_button = Button(viewreport_frame, text = "Export", font = fonts, bg = wc, bd = 0, width = 95, image = buttonicon, compound = CENTER, cursor = "hand2", command = self.reportprinter)
            printreport_button.grid(column = 5, row = 1, padx = 2)
            printreport_button.bind("<Return>", self.reportprinter)
            printreport_button.config(state = DISABLED)

            closeanalysis_button = Button(viewreport_frame, text = "Close", font = fonts, bg = wc, bd = 0, width = 95, image = buttonicon, compound = CENTER, cursor = "hand2", command = self.closeanal)
            closeanalysis_button.grid(column = 6, row = 1, padx = 2)
            closeanalysis_button.bind("<Return>", self.closeanal)

        elif report_list.get() == "Finder":
            finder_frame = LabelFrame(reports_frame, text = "Finder", font = fonts, bg = wc, fg = fc)
            finder_frame.grid(column = 0, row = 1, sticky = NW)

            finder_subframe = Frame(finder_frame, bg = wc)
            finder_subframe.grid(column = 0, row = 0, sticky = NW)

            scroll_frame = Frame(finder_frame, bg = wc)
            scroll_frame.grid(column = 0, row = 1, sticky = NW)

            total_frame = Frame(finder_frame, bg = wc)
            total_frame.grid(column = 0, row = 2, sticky = NW)

            column_label = Label(finder_subframe, text = "Column", font = fonts, width = 10, bg = wc, fg = fc)
            column_label.grid(column = 0, row = 0, padx = 2)

            month_label = Label(finder_subframe, text = "Month", font = fonts, width = 5, bg = wc, fg = fc)
            month_label.grid(column = 1, row = 0, padx = 2)

            item_label = Label(finder_subframe, text = "Item", font = fonts, width = 25, bg = wc, fg = fc)
            item_label.grid(column = 2, row = 0, padx = 2)

            columns = ["reference","particulars","client"]
            column_list = tk.Combobox(finder_subframe, font = fonts, values = columns, width = 10)
            column_list.grid(column = 0, row = 1, padx = 10)

            month_list = tk.Combobox(finder_subframe, font = fonts, width = 5)
            month_list.grid(column = 1, row = 1, padx = 10)
            
            finder_entry = Entry(finder_subframe, font = fonts, relief = SUNKEN, width = 25, bd = 3, justify = LEFT)
            finder_entry.grid(column = 2, row = 1, padx = 2)

            finder_button = Button(finder_subframe, text = "View", font = fonts, bg = wc, bd = 0, width = 95, image = buttonicon, compound = CENTER, cursor = "hand2", command = self.finderfinder)
            finder_button.grid(column = 3, row = 1, padx = 2)
            finder_button.bind("<Return>", self.finderfinder)

            printreport_button = Button(finder_subframe, text = "Export", font = fonts, bg = wc, bd = 0, width = 95, image = buttonicon, compound = CENTER, cursor = "hand2", command = self.reportprinter)
            printreport_button.grid(column = 4, row = 1, padx = 2)
            printreport_button.bind("<Return>", self.reportprinter)
            printreport_button.config(state = NORMAL)

            closefinder_button = Button(finder_subframe, text = "Close", font = fonts, bg = wc, bd = 0, width = 95, image = buttonicon, compound = CENTER, cursor = "hand2", command = self.closefinder)
            closefinder_button.grid(column = 5, row = 1, padx = 2)
            closefinder_button.bind("<Return>", self.closefinder)
            
        elif report_list.get() == "Bookkeeping":
            book_frame = LabelFrame(reports_frame, text = "Bookkeeping", font = fonts, bg = wc, fg = fc)
            book_frame.grid(column = 0, row = 1, sticky = NW)

            keeper_label = Label(book_frame, text = "Book", font = fonts, width = 5, bg = wc, fg = fc)
            keeper_label.grid(column = 0, row = 0, padx = 2)

            year_label = Label(book_frame, text = "Year", font = fonts, width = 5, bg = wc, fg = fc)
            year_label.grid(column = 1, row = 0, padx = 2)

            month_label = Label(book_frame, text = "Month", font = fonts, width = 5, bg = wc, fg = fc)
            month_label.grid(column = 2, row = 0, padx = 2)

            book_list = tk.Combobox(book_frame, font = fonts, width = 5)
            book_list.grid(column = 0, row = 1, padx = 2)

            year_entry = tk.Combobox(book_frame, font = fonts, width = 5)
            year_entry.grid(column = 1, row = 1, padx = 2)
            year_entry.bind("<<ComboboxSelected>>", self.yearlister)

            month_list = tk.Combobox(book_frame, font = fonts, width = 5)
            month_list.grid(column = 2, row = 1, padx = 2)

            book_printer = Button(book_frame, text = "Export", font = fonts, bg = wc, bd = 0, width = 95, image = buttonicon, compound = CENTER, cursor = "hand2", command = self.reportprinter)
            book_printer.grid(column = 3, row = 1, padx = 2)
            book_printer.bind("<Return>", self.reportprinter)
            book_printer.config(state = NORMAL)
            self.booklister()

            closebookkeeping = Button(book_frame, text = "Close", font = fonts, bg = wc, bd = 0, width = 95, image = buttonicon, compound = CENTER, cursor = "hand2", command = self.closebook)
            closebookkeeping.grid(column = 5, row = 1, padx = 2)
            closebookkeeping.bind("<Return>", self.closebook)

        elif report_list.get() == "Trial Balance":
            tb_frame = LabelFrame(reports_frame, text = "Trial Balance", font = fonts, bg = wc, fg = fc)
            tb_frame.grid(column = 0, row = 1, sticky = NW)

            tb_subframe = Frame(tb_frame, bg = wc)
            tb_subframe.grid(column = 0, row = 0, sticky = NW)

            scroll_frame = Frame(tb_frame, bg = wc)
            scroll_frame.grid(column = 0, row = 1, sticky = NW)

            total_frame = Frame(tb_frame, bg = wc)
            total_frame.grid(column = 0, row = 2, sticky = NW)

            year_label = Label(tb_subframe, text = "Year", font = fonts, width = 5, bg = wc, fg = fc)
            year_label.grid(column = 0, row = 0, padx = 2)

            month_label = Label(tb_subframe, text = "Month", font = fonts, width = 8, bg = wc, fg = fc)
            month_label.grid(column = 1, row = 0, padx = 2)

            year_entry = tk.Combobox(tb_subframe, font = fonts, width = 5)
            year_entry.grid(column = 0, row = 1, padx = 2)
            year_entry.bind("<<ComboboxSelected>>", self.yearlister)

            month_list = tk.Combobox(tb_subframe, font = fonts, width = 8)
            month_list.grid(column = 1, row = 1, padx = 2)

            extract_button = Button(tb_subframe, text = "View", font = fonts, bg = wc, bd = 0, width = 95, image = buttonicon, compound = CENTER, cursor = "hand2", command = self.extracetbresults)
            extract_button.grid(column = 2, row = 1, padx = 2)
            extract_button.bind("<Return>", self.extracetbresults)

            printtb_button = Button(tb_subframe, text = "Export", font = fonts, bg = wc, bd = 0, width = 95, image = buttonicon, compound = CENTER, cursor = "hand2", command = self.reportprinter)
            printtb_button.grid(column = 3, row = 1, padx = 2)
            printtb_button.bind("<Return>", self.reportprinter)
            printtb_button.config(state = DISABLED)

            closetb_button = Button(tb_subframe, text = "Close", font = fonts, bg = wc, bd = 0, width = 95, image = buttonicon, compound = CENTER, cursor = "hand2", command = self.closetb)
            closetb_button.grid(column = 4, row = 1, padx = 2)
            closetb_button.bind("<Return>", self.closetb)

        elif report_list.get() == "GL Excel":
            excel_frame = LabelFrame(reports_frame, text = "GL Excel", font = fonts, bg = wc, fg = fc)
            excel_frame.grid(column = 0, row = 1, sticky = NW)

            year_label = Label(excel_frame, text = "Year", font = fonts, width = 5, bg = wc, fg = fc)
            year_label.grid(column = 0, row = 0, padx = 2)

            year_entry = tk.Combobox(excel_frame, font = fonts, width = 5)
            year_entry.grid(column = 0, row = 1, padx = 2)

            export_button = Button(excel_frame, text = "Export", font = fonts, bg = wc, bd = 0, width = 95, image = buttonicon, compound = CENTER, cursor = "hand2", command = self.excelexport)
            export_button.grid(column = 1, row = 1, padx = 2)
            export_button.bind("<Return>", self.excelexport)

            closeexcel_button = Button(excel_frame, text = "Close", font = fonts, bg = wc, bd = 0, width = 95, image = buttonicon, compound = CENTER, cursor = "hand2", command = self.closeexcel)
            closeexcel_button.grid(column = 2, row = 1, padx = 2)
            closeexcel_button.bind("<Return>", self.closeexcel)

        elif report_list.get() == "Financial Statements":
            fs_frame = LabelFrame(reports_frame, text = "Financial Statements", font = fonts, bg = wc, fg = fc)
            fs_frame.grid(column = 0, row = 1, sticky = NW)

            fs_label = Label(fs_frame, text = "FS", font = fonts, width = 20, bg = wc, fg = fc)
            fs_label.grid(column = 0, row = 0, padx = 2)

            year_label = Label(fs_frame, text = "Year", font = fonts, width = 5, bg = wc, fg = fc)
            year_label.grid(column = 1, row = 0, padx = 2)

            month_label = Label(fs_frame, text = "Month", font = fonts, width = 8, bg = wc, fg = fc)
            month_label.grid(column = 2, row = 0, padx = 2)

            fs = ["Statement of Financial Position", "Income Statement"] #"Statement of Financial Position", "Income Statement"
            fs_list = tk.Combobox(fs_frame, values = fs, font = fonts, width = 20)
            fs_list.grid(column = 0, row = 1, padx = 2)

            year_entry = tk.Combobox(fs_frame, font = fonts, width = 5)
            year_entry.grid(column = 1, row = 1, padx = 2)
            year_entry.bind("<<ComboboxSelected>>", self.yearlister)

            month_list = tk.Combobox(fs_frame, font = fonts, width = 8)
            month_list.grid(column = 2, row = 1, padx = 2)
            
            printbs_button = Button(fs_frame, text = "Export", font = fonts, bg = wc, bd = 0, width = 95, image = buttonicon, compound = CENTER, cursor = "hand2", command = self.printfs)
            printbs_button.grid(column = 3, row = 1, padx = 2)
            printbs_button.bind("<Return>", self.printfs)

            closebs_button = Button(fs_frame, text = "Close", font = fonts, bg = wc, bd = 0, width = 95, image = buttonicon, compound = CENTER, cursor = "hand2", command = self.closebs)
            closebs_button.grid(column = 4, row = 1, padx = 2)
            closebs_button.bind("<Return>", self.closebs)

        elif report_list.get() == "Unaccounted GV/DM":
            self.showunforwardedreport(reports_frame)

        elif report_list.get() == "Advanced SSS Claims":
            self.showsssclaims(reports_frame)

        self.yearlister()

    def showsssclaims(self, frame):
        global sss_frame, year_entry, month_list, scroll_frame, printreport_button
        sss_frame = LabelFrame(frame, text = "Advanced SSS Claims", font = fonts, bg = wc, fg = fc)
        sss_frame.grid(column = 0, row = 1, sticky = NW)

        sub_frame = Frame(sss_frame, bg = wc)
        sub_frame.grid(column = 0, row = 0, sticky = NW)

        scroll_frame = Frame(sss_frame, bg = wc)
        scroll_frame.grid(column = 0, row = 1, sticky = NW)

        total_frame = Frame(sss_frame, bg = wc)
        total_frame.grid(column = 0, row = 2, sticky = NW)

        year_label = Label(sub_frame, text = "Year", font = fonts, width = 6, bg = wc, fg = fc)
        year_label.grid(column = 0, row = 1, padx = 2)

        month_label = Label(sub_frame, text = "Month", font = fonts, width = 4, bg = wc, fg = fc)
        month_label.grid(column = 1, row = 1, padx = 2)
        
        year_entry = tk.Combobox(sub_frame, font = fonts, width = 5)
        year_entry.grid(column = 0, row = 2, padx = pad)
        year_entry.bind("<<ComboboxSelected>>", self.yearlister)

        month_list = tk.Combobox(sub_frame, font = fonts, width = 5)
        month_list.grid(column = 1, row = 2, padx = pad)

        extract_button = Button(sub_frame, text = "View", font = fonts, bg = wc, bd = 0, width = 95, image = buttonicon, compound = CENTER, cursor = "hand2", command = self.extractsssclaims)
        extract_button.grid(column = 2, row = 2, padx = 2)
        extract_button.bind("<Return>", self.extractsssclaims)

        printreport_button = Button(sub_frame, text = "Export", font = fonts, bg = wc, bd = 0, width = 95, image = buttonicon, compound = CENTER, cursor = "hand2", state = DISABLED, command = self.printsssclaims)
        printreport_button.grid(column = 3, row = 2, padx = 2)
        printreport_button.bind("<Return>", self.printsssclaims)

        close_button = Button(sub_frame, text = "Close", font = fonts, bg = wc, bd = 0, width = 95, image = buttonicon, compound = CENTER, cursor = "hand2", command = self.closesss)
        close_button.grid(column = 4, row = 2, padx = 2)
        close_button.bind("<Return>", self.closesss)

    def extractsssclaims(self, *args):
        global result
        if company == "DBPSC":
            select = "SELECT comment, debit, number, date FROM DBPSC WHERE year = ? AND month = ? AND code = '1303'"
        else:
            select = "SELECT comment, debit, number, date FROM DSSI WHERE year = ? AND month = ? AND code = '1303'"
        c.execute(select, [year_entry.get(), month_list.get()])
        result = c.fetchall()
        if result:
            sss_tree = tk.Treeview(scroll_frame)
            sss_tree["columns"] = ("Number","Date","Assignment","Last Name","First Name","Claim Type","Date1","Date2","Amount")
            sss_tree.column("#0", width=0, minwidth=0, stretch = True)
            sss_tree.column("Number", anchor = W, width = 75)
            sss_tree.column("Date", anchor = W, width = 75)
            sss_tree.column("Assignment", anchor = W, width = 100)
            sss_tree.column("Last Name", anchor = W, width = 100)
            sss_tree.column("First Name", anchor = W, width = 100)
            sss_tree.column("Claim Type", anchor = W, width = 75)
            sss_tree.column("Date1", anchor = W, width = 75)
            sss_tree.column("Date2", anchor = W, width = 75)
            sss_tree.column("Amount", anchor = E, width = 100)
            sss_tree.heading("#0", text = "", anchor = W)
            sss_tree.heading("Number", text = "Number", anchor = N)
            sss_tree.heading("Date", text = "Date", anchor = N)
            sss_tree.heading("Assignment", text = "Assignment", anchor = N)
            sss_tree.heading("Last Name", text = "Last Name", anchor = N)
            sss_tree.heading("First Name", text = "First Name", anchor = N)
            sss_tree.heading("Claim Type", text = "Claim Type", anchor = N)
            sss_tree.heading("Date1", text = "Date1", anchor = N)
            sss_tree.heading("Date2", text = "Date2", anchor = N)
            sss_tree.heading("Amount", text = "Amount", anchor = N)
            for i in result:
                try:
                    if len(i[0].split("_")) == 6:
                        sss_tree.insert("", "end", values = (i[2],i[3],i[0].split("_")[0],i[0].split("_")[1],i[0].split("_")[2],i[0].split("_")[3],i[0].split("_")[4],i[0].split("_")[5],format(i[1], ",.2f")))
                    elif len(i[0].split("_")) == 5:
                        sss_tree.insert("", "end", values = (i[2],i[3],i[0].split("_")[0],i[0].split("_")[1],i[0].split("_")[2],i[0].split("_")[3],i[0].split("_")[4],"",format(i[1], ",.2f")))
                    else:
                        sss_tree.insert("", "end", values = (i[2],i[3],i[0].split("_")[0],i[0].split("_")[1],i[0].split("_")[2],i[0].split("_")[3],"","",format(i[1], ",.2f")))
                except:
                    pass
            sss_tree.pack()
            printreport_button.config(state = NORMAL)

    def printsssclaims(self, *args):
        wb = openpyxl.load_workbook(path + "report.xlsx")
        st = wb.active
        st.append([company])
        st.append([f"Advanced SSS Claims for the month of {month_list.get()}-{year_entry.get()}"])
        st.append(["Company","Number","Date","Assignment","Last Name","First Name","Claim Type","Date1","Date2","Amount"])
        for i in result:
            try:
                if len(i[0].split("_")) == 6:
                    st.append([company,i[2],i[3],i[0].split("_")[0],i[0].split("_")[1],i[0].split("_")[2],i[0].split("_")[3],i[0].split("_")[4],i[0].split("_")[5],i[1]])
                elif len(i[0].split("_")) == 5:
                    st.append([company,i[2],i[3],i[0].split("_")[0],i[0].split("_")[1],i[0].split("_")[2],i[0].split("_")[3],i[0].split("_")[4],"",i[1]])
                else:
                    st.append([company,i[2],i[3],i[0].split("_")[0],i[0].split("_")[1],i[0].split("_")[2],i[0].split("_")[3],"","",i[1]])
            except:
                pass
        wb.save(savepath + "Advanced SSS Claims" + company + ".xlsx")
        os.startfile(savepath + "Advanced SSS Claims" + company + ".xlsx", "open")

    def showunforwardedreport(self, frame):
        global ud_frame, year_entry, month_list, scroll_frame, print_button, close_button
        ud_frame = LabelFrame(frame, text = "Unaccounted GV/DM", font = fonts, bg = wc, fg = fc)
        ud_frame.grid(column = 0, row = 1, sticky = NW)

        scroll_frame = Frame(ud_frame, bg = wc)
        scroll_frame.grid(column = 0, row = 2, padx = 2)

        year_label = Label(ud_frame, text = "Year", font = fonts, width = 5, bg = wc, fg = fc)
        year_label.grid(column = 0, row = 0, padx = 2)

        month_label = Label(ud_frame, text = "Month", font = fonts, width = 8, bg = wc, fg = fc)
        month_label.grid(column = 1, row = 0, padx = 2)

        year_entry = tk.Combobox(ud_frame, font = fonts, width = 5)
        year_entry.grid(column = 0, row = 1, padx = 2)
        year_entry.bind("<<ComboboxSelected>>", self.fmdmonthlister2)

        month_list = tk.Combobox(ud_frame, font = fonts, width = 8)
        month_list.grid(column = 1, row = 1, padx = 2)
        
        view_button = Button(ud_frame, text = "View", font = fonts, bg = wc, bd = 0, width = 95, image = buttonicon, compound = CENTER, cursor = "hand2", command = self.viewunforwarded)
        # view_button.grid(column = 3, row = 1, padx = 2)
        view_button.bind("<Return>", self.viewunforwarded)

        print_button = Button(ud_frame, text = "Export", font = fonts, bg = wc, bd = 0, width = 95, image = buttonicon, compound = CENTER, cursor = "hand2", command = self.printunforwarded)
        print_button.grid(column = 4, row = 1, padx = 2)
        print_button.bind("<Return>", self.printunforwarded)

        close_button = Button(ud_frame, text = "Close", font = fonts, bg = wc, bd = 0, width = 95, image = buttonicon, compound = CENTER, cursor = "hand2", command = self.closeud)
        close_button.grid(column = 5, row = 1, padx = 2)
        close_button.bind("<Return>", self.closeud)

    def viewunforwarded(self, *args):
        global unforwarded
        if company == "DBPSC":
            finder = "SELECT reference FROM DBPSC WHERE year = ?"
        else:
            finder = "SELECT reference FROM DSSI WHERE year = ?"
        fmdlist, asdlist = [], []
        fmdfinder = "SELECT number FROM disbursements WHERE company = ? AND year = ? AND month = ?"
        c.execute(fmdfinder, [company, year_entry.get(), month_list.get()])
        fmdresult = c.fetchall()
        for i in fmdresult:
            fmdlist.append(i[0])
        c.execute(finder, [year_entry.get()])
        asdresult = c.fetchall()
        for x in asdresult:
            asdlist.append(x[0])
        doclist = []
        fmdlistfiltered = list(dict.fromkeys(sorted(fmdlist)))
        asdlistfiltered = list(dict.fromkeys(sorted(asdlist)))
        for y in fmdlistfiltered:
            if y not in asdlistfiltered:
                doclist.append(y)
        finalfinder = "SELECT company, date, number, payee, particulars, bank, mode, net, status FROM disbursements WHERE number = ? AND company = ? AND status = 'VALID' ORDER BY number"
        unforwarded = []
        for z in doclist:
            c.execute(finalfinder, [z, company])
            result = c.fetchall()
            for r in result:
                unforwarded.append(r)
        try:
            labelframe.destroy()
            scrollbox.destroy()
        except:
            pass
        if len(unforwarded) == 0:
            messagebox.showinfo("Unaccounted Report", "No Unaccounted documents!")
        else:
            pass
            # self.showdisblabels(scroll_frame, 0, 0)
            # self.scrollreports(scroll_frame, unforwarded, 0, 1)
            # print_button.config(state = NORMAL)

    def printunforwarded(self, *args):
        self.viewunforwarded()
        wb = openpyxl.load_workbook(path + "report.xlsx")
        st = wb.active
        st.append([company])
        st.append([f"Unaccounted GV/DM for the month of {month_list.get()}-{year_entry.get()}"])
        st.append(["Company", "Date", "Number", "Payee", "Particulars", "Bank", "Mode", "Net", "Status"])
        for i in unforwarded:
            st.append(i)
        wb.save(savepath + "UnaccountedDocuments_" + company + ".xlsx")
        os.startfile(savepath + "UnaccountedDocuments_" + company + ".xlsx", "open")

    def excelexport(self, *args):
        if company == "DBPSC":
            selecter = "SELECT code,title,number,date,debit,credit,reference,comment,client,particulars,user FROM DBPSC WHERE year = ? ORDER BY number"
        else:
            selecter = "SELECT code,title,number,date,debit,credit,reference,comment,client,particulars,user FROM DSSI WHERE year = ? ORDER BY number"    
        c.execute(selecter, [year_entry.get()])
        result = c.fetchall()
        wb = openpyxl.load_workbook(path + "report.xlsx") #path
        sheet = wb.active
        sheet.append(["code","title","number","date","debit","credit","reference","comment","client","particulars","users"])
        for i in result:
            sheet.append([i[0],i[1],i[2],i[3].replace("-","/"),i[4],i[5],i[6],i[7],i[8],i[9],i[10]])
        wb.save(savepath + "GLEXCEL_" + company + ".xlsx") #path
        os.startfile(savepath + "GLEXCEL_" + company + ".xlsx", "open") #path

    def printfs(self, *args):
        if fs_list.get() == "Statement of Financial Position":
            if company == "DBPSC":
                selecter = "SELECT class, mainclass, subclass, subtitle, title, sum(debit), sum(credit) FROM (SELECT chart.class, mainclass, subclass, subtitle, chart.title, debit, credit FROM DBPSC INNER JOIN chart ON DBPSC.code = chart.code WHERE year = ? AND month <= ? UNION ALL SELECT chart.class, mainclass, subclass, subtitle, chart.title, debit, credit FROM begbal INNER JOIN chart ON begbal.code = chart.code) GROUP BY subtitle ORDER BY mainclass"
                selecter2 = "SELECT * FROM (SELECT chart.class, mainclass, subclass, subtitle, SUM(debit), SUM(credit) FROM DBPSC INNER JOIN chart ON DBPSC.code = chart.code WHERE year = ? AND month <= ? GROUP BY subtitle) WHERE class IN (?,?,?) ORDER BY mainclass"
                companytitle = "DBP SERVICE CORPORATION"
            else:
                selecter = "SELECT class, mainclass, subclass, subtitle, title, sum(debit), sum(credit) FROM (SELECT dssichart.class, mainclass, subclass, subtitle, dssichart.title, debit, credit FROM DSSI INNER JOIN dssichart ON DSSI.code = dssichart.code WHERE year = ? AND month <= ? UNION ALL SELECT dssichart.class, mainclass, subclass, subtitle, dssichart.title, debit, credit FROM dssibegbal INNER JOIN dssichart ON dssibegbal.code = dssichart.code) GROUP BY subtitle ORDER BY mainclass"
                selecter2 = "SELECT * FROM (SELECT dssichart.class, mainclass, subclass, subtitle, SUM(debit), SUM(credit) FROM DSSI INNER JOIN dssichart ON DSSI.code = dssichart.code WHERE year = ? AND month <= ? GROUP BY subtitle) WHERE class IN (?,?,?) ORDER BY mainclass"
                companytitle = "DBPSC SECURITY SERVICE, INC."
            c.execute(selecter, [year_entry.get(), month_list.get()])
            result = c.fetchall()
            c.execute(selecter2, [year_entry.get(), month_list.get(), "E", "I", "T"])
            result2 = c.fetchall()
            totalincome, totalexpense = [], []
            for i in result2:
                if i[0] == "I":
                    totalincome.append(i[5]-i[4])
                elif i[0] != "I":
                    totalexpense.append(i[4]-i[5])
            wb = openpyxl.load_workbook(path + "report.xlsx") #path
            sheet = wb.active
            currentassets, noncurrentassets, currentliabilities, noncurrentliabilities, equity = [], [], [], [], []
            sheet.append([companytitle])
            sheet.append(["Statement of Financial Position"])
            sheet.append([f"{month_list.get()}-{year_entry.get()}"])
            sheet.append([])
            row = 5
            for i in result:
                if i[0] == "A" and i[1] == "CASH AND CASH EQUIVALENTS":
                    sheet.append([i[1], i[3], i[5]-i[6]])
                    currentassets.append(i[5]-i[6])
                    row += 1
            for i in result:
                if i[0] == "A" and i[1] == "RECEIVABLES":
                    sheet.append([i[1], i[3], i[5]-i[6]])
                    currentassets.append(i[5]-i[6])
                    row += 1
            for i in result:
                if i[0] == "A" and i[1] == "FINANCIAL ASSETS AT FAIR VALUE THROUGH PROFIT OR LOSS":
                    sheet.append([i[1], i[3], i[5]-i[6]])
                    currentassets.append(i[5]-i[6])
                    row += 1
            for i in result:
                if i[0] == "A" and i[1] == "PREPAYMENTS AND OTHER CURRENT ASSETS":
                    sheet.append([i[1], i[3], i[5]-i[6]])
                    currentassets.append(i[5]-i[6])
                    row += 1
            sheet.append(["Total Current Assets","",sum(currentassets)])
            sheet.append([])
            row += 2
            for i in result:
                if i[0] == "A" and i[1] == "PROPERTY AND EQUIPMENT":
                    sheet.append([i[1], i[3], i[5]-i[6]])
                    noncurrentassets.append(i[5]-i[6])
                    row += 1
            for i in result:
                if i[0] == "A" and i[1] == "OTHER NONCURRENT ASSETS":
                    sheet.append([i[1], i[3], i[5]-i[6]])
                    noncurrentassets.append(i[5]-i[6])
                    row += 1
            sheet.append(["Total Non-Current Assets","",sum(noncurrentassets)])
            sheet.append(["Total Assets","",sum(currentassets)+sum(noncurrentassets)])
            sheet.append([])
            row += 3
            for i in result:
                if i[0] == "L" and i[1] == "TRADE AND OTHER PAYABLES":
                    sheet.append([i[1], i[3], i[6]-i[5]])
                    currentliabilities.append(i[6]-i[5])
                    row += 1
            sheet.append(["Total Current Liabilities","",sum(currentliabilities)])
            sheet.append([])
            row += 2
            for i in result:
                if i[0] == "L" and i[1] != "TRADE AND OTHER PAYABLES":
                    sheet.append([i[1], i[3], i[6]-i[5]])
                    noncurrentliabilities.append(i[6]-i[5])
                    row += 1
            sheet.append(["Total Non-Current Liabilities","",sum(noncurrentliabilities)])
            sheet.append(["Total Liabilities","",sum(currentliabilities)+sum(noncurrentliabilities)])
            sheet.append([])
            row += 3
            for i in result:
                if i[0] == "S":
                    sheet.append([i[1], i[3], i[6]-i[5]])
                    equity.append(i[6]-i[5])
                    row += 1
            sheet.append(["Net Income","",sum(totalincome)-sum(totalexpense)])
            sheet.append(["Total Equity","",sum(equity)+sum(totalincome)-sum(totalexpense)])
            sheet.append(["Total Liabilities and Equity","",sum(currentliabilities)+sum(noncurrentliabilities)+sum(equity)+sum(totalincome)-sum(totalexpense)])

            wb.save(savepath + "financialposition.xlsx") #path
            os.startfile(savepath + "financialposition.xlsx", "open") #path

        elif fs_list.get() == "Income Statement":
            if company == "DBPSC":
                selecter = "SELECT * FROM (SELECT chart.class, mainclass, subclass, subtitle, SUM(debit), SUM(credit) FROM DBPSC INNER JOIN chart ON DBPSC.code = chart.code WHERE year = ? AND month = ? GROUP BY subtitle) WHERE class IN (?,?,?) ORDER BY mainclass"
                companytitle = "DBP SERVICE CORPORATION"
            else:
                selecter = "SELECT * FROM (SELECT dssichart.class, mainclass, subclass, subtitle, SUM(debit), SUM(credit) FROM DSSI INNER JOIN dssichart ON DSSI.code = dssichart.code WHERE year = ? AND month = ? GROUP BY subtitle) WHERE class IN (?,?,?) ORDER BY mainclass"
                companytitle = "DBPSC SECURITY SERVICE, INC."
            c.execute(selecter, [year_entry.get(), month_list.get(), "E", "I", "T"])
            getter = c.fetchall()
            wb = openpyxl.load_workbook(path + "report.xlsx") #path
            sheet = wb.active
            totalrevenue, totalcostofrevenue, totalexpense, totalotherincome = [], [], [], []
            sheet.append([companytitle])
            sheet.append(["Income Statement"])
            sheet.append([f"{month_list.get()}-{year_entry.get()}"])
            sheet.append([])
            row = 5
            for i in getter:
                if i[0] == "I" and i[3] == "SERVICE REVENUE":
                    sheet.append([i[1], i[3], i[5]-i[4]])
                    totalrevenue.append(i[5]-i[4])
                    row += 1
            sheet.cell(column = 1, row = row-1).border = Border(bottom = Side(style='thin'))
            sheet.cell(column = 2, row = row-1).border = Border(bottom = Side(style='thin'))
            sheet.cell(column = 3, row = row-1).border = Border(bottom = Side(style='thin'))
            sheet.append(["Total Revenue", "", sum(totalrevenue)])
            sheet.cell(column = 1, row = row).fill = PatternFill(start_color = fsbg, end_color = fsbg, fill_type = "solid")
            sheet.cell(column = 2, row = row).fill = PatternFill(start_color = fsbg, end_color = fsbg, fill_type = "solid")
            sheet.cell(column = 3, row = row).fill = PatternFill(start_color = fsbg, end_color = fsbg, fill_type = "solid")
            sheet.append([])
            row += 1
            for i in getter:
                if i[1] == "COST OF REVENUE":
                    sheet.append([i[1], i[3], (i[4]-i[5])])
                    totalcostofrevenue.append(i[4]-i[5])
                    row += 1
            sheet.cell(column = 1, row = row).border = Border(bottom = Side(style='thin'))
            sheet.cell(column = 2, row = row).border = Border(bottom = Side(style='thin'))
            sheet.cell(column = 3, row = row).border = Border(bottom = Side(style='thin'))
            sheet.append(["Total Cost of Revenue", "", sum(totalcostofrevenue)])
            sheet.cell(column = 1, row = row+1).fill = PatternFill(start_color = fsbg, end_color = fsbg, fill_type = "solid")
            sheet.cell(column = 2, row = row+1).fill = PatternFill(start_color = fsbg, end_color = fsbg, fill_type = "solid")
            sheet.cell(column = 3, row = row+1).fill = PatternFill(start_color = fsbg, end_color = fsbg, fill_type = "solid")
            sheet.cell(column = 1, row = row+1).border = Border(bottom = Side(style='thin'))
            sheet.cell(column = 2, row = row+1).border = Border(bottom = Side(style='thin'))
            sheet.cell(column = 3, row = row+1).border = Border(bottom = Side(style='thin'))
            sheet.append(["Gross Income", "",(sum(totalrevenue)-sum(totalcostofrevenue))])
            sheet.cell(column = 1, row = row+2).fill = PatternFill(start_color = fsbg, end_color = fsbg, fill_type = "solid")
            sheet.cell(column = 2, row = row+2).fill = PatternFill(start_color = fsbg, end_color = fsbg, fill_type = "solid")
            sheet.cell(column = 3, row = row+2).fill = PatternFill(start_color = fsbg, end_color = fsbg, fill_type = "solid")
            sheet.append([])
            row += 4
            for i in getter:
                if i[0] != "I" and i[1] != "COST OF REVENUE":
                    sheet.append([i[1], i[3], (i[4]-i[5])])
                    totalexpense.append(i[4]-i[5])
                    row += 1
            sheet.cell(column = 1, row = row-1).border = Border(bottom = Side(style='thin'))
            sheet.cell(column = 2, row = row-1).border = Border(bottom = Side(style='thin'))
            sheet.cell(column = 3, row = row-1).border = Border(bottom = Side(style='thin'))
            sheet.append(["Total Operating Expense", "", sum(totalexpense)])
            sheet.cell(column = 1, row = row).fill = PatternFill(start_color = fsbg, end_color = fsbg, fill_type = "solid")
            sheet.cell(column = 2, row = row).fill = PatternFill(start_color = fsbg, end_color = fsbg, fill_type = "solid")
            sheet.cell(column = 3, row = row).fill = PatternFill(start_color = fsbg, end_color = fsbg, fill_type = "solid")
            for i in getter:
                if i[0] == "I" and i[3] != "SERVICE REVENUE":
                    totalotherincome.append(i[5]-i[4])
            sheet.append(["Other Income", "", sum(totalotherincome)])
            row += 1
            sheet.cell(column = 1, row = row).border = Border(bottom = Side(style='thin'))
            sheet.cell(column = 2, row = row).border = Border(bottom = Side(style='thin'))
            sheet.cell(column = 3, row = row).border = Border(bottom = Side(style='thin'))
            sheet.append(["Net Income", "", (sum(totalrevenue)-sum(totalcostofrevenue)-sum(totalexpense)+sum(totalotherincome))])
            sheet.cell(column = 1, row = row+1).fill = PatternFill(start_color = fsbg, end_color = fsbg, fill_type = "solid")
            sheet.cell(column = 2, row = row+1).fill = PatternFill(start_color = fsbg, end_color = fsbg, fill_type = "solid")
            sheet.cell(column = 3, row = row+1).fill = PatternFill(start_color = fsbg, end_color = fsbg, fill_type = "solid")
            sheet.cell(column = 1, row = row+1).border = Border(bottom = Side(style='double'))
            sheet.cell(column = 2, row = row+1).border = Border(bottom = Side(style='double'))
            sheet.cell(column = 3, row = row+1).border = Border(bottom = Side(style='double'))
            wb.save(savepath + "incomestatement.xlsx") #path
            os.startfile(savepath + "incomestatement.xlsx", "open")   #path

    def booklister(self):
        books = []
        if company == "DBPSC":
            bookfinder = "SELECT book FROM DBPSC"
        else:
            bookfinder = "SELECT book FROM DSSI"
        c.execute(bookfinder)
        booker = c.fetchall()
        for b in booker:
            books.append(b[0])
        book_list.config(values = list(dict.fromkeys(books)))
        books.sort()

    def yearlister(self, *args):
        years = []
        if report_list.get() == "Unforwarded GV/DM":
            periodfinder = "SELECT year FROM disbursements WHERE company = ?"
            c.execute(periodfinder, [company])
            periods = c.fetchall()
            for p in periods:
                years.append(p[0])
            year_entry.config(values = list(dict.fromkeys(sorted(years))))
            self.monthlister()
        elif report_list.get() == "Finder":
            if company == "DBPSC":
                periodfinder = "SELECT month FROM DBPSC WHERE year = 2020"
            else:
                periodfinder = "SELECT month FROM DSSI WHERE year = 2020"
            c.execute(periodfinder)
            periods = c.fetchall()
            for p in periods:
                years.append(p[0])
            month_list.config(values = list(dict.fromkeys(sorted(years))))
        elif report_list.get() == "GL Excel":
            if company == "DBPSC":
                periodfinder = "SELECT year FROM DBPSC"
            else:
                periodfinder = "SELECT year FROM DSSI"
            c.execute(periodfinder)
            periods = c.fetchall()
            for p in periods:
                years.append(p[0])
            year_entry.config(values = list(dict.fromkeys(sorted(years))))
        else:
            if company == "DBPSC":
                periodfinder = "SELECT year FROM DBPSC"
            else:
                periodfinder = "SELECT year FROM DSSI"
            c.execute(periodfinder)
            periods = c.fetchall()
            for p in periods:
                years.append(p[0])
            year_entry.config(values = list(dict.fromkeys(sorted(years))))
            self.monthlister()

    def monthlister(self):
        months = []
        if company == "DBPSC":
            periodfinder = "SELECT month FROM DBPSC WHERE year = ?"
        else:
            periodfinder = "SELECT month FROM DSSI WHERE year = ?"
        if report_list.get() == "Trial Balance" or report_list.get() == "Financial Statements":
            c.execute(periodfinder, [year_entry.get()])
            periods = c.fetchall()
            for p in periods:
                months.append(p[0])
            month_list.config(values = list(dict.fromkeys(sorted(months))))
        elif report_list.get() == "GL Excel":
            pass
        else:
            c.execute(periodfinder, [year_entry.get()])
            periods = c.fetchall()
            for p in periods:
                months.append(p[0])
            month_list.config(values = list(dict.fromkeys(sorted(months))))

    def fmdmonthlister2(self, *args):
        lister = "SELECT month FROM disbursements WHERE year = ? AND company = ?"
        c.execute(lister, [year_entry.get(), company])
        listresult = c.fetchall()
        months = []
        if listresult:
            for i in listresult:
                months.append(i[0])
            month_list.config(values = list(dict.fromkeys(sorted(months))))

    def fmdyearlister2(self):
        lister = "SELECT year FROM disbursements WHERE company = ?"
        c.execute(lister, [company])
        listresult = c.fetchall()
        years = []
        if listresult:
            for i in listresult:
                years.append(i[0])
            year_entry.config(values = list(dict.fromkeys(sorted(years))))

    def extracetbresults(self, *args):
        global tbresults, reportmonth, reportyear, tbresults
        reportyear = year_entry.get()
        reportmonth = month_list.get()
        if company == "DBPSC":
            extracter = "SELECT code, title, SUM(debit), SUM(credit) FROM (SELECT code, title, debit, credit FROM begbal UNION ALL SELECT code, title, debit, credit FROM DBPSC WHERE year <= ? AND month <= ?) GROUP BY code, title"
        else:
            extracter = "SELECT code, title, SUM(debit), SUM(credit) FROM DSSI WHERE year = ? AND month = ? GROUP BY code, title"
        c.execute(extracter, [reportyear, reportmonth])
        tbresults = c.fetchall()
        if tbresults:
            self.showtbresults(scroll_frame, tbresults)
            self.reporttotal(total_frame, tbresults, 2, 3)
            printtb_button.config(state = NORMAL)
        else:
            messagebox.showerror("Error", "No record found!")

    def showtbresults(self, master, result):
        global scrollbox
        scrollbox = Frame(master)
        container = Frame(scrollbox)
        canvas = Canvas(container, bg = wc)
        scrollbar = Scrollbar(container, orient = "vertical", width = 25, command = canvas.yview)
        scrollable_frame = Frame(canvas)
        scrollable_frame.bind("<Configure>", lambda e: canvas.configure(scrollregion = canvas.bbox("all")))
        canvas.create_window((0,0), window = scrollable_frame, anchor = "nw")
        canvas.configure(yscrollcommand = scrollbar.set)

        row = 0
        for i in result:
            if i[2]+i[3] != 0:
                if i[0][0] == "1":
                    self.reportlinemaster(scrollable_frame, row, i[0], i[1], format(round((i[2]-i[3]),2), ",.2f"), float(0.00))
                    row += 1
                elif i[0][0] == "6":
                    self.reportlinemaster(scrollable_frame, row, i[0], i[1], format(round((i[2]-i[3]),2), ",.2f"), float(0.00))
                    row += 1
                elif i[0][0] == "2":
                    self.reportlinemaster(scrollable_frame, row, i[0], i[1], float(0.00), format(round((i[3]-i[2]),2), ",.2f"))
                    row += 1
                elif i[0][0] == "3":
                    self.reportlinemaster(scrollable_frame, row, i[0], i[1], float(0.00), format(round((i[3]-i[2]),2), ",.2f"))
                    row += 1
                elif i[0][0] == "5":
                    self.reportlinemaster(scrollable_frame, row, i[0], i[1], float(0.00), format(round((i[3]-i[2]),2), ",.2f"))
                    row += 1
                elif i[0][0] == "7":
                    self.reportlinemaster(scrollable_frame, row, i[0], i[1], float(0.00), format(round((i[3]-i[2]),2), ",.2f"))
                    row += 1
                elif i[0][0] == "4":
                    self.reportlinemaster(scrollable_frame, row, i[0], i[1], float(0.00), format(round((i[3]-i[2]),2), ",.2f"))
                    row += 1
            else:
                pass
        scrollbox.grid(column = 0, row = 2, sticky = NW)
        container.pack()
        canvas.pack(side = "left", fill = "both", expand = True)
        canvas.config(width = 562, height = 500)
        scrollbar.pack(side = "right", fill = "y")

    def finderfinder(self, *args):
        global reportmonth, reportyear, finder
        reportmonth = month_list.get()
        reportyear = "2020"
        if company == "DBPSC":
            if column_list.get() == "reference":
                finderfinder = "SELECT number, date, code, title, debit, credit, comment, reference, client, particulars FROM DBPSC WHERE reference LIKE ? AND month = ? ORDER BY number"
            elif column_list.get() == "particulars":
                finderfinder = "SELECT number, date, code, title, debit, credit, comment, reference, client, particulars FROM DBPSC WHERE particulars LIKE ? AND month = ? ORDER BY number"
            elif column_list.get() == "client":
                finderfinder = "SELECT number, date, code, title, debit, credit, comment, reference, client, particulars FROM DBPSC WHERE client LIKE ? AND month = ? ORDER BY number"
        else:
            if column_list.get() == "reference":
                finderfinder = "SELECT number, date, code, title, debit, credit, comment, reference, client, particulars FROM DSSI WHERE reference LIKE ? AND month = ? ORDER BY number"
            elif column_list.get() == "particulars":
                finderfinder = "SELECT number, date, code, title, debit, credit, comment, reference, client, particulars FROM DSSI WHERE particulars LIKE ? AND month = ? ORDER BY number"
            elif column_list.get() == "client":
                finderfinder = "SELECT number, date, code, title, debit, credit, comment, reference, client, particulars FROM DSSI WHERE client LIKE ? AND month = ? ORDER BY number"
        c.execute(finderfinder, ["%"+finder_entry.get()+"%",month_list.get()])
        finder = c.fetchall()
        if finder:
            for p in finder:
                self.showanalysis(scroll_frame, finder)
                self.reporttotal(total_frame, finder, 4, 5)
                printreport_button.config(state = NORMAL)
        
    def closereports(self, *args):
        reports_frame.destroy()
        self.menubuttons(NORMAL)
        self.colorswitch(reports_button, fc)
    
    def accountanalysis(self, *args):
        global analysis, reportmonth, reportyear
        reportyear = year_entry.get()
        reportmonth = month_list.get()
        if company == "DBPSC":
            extracter = "SELECT number, date, code, title, debit, credit, comment, reference, client, particulars FROM DBPSC WHERE year = ? AND month = ? AND code = ? ORDER BY number"
        else:
            extracter = "SELECT number, date, code, title, debit, credit, comment, reference, client, particulars FROM DSSI WHERE year = ? AND month = ? AND code = ? ORDER BY number"
        c.execute(extracter, [reportyear,reportmonth,code_entry.get()])
        analysis = c.fetchall()
        if analysis:
            self.showanalysis(scroll_frame, analysis)
            self.reporttotal(total_frame, analysis, 4, 5)
            printreport_button.config(state = NORMAL)
        else:
            messagebox.showerror("Account Analyis","No record found!")
    
    def showanalysis(self, master, result):
        global scrollbox
        scrollbox = Frame(master)
        container = Frame(scrollbox)
        canvas = Canvas(container, bg = wc)
        scrollbar = Scrollbar(container, orient = "vertical", width = 25, command = canvas.yview)
        scrollable_frame = Frame(canvas)
        scrollable_frame.bind("<Configure>", lambda e: canvas.configure(scrollregion = canvas.bbox("all")))
        canvas.create_window((0,0), window = scrollable_frame, anchor = "nw")
        canvas.configure(yscrollcommand = scrollbar.set)

        row = 0
        for i in result:
            self.anallinemaster(scrollable_frame, row, i[0], i[1], format(i[4],",.2f"), format(i[5],",.2f"), i[7], i[9])
            row += 1

        scrollbox.grid(column = 0, row = 2, sticky = NW)
        container.pack()
        canvas.pack(side = "left", fill = "both", expand = True)
        canvas.config(width = 760, height = 500) 
        scrollbar.pack(side = "right", fill = "y")

    def journalcode_val(self, *args):
        global result
        codeinput = code_entry.get()
        if len(codeinput) == 4 and codeinput.isdigit():
            if company == "DBPSC":
                find_code = "SELECT title FROM chart WHERE code = ?"
            else:
                find_code = "SELECT title FROM dssichart WHERE code = ?"
            c.execute(find_code, [codeinput])
            result = c.fetchall()
            if result:
                title_label.config(text = result[0][0])
            else:
                messagebox.showerror("Code Error", "Code not in Chart!")
                code_entry.delete(0, END)
        else:
            code_entry.delete(0, END)

    def extracteom(self, *args):
        global eomresults, reportmonth, reportyear
        reportyear = year_entry.get()
        reportmonth = month_list.get()
        if company == "DBPSC":
            extracter = "SELECT code, title, SUM(debit), SUM(credit) FROM DBPSC WHERE year = ? AND month = ? GROUP BY code, title"
        else:
            extracter = "SELECT code, title, SUM(debit), SUM(credit) FROM DSSI WHERE year = ? AND month = ? GROUP BY code, title"
        c.execute(extracter, [reportyear, reportmonth])
        eomresults = c.fetchall()
        if eomresults:
            self.showeomresults(scroll_frame)
            self.reporttotal(total_frame, eomresults, 2, 3)
            printreport_button.config(state = NORMAL)
        else:
            messagebox.showerror("Error", "No record found!")

    def showeomresults(self, master):
        global scrollbox
        scrollbox = Frame(master)
        container = Frame(scrollbox)
        canvas = Canvas(container, bg = wc)
        scrollbar = Scrollbar(container, orient = "vertical", width = 25, command = canvas.yview)
        scrollable_frame = Frame(canvas)
        scrollable_frame.bind("<Configure>", lambda e: canvas.configure(scrollregion = canvas.bbox("all")))
        canvas.create_window((0,0), window = scrollable_frame, anchor = "nw")
        canvas.configure(yscrollcommand = scrollbar.set)

        row = 0
        for i in eomresults:
            if i[0][0] == "1":
                self.reportlinemaster(scrollable_frame, row, i[0], i[1], format(round((i[2]-i[3]),2), ",.2f"), float(0.00))
                row += 1
            elif i[0][0] == "6":
                self.reportlinemaster(scrollable_frame, row, i[0], i[1], format(round((i[2]-i[3]),2), ",.2f"), float(0.00))
                row += 1
            elif i[0][0] == "2":
                self.reportlinemaster(scrollable_frame, row, i[0], i[1], float(0.00), format(round((i[3]-i[2]),2), ",.2f"))
                row += 1
            elif i[0][0] == "3":
                self.reportlinemaster(scrollable_frame, row, i[0], i[1], float(0.00), format(round((i[3]-i[2]),2), ",.2f"))
                row += 1
            elif i[0][0] == "5":
                self.reportlinemaster(scrollable_frame, row, i[0], i[1], float(0.00), format(round((i[3]-i[2]),2), ",.2f"))
                row += 1
            elif i[0][0] == "7":
                self.reportlinemaster(scrollable_frame, row, i[0], i[1], float(0.00), format(round((i[3]-i[2]),2), ",.2f"))
                row += 1
            elif i[0][0] == "4":
                self.reportlinemaster(scrollable_frame, row, i[0], i[1], float(0.00), format(round((i[3]-i[2]),2), ",.2f"))
                row += 1

        scrollbox.grid(column = 0, row = 2, sticky = NW)
        container.pack()
        canvas.pack(side = "left", fill = "both", expand = True)
        canvas.config(width = 562, height = 500)
        scrollbar.pack(side = "right", fill = "y")

    def reportlinemaster(self, master, linerow, var1, var2, var3, var4):
        eomframe = Frame(master, bg = wc)
        eomframe.grid(column = 0, row = linerow)

        eomcode = Entry(eomframe, font = fonts, width = 4)
        eomcode.grid(column = 0, row = 0, padx = 1)
        eomcode.insert(0, var1)
        
        eomtitle = Entry(eomframe, font = fonts, width = 35)
        eomtitle.grid(column = 1, row = 0, padx = 1)
        eomtitle.insert(0, var2)

        eomdebit = Entry(eomframe, font = fonts, width = 14, justify = RIGHT)
        eomdebit.grid(column = 2, row = 0, padx = 1)
        eomdebit.insert(0, var3)

        eomcredit = Entry(eomframe, font = fonts, width = 14, justify = RIGHT)
        eomcredit.grid(column = 3, row = 0, padx = 1)
        eomcredit.insert(0, var4)
    
    def anallinemaster(self, master, linerow, var1, var2, var3, var4, var5, var6):
        anal_frame = Frame(master, bg = wc)
        anal_frame.grid(column = 0, row = linerow)

        analnum = Entry(anal_frame, font = fonts, width = 8)
        analnum.grid(column = 0, row = 0, padx = 1)
        analnum.insert(0, var1)
        
        analdat = Entry(anal_frame, font = fonts, width = 10)
        analdat.grid(column = 1, row = 0, padx = 1)
        analdat.insert(0, var2)

        analdebit = Entry(anal_frame, font = fonts, width = 14, justify = RIGHT)
        analdebit.grid(column = 4, row = 0, padx = 1)
        analdebit.insert(0, var3)

        analcredit = Entry(anal_frame, font = fonts, width = 14, justify = RIGHT)
        analcredit.grid(column = 5, row = 0, padx = 1)
        analcredit.insert(0, var4)

        analref = Entry(anal_frame, font = fonts, width = 10, justify = LEFT)
        analref.grid(column = 6, row = 0, padx = 1)
        analref.insert(0, var5)

        analpar = Entry(anal_frame, font = fonts, width = 38, justify = LEFT)
        analpar.grid(column = 7, row = 0, padx = 1)
        analpar.insert(0, var6)

    def reporttotal(self, master, result, dr, cr, *args):
        global debit_total, credit_total, reporttotalframe, variance, resultdr, resultcr
        resultdr = []
        resultcr = []
        if report_list.get() == "Account Analysis":
            padding = 19
            for r in result:
                resultdr.append(r[dr])
                resultcr.append(r[cr])
        elif report_list.get() == "Finder":
            padding = 19
            for r in result:
                resultdr.append(r[dr])
                resultcr.append(r[cr])
        else:
            padding = 104
            for r in result:
                if r[0][0] == "1":
                    resultdr.append(r[dr]-r[cr])
                elif r[0][0] == "6":
                    resultdr.append(r[dr]-r[cr])
                elif r[0][0] == "2":
                    resultcr.append(r[cr]-r[dr])
                elif r[0][0] == "3":
                    resultcr.append(r[cr]-r[dr])
                elif r[0][0] == "5":
                    resultcr.append(r[cr]-r[dr])
                elif r[0][0] == "7":
                    resultcr.append(r[cr]-r[dr])
                elif r[0][0] == "4":
                    resultcr.append(r[cr]-r[dr])
            
        reporttotalframe = Frame(master, bg = wc)
        reporttotalframe.grid(column = 0, row = 0)

        variance = Label(reporttotalframe, text = format((sum(resultdr)-sum(resultcr)), ",.2f"), font = fonts, anchor = E, width = 14, bg = wc, fg = fc)
        variance.grid(column = 0, row = 0, padx = padding)

        debit_total = Entry(reporttotalframe, font = fonts, relief = RIDGE, width = 14, justify = RIGHT)
        debit_total.grid(column = 1, row = 0, padx = 2)
        debit_total.delete(0, END)
        debit_total.insert(0, format(round(sum(resultdr),2), ",.2f"))

        credit_total = Entry(reporttotalframe, font = fonts, relief = RIDGE, width = 14, justify = RIGHT)
        credit_total.grid(column = 2, row = 0)
        credit_total.delete(0, END)
        credit_total.insert(0, format(round(sum(resultcr),2), ",.2f"))  

    def reportprinter(self, *args):
        try:
            reportyear = year_entry.get()
        except:
            reportyear = "2020"
        wb = openpyxl.load_workbook(path + "report.xlsx") #path
        sheet = wb.active
        sheet.append([company])
        sheet.append([report_list.get()])
        sheet.append([month_list.get() + "-" + reportyear])
        if report_list.get() == "End-of-month":
            sheet.append(["code","title","debit","credit"])
            for i in eomresults:
                if i[0][0] == "1":
                    sheet.append([i[0],i[1],round((i[2]-i[3]),2),0])
                elif i[0][0] == "6":
                    sheet.append([i[0],i[1],round((i[2]-i[3]),2),0])
                elif i[0][0] == "2":
                    sheet.append([i[0],i[1],0,round((i[3]-i[2]),2)])
                elif i[0][0] == "3":
                    sheet.append([i[0],i[1],0,round((i[3]-i[2]),2)])
                elif i[0][0] == "5":
                    sheet.append([i[0],i[1],0,round((i[3]-i[2]),2)])
                elif i[0][0] == "7":
                    sheet.append([i[0],i[1],0,round((i[3]-i[2]),2)])
                elif i[0][0] == "4":
                    sheet.append([i[0],i[1],0,round((i[3]-i[2]),2)])
            sheet.append(["Total:","",round(sum(resultdr),2),round(sum(resultcr),2)])

        elif report_list.get() == "Trial Balance":
            sheet.append(["code","title","debit","credit"])
            for i in tbresults:
                if i[2]+i[3] != 0:
                    if i[0][0] == "1":
                        sheet.append([i[0],i[1],round((i[2]-i[3]),2),0])
                    elif i[0][0] == "6":
                        sheet.append([i[0],i[1],round((i[2]-i[3]),2),0])
                    elif i[0][0] == "2":
                        sheet.append([i[0],i[1],0,round((i[3]-i[2]),2)])
                    elif i[0][0] == "3":
                        sheet.append([i[0],i[1],0,round((i[3]-i[2]),2)])
                    elif i[0][0] == "5":
                        sheet.append([i[0],i[1],0,round((i[3]-i[2]),2)])
                    elif i[0][0] == "7":
                        sheet.append([i[0],i[1],0,round((i[3]-i[2]),2)])
                    elif i[0][0] == "4":
                        sheet.append([i[0],i[1],0,round((i[3]-i[2]),2)])
            sheet.append(["Total:","",round(sum(resultdr),2),round(sum(resultcr),2)])

        elif report_list.get() == "Account Analysis":
            sheet.append(["jv no.","date","code","title","debit","credit","comment","reference","client","particulars"])
            for i in analysis:
                sheet.append(i)
            sheet.append(["Total:","","","",round(sum(resultdr),2),round(sum(resultcr),2)])
        elif report_list.get() == "Finder":
            sheet.append(["jv no.","date","code","title","debit","credit","comment","reference","client","particulars"])
            for i in finder:
                sheet.append(i)
            sheet.append(["Total:","","","",round(sum(resultdr),2),round(sum(resultcr),2)])
        elif report_list.get() == "Bookkeeping":
            sheet.append(["book","jv no.","date","code","title","debit","credit","comment","reference","client","particulars"])
            if company == "DBPSC":
                keeper = "SELECT book,number,date,code,title,debit,credit,reference,particulars FROM DBPSC WHERE month = ? AND year = ? AND book = ? ORDER BY date"
            else:
                keeper = "SELECT book,number,date,code,title,debit,credit,reference,particulars FROM DSSI WHERE month = ? AND year = ? AND book = ? ORDER BY date"
            c.execute(keeper, [month_list.get(),year_entry.get(),book_list.get()])
            booker = c.fetchall()
            if booker:
                for b in booker:
                    sheet.append(b)
        sheet.append(["Printed by: " + self.username.get()])
        sheet.append(["Date printed: " + today.strftime('%m-%d-%Y')])
        wb.save(savepath + "charliereports.xlsx") #path
        os.startfile(savepath + "charliereports.xlsx", "open") #path

    def closefinder(self, *args):
        finder_frame.destroy()
        report_list.config(state = NORMAL)
        report_list.delete(0, END)

    def closeanal(self, *args):
        analysis_frame.destroy()
        report_list.config(state = NORMAL)
        report_list.delete(0, END)

    def closeeom(self, *args):
        monthend_frame.destroy()
        report_list.config(state = NORMAL)
        report_list.delete(0, END)

    def closetb(self, *args):
        tb_frame.destroy()
        report_list.config(state = NORMAL)
        report_list.delete(0, END)

    def closeexcel(self, *args):
        excel_frame.destroy()
        report_list.config(state = NORMAL)
        report_list.delete(0, END)

    def closebook(self, *args):
        book_frame.destroy()
        report_list.config(state = NORMAL)
        report_list.delete(0, END)

    def closebs(self, *args):
        fs_frame.destroy()
        report_list.config(state = NORMAL)
        report_list.delete(0, END)

    def closeud(self, *args):
        ud_frame.destroy()
        report_list.config(state = NORMAL)
        report_list.delete(0, END)

    def closesss(self, *args):
        sss_frame.destroy()
        report_list.config(state = NORMAL)
        report_list.delete(0, END)

### payroll factory ###
    def showpayroll(self, *args):
        global client_type
        if company == "DBPSC":
            clienttype = ['clerical', 'janitorial', 'jfcc', 'jfcj']
        else:
            clienttype = ['security']
        self.menubuttons(DISABLED)
        self.colorswitch(payroll_button, fc)
        global newtype, payroll_frame, book
        newtype = "D"
        book = "GJ"
        payroll_frame = LabelFrame(self.master, text = "\nPayroll Importer", font = fonts, bg = wc, fg = fc)
        payroll_frame.grid(column = 1, row = 0, sticky = NW)

        import_frame = LabelFrame(payroll_frame, text = "Import", font = fonts, bg = wc, fg = fc)
        import_frame.grid(column = 0, row = 0, sticky = NW)

        global paydate_type
        paydate = os.listdir(savepath + "payroll_files") #path
        paydate_type = tk.Combobox(import_frame, font = fonts, values = paydate, width = 9)
        paydate_type.grid(column = 0, row = 0, sticky = W, padx = 3)
        paydate_type.insert(0, "paydate")
        paydate_type.bind("<<ComboboxSelected>>", self.clientdir)
        paydate_type.bind("<FocusOut>", self.clientdir)
        
        global client_name
        client_name = tk.Combobox(import_frame, font = fonts, width = 18)
        client_name.grid(column = 1, row = 0, padx = 3)
        client_name.insert(0, "client")

        client_type = tk.Combobox(import_frame, font = fonts, values = clienttype, width = 7)
        client_type.grid(column = 2, row = 0, sticky = W, padx = 3)
        client_type.insert(0, 'type')

        import_button = Button(import_frame, text = "Import", font = fonts, bg = wc, bd = 0, width = 95, image = buttonicon, compound = CENTER, cursor = "hand2", command = self.importer)
        import_button.grid(column = 3, row = 0, padx = 6)
        import_button.bind("<Return>", self.importer)
        
        global post_button
        post_button = Button(import_frame, text = "Post", font = fonts, bg = wc, bd = 0, width = 95, image = buttonicon, compound = CENTER, cursor = "hand2", command = self.debitcreditchecker)
        post_button.grid(column = 4, row = 0, padx = 5)
        post_button.bind("<Return>", self.debitcreditchecker)
        post_button.config(state = DISABLED)

        global print_button
        print_button = Button(import_frame, text = "Print", font = fonts, bg = wc, bd = 0, width = 95, image = buttonicon, compound = CENTER, cursor = "hand2", command = self.printer)
        print_button.grid(column = 5, row = 0, padx = 5)
        print_button.bind("<Return>", self.printer)
        print_button.config(state = DISABLED)
        
        details_back = Button(import_frame, text = "Close", font = fonts, bg = wc, bd = 0, width = 95, image = buttonicon, compound = CENTER, cursor = "hand2", command = self.closepayroll)
        details_back.grid(column = 6, row = 0, padx = 5)
        details_back.bind("<Return>", self.closepayroll)

    def clientdir(self, *args):
        try:
            client_name.config(values = os.listdir(savepath + "payroll_files/" + paydate_type.get())) #path
        except:
            client_name.config(values = ["No records found"])

    def closepayroll(self, *args):
        payroll_frame.destroy()
        self.menubuttons(NORMAL)
        self.colorswitch(payroll_button, fc)
         
    def showpayrolldetails(self, frame):
        global details_frame
        details_frame = LabelFrame(frame, text = "Details", font = fonts, bg = wc, fg = fc)
        details_frame.grid(column = 0, row = 1, sticky = NW)

        voucher_label = Label(details_frame, text = "Voucher No.", font = fonts, anchor = W, width = 20, bg = wc, fg = fc)
        voucher_label.grid(column = 0, row = 0)
        
        date_label = Label(details_frame, text = "Date", font = fonts, anchor = W, width = 20, bg = wc, fg = fc)
        date_label.grid(column = 0, row = 1)

        client_label = Label(details_frame, text = "Client/Supplier", font = fonts, anchor = W, width = 20, bg = wc, fg = fc)
        client_label.grid(column = 0, row = 2)

        particulars_label = Label(details_frame, text = "Particulars", font = fonts, anchor = W, width = 20, bg = wc, fg = fc)
        particulars_label.grid(column = 0, row = 3)

        reference_label = Label(details_frame, text = "Reference", font = fonts, anchor = W, width = 20, bg = wc, fg = fc)
        reference_label.grid(column = 0, row = 4)

        global voucher_number
        voucher_number = Text(details_frame, font = fonts, height = 1, bd = 3, relief = SUNKEN, width = 9)
        voucher_number.grid(column = 1, row = 0, sticky = W)
        voucher_number.insert(1.0, "")
        voucher_number.config(state = DISABLED)

        global date
        date = StringVar()
        global date_entry
        date_entry = Entry(details_frame, textvariable = date, font = fonts, bd = 3, relief = SUNKEN, width = 10)
        date_entry.grid(column = 1, row = 1, sticky = W)
        date_entry.bind("<Tab>", self.formatdate)

        global client_entry
        client_entry = Entry(details_frame, font = fonts, width = 50)
        client_entry.grid(column = 1, row = 2, sticky = W)
        client_entry.bind("<FocusOut>", lambda e: self.uppercase(client_entry))

        global particulars_entry
        particulars_entry = Entry(details_frame, font = fonts, width = 75, bd = 3, relief = SUNKEN)
        particulars_entry.grid(column = 1, row = 3, sticky = W)
        particulars_entry.bind("<FocusOut>", lambda e: self.uppercase(particulars_entry))

        global reference_entry
        reference_entry = Entry(details_frame, font = fonts, width = 20, bd = 3, relief = SUNKEN)
        reference_entry.grid(column = 1, row = 4, sticky = W)
        reference_entry.insert(0, "EIS Payroll Register")
        
    def importer(self, *args):
        try:
            details_frame.destroy()
            details.destroy()
            entries_frame.destroy()
            status_frame.destroy()
            scrollbox.desroy()
        except:
            pass
        finally:
            global addlist, clerical, janitorial, jfcc, jfcj, security, gettype, getclient
            addlist = []
            clerical = []
            janitorial = []
            jfcc = []
            jfcj =  []
            security = []
            getclient = str(client_name.get())
            gettype = client_type.get()
            if gettype == 'clerical':
                clerical.append(getclient)
                addlist.append(getclient)
            elif gettype == 'janitorial':
                janitorial.append(getclient)
                addlist.append(getclient)
            elif gettype == 'jfcc':
                jfcc.append(getclient)
                addlist.append(getclient)
            elif gettype == 'jfcj':
                jfcj.append(getclient)
                addlist.append(getclient)
            elif gettype == 'security':
                jfcj.append(getclient)
                addlist.append(getclient)
                    
            self.showpayrolldetails(payroll_frame)  
            self.payrollmaster(addlist, clerical, janitorial, jfcc, jfcj, security)
            self.labels(payroll_frame, 2)
            self.showalllines(payroll_frame, NORMAL)
            self.arranger()
            self.lister()
            code_entry1.focus()

    def payrollmaster(self, addlist, clerical, janitorial, jfcc, jfcj, security):
        global pclient,pperiod,basicpay,overtimepay,holidaypay,ssspremium,hdmfpremium,philhealthpremium,sssecc,coverup,ecola,primetime,lwp,nightdifferential,sea,ctpa,nthmonthpay,hazardpay,refund,monetization,communication,deminimis,family,gas,maintenance,meal,motorcycle,travel,uniform,adjbasicpay,adjovertimepay,adjlwp,apeccpremium,apwithholdingtax,aphdmfpremium,apphilhealthpremium,apssspremium,apssscalamity,apssssalary,apsssinvestment,apssshousing,aphdmfcalamity,aphdmfmp2,aphdmfmpl,aphdmfpivol,apcashbond,apewt,aptriplehgc,aptriplehgrocery,aptriplehprepaid,aptriplehhmo,aptriplehmfl,aptriplehabuloy,aptriplehothers,arcashadvance,artelephonebill,arrental,arrepairs,idorsupplies,otherdeductions,aratmcharges,sssclaims,adjbasicpay2,adjovertimepay2,adjlwp2,bdo,cash,rcbc,payroll,lbp
        for payroll in addlist:
            wb = openpyxl.load_workbook(savepath + "payroll_files/" + paydate_type.get() + "/" + payroll) #path
            s1 = wb.active
            pclient = s1['A5'].value.split('NAME: ')[1]
            pperiod = s1['A6'].value.split('Period : ')[1]
            basicpay = s1['H29'].value   #          ### DEBIT
            overtimepay = s1['H30'].value #
            holidaypay = s1['H31'].value #
            ssspremium = s1['H32'].value #
            hdmfpremium = s1['H33'].value #
            philhealthpremium = s1['H34'].value #
            sssecc = s1['H35'].value #
            coverup = s1['H36'].value #
            ecola = s1['H37'].value #
            primetime = s1['H38'].value #
            lwp = s1['H39'].value #
            nightdifferential = s1['H40'].value #
            sea = s1['H41'].value #???
            ctpa = s1['H42'].value #???
            nthmonthpay = s1['H43'].value #
            hazardpay = s1['H44'].value #
            refund = s1['H45'].value #
            monetization = s1['H46'].value #
            communication = s1['H48'].value #
            deminimis = s1['H49'].value #
            family = s1['H50'].value #
            gas = s1['H51'].value #
            maintenance = s1['H52'].value #
            meal = s1['H53'].value #
            motorcycle = s1['H54'].value #
            travel = s1['H55'].value #
            uniform = s1['H56'].value #
            adjbasicpay = s1['H59'].value #
            adjovertimepay = s1['H60'].value #
            adjlwp = s1['H61'].value #
            apeccpremium = s1['P30'].value #           ### CREDIT
            apwithholdingtax = s1['P31'].value #
            apssspremium = s1['P32'].value #
            aphdmfpremium = s1['P33'].value #
            apphilhealthpremium = s1['P34'].value #
            apssscalamity = s1['P36'].value #
            apssssalary = s1['P37'].value #
            apsssinvestment = s1['P38'].value #
            apssshousing = s1['P39'].value #
            aphdmfcalamity = s1['P41'].value #
            aphdmfmp2 = s1['P42'].value #
            aphdmfmpl = s1['P43'].value #
            aphdmfpivol = s1['P44'].value #
            apcashbond = s1['P45'].value #
            apewt = s1['P46'].value #
            aptriplehgc = s1['O19'].value #
            aptriplehgrocery = s1['O20'].value #
            aptriplehprepaid = s1['O21'].value #
            aptriplehhmo = s1['P19'].value #
            aptriplehmfl = s1['P20'].value #
            aptriplehabuloy = s1['P21'].value #
            aptriplehothers = s1['P22'].value #
            arcashadvance = s1['P49'].value #
            artelephonebill = s1['P50'].value #
            arrental = s1['P51'].value #
            arrepairs = s1['P52'].value #
            idorsupplies = s1['P54'].value #
            otherdeductions = s1['P55'].value #
            aratmcharges = s1['P56'].value #
            sssclaims = s1['P57'].value #
            adjbasicpay2 = s1['P59'].value #
            adjovertimepay2 = s1['P60'].value #
            adjlwp2 = s1['P61'].value #
            bdo = s1['T31'].value #
            cash = s1['T32'].value #
            rcbc = s1['T33'].value #
            lbp = s1['T34'].value #
            wb.close()
            
    def lister(self):
        if company == "DBPSC":
            commoncredit = ['2116','2111','2115','2115','2115','2112','2114','2114','2114','2117','2113','2101','2103','2307','2309','2305','1320','1320','1320','1320','1320','2117','2120','2321','1329','1303','6256','6259','6283','2307','2307','2307']
            commoncomment = ['','','PREMIUMS','MP2','PIVOL','','SALARY LOAN','CALAMITY LOAN','INVESTMENT','MPL','','','','ABULOY','','','CASH ADVANCE','REPAIRS','TELEPHONE','RENTAL','ATM','','CALAMITY LOAN','','','SSS CLAIMS','BASIC ADJUSTMENT','OT ADJUSTMENT','LEAVE ADJUSTMENT','CASH','BDO','RCBC']
        else:
            commoncredit = ['2116','2111','2115','2115','2115','2112','2114','2114','2114','2117','2113','2101','2103','2307','2309','2305','1320','1320','1320','1320','1320','2117','2120','2321','1329','1303','6258','6261','6264','2307','2417','2417','2307']
            commoncomment = ['','','PREMIUMS','MP2','PIVOL','','SALARY LOAN','CALAMITY LOAN','INVESTMENT','MPL','','','','ABULOY','','','CASH ADVANCE','REPAIRS','TELEPHONE','RENTAL','ATM','','CALAMITY LOAN','','','SSS CLAIMS','BASIC ADJUSTMENT','OT ADJUSTMENT','LEAVE ADJUSTMENT','CASH','BDO','RCBC','LBP']
        cdebit = ['6256','6259','6262','6265','6276','6268','6271','6281','6283','6411','6229','1345','6277','1329']
        jdebit = ['6257','6260','6263','6266','6274','6269','6272','6282','6284','6411','6229','1345','6278','1329']
        jfccdebit = ['6321','6325','6328','6329','6335','6333','6331','6323','6339','6411','6343','1345','6337','1329']
        jfcjdebit = ['6322','6326','6328','6330','6336','6334','6332','6324','6284','6411','6418','1345','6338','1329']
        securitydebit = ['6258','6261','6264','6267','6275','6270','6273','6295','6285','6229','6229','1345','6279','1329']
        jj = ['self.journalcode_val1()','self.journalcode_val2()','self.journalcode_val3()','self.journalcode_val4()','self.journalcode_val5()','self.journalcode_val6()','self.journalcode_val7()','self.journalcode_val8()','self.journalcode_val9()','self.journalcode_val10()','self.journalcode_val11()','self.journalcode_val12()','self.journalcode_val13()','self.journalcode_val14()','self.journalcode_val15()','self.journalcode_val16()','self.journalcode_val17()','self.journalcode_val18()','self.journalcode_val19()','self.journalcode_val20()','self.journalcode_val21()','self.journalcode_val22()','self.journalcode_val23()','self.journalcode_val24()','self.journalcode_val25()','self.journalcode_val26()','self.journalcode_val27()','self.journalcode_val28()','self.journalcode_val29()','self.journalcode_val30()']
        self.allvars()
        client_entry.insert(0, pclient)
        particulars_entry.insert(0, pperiod)
        if gettype == "clerical":
            i = 0
            e = 0
            for d in dd:
                if d != 0:
                    codelist[i].insert(0, cdebit[e])
                    eval(jj[i])
                    drs[i].insert(0, format(d, ",.2f"))
                    crs[i].insert(0, 0)
                    i += 1
                    e += 1
                else:
                    e += 1
            s = []
            for code in codelist:
                if code.get() != "":
                    s.append(1)
                    i = len(s)
                    e = 0
            for c in cc:
                if c != 0:
                    codelist[i].insert(0, commoncredit[e])
                    eval(jj[i])
                    crs[i].insert(0, format(c, ",.2f"))
                    drs[i].insert(0, 0)
                    comments[i].insert(0, commoncomment[e])
                    i += 1
                    e += 1
                else:
                    e += 1
        elif gettype == "janitorial":
            i = 0
            e = 0
            for d in dd:
                if d != 0:
                    codelist[i].insert(0, jdebit[e])
                    eval(jj[i])
                    drs[i].insert(0, format(d, ",.2f"))
                    crs[i].insert(0, 0)
                    i += 1
                    e += 1
                else:
                    e += 1
            s = []
            for code in codelist:
                if code.get() != "":
                    s.append(1)
                    i = len(s)
                    e = 0
            for c in cc:
                if c != 0:
                    codelist[i].insert(0, commoncredit[e])
                    eval(jj[i])
                    crs[i].insert(0, format(c, ",.2f"))
                    drs[i].insert(0, 0)
                    comments[i].insert(0, commoncomment[e])
                    i += 1
                    e += 1
                else:
                    e += 1
        elif gettype == "jfcc":
            i = 0
            e = 0
            for d in dd:
                if d != 0:
                    codelist[i].insert(0, jfccdebit[e])
                    eval(jj[i])
                    drs[i].insert(0, format(d, ",.2f"))
                    crs[i].insert(0, 0)
                    i += 1
                    e += 1
                else:
                    e += 1
            s = []
            for code in codelist:
                if code.get() != "":
                    s.append(1)
                    i = len(s)
                    e = 0
            for c in cc:
                if c != 0:
                    codelist[i].insert(0, commoncredit[e])
                    eval(jj[i])
                    crs[i].insert(0, format(c, ",.2f"))
                    drs[i].insert(0, 0)
                    comments[i].insert(0, commoncomment[e])
                    i += 1
                    e += 1
                else:
                    e += 1
        elif gettype == "jfcj":
            i = 0
            e = 0
            for d in dd:
                if d != 0:
                    codelist[i].insert(0, jfcjdebit[e])
                    eval(jj[i])
                    drs[i].insert(0, format(d, ",.2f"))
                    crs[i].insert(0, 0)
                    i += 1
                    e += 1
                else:
                    e += 1
            s = []
            for code in codelist:
                if code.get() != "":
                    s.append(1)
                    i = len(s)
                    e = 0
            for c in cc:
                if c != 0:
                    codelist[i].insert(0, commoncredit[e])
                    eval(jj[i])
                    crs[i].insert(0, format(c, ",.2f"))
                    drs[i].insert(0, 0)
                    comments[i].insert(0, commoncomment[e])
                    i += 1
                    e += 1
                else:
                    e += 1
        elif gettype == "security":
            i = 0
            e = 0
            for d in dd:
                if d != 0:
                    codelist[i].insert(0, securitydebit[e])
                    eval(jj[i])
                    drs[i].insert(0, format(d, ",.2f"))
                    crs[i].insert(0, 0)
                    i += 1
                    e += 1
                else:
                    e += 1
            s = []
            for code in codelist:
                if code.get() != "":
                    s.append(1)
                    i = len(s)
                    e = 0
            for c in cc:
                if c != 0:
                    codelist[i].insert(0, commoncredit[e])
                    eval(jj[i])
                    crs[i].insert(0, format(c, ",.2f"))
                    drs[i].insert(0, 0)
                    comments[i].insert(0, commoncomment[e])
                    i += 1
                    e += 1
                else:
                    e += 1
        
    def arranger(self):
        global dd, cc
        salary = sum([basicpay, coverup, nightdifferential, hazardpay, adjbasicpay])
        overtime = sum([overtimepay, primetime, adjovertimepay])
        holiday = holidaypay
        sssp = ssspremium
        hdmfp = hdmfpremium
        phicp = philhealthpremium
        ssse = sssecc
        ecol = ecola
        lwpay = sum([lwp, adjlwp, monetization])
        allow = deminimis
        travels = sum([communication, gas, maintenance, meal, motorcycle, travel, uniform])
        fallow = family
        mp13 = nthmonthpay
        refunds = refund
        
        dd = [salary,overtime,holiday,sssp,hdmfp,phicp,ssse,ecol,lwpay,allow,travels,fallow,mp13,refunds]
       
        aphl = apssshousing
        apsssp = apssspremium
        aphdmfpr = aphdmfpremium
        aphdmfm2 = aphdmfmp2
        aphdmfpvl = aphdmfpivol
        apphicp = apphilhealthpremium
        apsssl = apssssalary
        apssslc = apssscalamity
        apsssli = apsssinvestment
        aphdmfl = aphdmfmpl
        apssse = apeccpremium
        apwtx = apwithholdingtax
        apewtx = apewt
        apabu = aptriplehabuloy
        ap3h = sum([aptriplehgc, aptriplehgrocery, aptriplehprepaid, aptriplehhmo, aptriplehothers])
        apcb = apcashbond
        arca = sum([arcashadvance, idorsupplies])
        arcarl = arrepairs
        arcatb = artelephonebill
        arcar = arrental
        arcaatm = aratmcharges
        aphdmfl2 = 0
        aphdmfcl = aphdmfcalamity
        ap3hmfl = aptriplehmfl
        arothrs = otherdeductions
        ssscr = sssclaims
        adjsalary = adjbasicpay2
        adjot = adjovertimepay2
        adjleave = adjlwp2
        apoewo = cash
        apoebdo = bdo
        apoercbc = rcbc
        aplbp = lbp
        
        cc = [aphl,apsssp,aphdmfpr,aphdmfm2,aphdmfpvl,apphicp,apsssl,apssslc,apsssli,aphdmfl,apssse,apwtx,apewtx,apabu,ap3h,apcb,arca,arcarl,arcatb,arcar,arcaatm,aphdmfl2,aphdmfcl,ap3hmfl,arothrs,ssscr,adjsalary,adjot,adjleave,apoewo,apoebdo,apoercbc,aplbp]
    
### chart factory ###

    def chart_close(self, *args):
        chart_frame.destroy()
        self.menubuttons(NORMAL)
        self.colorswitch(chart_button, fc)

    def chart(self, *args):
        self.menubuttons(DISABLED)
        self.colorswitch(chart_button, fc)
        global chart_frame, addcode_button, update_button, delete_account, chart_back
        chart_frame = LabelFrame(self.master, text = "\nChart of Accounts", font = fonts, bg = wc, fg = fc)
        chart_frame.grid(column = 1, row = 0, ipadx = 22, sticky = NW)

        button_frame = Frame(chart_frame, bg = wc)
        button_frame.grid(column = 0, row = 0, sticky = NW)

        addcode_button = Button(button_frame, text = "Add", font = fonts, bg = wc, bd = 0, width = 95, image = buttonicon, compound = CENTER, cursor = "hand2", command = self.addcode)
        addcode_button.grid(column = 0, row = 0, padx = 5)
        addcode_button.bind("<Return>", self.addcode)
        addcode_button.focus()

        update_button = Button(button_frame, text = "Update", font = fonts, bg = wc, bd = 0, width = 95, image = buttonicon, compound = CENTER, cursor = "hand2", command = self.updcode)
        update_button.grid(column = 1, row = 0, padx = 5)
        update_button.bind("<Return>", self.updcode)

        delete_account = Button(button_frame, text = "Delete", font = fonts, bg = wc, bd = 0, width = 95, image = buttonicon, compound = CENTER, cursor = "hand2", command = self.showdeleteaccount)
        delete_account.grid(column = 2, row = 0, padx = 5)
        delete_account.bind("<Return>", self.showdeleteaccount)
        
        chart_back = Button(button_frame, text = "Close", font = fonts, bg = wc, bd = 0, width = 95, image = buttonicon, compound = CENTER, cursor = "hand2", command = self.chart_close)
        chart_back.grid(column = 3, row = 0, padx = 5)
        chart_back.bind("<Return>", self.chart_close)

    def updcode(self, *args):
        addcode_button.config(state = DISABLED)
        update_button.config(state = DISABLED)
        delete_account.config(state = DISABLED)
        chart_back.config(state = DISABLED)
        self.coalister()
        global updcode_frame, code_list
        updcode_frame = LabelFrame(chart_frame, text = "Update Account", font = fonts, bg = wc, fg = fc)
        updcode_frame.grid(column = 0, row = 1, ipadx = 32, sticky = NW)

        code_list = tk.Combobox(updcode_frame, values = coa, font = fonts, width = 40)
        code_list.grid(column = 1, row = 0, pady = 1, sticky = NW)
        code_list.bind("<<ComboboxSelected>>", self.coafiller)

        updcode_label = Label(updcode_frame, text = "Account Code", font = fonts, anchor = W, width = 20, bg = wc, fg = fc)
        updcode_label.grid(column = 0, row = 1, pady = 1)

        updtitle_label = Label(updcode_frame, text = "Account Title", font = fonts, anchor = W, width = 20, bg = wc, fg = fc)
        updtitle_label.grid(column = 0, row = 2, pady = 1)

        updclass_label = Label(updcode_frame, text = "Class", font = fonts, anchor = W, width = 20, bg = wc, fg = fc)
        updclass_label.grid(column = 0, row = 3, pady = 1)

        global code
        code = StringVar()
        global addcode_entry
        addcode_entry = Entry(updcode_frame, textvariable = code, font = fonts, bd = 3, relief = SUNKEN, width = 4)
        addcode_entry.grid(column = 1, row = 1, sticky = W)
        addcode_entry.bind("<Key>", self.codedigit_validation)
        addcode_entry.bind("<FocusOut>", self.codedigit_validation)
        addcode_entry.bind("<FocusOut>", self.codelen_validation)
        addcode_entry.focus()

        global title
        title = StringVar()
        global addtitle_entry
        addtitle_entry = Entry(updcode_frame, textvariable = title, font = fonts, bd = 3, relief = SUNKEN, width = 40)
        addtitle_entry.grid(column = 1, row = 2, sticky = W)
        addtitle_entry.bind("<FocusOut>", lambda e: self.uppercase(addtitle_entry))

        global addclass_entry
        addclass_entry = Label(updcode_frame, text = "", font = fonts, bd = 3, relief = SUNKEN, width = 2)
        addclass_entry.grid(column = 1, row = 3, sticky = W)

        global upd_account
        upd_account = Button(updcode_frame, text = "Update", font = fonts, bg = wc, bd = 0, width = 95, image = buttonicon, compound = CENTER, cursor = "hand2", command = self.updateaccount)
        upd_account.grid(column = 1, row = 4, pady = 5)
        upd_account.bind("<Return>", self.updateaccount)

        global account_back
        account_back = Button(updcode_frame, text = "Close", font = fonts, bg = wc, bd = 0, width = 95, image = buttonicon, compound = CENTER, cursor = "hand2", command = self.updcode_close)
        account_back.grid(column = 1, row = 5, pady = 5)
        account_back.bind("<Return>", self.updcode_close) 

    def coafinder(self):
        global lister
        if company == "DBPSC":
            c.execute("SELECT * FROM chart")
        else:
            c.execute("SELECT * FROM dssichart")
        lister = c.fetchall()

    def coafiller(self, *args):
        global codetoupdate
        codetoupdate = code_list.get()[0:4]
        if company == "DBPSC":
            coaselected = "SELECT * FROM chart WHERE code = ?"
        else:
            coaselected = "SELECT * FROM dssichart WHERE code = ?"
        c.execute(coaselected, [codetoupdate])
        fill = c.fetchall()[0]
        addcode_entry.delete(0, END)
        addtitle_entry.delete(0, END)
        addcode_entry.insert(0, fill[0])
        addtitle_entry.insert(0, fill[1])
        self.addcodeclass()

    def updateaccount(self, *args):
        updecide = messagebox.askyesno("Update Protocol", "Are you sure?")
        if updecide == True:
            if company == "DBPSC":
                coaupdater = "UPDATE chart SET code = ?, title = ?, class = ? WHERE code = ?"
            else:
                coaupdater = "UPDATE dssichart SET code = ?, title = ?, class = ? WHERE code = ?"
            c.execute(coaupdater, [addcode_entry.get(), addtitle_entry.get(), addclass_entry.cget("text"), codetoupdate])
            conn.commit()
            messagebox.showinfo("Success!","Account has been updated!")
            self.coalister()
        else:
            pass

    def coalister(self):
        global coa
        coa = []
        self.coafinder()
        for l in lister:
            coa.append(l)
        coa.sort()

    def codelen_validation(self, *args):
        self.addcodeclass()
        global codeinput
        codeinput = code.get()
        if len(codeinput) == 4:
            return True
        elif len(codeinput) == 0:
            return True
        else:
            addcode_entry.delete(0, END)
            
    def codedigit_validation(self, *args):
        codeinput = code.get()
        if codeinput.isdigit():
            return True
        else:
            addcode_entry.delete(0, END)       

    def addaccount(self, *args):
        if len(code.get()) == 4 and len(addclass_entry.cget("text")) == 1 and len(title.get()) != 0:
            codeinput = code.get()
            if company == "DBPSC":
                find_code = "SELECT * FROM chart WHERE code = ?"
                addtochart = "INSERT INTO chart (code, title, class) VALUES (?,?,?)"
            else:
                find_code = "SELECT * FROM dssichart WHERE code = ?"
                addtochart = "INSERT INTO dssichart (code, title, class) VALUES (?,?,?)"  
            c.execute(find_code, [codeinput])
            result = c.fetchall()
            if result:
                messagebox.showerror("Error!", "Account Code already exists!")
            else:
                addcodesure = messagebox.askyesno("Add Code Protocol", "Are you sure?")
                if addcodesure == True:    
                    c.execute(addtochart, [code.get(), title.get(), addclass_entry.cget("text")])
                    conn.commit()
                    messagebox.showinfo("Success!", "A new account has been added!" + "\n Account Code: " + code.get() + "\n Account Title: " + title.get() + "\n Account Class: " + addclass_entry.cget("text"))
                    addcode_entry.delete(0, END)
                    addtitle_entry.delete(0, END)
                    addclass_entry.config(text = "")
                else:
                    addcode_entry.focus()  
        else:
            messagebox.showerror("Error!", "Please complete the required details!")
            
    def addcode_close(self, *args):
        addcode_frame.destroy()
        addcode_button.config(state = NORMAL)
        update_button.config(state = NORMAL)
        delete_account.config(state = NORMAL)
        chart_back.config(state = NORMAL)
        addcode_button.focus()
    
    def updcode_close(self, *args):
        updcode_frame.destroy()
        addcode_button.config(state = NORMAL)
        update_button.config(state = NORMAL)
        delete_account.config(state = NORMAL)
        chart_back.config(state = NORMAL)
        addcode_button.focus()

    def delcode_close(self, *args):
        delcode_frame.destroy()
        addcode_button.config(state = NORMAL)
        update_button.config(state = NORMAL)
        delete_account.config(state = NORMAL)
        chart_back.config(state = NORMAL)
        addcode_button.focus()

    def addcode(self, *args):
        addcode_button.config(state = DISABLED)
        update_button.config(state = DISABLED)
        delete_account.config(state = DISABLED)
        chart_back.config(state = DISABLED)
        global addcode_frame
        addcode_frame = LabelFrame(chart_frame, text = "Add Account", font = fonts, bg = wc, fg = fc)
        addcode_frame.grid(column = 0, row = 2, ipadx = 32, sticky = NW)

        addcode_label = Label(addcode_frame, text = "Account Code", font = fonts, anchor = W, width = 20, bg = wc, fg = fc)
        addcode_label.grid(column = 0, row = 0, pady = 1)

        addtitle_label = Label(addcode_frame, text = "Account Title", font = fonts, anchor = W, width = 20, bg = wc, fg = fc)
        addtitle_label.grid(column = 0, row = 1, pady = 1)

        class_label = Label(addcode_frame, text = "Class", font = fonts, anchor = W, width = 20, bg = wc, fg = fc)
        class_label.grid(column = 0, row = 2, pady = 1)

        global code
        code = StringVar()
        global addcode_entry
        addcode_entry = Entry(addcode_frame, textvariable = code, font = fonts, bd = 3, relief = SUNKEN, width = 4)
        addcode_entry.grid(column = 1, row = 0, sticky = W)
        addcode_entry.bind("<Key>", self.codedigit_validation)
        addcode_entry.bind("<FocusOut>", self.codedigit_validation)
        addcode_entry.bind("<FocusOut>", self.codelen_validation)
        addcode_entry.focus()

        global title
        title = StringVar()
        global addtitle_entry
        addtitle_entry = Entry(addcode_frame, textvariable = title, font = fonts, bd = 3, relief = SUNKEN, width = 40)
        addtitle_entry.grid(column = 1, row = 1, sticky = W)
        addtitle_entry.bind("<FocusOut>", lambda e: self.uppercase(addtitle_entry))

        global addclass_entry
        addclass_entry = Label(addcode_frame, text = "", font = fonts, bd = 3, relief = SUNKEN, width = 2)
        addclass_entry.grid(column = 1, row = 2, sticky = W)

        global add_account
        add_account = Button(addcode_frame, text = "Add Account", font = fonts, bg = wc, bd = 0, width = 95, image = buttonicon, compound = CENTER, cursor = "hand2", command = self.addaccount)
        add_account.grid(column = 1, row = 3, pady = 5)
        add_account.bind("<Return>", self.addaccount)

        global account_back
        account_back = Button(addcode_frame, text = "Close", font = fonts, bg = wc, bd = 0, width = 95, image = buttonicon, compound = CENTER, cursor = "hand2", command = self.addcode_close)
        account_back.grid(column = 1, row = 4, pady = 5)
        account_back.bind("<Return>", self.addcode_close)   

    def addcodeclass(self, *args):
        if len(code.get()) == 4:
            if code.get()[0] == "1":
                addclass_entry.config(text = "A")
            elif code.get()[0] == "2":
                addclass_entry.config(text = "L")
            elif code.get()[0] == "5":
                addclass_entry.config(text = "I")
            elif code.get()[0] == "6":
                addclass_entry.config(text = "E")
            elif code.get()[0] == "3":
                addclass_entry.config(text = "S")
            elif code.get()[0] == "4":
                addclass_entry.config(text = "N")
            elif code.get()[0] == "7":
                addclass_entry.config(text = "T")
            else:
                messagebox.showerror("Account Code Error", "Account Code Prefix not recognized!")
                addcode_entry.delete(0, END)
                addclass_entry.config(text = "")
        else:
            addcode_entry.delete(0, END)
            addclass_entry.config(text = "")

    def delcodelen_validation(self, *args):
        global codeinput
        codeinput = code2.get()
        if len(codeinput) == 4:
            return True
        elif len(codeinput) == 0:
            return True
        else:
            delview_entry.delete(0, END)
            
    def delcodedigit_validation(self, *args):
        codeinput = code2.get()
        if codeinput.isdigit():
            return True
        else:
            delview_entry.delete(0, END)

    def showdeleteaccount(self, *args):
        addcode_button.config(state = DISABLED)
        update_button.config(state = DISABLED)
        chart_back.config(state = DISABLED)
        delete_account.config(state = DISABLED)
        global delcode_frame
        delcode_frame = LabelFrame(chart_frame, text = "Delete Account", font = fonts, bg = wc, fg = fc)
        delcode_frame.grid(column = 0, row = 3, ipadx = 32, sticky = NW)

        delcode_label = Label(delcode_frame, text = "Account Code", font = fonts, anchor = W, width = 20, bg = wc, fg = fc)
        delcode_label.grid(column = 0, row = 1, pady = 1)

        deltitle_label = Label(delcode_frame, text = "Account Title", font = fonts, anchor = W, width = 20, bg = wc, fg = fc)
        deltitle_label.grid(column = 0, row = 2, pady = 1)

        class_label = Label(delcode_frame, text = "Class", font = fonts, anchor = W, width = 20, bg = wc, fg = fc)
        class_label.grid(column = 0, row = 3, pady = 1)

        global delcode_entry
        delcode_entry = Entry(delcode_frame, font = fonts, bd = 3, relief = SUNKEN, width = 4)
        delcode_entry.grid(column = 1, row = 1, sticky = W)
        delcode_entry.config(state = DISABLED)
        
        global deltitle_entry
        deltitle_entry = Entry(delcode_frame, font = fonts, bd = 3, relief = SUNKEN, width = 40)
        deltitle_entry.grid(column = 1, row = 2, sticky = W)
        deltitle_entry.config(state = DISABLED)

        global delclass_entry
        delclass_entry = Entry(delcode_frame, font = fonts, bd = 3, relief = SUNKEN, width = 2)
        delclass_entry.grid(column = 1, row = 3, sticky = W)
        delclass_entry.config(state = DISABLED)

        global code2
        code2 = StringVar()
        global delview_entry
        delview_entry = Entry(delcode_frame, textvariable = code2, font = fonts, bd = 3, relief = SUNKEN, width = 4)
        delview_entry.grid(column = 1, row = 0, sticky = W)
        delview_entry.bind("<Key>", self.delcodedigit_validation)
        delview_entry.bind("<FocusOut>", self.delcodedigit_validation)
        delview_entry.bind("<FocusOut>", self.delcodelen_validation)
        delview_entry.bind("<Return>", self.delview)
        delview_entry.focus()

        delview_button = Button(delcode_frame, text = "View", font = fonts, bg = wc, bd = 0, width = 95, image = buttonicon, compound = CENTER, cursor = "hand2", command = self.delview)
        delview_button.grid(column = 1, row = 0, padx = 5)
        delview_button.bind("<Return>", self.delview)

        delaccount_button = Button(delcode_frame, text = "Delete", font = fonts, bg = wc, bd = 0, width = 95, image = buttonicon, compound = CENTER, cursor = "hand2", command = self.deleteaccount)
        delaccount_button.grid(column = 1, row = 4, pady = 5)
        delaccount_button.bind("<Return>", self.deleteaccount)

        delaccount_button = Button(delcode_frame, text = "Close", font = fonts, bg = wc, bd = 0, width = 95, image = buttonicon, compound = CENTER, cursor = "hand2", command = self.delcode_close)
        delaccount_button.grid(column = 1, row = 5, pady = 5)
        delaccount_button.bind("<Return>", self.delcode_close)

    def delview(self, *args):
        if len(code2.get()) == 4:
            codeinput = code2.get()
            if company == "DBPSC":
                find_code = "SELECT * FROM chart WHERE code = ?"
            else:
                find_code = "SELECT * FROM dssichart WHERE code = ?"
            c.execute(find_code, [codeinput])
            result = c.fetchall()
            if result:
                delcode_entry.config(state = NORMAL)
                delcode_entry.insert(0, result[0][0])
                delcode_entry.config(state = DISABLED)
                deltitle_entry.config(state = NORMAL)
                deltitle_entry.insert(0, result[0][1])
                deltitle_entry.config(state = DISABLED)
                delclass_entry.config(state = NORMAL)
                delclass_entry.insert(0, result[0][2])
                delclass_entry.config(state = DISABLED)   
            else:
                messagebox.showerror("Chart Protocol", "Account Code does not exist!")
                try:
                    delcode_entry.config(state = NORMAL)
                    deltitle_entry.config(state = NORMAL)
                    delclass_entry.config(state = NORMAL)
                    delcode_entry.delete(0, END)
                    deltitle_entry.delete(0, END)
                    delclass_entry.delete(0, END)
                    delcode_entry.config(state = DISABLED)
                    deltitle_entry.config(state = DISABLED)
                    delclass_entry.config(state = DISABLED)
                except:
                    pass
                finally:
                    delview_entry.focus()
                    
    def deleteaccount(self, *args):
        delcodesure = messagebox.askyesno("Delete Account Protocol", "Are you sure?")
        if delcodesure == True:
            if company == "DBPSC":
                delfrchart = "DELETE FROM chart WHERE code = ?"
            else:
                delfrchart = "DELETE FROM dssichart WHERE code = ?" 
            c.execute(delfrchart, [code2.get()])
            conn.commit()
            messagebox.showinfo("Success!", "An account has been deleted!" + "\n Account Code: " + delcode_entry.get() + "\n Account Title: " + deltitle_entry.get() + "\n Account Class: " + delclass_entry.get())
            delcode_entry.config(state = NORMAL)
            deltitle_entry.config(state = NORMAL)
            delclass_entry.config(state = NORMAL)
            delcode_entry.delete(0, END)
            deltitle_entry.delete(0, END)
            delclass_entry.delete(0, END)
            delcode_entry.config(state = DISABLED)
            deltitle_entry.config(state = DISABLED)
            delclass_entry.config(state = DISABLED)
        else:
            addcode_entry.focus()

### view/update record factory ###
            
    def view_journal(self, *args):
        self.menubuttons(DISABLED)
        self.colorswitch(viewrecord_button, fc)
        global view_details
        view_details = LabelFrame(self.master, text = "\n View/Update Record", font = fonts, bg = wc, fg = fc)
        view_details.grid(column = 1, row = 0, sticky = NW)

        search_frame = Frame(view_details, bg = wc)
        search_frame.grid(column = 0, row = 0, sticky = NW)
        
        viewnumber_label = Label(search_frame, text = "Year/Voucher No.", font = fonts, anchor = W, width = 15, bg = wc, fg = fc)
        viewnumber_label.grid(column = 0, row = 0, padx = pad)

        global year_entry
        year_entry = tk.Combobox(search_frame, font = fonts, width = 5)
        year_entry.grid(column = 1, row = 0, sticky = W, padx = 8)
        self.masteryearlister(year_entry)
        year_entry.insert(0, '2020')

        global viewnumber_entry, viewnumber
        viewnumber = StringVar()
        viewnumber_entry = Entry(search_frame, textvariable = viewnumber, font = fonts, bd = 3, relief = SUNKEN, width = 8, justify = LEFT)
        viewnumber_entry.grid(column = 2, row = 0, sticky = W, padx = 10)
        viewnumber_entry.bind("<Return>", self.viewer)
    
        search_button = Button(search_frame, bg = wc, fg = fc, bd = 0, image = viewicon, cursor = "hand2", command = self.viewer)
        search_button.grid(column = 3, row = 0, sticky = W, padx = 10)
        search_button.bind("<Return>", self.viewer)

        prev_button = Button(search_frame, bg = wc, fg = fc, bd = 0, image = previousicon, cursor = "hand2", command = self.prevjv)
        prev_button.grid(column = 4, row = 0, padx = 10)
        prev_button.bind("<Return>", self.prevjv)

        next_button = Button(search_frame, bg = wc, fg = fc, bd = 0, image = nexticon, cursor = "hand2", command = self.nextjv)
        next_button.grid(column = 5, row = 0, padx = 10)
        next_button.bind("<Return>", self.nextjv)

        global post_button
        post_button = Button(search_frame, text = "Update", font = fonts, bg = wc, bd = 0, width = 95, image = buttonicon, compound = CENTER, cursor = "hand2", command = self.updatedebitcreditchecker)
        post_button.grid(column = 7, row = 0, padx = 6)
        post_button.bind("<Return>", self.updatedebitcreditchecker)
        post_button.config(state = DISABLED)

        print_button = Button(search_frame, text = "Print", font = fonts, bg = wc, bd = 0, width = 95, image = buttonicon, compound = CENTER, cursor = "hand2", command = self.printer2)
        print_button.grid(column = 8, row = 0, padx = 6)
        print_button.bind("<Return>", self.printer2)
        
        view_back = Button(search_frame, text = "Close", font = fonts, bg = wc, bd = 0, width = 95, image = buttonicon, compound = CENTER, cursor = "hand2", command = self.viewback)
        view_back.grid(column = 9, row = 0, padx = 6)
        view_back.bind("<Return>", self.viewback)

        viewnumber_entry.focus()

    def nextjv(self, *args):
        try:
            if len(viewnumber.get()) == 8:
                current = str(int(viewnumber.get().split("-")[1].lstrip("0")) + 1).zfill(4)
                nextvalue = viewnumber.get().split("-")[0] + "-" + current
            elif len(viewnumber.get()) == 7:
                current = str(int(viewnumber.get().split("-")[1].lstrip("0")) + 1).zfill(3)
                nextvalue = viewnumber.get().split("-")[0] + "-" + current
        except:
            messagebox.showerror("Search Protocol", "Voucher Number does not exist!")
            viewnumber_entry.delete(0, END)
        finally:
            viewnumber_entry.delete(0, END)
            viewnumber_entry.insert(0, nextvalue)
            self.boxupdate()
            self.viewer()

    def prevjv(self, *args):
        try:
            if len(viewnumber.get()) == 8:
                current = str(int(viewnumber.get().split("-")[1].lstrip("0")) - 1).zfill(4)
                nextvalue = viewnumber.get().split("-")[0] + "-" + current
            elif len(viewnumber.get()) == 7:
                current = str(int(viewnumber.get().split("-")[1].lstrip("0")) - 1).zfill(3)
                nextvalue = viewnumber.get().split("-")[0] + "-" + current
        except:
            messagebox.showerror("Search Protocol", "Voucher Number does not exist!")
            viewnumber_entry.delete(0, END)
        finally:
            viewnumber_entry.delete(0, END)
            viewnumber_entry.insert(0, nextvalue)
            self.boxupdate()
            self.viewer()

    def viewback(self, *args):
        view_details.destroy()
        self.menubuttons(NORMAL)
        self.colorswitch(viewrecord_button, fc)

    def viewer(self, *args):
        global v, v2
        if company == "DBPSC":
            finder = "SELECT * FROM DBPSC WHERE number = ? AND year = ?"
        else:
            finder = "SELECT * FROM DSSI WHERE number = ? AND year = ?"
        c.execute(finder, [viewnumber.get(), year_entry.get()])
        v = c.fetchone()
        c.execute(finder, [viewnumber.get(), year_entry.get()])
        v2 = c.fetchall()
        try:
            details.destroy()
            label_frame.destroy()
            viewstatframe.destroy()
            scrollbox.destroy()
        except:
            pass
        finally:
            try:
                self.boxupdate()
                self.viewlabels()
                self.showalllines(view_details, NORMAL)
                self.showentries()
            except:
                messagebox.showerror("Search Protocol", "Voucher Number does not exist!")
                details.destroy()
        
    def boxupdate(self, *args):
        global box
        box = []
        boxtype = ""
        gettype = viewnumber.get()[0]
        if gettype == "D":
            boxtype = "supplier"
        elif gettype == "R":
            boxtype = "client"
        
        clientfinder = "SELECT * FROM clients WHERE type = ?"
        c.execute(clientfinder, [boxtype])
        clientresult = c.fetchall()
        for x in clientresult:
            box.append(x[1])

    def masteryearlister(self, combo):
        if company == "DBPSC":
            lister = "SELECT year FROM DBPSC WHERE company = ?"
        else:
            lister = "SELECT year FROM DSSI WHERE company = ?"
        c.execute(lister, [company])
        listresult = c.fetchall()
        years = []
        if listresult:
            for i in listresult:
                years.append(i[0])
            combo.config(values = list(dict.fromkeys(sorted(years))))

    def viewlabels(self):
        global details
        details = LabelFrame(view_details, text = "Details", font = fonts, bg = wc, fg = fc)
        details.grid(column = 0, row = 1, sticky = NW)
     
        voucher_label = Label(details, text = "Voucher No.", font = fonts, anchor = W, width = 20, bg = wc, fg = fc)
        voucher_label.grid(column = 0, row = 0)

        date_label = Label(details, text = "Date", font = fonts, anchor = W, width = 20, bg = wc, fg = fc)
        date_label.grid(column = 0, row = 1)

        client_label = Label(details, text = "Client/Supplier", font = fonts, anchor = W, width = 20, bg = wc, fg = fc)
        client_label.grid(column = 0, row = 2)

        particulars_label = Label(details, text = "Particulars", font = fonts, anchor = W, width = 21, bg = wc, fg = fc)
        particulars_label.grid(column = 0, row = 3)

        reference_label = Label(details, text = "Reference", font = fonts, anchor = W, width = 20, bg = wc, fg = fc)
        reference_label.grid(column = 0, row = 4)

        global vouchernumber
        vouchernumber = Text(details, font = fonts, height = 1, bd = 3, relief = SUNKEN, width = 9)
        vouchernumber.grid(column = 1, row = 0, sticky = W)
        vouchernumber.insert(1.0, v[2])
        vouchernumber.config(state = DISABLED, fg = "grey")

        global date_entry, date
        date = StringVar()
        date_entry = Entry(details, textvariable = date, font = fonts, bd = 3, relief = SUNKEN, width = 10)
        date_entry.grid(column = 1, row = 1, sticky = W)
        date_entry.insert(0, v[4])
        date_entry.config(state = DISABLED)

        global clients, client_entry
        clients = box
        client_entry = tk.Combobox(details, values = clients, font = fonts, width = 50)
        client_entry.grid(column = 1, row = 2, sticky = W)
        client_entry.bind("<FocusOut>", lambda e: self.uppercase(client_entry))
        client_entry.insert(0, v[16])

        global particulars_entry
        particulars_entry = Entry(details, width = 74, font = fonts, bd = 3, relief = SUNKEN)
        particulars_entry.grid(column = 1, row = 3, sticky = W)
        particulars_entry.bind("<FocusOut>", lambda e: self.uppercase(particulars_entry))
        particulars_entry.insert(0, v[12])

        global reference_entry
        reference_entry = Entry(details, width = 20, font = fonts, bd = 3, relief = SUNKEN)
        reference_entry.grid(column = 1, row = 4, sticky = W)
        reference_entry.insert(0, v[11])

        global label_frame
        label_frame = LabelFrame(view_details, text = "Journal Entries", font = fonts, bg = wc, fg = fc)
        label_frame.grid(column = 0, row = 2, sticky = W)
        
        code_label = Label(label_frame, text = "Code", font = fonts, relief = RIDGE, width = 5, bg = wc, fg = fc)
        code_label.grid(column = 0, row = 0)

        title_label = Label(label_frame, text = "Title", font = fonts, relief = RIDGE, width = 33, bg = wc, fg = fc)
        title_label.grid(column = 1, row = 0)
    
        debit_label = Label(label_frame, text = "Debit", font = fonts, relief = RIDGE, width = 14, bg = wc, fg = fc)
        debit_label.grid(column = 2, row = 0)

        credit_label = Label(label_frame, text = "Credit", font = fonts, relief = RIDGE, width = 14, bg = wc, fg = fc)
        credit_label.grid(column = 3, row = 0)

        comment_label = Label(label_frame, text = "Comment", font = fonts, relief = RIDGE, width = 20, bg = wc, fg = fc)
        comment_label.grid(column = 4, row = 0)

        repeat_label = Label(label_frame, text = "<<", font = fonts, relief = RIDGE, width = 4, bg = wc, fg = fc)
        repeat_label.grid(column = 5, row = 0)

        global viewstatframe, variance, debit_total, credit_total
        viewstatframe = Frame(view_details)
        viewstatframe.grid(column = 0, row = 7, sticky = W)

        variance = Label(viewstatframe, text = float(), font = fonts, anchor = E, width = 20, bg = wc, fg = fc)
        variance.grid(column = 0, row = 0, ipadx = 74)

        debit_total = Label(viewstatframe, text = float(), font = fonts, relief = RIDGE, width = 14, anchor = E, bg = totalbg, fg = totalfg)
        debit_total.grid(column = 1, row = 0, ipadx = 1)

        credit_total = Label(viewstatframe, text = float(), font = fonts, relief = RIDGE, width = 14, anchor = E, bg = totalbg, fg = totalfg)
        credit_total.grid(column = 2, row = 0)

    def allvars(self):
        global titles, repeats, codelist, drs, crs, comments
        titles = [title_label1,title_label2,title_label3,title_label4,title_label5,title_label6,title_label7,title_label8,title_label9,title_label10,title_label11,title_label12,title_label13,title_label14,title_label15,title_label16,title_label17,title_label18,title_label19,title_label20,title_label21,title_label22,title_label23,title_label24,title_label25,title_label26,title_label27,title_label28,title_label29,title_label30]
        repeats = [repeat_button1,repeat_button2,repeat_button3,repeat_button4,repeat_button5,repeat_button6,repeat_button7,repeat_button8,repeat_button9,repeat_button10,repeat_button11,repeat_button12,repeat_button13,repeat_button14,repeat_button15,repeat_button16,repeat_button17,repeat_button18,repeat_button19,repeat_button20,repeat_button21,repeat_button22,repeat_button23,repeat_button24,repeat_button25,repeat_button26,repeat_button27,repeat_button28,repeat_button29,repeat_button30]
        codelist = [code_entry1,code_entry2,code_entry3,code_entry4,code_entry5,code_entry6,code_entry7,code_entry8,code_entry9,code_entry10,code_entry11,code_entry12,code_entry13,code_entry14,code_entry15,code_entry16,code_entry17,code_entry18,code_entry19,code_entry20,code_entry21,code_entry22,code_entry23,code_entry24,code_entry25,code_entry26,code_entry27,code_entry28,code_entry29,code_entry30]
        drs = [debit_entry1,debit_entry2,debit_entry3,debit_entry4,debit_entry5,debit_entry6,debit_entry7,debit_entry8,debit_entry9,debit_entry10,debit_entry11,debit_entry12,debit_entry13,debit_entry14,debit_entry15,debit_entry16,debit_entry17,debit_entry18,debit_entry19,debit_entry20,debit_entry21,debit_entry22,debit_entry23,debit_entry24,debit_entry25,debit_entry26,debit_entry27,debit_entry28,debit_entry29,debit_entry30]
        crs = [credit_entry1,credit_entry2,credit_entry3,credit_entry4,credit_entry5,credit_entry6,credit_entry7,credit_entry8,credit_entry9,credit_entry10,credit_entry11,credit_entry12,credit_entry13,credit_entry14,credit_entry15,credit_entry16,credit_entry17,credit_entry18,credit_entry19,credit_entry20,credit_entry21,credit_entry22,credit_entry23,credit_entry24,credit_entry25,credit_entry26,credit_entry27,credit_entry28,credit_entry29,credit_entry30]
        comments = [comment_entry1,comment_entry2,comment_entry3,comment_entry4,comment_entry5,comment_entry6,comment_entry7,comment_entry8,comment_entry9,comment_entry10,comment_entry11,comment_entry12,comment_entry13,comment_entry14,comment_entry15,comment_entry16,comment_entry17,comment_entry18,comment_entry19,comment_entry20,comment_entry21,comment_entry22,comment_entry23,comment_entry24,comment_entry25,comment_entry26,comment_entry27,comment_entry28,comment_entry29,comment_entry30]
    
    def showentries(self):
        self.allvars()
        jj = ['self.journalcode_val1()','self.journalcode_val2()','self.journalcode_val3()','self.journalcode_val4()','self.journalcode_val5()','self.journalcode_val6()','self.journalcode_val7()','self.journalcode_val8()','self.journalcode_val9()','self.journalcode_val10()','self.journalcode_val11()','self.journalcode_val12()','self.journalcode_val13()','self.journalcode_val14()','self.journalcode_val15()','self.journalcode_val16()','self.journalcode_val17()','self.journalcode_val18()','self.journalcode_val19()','self.journalcode_val20()','self.journalcode_val21()','self.journalcode_val22()','self.journalcode_val23()','self.journalcode_val24()','self.journalcode_val25()','self.journalcode_val26()','self.journalcode_val27()','self.journalcode_val28()','self.journalcode_val29()','self.journalcode_val30()']
        for x in v2:
            for y in range(1,31):
                if x[15] == y:
                    codelist[y-1].insert(0, x[6])
                    drs[y-1].insert(0, format(x[8], ",.2f"))
                    crs[y-1].insert(0, format(x[9], ",.2f"))
                    comments[y-1].insert(0, x[10])
                    eval(jj[y-1])
        self.updatetotal()
        
    def linedestroyer(self):
        dd = ['line_frame1.destroy()','line_frame2.destroy()','line_frame3.destroy()','line_frame4.destroy()','line_frame5.destroy()','line_frame6.destroy()','line_frame7.destroy()','line_frame8.destroy()','line_frame9.destroy()','line_frame10.destroy()','line_frame11.destroy()','line_frame12.destroy()','line_frame13.destroy()','line_frame14.destroy()','line_frame15.destroy()','line_frame16.destroy()','line_frame17.destroy()','line_frame18.destroy()','line_frame19.destroy()','line_frame20.destroy()','line_frame21.destroy()','line_frame22.destroy()','line_frame23.destroy()','line_frame24.destroy()','line_frame25.destroy()','line_frame26.destroy()','line_frame27.destroy()','line_frame28.destroy()','line_frame29.destroy()','line_frame30.destroy()']
        for x in range(30):
            if codelist[x].get() == "":
                eval(dd[x])

    def linedisabler(self):
        client_entry.config(state = DISABLED)
        particulars_entry.config(state = DISABLED)
        reference_entry.config(state = DISABLED)
        for x in range(30):
            codelist[x].config(state = DISABLED)
            titles[x].config(fg = "grey")
            drs[x].config(state = DISABLED)
            crs[x].config(state = DISABLED)
            comments[x].config(state = DISABLED)
            repeats[x].config(state = DISABLED)

    def codeclass(self, varcode):
        global classcode
        classcode = ""
        if len(varcode.get()) == 4:
            if varcode.get()[0] == "1":
                classcode = "A"
            elif varcode.get()[0] == "2":
                classcode = "L"
            elif varcode.get()[0] == "5":
                classcode = "I"
            elif varcode.get()[0] == "6":
                classcode = "E"
            elif varcode.get()[0] == "3":
                classcode = "S"
            elif varcode.get()[0] == "4":
                classcode = "N"
            elif varcode.get()[0] == "7":
                classcode = "T"
            else:
                classcode = ""
        else:
            classcode = ""

    def updatedebitcreditchecker(self, *args):
        self.allvars()
        errorlines = []
        for x in range(30):
            if len(codelist[x].get()) == 4:
                try:
                    if float(drs[x].get().replace(",","")) > 0 and float(crs[x].get().replace(",","")) > 0:
                        errorlines.append("Please check line " + str(x + 1) + " with code " + codelist[x].get())
                    if drs[x].get() == "0.00" and crs[x].get() == "0.00":
                        errorlines.append("Please check line " + str(x + 1) + " with code " + codelist[x].get())
                    if drs[x].get() == "" and crs[x].get() == "":
                        errorlines.append("Please check line " + str(x + 1) + " with code " + codelist[x].get())
                except:
                    errorlines.append("Please check line " + str(x + 1) + " with code " + codelist[x].get())
        if len(errorlines) == 0:
            self.lineupdatermaster()
        else:
            messagebox.showerror("Posting Protocol", errorlines)

    def lineupdatermaster(self, *args):
        self.lockchecker()
        if datelockresult == "open":
            global validlines
            self.updatetotal()
            if round(sum(drvalues),2) == round(sum(crvalues),2) and round(sum(drvalues),2) + round(sum(crvalues),2) > 0:
                update = messagebox.askyesno("Update Record Protocol", "Are you sure?")
                if update == True:
                    v = vouchernumber.get("1.0", 'end-1c')
                    d = date_entry.get()
                    cl = client_entry.get()
                    p = particulars_entry.get()
                    r = reference_entry.get()
                    m = today.strftime('%m-%d-%Y') #modified
                    u = self.username.get()
                    co = company
                    if company == "DBPSC":
                        selector = "SELECT * FROM DBPSC WHERE number = ? AND date = ?"
                        deleter = "DELETE FROM DBPSC WHERE number = ? AND date = ?"
                    else:
                        selector = "SELECT * FROM DSSI WHERE number = ? AND date = ?"
                        deleter = "DELETE FROM DSSI WHERE number = ? AND date = ?"
                    c.execute(selector, [v, d])
                    s = c.fetchall()[0]
                    i = s[1] #index
                    t = s[3] #type
                    cd = s[19] #created
                    dm = s[20]
                    dd = s[21]
                    dy = s[22]
                    bk = s[23] #book
                    c.execute(deleter, [v, d])
                    conn.commit()
                    validlines = []
                    lin = 1
                    for y in range(30):
                        if len(codelist[y].get()) == 4:
                            self.codeclass(codelist[y])
                            validlines.append([lin, i, v, t, d, r, p, u, classcode, codelist[y].get(), titles[y].cget("text"), float(drs[y].get().replace(",", "")), float(crs[y].get().replace(",", "")), comments[y].get(), m, cl, "posted", co, cd, dm, dd, dy, bk])
                            lin += 1
                    self.updatemaster()
        else:
            messagebox.showerror("Period Locker", "Date entered is locked!")

    def updatemaster(self):
        for x in validlines:
            if company == "DBPSC":
                inserter = """INSERT INTO DBPSC (
                                   line, indexing, number, type, date,
                                   reference, particulars, user, class, code,
                                   title, debit, credit, comment, modified,
                                   client, status, company, created, month,
                                   day, year, book)
                                   VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)"""
            else:
                inserter = """INSERT INTO DSSI (
                                   line, indexing, number, type, date,
                                   reference, particulars, user, class, code,
                                   title, debit, credit, comment, modified,
                                   client, status, company, created, month,
                                   day, year, book)
                                   VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)"""
            c.execute(inserter, x)
            conn.commit()
        post_button.config(state = DISABLED)
        messagebox.showinfo("Update Status", "Update Successful! Journal record has been updated!")
        self.linedisabler()
        self.linedestroyer()        
              
### add record factory ###

    def dtypeswitch(self, *args):
        global newtype, boxtype, book
        newtype = "D"
        boxtype = "supplier"
        book = "CDJ"
        self.menubuttons(DISABLED)
        self.colorswitch(addrecord_buttond, fc)
        self.journal()
        self.labels(journal_details, 5)
        self.showalllines(journal_details, DISABLED)
        
    def rtypeswitch(self, *args):
        global newtype, boxtype, book
        newtype = "R"
        boxtype = "client"
        book = "CRJ"
        self.menubuttons(DISABLED)
        self.colorswitch(addrecord_buttonr, fc)
        self.journal()
        self.labels(journal_details, 5)
        self.showalllines(journal_details, DISABLED)

    def stypeswitch(self, *args):
        global newtype, boxtype, book
        newtype = "R"
        boxtype = "client"
        book = "SJ"
        self.menubuttons(DISABLED)
        self.colorswitch(addrecord_buttons, fc)
        self.journal()
        self.labels(journal_details, 5)
        self.showalllines(journal_details, DISABLED)

    def gtypeswitch(self, *args):
        global newtype, boxtype, book
        newtype = "D"
        boxtype = "client"
        book = "GJ"
        self.menubuttons(DISABLED)
        self.colorswitch(addrecord_buttong, fc)
        self.journal()
        self.labels(journal_details, 5)
        self.showalllines(journal_details, DISABLED)
    
    def lastdatechecker(self, dateformat):
        dateinput = dateformat.strftime('%m-%d-%Y')
        if company == "DBPSC":
            checker = "SELECT MAX(day), date FROM DBPSC WHERE month = ? AND year = ? AND type = ?"
        else:
            checker = "SELECT MAX(day), date FROM DSSI WHERE month = ? AND year = ? AND type = ?"
        c.execute(checker, [dateinput.split("-")[0], dateinput.split("-")[2], newtype])
        lastdateresult = c.fetchone()
        if lastdateresult[0]:
            if dateinput.split("-")[1] >= lastdateresult[0]:
                pass
            else:
                messagebox.showerror("Period Locker", "Date entered is not available. Last posting date will be used!")
                date.set(lastdateresult[1])

    def formatdate(self, *args):
        global date_input
        date_input = date.get()
        try:
            if len(date_input) == 10:
                try:
                    date_format = datetime.datetime.strptime(date_input, '%m-%d-%Y')
                except:
                    date_format = datetime.datetime.strptime(date_input, '%m/%d/%Y')
                finally:
                    date_entry.delete(0, END)
                    date_entry.insert(0, date_format.strftime('%m-%d-%Y'))
                    datelockcheck = "SELECT * FROM locker WHERE year = ? AND month = ? AND company = ?"
                    c.execute(datelockcheck, [date_entry.get().split("-")[2],date_entry.get().split("-")[0],company])
                    datelockresult = c.fetchone()[2]
                    if datelockresult:
                        if datelockresult == "open":
                            self.lastdatechecker(date_format)
                        else:
                            messagebox.showerror("Period Locker", "Date entered is locked!")
                            date_entry.delete(0, END)
                    else:
                        messagebox.showerror("Date Checker", "Date entered is not in range!")
                        date_entry.delete(0, END)
            else:
                messagebox.showerror("Date Checker", "Date entered is not in range!")
                date_entry.delete(0, END)
        except Exception as e:
            messagebox.showerror("Date Checker", e)
            date_entry.delete(0, END)

    def numbermaster(self):
        global newrecord, plusone, splitdate
        newrecord = ""
        splitdate = date.get().split('-')
        if company == "DBPSC":
            last = "SELECT MAX(indexing) FROM DBPSC WHERE type = ? AND month = ? AND year = ?"
        else:
            last = "SELECT MAX(indexing) FROM DSSI WHERE type = ? AND month = ? AND year = ?"
        c.execute(last, [newtype,splitdate[0],splitdate[2]])
        lastresult = c.fetchall()
        if lastresult[0][0] == None:
            plusone = 1
            newnumber = newtype + splitdate[0] + "-" + str(plusone).zfill(4)
            newrecord = newnumber
            voucher_number.config(state = NORMAL)
            voucher_number.insert(1.0, newnumber) 
            voucher_number.config(state = DISABLED)   
        else:
            plusone = lastresult[0][0] + 1
            newnumber = newtype + splitdate[0] + "-" + str(plusone).zfill(4)
            newrecord = newnumber
            voucher_number.config(state = NORMAL)
            voucher_number.insert(1.0, newrecord)
            voucher_number.config(state = DISABLED)
        
    def clientbox(self):
        global box
        box = []
        clientfinder = "SELECT name FROM clients WHERE type = ?"
        c.execute(clientfinder, [boxtype])
        clientresult = c.fetchall()
        for x in clientresult:
            box.append(x[0])
        box.sort()

    def journal(self, *args):
        global journal_details
        journal_details = LabelFrame(self.master, text = "\nAdd Record - " + book, font = fonts, bg = wc, fg = fc)
        journal_details.grid(column = 1, row = 0, sticky = NW)

        global details_frame
        details_frame = LabelFrame(journal_details, text = "Details", font = fonts, bg = wc, fg = fc)
        details_frame.grid(column = 0, row = 0, sticky = NW)

        voucher_label = Label(details_frame, text = "Voucher No.", font = fonts, anchor = W, width = 20, bg = wc, fg = fc)
        voucher_label.grid(column = 0, row = 0)
        
        date_label = Label(details_frame, text = "Date", font = fonts, anchor = W, width = 20, bg = wc, fg = fc)
        date_label.grid(column = 0, row = 1)

        client_label = Label(details_frame, text = "Client/Supplier", font = fonts, anchor = W, width = 20, bg = wc, fg = fc)
        client_label.grid(column = 0, row = 2)

        particulars_label = Label(details_frame, text = "Particulars", font = fonts, anchor = W, width = 20, bg = wc, fg = fc)
        particulars_label.grid(column = 0, row = 3)

        reference_label = Label(details_frame, text = "Reference", font = fonts, anchor = W, width = 20, bg = wc, fg = fc)
        reference_label.grid(column = 0, row = 4)

        global voucher_number
        voucher_number = Text(details_frame, font = fonts, height = 1, bd = 3, relief = SUNKEN, width = 9)
        voucher_number.grid(column = 1, row = 0, sticky = W)
        voucher_number.insert(1.0, "")
        voucher_number.config(state = DISABLED)
        
        global date
        date = StringVar()
        global date_entry
        date_entry = Entry(details_frame, textvariable = date, font = fonts, bd = 3, relief = SUNKEN, width = 10)
        date_entry.grid(column = 1, row = 1, sticky = W)
        date_entry.bind("<FocusOut>", self.formatdate)

        self.clientbox()
        global clients
        clients = box
        global client_entry
        client_entry = tk.Combobox(details_frame, values = clients, font = fonts, width = 50)
        client_entry.grid(column = 1, row = 2, sticky = W)
        client_entry.bind("<FocusOut>", lambda e: self.uppercase(client_entry))

        global particulars_entry
        particulars_entry = Entry(details_frame, font = fonts, width = 77, bd = 3, relief = SUNKEN)
        particulars_entry.grid(column = 1, row = 3, sticky = W)
        particulars_entry.bind("<FocusOut>", lambda e: self.uppercase(particulars_entry))

        global reference_entry
        reference_entry = Entry(details_frame, font = fonts, width = 20, bd = 3, relief = SUNKEN)
        reference_entry.grid(column = 1, row = 4, sticky = W)

        refresh_button = Button(details_frame, text = "Refresh", font = fonts, bg = wc, bd = 0, width = 95, image = buttonicon, compound = CENTER, cursor = "hand2", command = self.refresh)
        refresh_button.grid(column = 1, row = 5)
        refresh_button.bind("<Return>", self.refresh)

        global post_button
        post_button = Button(details_frame, text = "Post", font = fonts, bg = wc, bd = 0, width = 95, image = buttonicon, compound = CENTER, cursor = "hand2", command = self.debitcreditchecker)
        post_button.place(x = 538, y = 133)
        post_button.bind("<Return>", self.debitcreditchecker)
        post_button.config(state = DISABLED)

        global print_button
        print_button = Button(details_frame, text = "Print", font = fonts, bg = wc, bd = 0, width = 95, image = buttonicon, compound = CENTER, cursor = "hand2", command = self.printer)
        print_button.place(x = 650, y = 133)
        print_button.bind("<Return>", self.printer)
        print_button.config(state = DISABLED)
    
        details_back = Button(details_frame, text = "Close", font = fonts, bg = wc, bd = 0, width = 95, image = buttonicon, compound = CENTER, cursor = "hand2", command = self.details_back)
        details_back.place(x = 650, y = 0)
        details_back.bind("<Return>", self.details_back)

        date_entry.focus()

    def refresh(self, *args):
        journal_details.destroy()
        self.journal()
        self.labels(journal_details, 5)
        self.showalllines(journal_details, DISABLED)
        
    def details_back(self, *args):
        if len(voucher_number.get("1.0", "end-1c")) == 0 and len(date.get()) != 0:
            recordclose = messagebox.askyesno("Close Add Record", "Your progress has not been saved. Are you sure?")
            if recordclose == True:
                for widgets in self.master.winfo_children():
                    widgets.destroy()
                self.menu()
                self.general_journal()
        else:    
            recordclose = messagebox.askyesno("Close Add Record", "Are you sure?")
            if recordclose == True:
                for widgets in self.master.winfo_children():
                    widgets.destroy()
                self.menu()
                self.general_journal()

    def labels(self, master, row):
        global entries_frame
        entries_frame = LabelFrame(master, text = "Journal Entries", font = fonts, bg = wc, fg = fc)
        entries_frame.grid(column = 0, row = row, sticky = NW)
        
        code_label = Label(entries_frame, text = "Code", font = fonts, relief = RIDGE, width = 5, bg = wc, fg = fc)
        code_label.grid(column = 0, row = 0)

        title_label = Label(entries_frame, text = "Title", font = fonts, relief = RIDGE, width = 33, bg = wc, fg = fc)
        title_label.grid(column = 1, row = 0)
    
        debit_label = Label(entries_frame, text = "Debit", font = fonts, relief = RIDGE, width = 14, bg = wc, fg = fc)
        debit_label.grid(column = 2, row = 0)

        credit_label = Label(entries_frame, text = "Credit", font = fonts, relief = RIDGE, width = 14, bg = wc, fg = fc)
        credit_label.grid(column = 3, row = 0)

        comment_label = Label(entries_frame, text = "Comment", font = fonts, relief = RIDGE, width = 20, bg = wc, fg = fc)
        comment_label.grid(column = 4, row = 0)

        repeat_label = Label(entries_frame, text = "<<", font = fonts, relief = RIDGE, width = 4, bg = wc, fg = fc)
        repeat_label.grid(column = 5, row = 0)

        global status_frame, variance, debit_total, credit_total
        status_frame = Frame(master)
        status_frame.grid(column = 0, row = 30, sticky = W)

        variance = Label(status_frame, text = float(), font = fonts, anchor = E, width = 20, bg = wc, fg = fc)
        variance.grid(column = 0, row = 0, ipadx = 74)

        debit_total = Label(status_frame, text = float(), font = fonts, relief = RIDGE, width = 14, anchor = E, bg = totalbg, fg = totalfg, justify = RIGHT)
        debit_total.grid(column = 1, row = 0, ipadx = 1)

        credit_total = Label(status_frame, text = float(), font = fonts, relief = RIDGE, width = 14, anchor = E, bg = totalbg, fg = totalfg, justify = RIGHT)
        credit_total.grid(column = 2, row = 0)
        
    def showalllines(self, master, status):
        global scrollbox
        scrollbox = Frame(master)
        container = Frame(scrollbox)
        canvas = Canvas(container, bg = wc)
        scrollbar = Scrollbar(container, orient = "vertical", width = 25, command = canvas.yview)
        scrollable_frame = Frame(canvas)
        scrollable_frame.bind("<Configure>", lambda e: canvas.configure(scrollregion = canvas.bbox("all")))
        canvas.create_window((0,0), window = scrollable_frame, anchor = "nw")
        canvas.configure(yscrollcommand = scrollbar.set)

        showframes = ['self.lineframe1(scrollable_frame)','self.lineframe2(scrollable_frame)','self.lineframe3(scrollable_frame)','self.lineframe4(scrollable_frame)','self.lineframe5(scrollable_frame)','self.lineframe6(scrollable_frame)','self.lineframe7(scrollable_frame)','self.lineframe8(scrollable_frame)','self.lineframe9(scrollable_frame)','self.lineframe10(scrollable_frame)','self.lineframe11(scrollable_frame)','self.lineframe12(scrollable_frame)','self.lineframe13(scrollable_frame)','self.lineframe14(scrollable_frame)','self.lineframe15(scrollable_frame)','self.lineframe16(scrollable_frame)','self.lineframe17(scrollable_frame)','self.lineframe18(scrollable_frame)','self.lineframe19(scrollable_frame)','self.lineframe20(scrollable_frame)','self.lineframe21(scrollable_frame)','self.lineframe22(scrollable_frame)','self.lineframe23(scrollable_frame)','self.lineframe24(scrollable_frame)','self.lineframe25(scrollable_frame)','self.lineframe26(scrollable_frame)','self.lineframe27(scrollable_frame)','self.lineframe28(scrollable_frame)','self.lineframe29(scrollable_frame)','self.lineframe30(scrollable_frame)']
        showcodes = ['self.show_code1()','self.show_code2()','self.show_code3()','self.show_code4()','self.show_code5()','self.show_code6()','self.show_code7()','self.show_code8()','self.show_code9()','self.show_code10()','self.show_code11()','self.show_code12()','self.show_code13()','self.show_code14()','self.show_code15()','self.show_code16()','self.show_code17()','self.show_code18()','self.show_code19()','self.show_code20()','self.show_code21()','self.show_code22()','self.show_code23()','self.show_code24()','self.show_code25()','self.show_code26()','self.show_code27()','self.show_code28()','self.show_code29()','self.show_code30()']
        showtitles = ['self.showtitle1()','self.showtitle2()','self.showtitle3()','self.showtitle4()','self.showtitle5()','self.showtitle6()','self.showtitle7()','self.showtitle8()','self.showtitle9()','self.showtitle10()','self.showtitle11()','self.showtitle12()','self.showtitle13()','self.showtitle14()','self.showtitle15()','self.showtitle16()','self.showtitle17()','self.showtitle18()','self.showtitle19()','self.showtitle20()','self.showtitle21()','self.showtitle22()','self.showtitle23()','self.showtitle24()','self.showtitle25()','self.showtitle26()','self.showtitle27()','self.showtitle28()','self.showtitle29()','self.showtitle30()']
        showdebits = ['self.showdebit1()','self.showdebit2()','self.showdebit3()','self.showdebit4()','self.showdebit5()','self.showdebit6()','self.showdebit7()','self.showdebit8()','self.showdebit9()','self.showdebit10()','self.showdebit11()','self.showdebit12()','self.showdebit13()','self.showdebit14()','self.showdebit15()','self.showdebit16()','self.showdebit17()','self.showdebit18()','self.showdebit19()','self.showdebit20()','self.showdebit21()','self.showdebit22()','self.showdebit23()','self.showdebit24()','self.showdebit25()','self.showdebit26()','self.showdebit27()','self.showdebit28()','self.showdebit29()','self.showdebit30()']
        showcredits = ['self.showcredit1()','self.showcredit2()','self.showcredit3()','self.showcredit4()','self.showcredit5()','self.showcredit6()','self.showcredit7()','self.showcredit8()','self.showcredit9()','self.showcredit10()','self.showcredit11()','self.showcredit12()','self.showcredit13()','self.showcredit14()','self.showcredit15()','self.showcredit16()','self.showcredit17()','self.showcredit18()','self.showcredit19()','self.showcredit20()','self.showcredit21()','self.showcredit22()','self.showcredit23()','self.showcredit24()','self.showcredit25()','self.showcredit26()','self.showcredit27()','self.showcredit28()','self.showcredit29()','self.showcredit30()']
        showcomments = ['self.showcomment1()','self.showcomment2()','self.showcomment3()','self.showcomment4()','self.showcomment5()','self.showcomment6()','self.showcomment7()','self.showcomment8()','self.showcomment9()','self.showcomment10()','self.showcomment11()','self.showcomment12()','self.showcomment13()','self.showcomment14()','self.showcomment15()','self.showcomment16()','self.showcomment17()','self.showcomment18()','self.showcomment19()','self.showcomment20()','self.showcomment21()','self.showcomment22()','self.showcomment23()','self.showcomment24()','self.showcomment25()','self.showcomment26()','self.showcomment27()','self.showcomment28()','self.showcomment29()','self.showcomment30()']
        showrepeats = ['self.showrepeat1()','self.showrepeat2()','self.showrepeat3()','self.showrepeat4()','self.showrepeat5()','self.showrepeat6()','self.showrepeat7()','self.showrepeat8()','self.showrepeat9()','self.showrepeat10()','self.showrepeat11()','self.showrepeat12()','self.showrepeat13()','self.showrepeat14()','self.showrepeat15()','self.showrepeat16()','self.showrepeat17()','self.showrepeat18()','self.showrepeat19()','self.showrepeat20()','self.showrepeat21()','self.showrepeat22()','self.showrepeat23()','self.showrepeat24()','self.showrepeat25()','self.showrepeat26()','self.showrepeat27()','self.showrepeat28()','self.showrepeat29()','self.showrepeat30()']
        showdebitconfigs = ['debit_entry1.config(state = status)','debit_entry2.config(state = status)','debit_entry3.config(state = status)','debit_entry4.config(state = status)','debit_entry5.config(state = status)','debit_entry6.config(state = status)','debit_entry7.config(state = status)','debit_entry8.config(state = status)','debit_entry9.config(state = status)','debit_entry10.config(state = status)','debit_entry11.config(state = status)','debit_entry12.config(state = status)','debit_entry13.config(state = status)','debit_entry14.config(state = status)','debit_entry15.config(state = status)','debit_entry16.config(state = status)','debit_entry17.config(state = status)','debit_entry18.config(state = status)','debit_entry19.config(state = status)','debit_entry20.config(state = status)','debit_entry21.config(state = status)','debit_entry22.config(state = status)','debit_entry23.config(state = status)','debit_entry24.config(state = status)','debit_entry25.config(state = status)','debit_entry26.config(state = status)','debit_entry27.config(state = status)','debit_entry28.config(state = status)','debit_entry29.config(state = status)','debit_entry30.config(state = status)']
        showcreditconfigs = ['credit_entry1.config(state = status)','credit_entry2.config(state = status)','credit_entry3.config(state = status)','credit_entry4.config(state = status)','credit_entry5.config(state = status)','credit_entry6.config(state = status)','credit_entry7.config(state = status)','credit_entry8.config(state = status)','credit_entry9.config(state = status)','credit_entry10.config(state = status)','credit_entry11.config(state = status)','credit_entry12.config(state = status)','credit_entry13.config(state = status)','credit_entry14.config(state = status)','credit_entry15.config(state = status)','credit_entry16.config(state = status)','credit_entry17.config(state = status)','credit_entry18.config(state = status)','credit_entry19.config(state = status)','credit_entry20.config(state = status)','credit_entry21.config(state = status)','credit_entry22.config(state = status)','credit_entry23.config(state = status)','credit_entry24.config(state = status)','credit_entry25.config(state = status)','credit_entry26.config(state = status)','credit_entry27.config(state = status)','credit_entry28.config(state = status)','credit_entry29.config(state = status)','credit_entry30.config(state = status)']
        showcommentconfigs = ['comment_entry1.config(state = status)','comment_entry2.config(state = status)','comment_entry3.config(state = status)','comment_entry4.config(state = status)','comment_entry5.config(state = status)','comment_entry6.config(state = status)','comment_entry7.config(state = status)','comment_entry8.config(state = status)','comment_entry9.config(state = status)','comment_entry10.config(state = status)','comment_entry11.config(state = status)','comment_entry12.config(state = status)','comment_entry13.config(state = status)','comment_entry14.config(state = status)','comment_entry15.config(state = status)','comment_entry16.config(state = status)','comment_entry17.config(state = status)','comment_entry18.config(state = status)','comment_entry19.config(state = status)','comment_entry20.config(state = status)','comment_entry21.config(state = status)','comment_entry22.config(state = status)','comment_entry23.config(state = status)','comment_entry24.config(state = status)','comment_entry25.config(state = status)','comment_entry26.config(state = status)','comment_entry27.config(state = status)','comment_entry28.config(state = status)','comment_entry29.config(state = status)','comment_entry30.config(state = status)']
        showrepeatconfigs = ['repeat_button1.config(state = status)','repeat_button2.config(state = status)','repeat_button3.config(state = status)','repeat_button4.config(state = status)','repeat_button5.config(state = status)','repeat_button6.config(state = status)','repeat_button7.config(state = status)','repeat_button8.config(state = status)','repeat_button9.config(state = status)','repeat_button10.config(state = status)','repeat_button11.config(state = status)','repeat_button12.config(state = status)','repeat_button13.config(state = status)','repeat_button14.config(state = status)','repeat_button15.config(state = status)','repeat_button16.config(state = status)','repeat_button17.config(state = status)','repeat_button18.config(state = status)','repeat_button19.config(state = status)','repeat_button20.config(state = status)','repeat_button21.config(state = status)','repeat_button22.config(state = status)','repeat_button23.config(state = status)','repeat_button24.config(state = status)','repeat_button25.config(state = status)','repeat_button26.config(state = status)','repeat_button27.config(state = status)','repeat_button28.config(state = status)','repeat_button29.config(state = status)','repeat_button30.config(state = status)']
        showlists = [showframes,showcodes,showtitles,showdebits,showcredits,showcomments,showrepeats,showdebitconfigs,showcreditconfigs,showcommentconfigs,showrepeatconfigs]

        for x in showlists:
            for y in range(30):
                eval(x[y])

        scrollbox.grid(column = 0, row = 6, sticky = NW)
        container.pack()
        canvas.pack(side = "left", fill = "both", expand = True)
        canvas.config(height = 350, width = 760)
        scrollbar.pack(side = "right", fill = "y")

    def updatetotal(self):
        global drvalues, crvalues, dr, cr
        drvalues, crvalues = [], []
        dr = [dr1.get(),dr2.get(),dr3.get(),dr4.get(),dr5.get(),dr6.get(),dr7.get(),dr8.get(),dr9.get(),dr10.get(),dr11.get(),dr12.get(),dr13.get(),dr14.get(),dr15.get(),dr16.get(),dr17.get(),dr18.get(),dr19.get(),dr20.get(),dr21.get(),dr22.get(),dr23.get(),dr24.get(),dr25.get(),dr26.get(),dr27.get(),dr28.get(),dr29.get(),dr30.get()]
        cr = [cr1.get(),cr2.get(),cr3.get(),cr4.get(),cr5.get(),cr6.get(),cr7.get(),cr8.get(),cr9.get(),cr10.get(),cr11.get(),cr12.get(),cr13.get(),cr14.get(),cr15.get(),cr16.get(),cr17.get(),cr18.get(),cr19.get(),cr20.get(),cr21.get(),cr22.get(),cr23.get(),cr24.get(),cr25.get(),cr26.get(),cr27.get(),cr28.get(),cr29.get(),cr30.get()]
        for d in dr:
            if d != "":
                drvalues.append(float(d.replace(",", "")))
        for c in cr:
            if c != "":
                crvalues.append(float(c.replace(",", "")))

        global td, tc
        td = round(sum(drvalues), 2)
        tc = round(sum(crvalues), 2)
        variance.config(text = format((td-tc), ',.2f'))
        debit_total.config(text = format(td, ',.2f'))
        credit_total.config(text = format(tc, ',.2f'))
        self.enablepost()
    
    def enablepost(self):
        if td == tc and (td + tc) != 0:
            post_button.config(state = NORMAL)
            debit_total.config(bg = click, fg = totalfg)
            credit_total.config(bg = click, fg = totalfg)
        else:
            post_button.config(state = DISABLED)
            debit_total.config(bg = totalbg, fg = totalfg)
            credit_total.config(bg = totalbg, fg = totalfg)
           
### line 1 ###

    def lineframe1(self, master):
        global line_frame1
        line_frame1 = Frame(master, bg = wc)
        line_frame1.grid(column = 0, row = 6, sticky = NW)

    def show_code1(self):
        global journalcode1
        journalcode1 = StringVar()
        global code_entry1
        code_entry1 = Entry(line_frame1, textvariable = journalcode1, font = fonts, relief = SUNKEN, bd = 2, width = 5, justify = "center")
        code_entry1.grid(column = 0, row = 0)
        code_entry1.bind("<Tab>", self.journalcode_val1)
        code_entry1.bind("<FocusOut>", self.journalcode_val1)
        code_entry1.bind("<Return>", self.journalcode_val1)
        
    def journalcode_val1(self, *args):
        global result1
        codeinput1 = journalcode1.get()
        if len(codeinput1) == 4 and codeinput1.isdigit():
            if company == "DBPSC":
                find_code1 = "SELECT title FROM chart WHERE code = ?"
            else:
                find_code1 = "SELECT title FROM dssichart WHERE code = ?"
            c.execute(find_code1, [codeinput1])
            result1 = c.fetchall()
            if result1:
                title_label1.config(text = result1[0][0])
                self.linestatenormal1()
                debit_entry1.focus()
                self.updatetotal()
            else:
                self.linestatedisabled1()
                code_entry1.delete(0, END)
                self.updatetotal()
        else:
            self.linestatedisabled1()
            code_entry1.delete(0, END)
            self.updatetotal()

    def linestatenormal1(self):
        debit_entry1.config(state = NORMAL)
        credit_entry1.config(state = NORMAL)
        comment_entry1.config(state = NORMAL)
        repeat_button1.config(state = NORMAL)

    def linestatedisabled1(self):
        title_label1.config(text = "")
        debit_entry1.delete(0, END)
        credit_entry1.delete(0, END)
        comment_entry1.delete(0, END)

        debit_entry1.config(state = DISABLED)
        credit_entry1.config(state = DISABLED)
        comment_entry1.config(state = DISABLED)
        repeat_button1.config(state = DISABLED)
    
    def showtitle1(self, *args):
        global title_label1
        title_label1 = Label(line_frame1, text = "", relief = SUNKEN, font = fonts, bd = 2, width = 33, anchor = W, bg = codetitlebg, fg = codetitlefg)
        title_label1.grid(column = 1, row = 0)

    def validate_float1(self, var1):
        global new_value1
        new_value1 = var1.get()
        try:
            new_value1 == "" or "," or float(new_value1)
            old_value1 = new_value1
        except:
            old_value1 = ""
            var1.set(old_value1)

    def validate_comma_dr1(self, *args):
        try:
            if debit_entry1.get() == "":
                dr1.set(format(float(0), ',.2f'))
            else:
                comma1 = format(float(dr1.get()), ',.2f')
                dr1.set(comma1)      
        except:
            debit_entry1.delete(0, END)
            dr1.set(format(float(0), ',.2f'))
        finally:
            self.updatetotal()

    def validate_comma_cr1(self, *args):
        try:
            if credit_entry1.get() == "":
                cr1.set(format(float(0), ',.2f'))
            else:
                comma1 = format(float(cr1.get()), ',.2f')
                cr1.set(comma1)
        except:
            credit_entry1.delete(0, END)
            cr1.set(format(float(0), ',.2f'))
        finally:
            self.updatetotal()
               
    def showdebit1(self, *args):
        global debit_entry1
        global dr1
        dr1 = StringVar()
        debit_entry1 = Entry(line_frame1, textvariable = dr1, font = fonts, relief = SUNKEN, bd = 2, width = 14, justify = "right")
        debit_entry1.grid(column = 2, row = 0)
        dr1.trace("w", lambda nm, idx, mode, var1 = dr1: self.validate_float1(var1))
        debit_entry1.bind("<Tab>", self.validate_comma_dr1)
        
    def showcredit1(self):
        global cr1
        cr1 = StringVar()
        global credit_entry1
        credit_entry1 = Entry(line_frame1, textvariable = cr1, font = fonts, relief = SUNKEN, bd = 2, width = 14, justify = "right")
        credit_entry1.grid(column = 3, row = 0)
        cr1.trace("w", lambda nm, idx, mode, var1 = cr1: self.validate_float1(var1)) 
        credit_entry1.bind("<Tab>", self.validate_comma_cr1)
        
    def showcomment1(self):  
        global comment1
        comment1 = StringVar()
        global comment_entry1
        comment_entry1 = Entry(line_frame1, textvariable = comment1, font = fonts, relief = SUNKEN, bd = 2, width = 20, justify = "left")
        comment_entry1.grid(column = 4, row = 0)

    def showrepeat1(self, *args):
        global repeat_button1
        repeat_button1 = Button(line_frame1, text = "clear", font = clearbuttonfont, bg = buttonbg, command = self.cleanentry1, bd =0)
        repeat_button1.grid(column = 5, row = 0, padx = 2)
        repeat_button1.bind("<Return>", self.cleanentry1)
 
    def cleanentry1(self, *args):
        code_entry1.delete(0, END)
        title_label1.config(text = "")
        debit_entry1.delete(0, END)
        credit_entry1.delete(0, END)
        comment_entry1.delete(0, END)

        debit_entry1.config(state = DISABLED)
        credit_entry1.config(state = DISABLED)
        comment_entry1.config(state = DISABLED)
        repeat_button1.config(state = DISABLED)

        self.updatetotal()

### line 2 ###
        
    def lineframe2(self, master):
        global line_frame2
        line_frame2 = Frame(master, bg = wc)
        line_frame2.grid(column = 0, row = 7, sticky = NW)

    def show_code2(self):
        global journalcode2
        journalcode2 = StringVar()
        global code_entry2
        code_entry2 = Entry(line_frame2, textvariable = journalcode2, font = fonts, relief = SUNKEN, bd = 2, width = 5, justify = "center")
        code_entry2.grid(column = 0, row = 0)
        code_entry2.bind("<Tab>", self.journalcode_val2)
        code_entry2.bind("<FocusOut>", self.journalcode_val2)
        code_entry2.bind("<Return>", self.journalcode_val2)
        
    def journalcode_val2(self, *args):
        global result2
        codeinput2 = journalcode2.get()
        if len(codeinput2) == 4 and codeinput2.isdigit():
            if company == "DBPSC":
                find_code2 = "SELECT title FROM chart WHERE code = ?"
            else:
                find_code2 = "SELECT title FROM dssichart WHERE code = ?"    
            c.execute(find_code2, [codeinput2])
            result2 = c.fetchall()
            if result2:
                title_label2.config(text = result2[0][0])
                self.linestatenormal2()
                debit_entry2.focus()
                self.updatetotal()
            else:
                self.linestatedisabled2()
                code_entry2.delete(0, END)
                self.updatetotal()
        else:
            self.linestatedisabled2()
            code_entry2.delete(0, END)
            self.updatetotal()
            
    def linestatenormal2(self):
        debit_entry2.config(state = NORMAL)
        credit_entry2.config(state = NORMAL)
        comment_entry2.config(state = NORMAL)
        repeat_button2.config(state = NORMAL)

    def linestatedisabled2(self):
        title_label2.config(text = "")
        debit_entry2.delete(0, END)
        credit_entry2.delete(0, END)
        comment_entry2.delete(0, END)

        debit_entry2.config(state = DISABLED)
        credit_entry2.config(state = DISABLED)
        comment_entry2.config(state = DISABLED)
        repeat_button2.config(state = DISABLED)
                
    def showtitle2(self, *args):
        global title_label2
        title_label2 = Label(line_frame2, text = "", relief = SUNKEN, font = fonts, bd = 2, width = 33, anchor = W, bg = codetitlebg, fg = codetitlefg)
        title_label2.grid(column = 1, row = 0)

    def validate_float2(self, var2):
        global new_value2
        new_value2 = var2.get()
        try:
            new_value2 == "" or "," or float(new_value2)
            old_value2 = new_value2
        except:
            old_value2 = ""
            var2.set(old_value2)

    def validate_comma_dr2(self, *args):
        try:
            if debit_entry2.get() == "":
                dr2.set(format(float(0), ',.2f'))
            else:
                comma2 = format(float(dr2.get()), ',.2f')
                dr2.set(comma2)
        except:
            debit_entry2.delete(0, END)
            dr2.set(format(float(0), ',.2f'))
        finally:
            self.updatetotal()

    def validate_comma_cr2(self, *args):
        try:
            if credit_entry2.get() == "":
                cr2.set(format(float(0), ',.2f'))
            else:
                comma2 = format(float(cr2.get()), ',.2f')
                cr2.set(comma2)
        except:
            credit_entry2.delete(0, END)
            cr2.set(format(float(0), ',.2f'))
        finally:
            self.updatetotal()
               
    def showdebit2(self, *args):
        global debit_entry2
        global dr2
        dr2 = StringVar()
        debit_entry2 = Entry(line_frame2, textvariable = dr2, relief = SUNKEN, font = fonts, bd = 2, width = 14, justify = "right")
        debit_entry2.grid(column = 2, row = 0)
        dr2.trace("w", lambda nm, idx, mode, var2 = dr2: self.validate_float2(var2))
        debit_entry2.bind("<Tab>", self.validate_comma_dr2)
                
    def showcredit2(self):
        global cr2
        cr2 = StringVar()
        global credit_entry2
        credit_entry2 = Entry(line_frame2, textvariable = cr2, relief = SUNKEN, font = fonts, bd = 2, width = 14, justify = "right")
        credit_entry2.grid(column = 3, row = 0)
        cr2.trace("w", lambda nm, idx, mode, var2 = cr2: self.validate_float2(var2)) 
        credit_entry2.bind("<Tab>", self.validate_comma_cr2)
        
    def showcomment2(self):  
        global comment2
        comment2 = StringVar()
        global comment_entry2
        comment_entry2 = Entry(line_frame2, textvariable = comment2, relief = SUNKEN, font = fonts, bd = 2, width = 20, justify = "left")
        comment_entry2.grid(column = 4, row = 0)

    def showrepeat2(self, *args):
        global repeat_button2
        repeat_button2 = Button(line_frame2, text = "clear", bg = buttonbg, font = clearbuttonfont, command = self.cleanentry2, bd =0)
        repeat_button2.grid(column = 5, row = 0, padx = 2)
        repeat_button2.bind("<Return>", self.cleanentry2)
 
    def cleanentry2(self, *args):
        code_entry2.delete(0, END)
        title_label2.config(text = "")
        debit_entry2.delete(0, END)
        credit_entry2.delete(0, END)
        comment_entry2.delete(0, END)

        debit_entry2.config(state = DISABLED)
        credit_entry2.config(state = DISABLED)
        comment_entry2.config(state = DISABLED)
        repeat_button2.config(state = DISABLED)

        self.updatetotal()

### line 3 ###
    def lineframe3(self, master):
        global line_frame3
        line_frame3 = Frame(master, bg = wc)
        line_frame3.grid(column = 0, row = 8, sticky = NW)

    def show_code3(self):
        global journalcode3
        journalcode3 = StringVar()
        global code_entry3
        code_entry3 = Entry(line_frame3, textvariable = journalcode3, font = fonts, relief = SUNKEN, bd = 2, width = 5, justify = "center")
        code_entry3.grid(column = 0, row = 0)
        code_entry3.bind("<Tab>", self.journalcode_val3)
        code_entry3.bind("<FocusOut>", self.journalcode_val3)
        code_entry3.bind("<Return>", self.journalcode_val3)
        
    def journalcode_val3(self, *args):
        global result3
        codeinput3 = journalcode3.get()
        if len(codeinput3) == 4 and codeinput3.isdigit():
            if company == "DBPSC":
                find_code3 = "SELECT title FROM chart WHERE code = ?"
            else:
                find_code3 = "SELECT title FROM dssichart WHERE code = ?"
            c.execute(find_code3, [codeinput3])
            result3 = c.fetchall()
            if result3:
                title_label3.config(text = result3[0][0])
                self.linestatenormal3()
                debit_entry3.focus()
                self.updatetotal()
            else:
                self.linestatedisabled3()
                code_entry3.delete(0, END)
                self.updatetotal()
        else:
            self.linestatedisabled3()
            code_entry3.delete(0, END)
            self.updatetotal()

    def linestatenormal3(self):
        debit_entry3.config(state = NORMAL)
        credit_entry3.config(state = NORMAL)
        comment_entry3.config(state = NORMAL)
        repeat_button3.config(state = NORMAL)

    def linestatedisabled3(self):
        title_label3.config(text = "")
        debit_entry3.delete(0, END)
        credit_entry3.delete(0, END)
        comment_entry3.delete(0, END)

        debit_entry3.config(state = DISABLED)
        credit_entry3.config(state = DISABLED)
        comment_entry3.config(state = DISABLED)
        repeat_button3.config(state = DISABLED)
               
    def showtitle3(self, *args):
        global title_label3
        title_label3 = Label(line_frame3, text = "", font = fonts, relief = SUNKEN, bd = 2, width = 33, anchor = W, bg = codetitlebg, fg = codetitlefg)
        title_label3.grid(column = 1, row = 0)

    def validate_float3(self, var3):
        global new_value3
        new_value3 = var3.get()
        try:
            new_value3 == "" or "," or float(new_value3)
            old_value3 = new_value3
        except:
            old_value3 = ""
            var3.set(old_value3)

    def validate_comma_dr3(self, *args):
        try:
            if debit_entry3.get() == "":
                dr3.set(format(float(0), ',.2f'))
            else:
                comma3 = format(float(dr3.get()), ',.2f')
                dr3.set(comma3)
        except:
            debit_entry3.delete(0, END)
            dr3.set(format(float(0), ',.2f'))
        finally:
            self.updatetotal()

    def validate_comma_cr3(self, *args):
        try:
            if credit_entry3.get() == "":
                cr3.set(format(float(0), ',.2f'))
            else:
                comma3 = format(float(cr3.get()), ',.2f')
                cr3.set(comma3)
        except:
            credit_entry3.delete(0, END)
            cr3.set(format(float(0), ',.2f'))
        finally:
            self.updatetotal()
     
    def showdebit3(self, *args):
        global debit_entry3
        global dr3
        dr3 = StringVar()
        debit_entry3 = Entry(line_frame3, textvariable = dr3, font = fonts, relief = SUNKEN, bd = 2, width = 14, justify = "right")
        debit_entry3.grid(column = 2, row = 0)
        dr3.trace("w", lambda nm, idx, mode, var3 = dr3: self.validate_float3(var3))
        debit_entry3.bind("<Tab>", self.validate_comma_dr3)
                
    def showcredit3(self):
        global cr3
        cr3 = StringVar()
        global credit_entry3
        credit_entry3 = Entry(line_frame3, textvariable = cr3, font = fonts, relief = SUNKEN, bd = 2, width = 14, justify = "right")
        credit_entry3.grid(column = 3, row = 0)
        cr3.trace("w", lambda nm, idx, mode, var3 = cr3: self.validate_float3(var3)) 
        credit_entry3.bind("<Tab>", self.validate_comma_cr3)
        
    def showcomment3(self):  
        global comment3
        comment3 = StringVar()
        global comment_entry3
        comment_entry3 = Entry(line_frame3, textvariable = comment3, font = fonts, relief = SUNKEN, bd = 2, width = 20, justify = "left")
        comment_entry3.grid(column = 4, row = 0)

    def showrepeat3(self, *args):
        global repeat_button3
        repeat_button3 = Button(line_frame3, text = "clear", bg = buttonbg, font = clearbuttonfont, command = self.cleanentry3, bd =0)
        repeat_button3.grid(column = 5, row = 0, padx = 2)
        repeat_button3.bind("<Return>", self.cleanentry3)
 
    def cleanentry3(self, *args):
        code_entry3.delete(0, END)
        title_label3.config(text = "")
        debit_entry3.delete(0, END)
        credit_entry3.delete(0, END)
        comment_entry3.delete(0, END)

        debit_entry3.config(state = DISABLED)
        credit_entry3.config(state = DISABLED)
        comment_entry3.config(state = DISABLED)
        repeat_button3.config(state = DISABLED)

        self.updatetotal()

### line 4 ###
    def lineframe4(self, master):
        global line_frame4
        line_frame4 = Frame(master, bg = wc)
        line_frame4.grid(column = 0, row = 9, sticky = NW)

    def show_code4(self):
        global journalcode4
        journalcode4 = StringVar()
        global code_entry4
        code_entry4 = Entry(line_frame4, textvariable = journalcode4, font = fonts, relief = SUNKEN, bd = 2, width = 5, justify = "center")
        code_entry4.grid(column = 0, row = 0)
        code_entry4.bind("<Tab>", self.journalcode_val4)
        code_entry4.bind("<FocusOut>", self.journalcode_val4)
        code_entry4.bind("<Return>", self.journalcode_val4)
        
    def journalcode_val4(self, *args):
        global result4
        codeinput4 = journalcode4.get()
        if len(codeinput4) == 4 and codeinput4.isdigit():
            if company == "DBPSC":
                find_code4 = "SELECT title FROM chart WHERE code = ?"
            else:
                find_code4 = "SELECT title FROM dssichart WHERE code = ?"
            c.execute(find_code4, [codeinput4])
            result4 = c.fetchall()
            if result4:
                title_label4.config(text = result4[0][0])
                self.linestatenormal4()
                debit_entry4.focus()
                self.updatetotal()
            else:
                self.linestatedisabled4()
                code_entry4.delete(0, END)
                self.updatetotal()
        else:
            self.linestatedisabled4()
            code_entry4.delete(0, END)
            self.updatetotal()

    def linestatenormal4(self):
        debit_entry4.config(state = NORMAL)
        credit_entry4.config(state = NORMAL)
        comment_entry4.config(state = NORMAL)
        repeat_button4.config(state = NORMAL)

    def linestatedisabled4(self):
        title_label4.config(text = "")
        debit_entry4.delete(0, END)
        credit_entry4.delete(0, END)
        comment_entry4.delete(0, END)

        debit_entry4.config(state = DISABLED)
        credit_entry4.config(state = DISABLED)
        comment_entry4.config(state = DISABLED)
        repeat_button4.config(state = DISABLED)
      
    def showtitle4(self, *args):
        global title_label4
        title_label4 = Label(line_frame4, text = "", font = fonts, relief = SUNKEN, bd = 2, width = 33, anchor = W, bg = codetitlebg, fg = codetitlefg)
        title_label4.grid(column = 1, row = 0)

    def validate_float4(self, var4):
        global new_value4
        new_value4 = var4.get()
        try:
            new_value4 == "" or "," or float(new_value4)
            old_value4 = new_value4
        except:
            old_value4 = ""
            var4.set(old_value4)

    def validate_comma_dr4(self, *args):
        try:
            if debit_entry4.get() == "":
                dr4.set(format(float(0), ',.2f'))
            else:
                comma4 = format(float(dr4.get()), ',.2f')
                dr4.set(comma4)
        except:
            debit_entry4.delete(0, END)
            dr4.set(format(float(0), ',.2f'))
        finally:
            self.updatetotal()
            
    def validate_comma_cr4(self, *args):
        try:
            if credit_entry4.get() == "":
                cr4.set(format(float(0), ',.2f'))
            else:
                comma4 = format(float(cr4.get()), ',.2f')
                cr4.set(comma4)
        except:
            credit_entry4.delete(0, END)
            cr4.set(format(float(0), ',.2f'))
        finally:
            self.updatetotal()
            
    def showdebit4(self, *args):
        global debit_entry4
        global dr4
        dr4 = StringVar()
        debit_entry4 = Entry(line_frame4, textvariable = dr4, font = fonts, relief = SUNKEN, bd = 2, width = 14, justify = "right")
        debit_entry4.grid(column = 2, row = 0)
        dr4.trace("w", lambda nm, idx, mode, var4 = dr4: self.validate_float4(var4))
        debit_entry4.bind("<Tab>", self.validate_comma_dr4)
                
    def showcredit4(self):
        global cr4
        cr4 = StringVar()
        global credit_entry4
        credit_entry4 = Entry(line_frame4, textvariable = cr4, font = fonts, relief = SUNKEN, bd = 2, width = 14, justify = "right")
        credit_entry4.grid(column = 3, row = 0)
        cr4.trace("w", lambda nm, idx, mode, var4 = cr4: self.validate_float4(var4)) 
        credit_entry4.bind("<Tab>", self.validate_comma_cr4)
        
    def showcomment4(self):  
        global comment4
        comment4 = StringVar()
        global comment_entry4
        comment_entry4 = Entry(line_frame4, textvariable = comment4, font = fonts, relief = SUNKEN, bd = 2, width = 20, justify = "left")
        comment_entry4.grid(column = 4, row = 0)

    def showrepeat4(self, *args):
        global repeat_button4
        repeat_button4 = Button(line_frame4, text = "clear", bg = buttonbg, font = clearbuttonfont, command = self.cleanentry4, bd =0)
        repeat_button4.grid(column = 5, row = 0, padx = 2)
        repeat_button4.bind("<Return>", self.cleanentry4)
 
    def cleanentry4(self, *args):
        code_entry4.delete(0, END)
        title_label4.config(text = "")
        debit_entry4.delete(0, END)
        credit_entry4.delete(0, END)
        comment_entry4.delete(0, END)

        debit_entry4.config(state = DISABLED)
        credit_entry4.config(state = DISABLED)
        comment_entry4.config(state = DISABLED)
        repeat_button4.config(state = DISABLED)

        self.updatetotal()

### line 5 ###
    def lineframe5(self, master):
        global line_frame5
        line_frame5 = Frame(master, bg = wc)
        line_frame5.grid(column = 0, row = 10, sticky = NW)

    def show_code5(self):
        global journalcode5
        journalcode5 = StringVar()
        global code_entry5
        code_entry5 = Entry(line_frame5, textvariable = journalcode5, font = fonts, relief = SUNKEN, bd = 2, width = 5, justify = "center")
        code_entry5.grid(column = 0, row = 0)
        code_entry5.bind("<Tab>", self.journalcode_val5)
        code_entry5.bind("<FocusOut>", self.journalcode_val5)
        code_entry5.bind("<Return>", self.journalcode_val5)
        
    def journalcode_val5(self, *args):
        global result5
        codeinput5 = journalcode5.get()
        if len(codeinput5) == 4 and codeinput5.isdigit():
            if company == "DBPSC":
                find_code5 = "SELECT title FROM chart WHERE code = ?"
            else:
                find_code5 = "SELECT title FROM dssichart WHERE code = ?"
            c.execute(find_code5, [codeinput5])
            result5 = c.fetchall()
            if result5:
                title_label5.config(text = result5[0][0])
                self.linestatenormal5()
                debit_entry5.focus()
                self.updatetotal()
            else:
                self.linestatedisabled5()
                code_entry5.delete(0, END)
                self.updatetotal()
        else:
            self.linestatedisabled5()
            code_entry5.delete(0, END)
            self.updatetotal()

    def linestatenormal5(self):
        debit_entry5.config(state = NORMAL)
        credit_entry5.config(state = NORMAL)
        comment_entry5.config(state = NORMAL)
        repeat_button5.config(state = NORMAL)

    def linestatedisabled5(self):
        title_label5.config(text = "")
        debit_entry5.delete(0, END)
        credit_entry5.delete(0, END)
        comment_entry5.delete(0, END)

        debit_entry5.config(state = DISABLED)
        credit_entry5.config(state = DISABLED)
        comment_entry5.config(state = DISABLED)
        repeat_button5.config(state = DISABLED)
      
    def showtitle5(self, *args):
        global title_label5
        title_label5 = Label(line_frame5, text = "", font = fonts, relief = SUNKEN, bd = 2, width = 33, anchor = W, bg = codetitlebg, fg = codetitlefg)
        title_label5.grid(column = 1, row = 0)

    def validate_float5(self, var5):
        global new_value5
        new_value5 = var5.get()
        try:
            new_value5 == "" or "," or float(new_value5)
            old_value5 = new_value5
        except:
            old_value5 = ""
            var5.set(old_value5)

    def validate_comma_dr5(self, *args):
        try:
            if debit_entry5.get() == "":
                dr5.set(format(float(0), ',.2f'))
            else:
                comma5 = format(float(dr5.get()), ',.2f')
                dr5.set(comma5)
        except:
            debit_entry5.delete(0, END)
            dr5.set(format(float(0), ',.2f'))
        finally:
            self.updatetotal()
            
    def validate_comma_cr5(self, *args):
        try:
            if credit_entry5.get() == "":
                cr5.set(format(float(0), ',.2f'))
            else:
                comma5 = format(float(cr5.get()), ',.2f')
                cr5.set(comma5)
        except:
            credit_entry5.delete(0, END)
            cr5.set(format(float(0), ',.2f'))
        finally:
            self.updatetotal()
            
    def showdebit5(self, *args):
        global debit_entry5
        global dr5
        dr5 = StringVar()
        debit_entry5 = Entry(line_frame5, textvariable = dr5, font = fonts, relief = SUNKEN, bd = 2, width = 14, justify = "right")
        debit_entry5.grid(column = 2, row = 0)
        dr5.trace("w", lambda nm, idx, mode, var5 = dr5: self.validate_float5(var5))
        debit_entry5.bind("<Tab>", self.validate_comma_dr5)
                
    def showcredit5(self):
        global cr5
        cr5 = StringVar()
        global credit_entry5
        credit_entry5 = Entry(line_frame5, textvariable = cr5, font = fonts, relief = SUNKEN, bd = 2, width = 14, justify = "right")
        credit_entry5.grid(column = 3, row = 0)
        cr5.trace("w", lambda nm, idx, mode, var5 = cr5: self.validate_float5(var5)) 
        credit_entry5.bind("<Tab>", self.validate_comma_cr5)
        
    def showcomment5(self):  
        global comment5
        comment5 = StringVar()
        global comment_entry5
        comment_entry5 = Entry(line_frame5, textvariable = comment5, font = fonts, relief = SUNKEN, bd = 2, width = 20, justify = "left")
        comment_entry5.grid(column = 4, row = 0)

    def showrepeat5(self, *args):
        global repeat_button5
        repeat_button5 = Button(line_frame5, text = "clear", bg = buttonbg, font = clearbuttonfont, command = self.cleanentry5, bd =0)
        repeat_button5.grid(column = 5, row = 0, padx = 2)
        repeat_button5.bind("<Return>", self.cleanentry5)
 
    def cleanentry5(self, *args):
        code_entry5.delete(0, END)
        title_label5.config(text = "")
        debit_entry5.delete(0, END)
        credit_entry5.delete(0, END)
        comment_entry5.delete(0, END)

        debit_entry5.config(state = DISABLED)
        credit_entry5.config(state = DISABLED)
        comment_entry5.config(state = DISABLED)
        repeat_button5.config(state = DISABLED)

        self.updatetotal()

### line 6 ###

    def lineframe6(self, master):
        global line_frame6
        line_frame6 = Frame(master, bg = wc)
        line_frame6.grid(column = 0, row = 11, sticky = NW)

    def show_code6(self):
        global journalcode6
        journalcode6 = StringVar()
        global code_entry6
        code_entry6 = Entry(line_frame6, textvariable = journalcode6, font = fonts, relief = SUNKEN, bd = 2, width = 5, justify = "center")
        code_entry6.grid(column = 0, row = 0)
        code_entry6.bind("<Tab>", self.journalcode_val6)
        code_entry6.bind("<FocusOut>", self.journalcode_val6)
        code_entry6.bind("<Return>", self.journalcode_val6)
        
    def journalcode_val6(self, *args):
        global result6
        codeinput6 = journalcode6.get()
        if len(codeinput6) == 4 and codeinput6.isdigit():
            if company == "DBPSC":
                find_code6 = "SELECT title FROM chart WHERE code = ?"
            else:
                find_code6 = "SELECT title FROM dssichart WHERE code = ?"
            c.execute(find_code6, [codeinput6])
            result6 = c.fetchall()
            if result6:
                title_label6.config(text = result6[0][0])
                self.linestatenormal6()
                debit_entry6.focus()
                self.updatetotal()
            else:
                self.linestatedisabled6()
                code_entry6.delete(0, END)
                self.updatetotal()
        else:
            self.linestatedisabled6()
            code_entry6.delete(0, END)
            self.updatetotal()

    def linestatenormal6(self):
        debit_entry6.config(state = NORMAL)
        credit_entry6.config(state = NORMAL)
        comment_entry6.config(state = NORMAL)
        repeat_button6.config(state = NORMAL)

    def linestatedisabled6(self):
        title_label6.config(text = "")
        debit_entry6.delete(0, END)
        credit_entry6.delete(0, END)
        comment_entry6.delete(0, END)

        debit_entry6.config(state = DISABLED)
        credit_entry6.config(state = DISABLED)
        comment_entry6.config(state = DISABLED)
        repeat_button6.config(state = DISABLED)
      
    def showtitle6(self, *args):
        global title_label6
        title_label6 = Label(line_frame6, text = "", font = fonts, relief = SUNKEN, bd = 2, width = 33, anchor = W, bg = codetitlebg, fg = codetitlefg)
        title_label6.grid(column = 1, row = 0)

    def validate_float6(self, var6):
        global new_value6
        new_value6 = var6.get()
        try:
            new_value6 == "" or "," or float(new_value6)
            old_value6 = new_value6
        except:
            old_value6 = ""
            var6.set(old_value6)

    def validate_comma_dr6(self, *args):
        try:
            if debit_entry6.get() == "":
                dr6.set(format(float(0), ',.2f'))
            else:
                comma6 = format(float(dr6.get()), ',.2f')
                dr6.set(comma6)
        except:
            debit_entry6.delete(0, END)
            dr6.set(format(float(0), ',.2f'))
        finally:
            self.updatetotal()
            
    def validate_comma_cr6(self, *args):
        try:
            if credit_entry6.get() == "":
                cr6.set(format(float(0), ',.2f'))
            else:
                comma6 = format(float(cr6.get()), ',.2f')
                cr6.set(comma6)
        except:
            credit_entry6.delete(0, END)
            cr6.set(format(float(0), ',.2f'))
        finally:
            self.updatetotal()
            
    def showdebit6(self, *args):
        global debit_entry6
        global dr6
        dr6 = StringVar()
        debit_entry6 = Entry(line_frame6, textvariable = dr6, font = fonts, relief = SUNKEN, bd = 2, width = 14, justify = "right")
        debit_entry6.grid(column = 2, row = 0)
        dr6.trace("w", lambda nm, idx, mode, var6 = dr6: self.validate_float6(var6))
        debit_entry6.bind("<Tab>", self.validate_comma_dr6)
                
    def showcredit6(self):
        global cr6
        cr6 = StringVar()
        global credit_entry6
        credit_entry6 = Entry(line_frame6, textvariable = cr6, font = fonts, relief = SUNKEN, bd = 2, width = 14, justify = "right")
        credit_entry6.grid(column = 3, row = 0)
        cr6.trace("w", lambda nm, idx, mode, var6 = cr6: self.validate_float6(var6)) 
        credit_entry6.bind("<Tab>", self.validate_comma_cr6)
        
    def showcomment6(self):  
        global comment6
        comment6 = StringVar()
        global comment_entry6
        comment_entry6 = Entry(line_frame6, textvariable = comment6, font = fonts, relief = SUNKEN, bd = 2, width = 20, justify = "left")
        comment_entry6.grid(column = 4, row = 0)

    def showrepeat6(self, *args):
        global repeat_button6
        repeat_button6 = Button(line_frame6, text = "clear", bg = buttonbg, font = clearbuttonfont, command = self.cleanentry6, bd =0)
        repeat_button6.grid(column = 5, row = 0, padx = 2)
        repeat_button6.bind("<Return>", self.cleanentry6)
 
    def cleanentry6(self, *args):
        code_entry6.delete(0, END)
        title_label6.config(text = "")
        debit_entry6.delete(0, END)
        credit_entry6.delete(0, END)
        comment_entry6.delete(0, END)

        debit_entry6.config(state = DISABLED)
        credit_entry6.config(state = DISABLED)
        comment_entry6.config(state = DISABLED)
        repeat_button6.config(state = DISABLED)

        self.updatetotal()

### line 7 ###

    def lineframe7(self, master):
        global line_frame7
        line_frame7 = Frame(master, bg = wc)
        line_frame7.grid(column = 0, row = 12, sticky = NW)

    def show_code7(self):
        global journalcode7
        journalcode7 = StringVar()
        global code_entry7
        code_entry7 = Entry(line_frame7, textvariable = journalcode7, font = fonts, relief = SUNKEN, bd = 2, width = 5, justify = "center")
        code_entry7.grid(column = 0, row = 0)
        code_entry7.bind("<Tab>", self.journalcode_val7)
        code_entry7.bind("<FocusOut>", self.journalcode_val7)
        code_entry7.bind("<Return>", self.journalcode_val7)
        
    def journalcode_val7(self, *args):
        global result7
        codeinput7 = journalcode7.get()
        if len(codeinput7) == 4 and codeinput7.isdigit():
            if company == "DBPSC":
                find_code7 = "SELECT title FROM chart WHERE code = ?"
            else:
                find_code7 = "SELECT title FROM dssichart WHERE code = ?"
            c.execute(find_code7, [codeinput7])
            result7 = c.fetchall()
            if result7:
                title_label7.config(text = result7[0][0])
                self.linestatenormal7()
                debit_entry7.focus()
                self.updatetotal()
            else:
                self.linestatedisabled7()
                code_entry7.delete(0, END)
                self.updatetotal()
        else:
            self.linestatedisabled7()
            code_entry7.delete(0, END)
            self.updatetotal()

    def linestatenormal7(self):
        debit_entry7.config(state = NORMAL)
        credit_entry7.config(state = NORMAL)
        comment_entry7.config(state = NORMAL)
        repeat_button7.config(state = NORMAL)

    def linestatedisabled7(self):
        title_label7.config(text = "")
        debit_entry7.delete(0, END)
        credit_entry7.delete(0, END)
        comment_entry7.delete(0, END)

        debit_entry7.config(state = DISABLED)
        credit_entry7.config(state = DISABLED)
        comment_entry7.config(state = DISABLED)
        repeat_button7.config(state = DISABLED)
      
    def showtitle7(self, *args):
        global title_label7
        title_label7 = Label(line_frame7, text = "", font = fonts, relief = SUNKEN, bd = 2, width = 33, anchor = W, bg = codetitlebg, fg = codetitlefg)
        title_label7.grid(column = 1, row = 0)

    def validate_float7(self, var7):
        global new_value7
        new_value7 = var7.get()
        try:
            new_value7 == "" or "," or float(new_value7)
            old_value7 = new_value7
        except:
            old_value7 = ""
            var7.set(old_value7)

    def validate_comma_dr7(self, *args):
        try:
            if debit_entry7.get() == "":
                dr7.set(format(float(0), ',.2f'))
            else:
                comma7 = format(float(dr7.get()), ',.2f')
                dr7.set(comma7)
        except:
            debit_entry7.delete(0, END)
            dr7.set(format(float(0), ',.2f'))
        finally:
            self.updatetotal()
            
    def validate_comma_cr7(self, *args):
        try:
            if credit_entry7.get() == "":
                cr7.set(format(float(0), ',.2f'))
            else:
                comma7 = format(float(cr7.get()), ',.2f')
                cr7.set(comma7)
        except:
            credit_entry7.delete(0, END)
            cr7.set(format(float(0), ',.2f'))
        finally:
            self.updatetotal()
            
    def showdebit7(self, *args):
        global debit_entry7
        global dr7
        dr7 = StringVar()
        debit_entry7 = Entry(line_frame7, textvariable = dr7, font = fonts, relief = SUNKEN, bd = 2, width = 14, justify = "right")
        debit_entry7.grid(column = 2, row = 0)
        dr7.trace("w", lambda nm, idx, mode, var7 = dr7: self.validate_float7(var7))
        debit_entry7.bind("<Tab>", self.validate_comma_dr7)
                
    def showcredit7(self):
        global cr7
        cr7 = StringVar()
        global credit_entry7
        credit_entry7 = Entry(line_frame7, textvariable = cr7, font = fonts, relief = SUNKEN, bd = 2, width = 14, justify = "right")
        credit_entry7.grid(column = 3, row = 0)
        cr7.trace("w", lambda nm, idx, mode, var7 = cr7: self.validate_float7(var7)) 
        credit_entry7.bind("<Tab>", self.validate_comma_cr7)
        
    def showcomment7(self):  
        global comment7
        comment7 = StringVar()
        global comment_entry7
        comment_entry7 = Entry(line_frame7, textvariable = comment7, font = fonts, relief = SUNKEN, bd = 2, width = 20, justify = "left")
        comment_entry7.grid(column = 4, row = 0)

    def showrepeat7(self, *args):
        global repeat_button7
        repeat_button7 = Button(line_frame7, text = "clear", bg = buttonbg, font = clearbuttonfont, command = self.cleanentry7, bd =0)
        repeat_button7.grid(column = 5, row = 0, padx = 2)
        repeat_button7.bind("<Return>", self.cleanentry7)
 
    def cleanentry7(self, *args):
        code_entry7.delete(0, END)
        title_label7.config(text = "")
        debit_entry7.delete(0, END)
        credit_entry7.delete(0, END)
        comment_entry7.delete(0, END)

        debit_entry7.config(state = DISABLED)
        credit_entry7.config(state = DISABLED)
        comment_entry7.config(state = DISABLED)
        repeat_button7.config(state = DISABLED)

        self.updatetotal()

### line 8 ###

    def lineframe8(self, master):
        global line_frame8
        line_frame8 = Frame(master, bg = wc)
        line_frame8.grid(column = 0, row = 13, sticky = NW)

    def show_code8(self):
        global journalcode8
        journalcode8 = StringVar()
        global code_entry8
        code_entry8 = Entry(line_frame8, textvariable = journalcode8, font = fonts, relief = SUNKEN, bd = 2, width = 5, justify = "center")
        code_entry8.grid(column = 0, row = 0)
        code_entry8.bind("<Tab>", self.journalcode_val8)
        code_entry8.bind("<FocusOut>", self.journalcode_val8)
        code_entry8.bind("<Return>", self.journalcode_val8)
        
    def journalcode_val8(self, *args):
        global result8
        codeinput8 = journalcode8.get()
        if len(codeinput8) == 4 and codeinput8.isdigit():
            if company == "DBPSC":
                find_code8 = "SELECT title FROM chart WHERE code = ?"
            else:
                find_code8 = "SELECT title FROM dssichart WHERE code = ?"
            c.execute(find_code8, [codeinput8])
            result8 = c.fetchall()
            if result8:
                title_label8.config(text = result8[0][0])
                self.linestatenormal8()
                debit_entry8.focus()
                self.updatetotal()
            else:
                self.linestatedisabled8()
                code_entry8.delete(0, END)
                self.updatetotal()
        else:
            self.linestatedisabled8()
            code_entry8.delete(0, END)
            self.updatetotal()

    def linestatenormal8(self):
        debit_entry8.config(state = NORMAL)
        credit_entry8.config(state = NORMAL)
        comment_entry8.config(state = NORMAL)
        repeat_button8.config(state = NORMAL)

    def linestatedisabled8(self):
        title_label8.config(text = "")
        debit_entry8.delete(0, END)
        credit_entry8.delete(0, END)
        comment_entry8.delete(0, END)

        debit_entry8.config(state = DISABLED)
        credit_entry8.config(state = DISABLED)
        comment_entry8.config(state = DISABLED)
        repeat_button8.config(state = DISABLED)
      
    def showtitle8(self, *args):
        global title_label8
        title_label8 = Label(line_frame8, text = "", font = fonts, relief = SUNKEN, bd = 2, width = 33, anchor = W, bg = codetitlebg, fg = codetitlefg)
        title_label8.grid(column = 1, row = 0)

    def validate_float8(self, var8):
        global new_value8
        new_value8 = var8.get()
        try:
            new_value8 == "" or "," or float(new_value8)
            old_value8 = new_value8
        except:
            old_value8 = ""
            var8.set(old_value8)

    def validate_comma_dr8(self, *args):
        try:
            if debit_entry8.get() == "":
                dr8.set(format(float(0), ',.2f'))
            else:
                comma8 = format(float(dr8.get()), ',.2f')
                dr8.set(comma8)
        except:
            debit_entry8.delete(0, END)
            dr8.set(format(float(0), ',.2f'))
        finally:
            self.updatetotal()
            
    def validate_comma_cr8(self, *args):
        try:
            if credit_entry8.get() == "":
                cr8.set(format(float(0), ',.2f'))
            else:
                comma8 = format(float(cr8.get()), ',.2f')
                cr8.set(comma8)
        except:
            credit_entry8.delete(0, END)
            cr8.set(format(float(0), ',.2f'))
        finally:
            self.updatetotal()
            
    def showdebit8(self, *args):
        global debit_entry8
        global dr8
        dr8 = StringVar()
        debit_entry8 = Entry(line_frame8, textvariable = dr8, font = fonts, relief = SUNKEN, bd = 2, width = 14, justify = "right")
        debit_entry8.grid(column = 2, row = 0)
        dr8.trace("w", lambda nm, idx, mode, var8 = dr8: self.validate_float8(var8))
        debit_entry8.bind("<Tab>", self.validate_comma_dr8)
                
    def showcredit8(self):
        global cr8
        cr8 = StringVar()
        global credit_entry8
        credit_entry8 = Entry(line_frame8, textvariable = cr8, font = fonts, relief = SUNKEN, bd = 2, width = 14, justify = "right")
        credit_entry8.grid(column = 3, row = 0)
        cr8.trace("w", lambda nm, idx, mode, var8 = cr8: self.validate_float8(var8)) 
        credit_entry8.bind("<Tab>", self.validate_comma_cr8)
        
    def showcomment8(self):  
        global comment8
        comment8 = StringVar()
        global comment_entry8
        comment_entry8 = Entry(line_frame8, textvariable = comment8, font = fonts, relief = SUNKEN, bd = 2, width = 20, justify = "left")
        comment_entry8.grid(column = 4, row = 0)

    def showrepeat8(self, *args):
        global repeat_button8
        repeat_button8 = Button(line_frame8, text = "clear", bg = buttonbg, font = clearbuttonfont, command = self.cleanentry8, bd =0)
        repeat_button8.grid(column = 5, row = 0, padx = 2)
        repeat_button8.bind("<Return>", self.cleanentry8)
 
    def cleanentry8(self, *args):
        code_entry8.delete(0, END)
        title_label8.config(text = "")
        debit_entry8.delete(0, END)
        credit_entry8.delete(0, END)
        comment_entry8.delete(0, END)

        debit_entry8.config(state = DISABLED)
        credit_entry8.config(state = DISABLED)
        comment_entry8.config(state = DISABLED)
        repeat_button8.config(state = DISABLED)

        self.updatetotal()

### line 9 ###
    def lineframe9(self, master):
        global line_frame9
        line_frame9 = Frame(master, bg = wc)
        line_frame9.grid(column = 0, row = 14, sticky = NW)

    def show_code9(self):
        global journalcode9
        journalcode9 = StringVar()
        global code_entry9
        code_entry9 = Entry(line_frame9, textvariable = journalcode9, font = fonts, relief = SUNKEN, bd = 2, width = 5, justify = "center")
        code_entry9.grid(column = 0, row = 0)
        code_entry9.bind("<Tab>", self.journalcode_val9)
        code_entry9.bind("<FocusOut>", self.journalcode_val9)
        code_entry9.bind("<Return>", self.journalcode_val9)
        
    def journalcode_val9(self, *args):
        global result9
        codeinput9 = journalcode9.get()
        if len(codeinput9) == 4 and codeinput9.isdigit():
            if company == "DBPSC":
                find_code9 = "SELECT title FROM chart WHERE code = ?"
            else:
                find_code9 = "SELECT title FROM dssichart WHERE code = ?"
            c.execute(find_code9, [codeinput9])
            result9 = c.fetchall()
            if result9:
                title_label9.config(text = result9[0][0])
                self.linestatenormal9()
                debit_entry9.focus()
                self.updatetotal()
            else:
                self.linestatedisabled9()
                code_entry9.delete(0, END)
                self.updatetotal()
        else:
            self.linestatedisabled9()
            code_entry9.delete(0, END)
            self.updatetotal()

    def linestatenormal9(self):
        debit_entry9.config(state = NORMAL)
        credit_entry9.config(state = NORMAL)
        comment_entry9.config(state = NORMAL)
        repeat_button9.config(state = NORMAL)

    def linestatedisabled9(self):
        title_label9.config(text = "")
        debit_entry9.delete(0, END)
        credit_entry9.delete(0, END)
        comment_entry9.delete(0, END)

        debit_entry9.config(state = DISABLED)
        credit_entry9.config(state = DISABLED)
        comment_entry9.config(state = DISABLED)
        repeat_button9.config(state = DISABLED)
      
    def showtitle9(self, *args):
        global title_label9
        title_label9 = Label(line_frame9, text = "", font = fonts, relief = SUNKEN, bd = 2, width = 33, anchor = W, bg = codetitlebg, fg = codetitlefg)
        title_label9.grid(column = 1, row = 0)

    def validate_float9(self, var9):
        global new_value9
        new_value9 = var9.get()
        try:
            new_value9 == "" or "," or float(new_value9)
            old_value9 = new_value9
        except:
            old_value9 = ""
            var9.set(old_value9)

    def validate_comma_dr9(self, *args):
        try:
            if debit_entry9.get() == "":
                dr9.set(format(float(0), ',.2f'))
            else:
                comma9 = format(float(dr9.get()), ',.2f')
                dr9.set(comma9)
        except:
            debit_entry9.delete(0, END)
            dr9.set(format(float(0), ',.2f'))
        finally:
            self.updatetotal()
            
    def validate_comma_cr9(self, *args):
        try:
            if credit_entry9.get() == "":
                cr9.set(format(float(0), ',.2f'))
            else:
                comma9 = format(float(cr9.get()), ',.2f')
                cr9.set(comma9)
        except:
            credit_entry9.delete(0, END)
            cr9.set(format(float(0), ',.2f'))
        finally:
            self.updatetotal()
            
    def showdebit9(self, *args):
        global debit_entry9
        global dr9
        dr9 = StringVar()
        debit_entry9 = Entry(line_frame9, textvariable = dr9, font = fonts, relief = SUNKEN, bd = 2, width = 14, justify = "right")
        debit_entry9.grid(column = 2, row = 0)
        dr9.trace("w", lambda nm, idx, mode, var9 = dr9: self.validate_float9(var9))
        debit_entry9.bind("<Tab>", self.validate_comma_dr9)
                
    def showcredit9(self):
        global cr9
        cr9 = StringVar()
        global credit_entry9
        credit_entry9 = Entry(line_frame9, textvariable = cr9, font = fonts, relief = SUNKEN, bd = 2, width = 14, justify = "right")
        credit_entry9.grid(column = 3, row = 0)
        cr9.trace("w", lambda nm, idx, mode, var9 = cr9: self.validate_float9(var9)) 
        credit_entry9.bind("<Tab>", self.validate_comma_cr9)
        
    def showcomment9(self):  
        global comment9
        comment9 = StringVar()
        global comment_entry9
        comment_entry9 = Entry(line_frame9, textvariable = comment9, font = fonts, relief = SUNKEN, bd = 2, width = 20, justify = "left")
        comment_entry9.grid(column = 4, row = 0)

    def showrepeat9(self, *args):
        global repeat_button9
        repeat_button9 = Button(line_frame9, text = "clear", bg = buttonbg, font = clearbuttonfont, command = self.cleanentry9, bd =0)
        repeat_button9.grid(column = 5, row = 0, padx = 2)
        repeat_button9.bind("<Return>", self.cleanentry9)
 
    def cleanentry9(self, *args):
        code_entry9.delete(0, END)
        title_label9.config(text = "")
        debit_entry9.delete(0, END)
        credit_entry9.delete(0, END)
        comment_entry9.delete(0, END)

        debit_entry9.config(state = DISABLED)
        credit_entry9.config(state = DISABLED)
        comment_entry9.config(state = DISABLED)
        repeat_button9.config(state = DISABLED)

        self.updatetotal()

### line 10 ###
    def lineframe10(self, master):
        global line_frame10
        line_frame10 = Frame(master, bg = wc)
        line_frame10.grid(column = 0, row = 15, sticky = NW)

    def show_code10(self):
        global journalcode10
        journalcode10 = StringVar()
        global code_entry10
        code_entry10 = Entry(line_frame10, textvariable = journalcode10, font = fonts, relief = SUNKEN, bd = 2, width = 5, justify = "center")
        code_entry10.grid(column = 0, row = 0)
        code_entry10.bind("<Tab>", self.journalcode_val10)
        code_entry10.bind("<FocusOut>", self.journalcode_val10)
        code_entry10.bind("<Return>", self.journalcode_val10)
        
    def journalcode_val10(self, *args):
        global result10
        codeinput10 = journalcode10.get()
        if len(codeinput10) == 4 and codeinput10.isdigit():
            if company == "DBPSC":
                find_code10 = "SELECT title FROM chart WHERE code = ?"
            else:
                find_code10 = "SELECT title FROM dssichart WHERE code = ?"
            c.execute(find_code10, [codeinput10])
            result10 = c.fetchall()
            if result10:
                title_label10.config(text = result10[0][0])
                self.linestatenormal10()
                debit_entry10.focus()
                self.updatetotal()
            else:
                self.linestatedisabled10()
                code_entry10.delete(0, END)
                self.updatetotal()
        else:
            self.linestatedisabled10()
            code_entry10.delete(0, END)
            self.updatetotal()

    def linestatenormal10(self):
        debit_entry10.config(state = NORMAL)
        credit_entry10.config(state = NORMAL)
        comment_entry10.config(state = NORMAL)
        repeat_button10.config(state = NORMAL)

    def linestatedisabled10(self):
        title_label10.config(text = "")
        debit_entry10.delete(0, END)
        credit_entry10.delete(0, END)
        comment_entry10.delete(0, END)

        debit_entry10.config(state = DISABLED)
        credit_entry10.config(state = DISABLED)
        comment_entry10.config(state = DISABLED)
        repeat_button10.config(state = DISABLED)
      
    def showtitle10(self, *args):
        global title_label10
        title_label10 = Label(line_frame10, text = "", font = fonts, relief = SUNKEN, bd = 2, width = 33, anchor = W, bg = codetitlebg, fg = codetitlefg)
        title_label10.grid(column = 1, row = 0)

    def validate_float10(self, var10):
        global new_value10
        new_value10 = var10.get()
        try:
            new_value10 == "" or "," or float(new_value10)
            old_value10 = new_value10
        except:
            old_value10 = ""
            var10.set(old_value10)

    def validate_comma_dr10(self, *args):
        try:
            if debit_entry10.get() == "":
                dr10.set(format(float(0), ',.2f'))
            else:
                comma10 = format(float(dr10.get()), ',.2f')
                dr10.set(comma10)
        except:
            debit_entry10.delete(0, END)
            dr10.set(format(float(0), ',.2f'))
        finally:
            self.updatetotal()
            
    def validate_comma_cr10(self, *args):
        try:
            if credit_entry10.get() == "":
                cr10.set(format(float(0), ',.2f'))
            else:
                comma10 = format(float(cr10.get()), ',.2f')
                cr10.set(comma10)
        except:
            credit_entry10.delete(0, END)
            cr10.set(format(float(0), ',.2f'))
        finally:
            self.updatetotal()
            
    def showdebit10(self, *args):
        global debit_entry10
        global dr10
        dr10 = StringVar()
        debit_entry10 = Entry(line_frame10, textvariable = dr10, font = fonts, relief = SUNKEN, bd = 2, width = 14, justify = "right")
        debit_entry10.grid(column = 2, row = 0)
        dr10.trace("w", lambda nm, idx, mode, var10 = dr10: self.validate_float10(var10))
        debit_entry10.bind("<Tab>", self.validate_comma_dr10)
                
    def showcredit10(self):
        global cr10
        cr10 = StringVar()
        global credit_entry10
        credit_entry10 = Entry(line_frame10, textvariable = cr10, font = fonts, relief = SUNKEN, bd = 2, width = 14, justify = "right")
        credit_entry10.grid(column = 3, row = 0)
        cr10.trace("w", lambda nm, idx, mode, var10 = cr10: self.validate_float10(var10)) 
        credit_entry10.bind("<Tab>", self.validate_comma_cr10)
        
    def showcomment10(self):  
        global comment10
        comment10 = StringVar()
        global comment_entry10
        comment_entry10 = Entry(line_frame10, textvariable = comment10, font = fonts, relief = SUNKEN, bd = 2, width = 20, justify = "left")
        comment_entry10.grid(column = 4, row = 0)

    def showrepeat10(self, *args):
        global repeat_button10
        repeat_button10 = Button(line_frame10, text = "clear", bg = buttonbg, font = clearbuttonfont, command = self.cleanentry10, bd =0)
        repeat_button10.grid(column = 5, row = 0, padx = 2)
        repeat_button10.bind("<Return>", self.cleanentry10)
 
    def cleanentry10(self, *args):
        code_entry10.delete(0, END)
        title_label10.config(text = "")
        debit_entry10.delete(0, END)
        credit_entry10.delete(0, END)
        comment_entry10.delete(0, END)

        debit_entry10.config(state = DISABLED)
        credit_entry10.config(state = DISABLED)
        comment_entry10.config(state = DISABLED)
        repeat_button10.config(state = DISABLED)

        self.updatetotal()

### line 11 ###
    def lineframe11(self, master):
        global line_frame11
        line_frame11 = Frame(master, bg = wc)
        line_frame11.grid(column = 0, row = 16, sticky = NW)

    def show_code11(self):
        global journalcode11
        journalcode11 = StringVar()
        global code_entry11
        code_entry11 = Entry(line_frame11, textvariable = journalcode11, font = fonts, relief = SUNKEN, bd = 2, width = 5, justify = "center")
        code_entry11.grid(column = 0, row = 0)
        code_entry11.bind("<Tab>", self.journalcode_val11)
        code_entry11.bind("<FocusOut>", self.journalcode_val11)
        code_entry11.bind("<Return>", self.journalcode_val11)
        
    def journalcode_val11(self, *args):
        global result11
        codeinput11 = journalcode11.get()
        if len(codeinput11) == 4 and codeinput11.isdigit():
            if company == "DBPSC":
                find_code11 = "SELECT title FROM chart WHERE code = ?"
            else:
                find_code11 = "SELECT title FROM dssichart WHERE code = ?"    
            c.execute(find_code11, [codeinput11])
            result11 = c.fetchall()
            if result11:
                title_label11.config(text = result11[0][0])
                self.linestatenormal11()
                debit_entry11.focus()
                self.updatetotal()
            else:
                self.linestatedisabled11()
                code_entry11.delete(0, END)
                self.updatetotal()
        else:
            self.linestatedisabled11()
            code_entry11.delete(0, END)
            self.updatetotal()

    def linestatenormal11(self):
        debit_entry11.config(state = NORMAL)
        credit_entry11.config(state = NORMAL)
        comment_entry11.config(state = NORMAL)
        repeat_button11.config(state = NORMAL)

    def linestatedisabled11(self):
        title_label11.config(text = "")
        debit_entry11.delete(0, END)
        credit_entry11.delete(0, END)
        comment_entry11.delete(0, END)

        debit_entry11.config(state = DISABLED)
        credit_entry11.config(state = DISABLED)
        comment_entry11.config(state = DISABLED)
        repeat_button11.config(state = DISABLED)
      
    def showtitle11(self, *args):
        global title_label11
        title_label11 = Label(line_frame11, text = "", font = fonts, relief = SUNKEN, bd = 2, width = 33, anchor = W, bg = codetitlebg, fg = codetitlefg)
        title_label11.grid(column = 1, row = 0)

    def validate_float11(self, var11):
        global new_value11
        new_value11 = var11.get()
        try:
            new_value11 == "" or "," or float(new_value11)
            old_value11 = new_value11
        except:
            old_value11 = ""
            var11.set(old_value11)

    def validate_comma_dr11(self, *args):
        try:
            if debit_entry11.get() == "":
                dr11.set(format(float(0), ',.2f'))
            else:
                comma11 = format(float(dr11.get()), ',.2f')
                dr11.set(comma11)
        except:
            debit_entry11.delete(0, END)
            dr11.set(format(float(0), ',.2f'))
        finally:
            self.updatetotal()
            
    def validate_comma_cr11(self, *args):
        try:
            if credit_entry11.get() == "":
                cr11.set(format(float(0), ',.2f'))
            else:
                comma11 = format(float(cr11.get()), ',.2f')
                cr11.set(comma11)
        except:
            credit_entry11.delete(0, END)
            cr11.set(format(float(0), ',.2f'))
        finally:
            self.updatetotal()
            
    def showdebit11(self, *args):
        global debit_entry11
        global dr11
        dr11 = StringVar()
        debit_entry11 = Entry(line_frame11, textvariable = dr11, font = fonts, relief = SUNKEN, bd = 2, width = 14, justify = "right")
        debit_entry11.grid(column = 2, row = 0)
        dr11.trace("w", lambda nm, idx, mode, var11 = dr11: self.validate_float11(var11))
        debit_entry11.bind("<Tab>", self.validate_comma_dr11)
                
    def showcredit11(self):
        global cr11
        cr11 = StringVar()
        global credit_entry11
        credit_entry11 = Entry(line_frame11, textvariable = cr11, font = fonts, relief = SUNKEN, bd = 2, width = 14, justify = "right")
        credit_entry11.grid(column = 3, row = 0)
        cr11.trace("w", lambda nm, idx, mode, var11 = cr11: self.validate_float11(var11)) 
        credit_entry11.bind("<Tab>", self.validate_comma_cr11)
        
    def showcomment11(self):  
        global comment11
        comment11 = StringVar()
        global comment_entry11
        comment_entry11 = Entry(line_frame11, textvariable = comment11, font = fonts, relief = SUNKEN, bd = 2, width = 20, justify = "left")
        comment_entry11.grid(column = 4, row = 0)

    def showrepeat11(self, *args):
        global repeat_button11
        repeat_button11 = Button(line_frame11, text = "clear", bg = buttonbg, font = clearbuttonfont, command = self.cleanentry11, bd =0)
        repeat_button11.grid(column = 5, row = 0, padx = 2)
        repeat_button11.bind("<Return>", self.cleanentry11)
 
    def cleanentry11(self, *args):
        code_entry11.delete(0, END)
        title_label11.config(text = "")
        debit_entry11.delete(0, END)
        credit_entry11.delete(0, END)
        comment_entry11.delete(0, END)

        debit_entry11.config(state = DISABLED)
        credit_entry11.config(state = DISABLED)
        comment_entry11.config(state = DISABLED)
        repeat_button11.config(state = DISABLED)

        self.updatetotal()

### line 12 ###
    def lineframe12(self, master):
        global line_frame12
        line_frame12 = Frame(master, bg = wc)
        line_frame12.grid(column = 0, row = 17, sticky = NW)

    def show_code12(self):
        global journalcode12
        journalcode12 = StringVar()
        global code_entry12
        code_entry12 = Entry(line_frame12, textvariable = journalcode12, font = fonts, relief = SUNKEN, bd = 2, width = 5, justify = "center")
        code_entry12.grid(column = 0, row = 0)
        code_entry12.bind("<Tab>", self.journalcode_val12)
        code_entry12.bind("<FocusOut>", self.journalcode_val12)
        code_entry12.bind("<Return>", self.journalcode_val12)
        
    def journalcode_val12(self, *args):
        global result12
        codeinput12 = journalcode12.get()
        if len(codeinput12) == 4 and codeinput12.isdigit():
            if company == "DBPSC":
                find_code12 = "SELECT title FROM chart WHERE code = ?"
            else:
                find_code12 = "SELECT title FROM dssichart WHERE code = ?"
            c.execute(find_code12, [codeinput12])
            result12 = c.fetchall()
            if result12:
                title_label12.config(text = result12[0][0])
                self.linestatenormal12()
                debit_entry12.focus()
                self.updatetotal()
            else:
                self.linestatedisabled12()
                code_entry12.delete(0, END)
                self.updatetotal()
        else:
            self.linestatedisabled12()
            code_entry12.delete(0, END)
            self.updatetotal()

    def linestatenormal12(self):
        debit_entry12.config(state = NORMAL)
        credit_entry12.config(state = NORMAL)
        comment_entry12.config(state = NORMAL)
        repeat_button12.config(state = NORMAL)

    def linestatedisabled12(self):
        title_label12.config(text = "")
        debit_entry12.delete(0, END)
        credit_entry12.delete(0, END)
        comment_entry12.delete(0, END)

        debit_entry12.config(state = DISABLED)
        credit_entry12.config(state = DISABLED)
        comment_entry12.config(state = DISABLED)
        repeat_button12.config(state = DISABLED)
      
    def showtitle12(self, *args):
        global title_label12
        title_label12 = Label(line_frame12, text = "", font = fonts, relief = SUNKEN, bd = 2, width = 33, anchor = W, bg = codetitlebg, fg = codetitlefg)
        title_label12.grid(column = 1, row = 0)

    def validate_float12(self, var12):
        global new_value12
        new_value12 = var12.get()
        try:
            new_value12 == "" or "," or float(new_value12)
            old_value12 = new_value12
        except:
            old_value12 = ""
            var12.set(old_value12)

    def validate_comma_dr12(self, *args):
        try:
            if debit_entry12.get() == "":
                dr12.set(format(float(0), ',.2f'))
            else:
                comma12 = format(float(dr12.get()), ',.2f')
                dr12.set(comma12)
        except:
            debit_entry12.delete(0, END)
            dr12.set(format(float(0), ',.2f'))
        finally:
            self.updatetotal()
            
    def validate_comma_cr12(self, *args):
        try:
            if credit_entry12.get() == "":
                cr12.set(format(float(0), ',.2f'))
            else:
                comma12 = format(float(cr12.get()), ',.2f')
                cr12.set(comma12)
        except:
            credit_entry12.delete(0, END)
            cr12.set(format(float(0), ',.2f'))
        finally:
            self.updatetotal()
            
    def showdebit12(self, *args):
        global debit_entry12
        global dr12
        dr12 = StringVar()
        debit_entry12 = Entry(line_frame12, textvariable = dr12, font = fonts, relief = SUNKEN, bd = 2, width = 14, justify = "right")
        debit_entry12.grid(column = 2, row = 0)
        dr12.trace("w", lambda nm, idx, mode, var12 = dr12: self.validate_float12(var12))
        debit_entry12.bind("<Tab>", self.validate_comma_dr12)
                
    def showcredit12(self):
        global cr12
        cr12 = StringVar()
        global credit_entry12
        credit_entry12 = Entry(line_frame12, textvariable = cr12, font = fonts, relief = SUNKEN, bd = 2, width = 14, justify = "right")
        credit_entry12.grid(column = 3, row = 0)
        cr12.trace("w", lambda nm, idx, mode, var12 = cr12: self.validate_float12(var12)) 
        credit_entry12.bind("<Tab>", self.validate_comma_cr12)
        
    def showcomment12(self):  
        global comment12
        comment12 = StringVar()
        global comment_entry12
        comment_entry12 = Entry(line_frame12, textvariable = comment12, font = fonts, relief = SUNKEN, bd = 2, width = 20, justify = "left")
        comment_entry12.grid(column = 4, row = 0)

    def showrepeat12(self, *args):
        global repeat_button12
        repeat_button12 = Button(line_frame12, text = "clear", bg = buttonbg, font = clearbuttonfont, command = self.cleanentry12, bd =0)
        repeat_button12.grid(column = 5, row = 0, padx = 2)
        repeat_button12.bind("<Return>", self.cleanentry12)
 
    def cleanentry12(self, *args):
        code_entry12.delete(0, END)
        title_label12.config(text = "")
        debit_entry12.delete(0, END)
        credit_entry12.delete(0, END)
        comment_entry12.delete(0, END)

        debit_entry12.config(state = DISABLED)
        credit_entry12.config(state = DISABLED)
        comment_entry12.config(state = DISABLED)
        repeat_button12.config(state = DISABLED)

        self.updatetotal()

### line 13 ###
    def lineframe13(self, master):
        global line_frame13
        line_frame13 = Frame(master, bg = wc)
        line_frame13.grid(column = 0, row = 18, sticky = NW)

    def show_code13(self):
        global journalcode13
        journalcode13 = StringVar()
        global code_entry13
        code_entry13 = Entry(line_frame13, textvariable = journalcode13, font = fonts, relief = SUNKEN, bd = 2, width = 5, justify = "center")
        code_entry13.grid(column = 0, row = 0)
        code_entry13.bind("<Tab>", self.journalcode_val13)
        code_entry13.bind("<FocusOut>", self.journalcode_val13)
        code_entry13.bind("<Return>", self.journalcode_val13)
        
    def journalcode_val13(self, *args):
        global result13
        codeinput13 = journalcode13.get()
        if len(codeinput13) == 4 and codeinput13.isdigit():
            if company == "DBPSC":
                find_code13 = "SELECT title FROM chart WHERE code = ?"
            else:
                find_code13 = "SELECT title FROM dssichart WHERE code = ?"
            c.execute(find_code13, [codeinput13])
            result13 = c.fetchall()
            if result13:
                title_label13.config(text = result13[0][0])
                self.linestatenormal13()
                debit_entry13.focus()
                self.updatetotal()
            else:
                self.linestatedisabled13()
                code_entry13.delete(0, END)
                self.updatetotal()
        else:
            self.linestatedisabled13()
            code_entry13.delete(0, END)
            self.updatetotal()

    def linestatenormal13(self):
        debit_entry13.config(state = NORMAL)
        credit_entry13.config(state = NORMAL)
        comment_entry13.config(state = NORMAL)
        repeat_button13.config(state = NORMAL)

    def linestatedisabled13(self):
        title_label13.config(text = "")
        debit_entry13.delete(0, END)
        credit_entry13.delete(0, END)
        comment_entry13.delete(0, END)

        debit_entry13.config(state = DISABLED)
        credit_entry13.config(state = DISABLED)
        comment_entry13.config(state = DISABLED)
        repeat_button13.config(state = DISABLED)
      
    def showtitle13(self, *args):
        global title_label13
        title_label13 = Label(line_frame13, text = "", font = fonts, relief = SUNKEN, bd = 2, width = 33, anchor = W, bg = codetitlebg, fg = codetitlefg)
        title_label13.grid(column = 1, row = 0)

    def validate_float13(self, var13):
        global new_value13
        new_value13 = var13.get()
        try:
            new_value13 == "" or "," or float(new_value13)
            old_value13 = new_value13
        except:
            old_value13 = ""
            var13.set(old_value13)

    def validate_comma_dr13(self, *args):
        try:
            if debit_entry13.get() == "":
                dr13.set(format(float(0), ',.2f'))
            else:
                comma13 = format(float(dr13.get()), ',.2f')
                dr13.set(comma13)
        except:
            debit_entry13.delete(0, END)
            dr13.set(format(float(0), ',.2f'))
        finally:
            self.updatetotal()
            
    def validate_comma_cr13(self, *args):
        try:
            if credit_entry13.get() == "":
                cr13.set(format(float(0), ',.2f'))
            else:
                comma13 = format(float(cr13.get()), ',.2f')
                cr13.set(comma13)
        except:
            credit_entry13.delete(0, END)
            cr13.set(format(float(0), ',.2f'))
        finally:
            self.updatetotal()
            
    def showdebit13(self, *args):
        global debit_entry13
        global dr13
        dr13 = StringVar()
        debit_entry13 = Entry(line_frame13, textvariable = dr13, font = fonts, relief = SUNKEN, bd = 2, width = 14, justify = "right")
        debit_entry13.grid(column = 2, row = 0)
        dr13.trace("w", lambda nm, idx, mode, var13 = dr13: self.validate_float13(var13))
        debit_entry13.bind("<Tab>", self.validate_comma_dr13)
                
    def showcredit13(self):
        global cr13
        cr13 = StringVar()
        global credit_entry13
        credit_entry13 = Entry(line_frame13, textvariable = cr13, font = fonts, relief = SUNKEN, bd = 2, width = 14, justify = "right")
        credit_entry13.grid(column = 3, row = 0)
        cr13.trace("w", lambda nm, idx, mode, var13 = cr13: self.validate_float13(var13)) 
        credit_entry13.bind("<Tab>", self.validate_comma_cr13)
        
    def showcomment13(self):  
        global comment13
        comment13 = StringVar()
        global comment_entry13
        comment_entry13 = Entry(line_frame13, textvariable = comment13, font = fonts, relief = SUNKEN, bd = 2, width = 20, justify = "left")
        comment_entry13.grid(column = 4, row = 0)

    def showrepeat13(self, *args):
        global repeat_button13
        repeat_button13 = Button(line_frame13, text = "clear", bg = buttonbg, font = clearbuttonfont, command = self.cleanentry13, bd =0)
        repeat_button13.grid(column = 5, row = 0, padx = 2)
        repeat_button13.bind("<Return>", self.cleanentry13)
 
    def cleanentry13(self, *args):
        code_entry13.delete(0, END)
        title_label13.config(text = "")
        debit_entry13.delete(0, END)
        credit_entry13.delete(0, END)
        comment_entry13.delete(0, END)

        debit_entry13.config(state = DISABLED)
        credit_entry13.config(state = DISABLED)
        comment_entry13.config(state = DISABLED)
        repeat_button13.config(state = DISABLED)

        self.updatetotal()

### line 14 ###
    def lineframe14(self, master):
        global line_frame14
        line_frame14 = Frame(master, bg = wc)
        line_frame14.grid(column = 0, row = 19, sticky = NW)

    def show_code14(self):
        global journalcode14
        journalcode14 = StringVar()
        global code_entry14
        code_entry14 = Entry(line_frame14, textvariable = journalcode14, font = fonts, relief = SUNKEN, bd = 2, width = 5, justify = "center")
        code_entry14.grid(column = 0, row = 0)
        code_entry14.bind("<Tab>", self.journalcode_val14)
        code_entry14.bind("<FocusOut>", self.journalcode_val14)
        code_entry14.bind("<Return>", self.journalcode_val14)
        
    def journalcode_val14(self, *args):
        global result14
        codeinput14 = journalcode14.get()
        if len(codeinput14) == 4 and codeinput14.isdigit():
            if company == "DBPSC":
                find_code14 = "SELECT title FROM chart WHERE code = ?"
            else:
                find_code14 = "SELECT title FROM dssichart WHERE code = ?"
            c.execute(find_code14, [codeinput14])
            result14 = c.fetchall()
            if result14:
                title_label14.config(text = result14[0][0])
                self.linestatenormal14()
                debit_entry14.focus()
                self.updatetotal()
            else:
                self.linestatedisabled14()
                code_entry14.delete(0, END)
                self.updatetotal()
        else:
            self.linestatedisabled14()
            code_entry14.delete(0, END)
            self.updatetotal()

    def linestatenormal14(self):
        debit_entry14.config(state = NORMAL)
        credit_entry14.config(state = NORMAL)
        comment_entry14.config(state = NORMAL)
        repeat_button14.config(state = NORMAL)

    def linestatedisabled14(self):
        title_label14.config(text = "")
        debit_entry14.delete(0, END)
        credit_entry14.delete(0, END)
        comment_entry14.delete(0, END)

        debit_entry14.config(state = DISABLED)
        credit_entry14.config(state = DISABLED)
        comment_entry14.config(state = DISABLED)
        repeat_button14.config(state = DISABLED)
      
    def showtitle14(self, *args):
        global title_label14
        title_label14 = Label(line_frame14, text = "", font = fonts, relief = SUNKEN, bd = 2, width = 33, anchor = W, bg = codetitlebg, fg = codetitlefg)
        title_label14.grid(column = 1, row = 0)

    def validate_float14(self, var14):
        global new_value14
        new_value14 = var14.get()
        try:
            new_value14 == "" or "," or float(new_value14)
            old_value14 = new_value14
        except:
            old_value14 = ""
            var14.set(old_value14)

    def validate_comma_dr14(self, *args):
        try:
            if debit_entry14.get() == "":
                dr14.set(format(float(0), ',.2f'))
            else:
                comma14 = format(float(dr14.get()), ',.2f')
                dr14.set(comma14)
        except:
            debit_entry14.delete(0, END)
            dr14.set(format(float(0), ',.2f'))
        finally:
            self.updatetotal()
            
    def validate_comma_cr14(self, *args):
        try:
            if credit_entry14.get() == "":
                cr14.set(format(float(0), ',.2f'))
            else:
                comma14 = format(float(cr14.get()), ',.2f')
                cr14.set(comma14)
        except:
            credit_entry14.delete(0, END)
            cr14.set(format(float(0), ',.2f'))
        finally:
            self.updatetotal()
            
    def showdebit14(self, *args):
        global debit_entry14
        global dr14
        dr14 = StringVar()
        debit_entry14 = Entry(line_frame14, textvariable = dr14, font = fonts, relief = SUNKEN, bd = 2, width = 14, justify = "right")
        debit_entry14.grid(column = 2, row = 0)
        dr14.trace("w", lambda nm, idx, mode, var14 = dr14: self.validate_float14(var14))
        debit_entry14.bind("<Tab>", self.validate_comma_dr14)
                
    def showcredit14(self):
        global cr14
        cr14 = StringVar()
        global credit_entry14
        credit_entry14 = Entry(line_frame14, textvariable = cr14, font = fonts, relief = SUNKEN, bd = 2, width = 14, justify = "right")
        credit_entry14.grid(column = 3, row = 0)
        cr14.trace("w", lambda nm, idx, mode, var14 = cr14: self.validate_float14(var14)) 
        credit_entry14.bind("<Tab>", self.validate_comma_cr14)
        
    def showcomment14(self):  
        global comment14
        comment14 = StringVar()
        global comment_entry14
        comment_entry14 = Entry(line_frame14, textvariable = comment14, font = fonts, relief = SUNKEN, bd = 2, width = 20, justify = "left")
        comment_entry14.grid(column = 4, row = 0)

    def showrepeat14(self, *args):
        global repeat_button14
        repeat_button14 = Button(line_frame14, text = "clear", bg = buttonbg, font = clearbuttonfont, command = self.cleanentry14, bd =0)
        repeat_button14.grid(column = 5, row = 0, padx = 2)
        repeat_button14.bind("<Return>", self.cleanentry14)
 
    def cleanentry14(self, *args):
        code_entry14.delete(0, END)
        title_label14.config(text = "")
        debit_entry14.delete(0, END)
        credit_entry14.delete(0, END)
        comment_entry14.delete(0, END)

        debit_entry14.config(state = DISABLED)
        credit_entry14.config(state = DISABLED)
        comment_entry14.config(state = DISABLED)
        repeat_button14.config(state = DISABLED)

        self.updatetotal()

### line 15 ###
    def lineframe15(self, master):
        global line_frame15
        line_frame15 = Frame(master, bg = wc)
        line_frame15.grid(column = 0, row = 20, sticky = NW)

    def show_code15(self):
        global journalcode15
        journalcode15 = StringVar()
        global code_entry15
        code_entry15 = Entry(line_frame15, textvariable = journalcode15, font = fonts, relief = SUNKEN, bd = 2, width = 5, justify = "center")
        code_entry15.grid(column = 0, row = 0)
        code_entry15.bind("<Tab>", self.journalcode_val15)
        code_entry15.bind("<FocusOut>", self.journalcode_val15)
        code_entry15.bind("<Return>", self.journalcode_val15)
        
    def journalcode_val15(self, *args):
        global result15
        codeinput15 = journalcode15.get()
        if len(codeinput15) == 4 and codeinput15.isdigit():
            if company == "DBPSC":
                find_code15 = "SELECT title FROM chart WHERE code = ?"
            else:
                find_code15 = "SELECT title FROM dssichart WHERE code = ?"
            c.execute(find_code15, [codeinput15])
            result15 = c.fetchall()
            if result15:
                title_label15.config(text = result15[0][0])
                self.linestatenormal15()
                debit_entry15.focus()
                self.updatetotal()
            else:
                self.linestatedisabled15()
                code_entry15.delete(0, END)
                self.updatetotal()
        else:
            self.linestatedisabled15()
            code_entry15.delete(0, END)
            self.updatetotal()

    def linestatenormal15(self):
        debit_entry15.config(state = NORMAL)
        credit_entry15.config(state = NORMAL)
        comment_entry15.config(state = NORMAL)
        repeat_button15.config(state = NORMAL)

    def linestatedisabled15(self):
        title_label15.config(text = "")
        debit_entry15.delete(0, END)
        credit_entry15.delete(0, END)
        comment_entry15.delete(0, END)

        debit_entry15.config(state = DISABLED)
        credit_entry15.config(state = DISABLED)
        comment_entry15.config(state = DISABLED)
        repeat_button15.config(state = DISABLED)
      
    def showtitle15(self, *args):
        global title_label15
        title_label15 = Label(line_frame15, text = "", font = fonts, relief = SUNKEN, bd = 2, width = 33, anchor = W, bg = codetitlebg, fg = codetitlefg)
        title_label15.grid(column = 1, row = 0)

    def validate_float15(self, var15):
        global new_value15
        new_value15 = var15.get()
        try:
            new_value15 == "" or "," or float(new_value15)
            old_value15 = new_value15
        except:
            old_value15 = ""
            var15.set(old_value15)

    def validate_comma_dr15(self, *args):
        try:
            if debit_entry15.get() == "":
                dr15.set(format(float(0), ',.2f'))
            else:
                comma15 = format(float(dr15.get()), ',.2f')
                dr15.set(comma15)
        except:
            debit_entry15.delete(0, END)
            dr15.set(format(float(0), ',.2f'))
        finally:
            self.updatetotal()
            
    def validate_comma_cr15(self, *args):
        try:
            if credit_entry15.get() == "":
                cr15.set(format(float(0), ',.2f'))
            else:
                comma15 = format(float(cr15.get()), ',.2f')
                cr15.set(comma15)
        except:
            credit_entry15.delete(0, END)
            cr15.set(format(float(0), ',.2f'))
        finally:
            self.updatetotal()
            
    def showdebit15(self, *args):
        global debit_entry15
        global dr15
        dr15 = StringVar()
        debit_entry15 = Entry(line_frame15, textvariable = dr15, font = fonts, relief = SUNKEN, bd = 2, width = 14, justify = "right")
        debit_entry15.grid(column = 2, row = 0)
        dr15.trace("w", lambda nm, idx, mode, var15 = dr15: self.validate_float15(var15))
        debit_entry15.bind("<Tab>", self.validate_comma_dr15)
                
    def showcredit15(self):
        global cr15
        cr15 = StringVar()
        global credit_entry15
        credit_entry15 = Entry(line_frame15, textvariable = cr15, font = fonts, relief = SUNKEN, bd = 2, width = 14, justify = "right")
        credit_entry15.grid(column = 3, row = 0)
        cr15.trace("w", lambda nm, idx, mode, var15 = cr15: self.validate_float15(var15)) 
        credit_entry15.bind("<Tab>", self.validate_comma_cr15)
        
    def showcomment15(self):  
        global comment15
        comment15 = StringVar()
        global comment_entry15
        comment_entry15 = Entry(line_frame15, textvariable = comment15, font = fonts, relief = SUNKEN, bd = 2, width = 20, justify = "left")
        comment_entry15.grid(column = 4, row = 0)

    def showrepeat15(self, *args):
        global repeat_button15
        repeat_button15 = Button(line_frame15, text = "clear", bg = buttonbg, font = clearbuttonfont, command = self.cleanentry15, bd =0)
        repeat_button15.grid(column = 5, row = 0, padx = 2)
        repeat_button15.bind("<Return>", self.cleanentry15)
 
    def cleanentry15(self, *args):
        code_entry15.delete(0, END)
        title_label15.config(text = "")
        debit_entry15.delete(0, END)
        credit_entry15.delete(0, END)
        comment_entry15.delete(0, END)

        debit_entry15.config(state = DISABLED)
        credit_entry15.config(state = DISABLED)
        comment_entry15.config(state = DISABLED)
        repeat_button15.config(state = DISABLED)

        self.updatetotal()

### line 16 ###
    def lineframe16(self, master):
        global line_frame16
        line_frame16 = Frame(master, bg = wc)
        line_frame16.grid(column = 0, row = 21, sticky = NW)

    def show_code16(self):
        global journalcode16
        journalcode16 = StringVar()
        global code_entry16
        code_entry16 = Entry(line_frame16, textvariable = journalcode16, font = fonts, relief = SUNKEN, bd = 2, width = 5, justify = "center")
        code_entry16.grid(column = 0, row = 0)
        code_entry16.bind("<Tab>", self.journalcode_val16)
        code_entry16.bind("<FocusOut>", self.journalcode_val16)
        code_entry16.bind("<Return>", self.journalcode_val16)
        
    def journalcode_val16(self, *args):
        global result16
        codeinput16 = journalcode16.get()
        if len(codeinput16) == 4 and codeinput16.isdigit():
            if company == "DBPSC":
                find_code16 = "SELECT title FROM chart WHERE code = ?"
            else:
                find_code16 = "SELECT title FROM dssichart WHERE code = ?"
            c.execute(find_code16, [codeinput16])
            result16 = c.fetchall()
            if result16:
                title_label16.config(text = result16[0][0])
                self.linestatenormal16()
                debit_entry16.focus()
                self.updatetotal()
            else:
                self.linestatedisabled16()
                code_entry16.delete(0, END)
                self.updatetotal()
        else:
            self.linestatedisabled16()
            code_entry16.delete(0, END)
            self.updatetotal()

    def linestatenormal16(self):
        debit_entry16.config(state = NORMAL)
        credit_entry16.config(state = NORMAL)
        comment_entry16.config(state = NORMAL)
        repeat_button16.config(state = NORMAL)

    def linestatedisabled16(self):
        title_label16.config(text = "")
        debit_entry16.delete(0, END)
        credit_entry16.delete(0, END)
        comment_entry16.delete(0, END)

        debit_entry16.config(state = DISABLED)
        credit_entry16.config(state = DISABLED)
        comment_entry16.config(state = DISABLED)
        repeat_button16.config(state = DISABLED)
      
    def showtitle16(self, *args):
        global title_label16
        title_label16 = Label(line_frame16, text = "", font = fonts, relief = SUNKEN, bd = 2, width = 33, anchor = W, bg = codetitlebg, fg = codetitlefg)
        title_label16.grid(column = 1, row = 0)

    def validate_float16(self, var16):
        global new_value16
        new_value16 = var16.get()
        try:
            new_value16 == "" or "," or float(new_value16)
            old_value16 = new_value16
        except:
            old_value16 = ""
            var16.set(old_value16)

    def validate_comma_dr16(self, *args):
        try:
            if debit_entry16.get() == "":
                dr16.set(format(float(0), ',.2f'))
            else:
                comma16 = format(float(dr16.get()), ',.2f')
                dr16.set(comma16)
        except:
            debit_entry16.delete(0, END)
            dr16.set(format(float(0), ',.2f'))
        finally:
            self.updatetotal()
            
    def validate_comma_cr16(self, *args):
        try:
            if credit_entry16.get() == "":
                cr16.set(format(float(0), ',.2f'))
            else:
                comma16 = format(float(cr16.get()), ',.2f')
                cr16.set(comma16)
        except:
            credit_entry16.delete(0, END)
            cr16.set(format(float(0), ',.2f'))
        finally:
            self.updatetotal()
            
    def showdebit16(self, *args):
        global debit_entry16
        global dr16
        dr16 = StringVar()
        debit_entry16 = Entry(line_frame16, textvariable = dr16, font = fonts, relief = SUNKEN, bd = 2, width = 14, justify = "right")
        debit_entry16.grid(column = 2, row = 0)
        dr16.trace("w", lambda nm, idx, mode, var16 = dr16: self.validate_float16(var16))
        debit_entry16.bind("<Tab>", self.validate_comma_dr16)
                
    def showcredit16(self):
        global cr16
        cr16 = StringVar()
        global credit_entry16
        credit_entry16 = Entry(line_frame16, textvariable = cr16, font = fonts, relief = SUNKEN, bd = 2, width = 14, justify = "right")
        credit_entry16.grid(column = 3, row = 0)
        cr16.trace("w", lambda nm, idx, mode, var16 = cr16: self.validate_float16(var16)) 
        credit_entry16.bind("<Tab>", self.validate_comma_cr16)
        
    def showcomment16(self):  
        global comment16
        comment16 = StringVar()
        global comment_entry16
        comment_entry16 = Entry(line_frame16, textvariable = comment16, font = fonts, relief = SUNKEN, bd = 2, width = 20, justify = "left")
        comment_entry16.grid(column = 4, row = 0)

    def showrepeat16(self, *args):
        global repeat_button16
        repeat_button16 = Button(line_frame16, text = "clear", bg = buttonbg, font = clearbuttonfont, command = self.cleanentry16, bd =0)
        repeat_button16.grid(column = 5, row = 0, padx = 2)
        repeat_button16.bind("<Return>", self.cleanentry16)
 
    def cleanentry16(self, *args):
        code_entry16.delete(0, END)
        title_label16.config(text = "")
        debit_entry16.delete(0, END)
        credit_entry16.delete(0, END)
        comment_entry16.delete(0, END)

        debit_entry16.config(state = DISABLED)
        credit_entry16.config(state = DISABLED)
        comment_entry16.config(state = DISABLED)
        repeat_button16.config(state = DISABLED)

        self.updatetotal()

### line 17 ###
    def lineframe17(self, master):
        global line_frame17
        line_frame17 = Frame(master, bg = wc)
        line_frame17.grid(column = 0, row = 22, sticky = NW)

    def show_code17(self):
        global journalcode17
        journalcode17 = StringVar()
        global code_entry17
        code_entry17 = Entry(line_frame17, textvariable = journalcode17, font = fonts, relief = SUNKEN, bd = 2, width = 5, justify = "center")
        code_entry17.grid(column = 0, row = 0)
        code_entry17.bind("<Tab>", self.journalcode_val17)
        code_entry17.bind("<FocusOut>", self.journalcode_val17)
        code_entry17.bind("<Return>", self.journalcode_val17)
        
    def journalcode_val17(self, *args):
        global result17
        codeinput17 = journalcode17.get()
        if len(codeinput17) == 4 and codeinput17.isdigit():
            if company == "DBPSC":
                find_code17 = "SELECT title FROM chart WHERE code = ?"
            else:
                find_code17 = "SELECT title FROM dssichart WHERE code = ?"
            c.execute(find_code17, [codeinput17])
            result17 = c.fetchall()
            if result17:
                title_label17.config(text = result17[0][0])
                self.linestatenormal17()
                debit_entry17.focus()
                self.updatetotal()
            else:
                self.linestatedisabled17()
                code_entry17.delete(0, END)
                self.updatetotal()
        else:
            self.linestatedisabled17()
            code_entry17.delete(0, END)
            self.updatetotal()

    def linestatenormal17(self):
        debit_entry17.config(state = NORMAL)
        credit_entry17.config(state = NORMAL)
        comment_entry17.config(state = NORMAL)
        repeat_button17.config(state = NORMAL)

    def linestatedisabled17(self):
        title_label17.config(text = "")
        debit_entry17.delete(0, END)
        credit_entry17.delete(0, END)
        comment_entry17.delete(0, END)

        debit_entry17.config(state = DISABLED)
        credit_entry17.config(state = DISABLED)
        comment_entry17.config(state = DISABLED)
        repeat_button17.config(state = DISABLED)
      
    def showtitle17(self, *args):
        global title_label17
        title_label17 = Label(line_frame17, text = "", font = fonts, relief = SUNKEN, bd = 2, width = 33, anchor = W, bg = codetitlebg, fg = codetitlefg)
        title_label17.grid(column = 1, row = 0)

    def validate_float17(self, var17):
        global new_value17
        new_value17 = var17.get()
        try:
            new_value17 == "" or "," or float(new_value17)
            old_value17 = new_value17
        except:
            old_value17 = ""
            var17.set(old_value17)

    def validate_comma_dr17(self, *args):
        try:
            if debit_entry17.get() == "":
                dr17.set(format(float(0), ',.2f'))
            else:
                comma17 = format(float(dr17.get()), ',.2f')
                dr17.set(comma17)
        except:
            debit_entry17.delete(0, END)
            dr17.set(format(float(0), ',.2f'))
        finally:
            self.updatetotal()
            
    def validate_comma_cr17(self, *args):
        try:
            if credit_entry17.get() == "":
                cr17.set(format(float(0), ',.2f'))
            else:
                comma17 = format(float(cr17.get()), ',.2f')
                cr17.set(comma17)
        except:
            credit_entry17.delete(0, END)
            cr17.set(format(float(0), ',.2f'))
        finally:
            self.updatetotal()
            
    def showdebit17(self, *args):
        global debit_entry17
        global dr17
        dr17 = StringVar()
        debit_entry17 = Entry(line_frame17, textvariable = dr17, font = fonts, relief = SUNKEN, bd = 2, width = 14, justify = "right")
        debit_entry17.grid(column = 2, row = 0)
        dr17.trace("w", lambda nm, idx, mode, var17 = dr17: self.validate_float17(var17))
        debit_entry17.bind("<Tab>", self.validate_comma_dr17)
                
    def showcredit17(self):
        global cr17
        cr17 = StringVar()
        global credit_entry17
        credit_entry17 = Entry(line_frame17, textvariable = cr17, font = fonts, relief = SUNKEN, bd = 2, width = 14, justify = "right")
        credit_entry17.grid(column = 3, row = 0)
        cr17.trace("w", lambda nm, idx, mode, var17 = cr17: self.validate_float17(var17)) 
        credit_entry17.bind("<Tab>", self.validate_comma_cr17)
        
    def showcomment17(self):  
        global comment17
        comment17 = StringVar()
        global comment_entry17
        comment_entry17 = Entry(line_frame17, textvariable = comment17, font = fonts, relief = SUNKEN, bd = 2, width = 20, justify = "left")
        comment_entry17.grid(column = 4, row = 0)

    def showrepeat17(self, *args):
        global repeat_button17
        repeat_button17 = Button(line_frame17, text = "clear", bg = buttonbg, font = clearbuttonfont, command = self.cleanentry17, bd =0)
        repeat_button17.grid(column = 5, row = 0, padx = 2)
        repeat_button17.bind("<Return>", self.cleanentry17)
 
    def cleanentry17(self, *args):
        code_entry17.delete(0, END)
        title_label17.config(text = "")
        debit_entry17.delete(0, END)
        credit_entry17.delete(0, END)
        comment_entry17.delete(0, END)

        debit_entry17.config(state = DISABLED)
        credit_entry17.config(state = DISABLED)
        comment_entry17.config(state = DISABLED)
        repeat_button17.config(state = DISABLED)

        self.updatetotal()

### line 18 ###
    def lineframe18(self, master):
        global line_frame18
        line_frame18 = Frame(master, bg = wc)
        line_frame18.grid(column = 0, row = 23, sticky = NW)

    def show_code18(self):
        global journalcode18
        journalcode18 = StringVar()
        global code_entry18
        code_entry18 = Entry(line_frame18, textvariable = journalcode18, font = fonts, relief = SUNKEN, bd = 2, width = 5, justify = "center")
        code_entry18.grid(column = 0, row = 0)
        code_entry18.bind("<Tab>", self.journalcode_val18)
        code_entry18.bind("<FocusOut>", self.journalcode_val18)
        code_entry18.bind("<Return>", self.journalcode_val18)
        
    def journalcode_val18(self, *args):
        global result18
        codeinput18 = journalcode18.get()
        if len(codeinput18) == 4 and codeinput18.isdigit():
            if company == "DBPSC":
                find_code18 = "SELECT title FROM chart WHERE code = ?"
            else:
                find_code18 = "SELECT title FROM dssichart WHERE code = ?"
            c.execute(find_code18, [codeinput18])
            result18 = c.fetchall()
            if result18:
                title_label18.config(text = result18[0][0])
                self.linestatenormal18()
                debit_entry18.focus()
                self.updatetotal()
            else:
                self.linestatedisabled18()
                code_entry18.delete(0, END)
                self.updatetotal()
        else:
            self.linestatedisabled18()
            code_entry18.delete(0, END)
            self.updatetotal()

    def linestatenormal18(self):
        debit_entry18.config(state = NORMAL)
        credit_entry18.config(state = NORMAL)
        comment_entry18.config(state = NORMAL)
        repeat_button18.config(state = NORMAL)

    def linestatedisabled18(self):
        title_label18.config(text = "")
        debit_entry18.delete(0, END)
        credit_entry18.delete(0, END)
        comment_entry18.delete(0, END)

        debit_entry18.config(state = DISABLED)
        credit_entry18.config(state = DISABLED)
        comment_entry18.config(state = DISABLED)
        repeat_button18.config(state = DISABLED)
      
    def showtitle18(self, *args):
        global title_label18
        title_label18 = Label(line_frame18, text = "", font = fonts, relief = SUNKEN, bd = 2, width = 33, anchor = W, bg = codetitlebg, fg = codetitlefg)
        title_label18.grid(column = 1, row = 0)

    def validate_float18(self, var18):
        global new_value18
        new_value18 = var18.get()
        try:
            new_value18 == "" or "," or float(new_value18)
            old_value18 = new_value18
        except:
            old_value18 = ""
            var18.set(old_value18)

    def validate_comma_dr18(self, *args):
        try:
            if debit_entry18.get() == "":
                dr18.set(format(float(0), ',.2f'))
            else:
                comma18 = format(float(dr18.get()), ',.2f')
                dr18.set(comma18)
        except:
            debit_entry18.delete(0, END)
            dr18.set(format(float(0), ',.2f'))
        finally:
            self.updatetotal()
            
    def validate_comma_cr18(self, *args):
        try:
            if credit_entry18.get() == "":
                cr18.set(format(float(0), ',.2f'))
            else:
                comma18 = format(float(cr18.get()), ',.2f')
                cr18.set(comma18)
        except:
            credit_entry18.delete(0, END)
            cr18.set(format(float(0), ',.2f'))
        finally:
            self.updatetotal()
            
    def showdebit18(self, *args):
        global debit_entry18
        global dr18
        dr18 = StringVar()
        debit_entry18 = Entry(line_frame18, textvariable = dr18, font = fonts, relief = SUNKEN, bd = 2, width = 14, justify = "right")
        debit_entry18.grid(column = 2, row = 0)
        dr18.trace("w", lambda nm, idx, mode, var18 = dr18: self.validate_float18(var18))
        debit_entry18.bind("<Tab>", self.validate_comma_dr18)
                
    def showcredit18(self):
        global cr18
        cr18 = StringVar()
        global credit_entry18
        credit_entry18 = Entry(line_frame18, textvariable = cr18, font = fonts, relief = SUNKEN, bd = 2, width = 14, justify = "right")
        credit_entry18.grid(column = 3, row = 0)
        cr18.trace("w", lambda nm, idx, mode, var18 = cr18: self.validate_float18(var18)) 
        credit_entry18.bind("<Tab>", self.validate_comma_cr18)
        
    def showcomment18(self):  
        global comment18
        comment18 = StringVar()
        global comment_entry18
        comment_entry18 = Entry(line_frame18, textvariable = comment18, font = fonts, relief = SUNKEN, bd = 2, width = 20, justify = "left")
        comment_entry18.grid(column = 4, row = 0)

    def showrepeat18(self, *args):
        global repeat_button18
        repeat_button18 = Button(line_frame18, text = "clear", bg = buttonbg, font = clearbuttonfont, command = self.cleanentry18, bd =0)
        repeat_button18.grid(column = 5, row = 0, padx = 2)
        repeat_button18.bind("<Return>", self.cleanentry18)
 
    def cleanentry18(self, *args):
        code_entry18.delete(0, END)
        title_label18.config(text = "")
        debit_entry18.delete(0, END)
        credit_entry18.delete(0, END)
        comment_entry18.delete(0, END)

        debit_entry18.config(state = DISABLED)
        credit_entry18.config(state = DISABLED)
        comment_entry18.config(state = DISABLED)
        repeat_button18.config(state = DISABLED)

        self.updatetotal()

### line 19 ###
    def lineframe19(self, master):
        global line_frame19
        line_frame19 = Frame(master, bg = wc)
        line_frame19.grid(column = 0, row = 24, sticky = NW)

    def show_code19(self):
        global journalcode19
        journalcode19 = StringVar()
        global code_entry19
        code_entry19 = Entry(line_frame19, textvariable = journalcode19, font = fonts, relief = SUNKEN, bd = 2, width = 5, justify = "center")
        code_entry19.grid(column = 0, row = 0)
        code_entry19.bind("<Tab>", self.journalcode_val19)
        code_entry19.bind("<FocusOut>", self.journalcode_val19)
        code_entry19.bind("<Return>", self.journalcode_val19)
        
    def journalcode_val19(self, *args):
        global result19
        codeinput19 = journalcode19.get()
        if len(codeinput19) == 4 and codeinput19.isdigit():
            if company == "DBPSC":
                find_code19 = "SELECT title FROM chart WHERE code = ?"
            else:
                find_code19 = "SELECT title FROM dssichart WHERE code = ?"
            c.execute(find_code19, [codeinput19])
            result19 = c.fetchall()
            if result19:
                title_label19.config(text = result19[0][0])
                self.linestatenormal19()
                debit_entry19.focus()
                self.updatetotal()
            else:
                self.linestatedisabled19()
                code_entry19.delete(0, END)
                self.updatetotal()
        else:
            self.linestatedisabled19()
            code_entry19.delete(0, END)
            self.updatetotal()

    def linestatenormal19(self):
        debit_entry19.config(state = NORMAL)
        credit_entry19.config(state = NORMAL)
        comment_entry19.config(state = NORMAL)
        repeat_button19.config(state = NORMAL)

    def linestatedisabled19(self):
        title_label19.config(text = "")
        debit_entry19.delete(0, END)
        credit_entry19.delete(0, END)
        comment_entry19.delete(0, END)

        debit_entry19.config(state = DISABLED)
        credit_entry19.config(state = DISABLED)
        comment_entry19.config(state = DISABLED)
        repeat_button19.config(state = DISABLED)
      
    def showtitle19(self, *args):
        global title_label19
        title_label19 = Label(line_frame19, text = "", font = fonts, relief = SUNKEN, bd = 2, width = 33, anchor = W, bg = codetitlebg, fg = codetitlefg)
        title_label19.grid(column = 1, row = 0)

    def validate_float19(self, var19):
        global new_value19
        new_value19 = var19.get()
        try:
            new_value19 == "" or "," or float(new_value19)
            old_value19 = new_value19
        except:
            old_value19 = ""
            var19.set(old_value19)

    def validate_comma_dr19(self, *args):
        try:
            if debit_entry19.get() == "":
                dr19.set(format(float(0), ',.2f'))
            else:
                comma19 = format(float(dr19.get()), ',.2f')
                dr19.set(comma19)
        except:
            debit_entry19.delete(0, END)
            dr19.set(format(float(0), ',.2f'))
        finally:
            self.updatetotal()
            
    def validate_comma_cr19(self, *args):
        try:
            if credit_entry19.get() == "":
                cr19.set(format(float(0), ',.2f'))
            else:
                comma19 = format(float(cr19.get()), ',.2f')
                cr19.set(comma19)
        except:
            credit_entry19.delete(0, END)
            cr19.set(format(float(0), ',.2f'))
        finally:
            self.updatetotal()
            
    def showdebit19(self, *args):
        global debit_entry19
        global dr19
        dr19 = StringVar()
        debit_entry19 = Entry(line_frame19, textvariable = dr19, font = fonts, relief = SUNKEN, bd = 2, width = 14, justify = "right")
        debit_entry19.grid(column = 2, row = 0)
        dr19.trace("w", lambda nm, idx, mode, var19 = dr19: self.validate_float19(var19))
        debit_entry19.bind("<Tab>", self.validate_comma_dr19)
                
    def showcredit19(self):
        global cr19
        cr19 = StringVar()
        global credit_entry19
        credit_entry19 = Entry(line_frame19, textvariable = cr19, font = fonts, relief = SUNKEN, bd = 2, width = 14, justify = "right")
        credit_entry19.grid(column = 3, row = 0)
        cr19.trace("w", lambda nm, idx, mode, var19 = cr19: self.validate_float19(var19)) 
        credit_entry19.bind("<Tab>", self.validate_comma_cr19)
        
    def showcomment19(self):  
        global comment19
        comment19 = StringVar()
        global comment_entry19
        comment_entry19 = Entry(line_frame19, textvariable = comment19, font = fonts, relief = SUNKEN, bd = 2, width = 20, justify = "left")
        comment_entry19.grid(column = 4, row = 0)

    def showrepeat19(self, *args):
        global repeat_button19
        repeat_button19 = Button(line_frame19, text = "clear", bg = buttonbg, font = clearbuttonfont, command = self.cleanentry19, bd =0)
        repeat_button19.grid(column = 5, row = 0, padx = 2)
        repeat_button19.bind("<Return>", self.cleanentry19)
 
    def cleanentry19(self, *args):
        code_entry19.delete(0, END)
        title_label19.config(text = "")
        debit_entry19.delete(0, END)
        credit_entry19.delete(0, END)
        comment_entry19.delete(0, END)

        debit_entry19.config(state = DISABLED)
        credit_entry19.config(state = DISABLED)
        comment_entry19.config(state = DISABLED)
        repeat_button19.config(state = DISABLED)

        self.updatetotal()

### line 20 ###
    def lineframe20(self, master):
        global line_frame20
        line_frame20 = Frame(master, bg = wc)
        line_frame20.grid(column = 0, row = 25, sticky = NW)

    def show_code20(self):
        global journalcode20
        journalcode20 = StringVar()
        global code_entry20
        code_entry20 = Entry(line_frame20, textvariable = journalcode20, font = fonts, relief = SUNKEN, bd = 2, width = 5, justify = "center")
        code_entry20.grid(column = 0, row = 0)
        code_entry20.bind("<Tab>", self.journalcode_val20)
        code_entry20.bind("<FocusOut>", self.journalcode_val20)
        code_entry20.bind("<Return>", self.journalcode_val20)
        
    def journalcode_val20(self, *args):
        global result20
        codeinput20 = journalcode20.get()
        if len(codeinput20) == 4 and codeinput20.isdigit():
            if company == "DBPSC":
                find_code20 = "SELECT title FROM chart WHERE code = ?"
            else:
                find_code20 = "SELECT title FROM dssichart WHERE code = ?"
            c.execute(find_code20, [codeinput20])
            result20 = c.fetchall()
            if result20:
                title_label20.config(text = result20[0][0])
                self.linestatenormal20()
                debit_entry20.focus()
                self.updatetotal()
            else:
                self.linestatedisabled20()
                code_entry20.delete(0, END)
                self.updatetotal()
        else:
            self.linestatedisabled20()
            code_entry20.delete(0, END)
            self.updatetotal()

    def linestatenormal20(self):
        debit_entry20.config(state = NORMAL)
        credit_entry20.config(state = NORMAL)
        comment_entry20.config(state = NORMAL)
        repeat_button20.config(state = NORMAL)

    def linestatedisabled20(self):
        title_label20.config(text = "")
        debit_entry20.delete(0, END)
        credit_entry20.delete(0, END)
        comment_entry20.delete(0, END)

        debit_entry20.config(state = DISABLED)
        credit_entry20.config(state = DISABLED)
        comment_entry20.config(state = DISABLED)
        repeat_button20.config(state = DISABLED)
      
    def showtitle20(self, *args):
        global title_label20
        title_label20 = Label(line_frame20, text = "", font = fonts, relief = SUNKEN, bd = 2, width = 33, anchor = W, bg = codetitlebg, fg = codetitlefg)
        title_label20.grid(column = 1, row = 0)

    def validate_float20(self, var20):
        global new_value20
        new_value20 = var20.get()
        try:
            new_value20 == "" or "," or float(new_value20)
            old_value20 = new_value20
        except:
            old_value20 = ""
            var20.set(old_value20)

    def validate_comma_dr20(self, *args):
        try:
            if debit_entry20.get() == "":
                dr20.set(format(float(0), ',.2f'))
            else:
                comma20 = format(float(dr20.get()), ',.2f')
                dr20.set(comma20)
        except:
            debit_entry20.delete(0, END)
            dr20.set(format(float(0), ',.2f'))
        finally:
            self.updatetotal()
            
    def validate_comma_cr20(self, *args):
        try:
            if credit_entry20.get() == "":
                cr20.set(format(float(0), ',.2f'))
            else:
                comma20 = format(float(cr20.get()), ',.2f')
                cr20.set(comma20)
        except:
            credit_entry20.delete(0, END)
            cr20.set(format(float(0), ',.2f'))
        finally:
            self.updatetotal()
            
    def showdebit20(self, *args):
        global debit_entry20
        global dr20
        dr20 = StringVar()
        debit_entry20 = Entry(line_frame20, textvariable = dr20, font = fonts, relief = SUNKEN, bd = 2, width = 14, justify = "right")
        debit_entry20.grid(column = 2, row = 0)
        dr20.trace("w", lambda nm, idx, mode, var20 = dr20: self.validate_float20(var20))
        debit_entry20.bind("<Tab>", self.validate_comma_dr20)
                
    def showcredit20(self):
        global cr20
        cr20 = StringVar()
        global credit_entry20
        credit_entry20 = Entry(line_frame20, textvariable = cr20, font = fonts, relief = SUNKEN, bd = 2, width = 14, justify = "right")
        credit_entry20.grid(column = 3, row = 0)
        cr20.trace("w", lambda nm, idx, mode, var20 = cr20: self.validate_float20(var20)) 
        credit_entry20.bind("<Tab>", self.validate_comma_cr20)
        
    def showcomment20(self):  
        global comment20
        comment20 = StringVar()
        global comment_entry20
        comment_entry20 = Entry(line_frame20, textvariable = comment20, font = fonts, relief = SUNKEN, bd = 2, width = 20, justify = "left")
        comment_entry20.grid(column = 4, row = 0)

    def showrepeat20(self, *args):
        global repeat_button20
        repeat_button20 = Button(line_frame20, text = "clear", bg = buttonbg, font = clearbuttonfont, command = self.cleanentry20, bd =0)
        repeat_button20.grid(column = 5, row = 0, padx = 2)
        repeat_button20.bind("<Return>", self.cleanentry20)
 
    def cleanentry20(self, *args):
        code_entry20.delete(0, END)
        title_label20.config(text = "")
        debit_entry20.delete(0, END)
        credit_entry20.delete(0, END)
        comment_entry20.delete(0, END)

        debit_entry20.config(state = DISABLED)
        credit_entry20.config(state = DISABLED)
        comment_entry20.config(state = DISABLED)
        repeat_button20.config(state = DISABLED)

        self.updatetotal()

### line 21 ###
    def lineframe21(self, master):
        global line_frame21
        line_frame21 = Frame(master, bg = wc)
        line_frame21.grid(column = 0, row = 26, sticky = NW)

    def show_code21(self):
        global journalcode21
        journalcode21 = StringVar()
        global code_entry21
        code_entry21 = Entry(line_frame21, textvariable = journalcode21, font = fonts, relief = SUNKEN, bd = 2, width = 5, justify = "center")
        code_entry21.grid(column = 0, row = 0)
        code_entry21.bind("<Tab>", self.journalcode_val21)
        code_entry21.bind("<FocusOut>", self.journalcode_val21)
        code_entry21.bind("<Return>", self.journalcode_val21)
        
    def journalcode_val21(self, *args):
        global result21
        codeinput21 = journalcode21.get()
        if len(codeinput21) == 4 and codeinput21.isdigit():
            if company == "DBPSC":
                find_code21 = "SELECT title FROM chart WHERE code = ?"
            else:
                find_code21 = "SELECT title FROM dssichart WHERE code = ?"
            c.execute(find_code21, [codeinput21])
            result21 = c.fetchall()
            if result21:
                title_label21.config(text = result21[0][0])
                self.linestatenormal21()
                debit_entry21.focus()
                self.updatetotal()
            else:
                self.linestatedisabled21()
                code_entry21.delete(0, END)
                self.updatetotal()
        else:
            self.linestatedisabled21()
            code_entry21.delete(0, END)
            self.updatetotal()

    def linestatenormal21(self):
        debit_entry21.config(state = NORMAL)
        credit_entry21.config(state = NORMAL)
        comment_entry21.config(state = NORMAL)
        repeat_button21.config(state = NORMAL)

    def linestatedisabled21(self):
        title_label21.config(text = "")
        debit_entry21.delete(0, END)
        credit_entry21.delete(0, END)
        comment_entry21.delete(0, END)

        debit_entry21.config(state = DISABLED)
        credit_entry21.config(state = DISABLED)
        comment_entry21.config(state = DISABLED)
        repeat_button21.config(state = DISABLED)
      
    def showtitle21(self, *args):
        global title_label21
        title_label21 = Label(line_frame21, text = "", font = fonts, relief = SUNKEN, bd = 2, width = 33, anchor = W, bg = codetitlebg, fg = codetitlefg)
        title_label21.grid(column = 1, row = 0)

    def validate_float21(self, var21):
        global new_value21
        new_value21 = var21.get()
        try:
            new_value21 == "" or "," or float(new_value21)
            old_value21 = new_value21
        except:
            old_value21 = ""
            var21.set(old_value21)

    def validate_comma_dr21(self, *args):
        try:
            if debit_entry21.get() == "":
                dr21.set(format(float(0), ',.2f'))
            else:
                comma21 = format(float(dr21.get()), ',.2f')
                dr21.set(comma21)
        except:
            debit_entry21.delete(0, END)
            dr21.set(format(float(0), ',.2f'))
        finally:
            self.updatetotal()
            
    def validate_comma_cr21(self, *args):
        try:
            if credit_entry21.get() == "":
                cr21.set(format(float(0), ',.2f'))
            else:
                comma21 = format(float(cr21.get()), ',.2f')
                cr21.set(comma21)
        except:
            credit_entry21.delete(0, END)
            cr21.set(format(float(0), ',.2f'))
        finally:
            self.updatetotal()
            
    def showdebit21(self, *args):
        global debit_entry21
        global dr21
        dr21 = StringVar()
        debit_entry21 = Entry(line_frame21, textvariable = dr21, font = fonts, relief = SUNKEN, bd = 2, width = 14, justify = "right")
        debit_entry21.grid(column = 2, row = 0)
        dr21.trace("w", lambda nm, idx, mode, var21 = dr21: self.validate_float21(var21))
        debit_entry21.bind("<Tab>", self.validate_comma_dr21)
                
    def showcredit21(self):
        global cr21
        cr21 = StringVar()
        global credit_entry21
        credit_entry21 = Entry(line_frame21, textvariable = cr21, font = fonts, relief = SUNKEN, bd = 2, width = 14, justify = "right")
        credit_entry21.grid(column = 3, row = 0)
        cr21.trace("w", lambda nm, idx, mode, var21 = cr21: self.validate_float21(var21)) 
        credit_entry21.bind("<Tab>", self.validate_comma_cr21)
        
    def showcomment21(self):  
        global comment21
        comment21 = StringVar()
        global comment_entry21
        comment_entry21 = Entry(line_frame21, textvariable = comment21, font = fonts, relief = SUNKEN, bd = 2, width = 20, justify = "left")
        comment_entry21.grid(column = 4, row = 0)

    def showrepeat21(self, *args):
        global repeat_button21
        repeat_button21 = Button(line_frame21, text = "clear", bg = buttonbg, font = clearbuttonfont, command = self.cleanentry21, bd =0)
        repeat_button21.grid(column = 5, row = 0, padx = 2)
        repeat_button21.bind("<Return>", self.cleanentry21)
 
    def cleanentry21(self, *args):
        code_entry21.delete(0, END)
        title_label21.config(text = "")
        debit_entry21.delete(0, END)
        credit_entry21.delete(0, END)
        comment_entry21.delete(0, END)

        debit_entry21.config(state = DISABLED)
        credit_entry21.config(state = DISABLED)
        comment_entry21.config(state = DISABLED)
        repeat_button21.config(state = DISABLED)

        self.updatetotal()

### line 22 ###
    def lineframe22(self, master):
        global line_frame22
        line_frame22 = Frame(master, bg = wc)
        line_frame22.grid(column = 0, row = 27, sticky = NW)

    def show_code22(self):
        global journalcode22
        journalcode22 = StringVar()
        global code_entry22
        code_entry22 = Entry(line_frame22, textvariable = journalcode22, font = fonts, relief = SUNKEN, bd = 2, width = 5, justify = "center")
        code_entry22.grid(column = 0, row = 0)
        code_entry22.bind("<Tab>", self.journalcode_val22)
        code_entry22.bind("<FocusOut>", self.journalcode_val22)
        code_entry22.bind("<Return>", self.journalcode_val22)
        
    def journalcode_val22(self, *args):
        global result22
        codeinput22 = journalcode22.get()
        if len(codeinput22) == 4 and codeinput22.isdigit():
            if company == "DBPSC":
                find_code22 = "SELECT title FROM chart WHERE code = ?"
            else:
                find_code22 = "SELECT title FROM dssichart WHERE code = ?"
            c.execute(find_code22, [codeinput22])
            result22 = c.fetchall()
            if result22:
                title_label22.config(text = result22[0][0])
                self.linestatenormal22()
                debit_entry22.focus()
                self.updatetotal()
            else:
                self.linestatedisabled22()
                code_entry22.delete(0, END)
                self.updatetotal()
        else:
            self.linestatedisabled22()
            code_entry22.delete(0, END)
            self.updatetotal()

    def linestatenormal22(self):
        debit_entry22.config(state = NORMAL)
        credit_entry22.config(state = NORMAL)
        comment_entry22.config(state = NORMAL)
        repeat_button22.config(state = NORMAL)

    def linestatedisabled22(self):
        title_label22.config(text = "")
        debit_entry22.delete(0, END)
        credit_entry22.delete(0, END)
        comment_entry22.delete(0, END)

        debit_entry22.config(state = DISABLED)
        credit_entry22.config(state = DISABLED)
        comment_entry22.config(state = DISABLED)
        repeat_button22.config(state = DISABLED)
      
    def showtitle22(self, *args):
        global title_label22
        title_label22 = Label(line_frame22, text = "", font = fonts, relief = SUNKEN, bd = 2, width = 33, anchor = W, bg = codetitlebg, fg = codetitlefg)
        title_label22.grid(column = 1, row = 0)

    def validate_float22(self, var22):
        global new_value22
        new_value22 = var22.get()
        try:
            new_value22 == "" or "," or float(new_value22)
            old_value22 = new_value22
        except:
            old_value22 = ""
            var22.set(old_value22)

    def validate_comma_dr22(self, *args):
        try:
            if debit_entry22.get() == "":
                dr22.set(format(float(0), ',.2f'))
            else:
                comma22 = format(float(dr22.get()), ',.2f')
                dr22.set(comma22)
        except:
            debit_entry22.delete(0, END)
            dr22.set(format(float(0), ',.2f'))
        finally:
            self.updatetotal()
            
    def validate_comma_cr22(self, *args):
        try:
            if credit_entry22.get() == "":
                cr22.set(format(float(0), ',.2f'))
            else:
                comma22 = format(float(cr22.get()), ',.2f')
                cr22.set(comma22)
        except:
            credit_entry22.delete(0, END)
            cr22.set(format(float(0), ',.2f'))
        finally:
            self.updatetotal()
            
    def showdebit22(self, *args):
        global debit_entry22
        global dr22
        dr22 = StringVar()
        debit_entry22 = Entry(line_frame22, textvariable = dr22, font = fonts, relief = SUNKEN, bd = 2, width = 14, justify = "right")
        debit_entry22.grid(column = 2, row = 0)
        dr22.trace("w", lambda nm, idx, mode, var22 = dr22: self.validate_float22(var22))
        debit_entry22.bind("<Tab>", self.validate_comma_dr22)
                
    def showcredit22(self):
        global cr22
        cr22 = StringVar()
        global credit_entry22
        credit_entry22 = Entry(line_frame22, textvariable = cr22, font = fonts, relief = SUNKEN, bd = 2, width = 14, justify = "right")
        credit_entry22.grid(column = 3, row = 0)
        cr22.trace("w", lambda nm, idx, mode, var22 = cr22: self.validate_float22(var22)) 
        credit_entry22.bind("<Tab>", self.validate_comma_cr22)
        
    def showcomment22(self):  
        global comment22
        comment22 = StringVar()
        global comment_entry22
        comment_entry22 = Entry(line_frame22, textvariable = comment22, font = fonts, relief = SUNKEN, bd = 2, width = 20, justify = "left")
        comment_entry22.grid(column = 4, row = 0)

    def showrepeat22(self, *args):
        global repeat_button22
        repeat_button22 = Button(line_frame22, text = "clear", bg = buttonbg, font = clearbuttonfont, command = self.cleanentry22, bd =0)
        repeat_button22.grid(column = 5, row = 0, padx = 2)
        repeat_button22.bind("<Return>", self.cleanentry22)
 
    def cleanentry22(self, *args):
        code_entry22.delete(0, END)
        title_label22.config(text = "")
        debit_entry22.delete(0, END)
        credit_entry22.delete(0, END)
        comment_entry22.delete(0, END)

        debit_entry22.config(state = DISABLED)
        credit_entry22.config(state = DISABLED)
        comment_entry22.config(state = DISABLED)
        repeat_button22.config(state = DISABLED)

        self.updatetotal()

### line 23 ###
    def lineframe23(self, master):
        global line_frame23
        line_frame23 = Frame(master, bg = wc)
        line_frame23.grid(column = 0, row = 28, sticky = NW)

    def show_code23(self):
        global journalcode23
        journalcode23 = StringVar()
        global code_entry23
        code_entry23 = Entry(line_frame23, textvariable = journalcode23, font = fonts, relief = SUNKEN, bd = 2, width = 5, justify = "center")
        code_entry23.grid(column = 0, row = 0)
        code_entry23.bind("<Tab>", self.journalcode_val23)
        code_entry23.bind("<FocusOut>", self.journalcode_val23)
        code_entry23.bind("<Return>", self.journalcode_val23)
        
    def journalcode_val23(self, *args):
        global result23
        codeinput23 = journalcode23.get()
        if len(codeinput23) == 4 and codeinput23.isdigit():
            if company == "DBPSC":
                find_code23 = "SELECT title FROM chart WHERE code = ?"
            else:
                find_code23 = "SELECT title FROM dssichart WHERE code = ?"
            c.execute(find_code23, [codeinput23])
            result23 = c.fetchall()
            if result23:
                title_label23.config(text = result23[0][0])
                self.linestatenormal23()
                debit_entry23.focus()
                self.updatetotal()
            else:
                self.linestatedisabled23()
                code_entry23.delete(0, END)
                self.updatetotal()
        else:
            self.linestatedisabled23()
            code_entry23.delete(0, END)
            self.updatetotal()

    def linestatenormal23(self):
        debit_entry23.config(state = NORMAL)
        credit_entry23.config(state = NORMAL)
        comment_entry23.config(state = NORMAL)
        repeat_button23.config(state = NORMAL)

    def linestatedisabled23(self):
        title_label23.config(text = "")
        debit_entry23.delete(0, END)
        credit_entry23.delete(0, END)
        comment_entry23.delete(0, END)

        debit_entry23.config(state = DISABLED)
        credit_entry23.config(state = DISABLED)
        comment_entry23.config(state = DISABLED)
        repeat_button23.config(state = DISABLED)
      
    def showtitle23(self, *args):
        global title_label23
        title_label23 = Label(line_frame23, text = "", font = fonts, relief = SUNKEN, bd = 2, width = 33, anchor = W, bg = codetitlebg, fg = codetitlefg)
        title_label23.grid(column = 1, row = 0)

    def validate_float23(self, var23):
        global new_value23
        new_value23 = var23.get()
        try:
            new_value23 == "" or "," or float(new_value23)
            old_value23 = new_value23
        except:
            old_value23 = ""
            var23.set(old_value23)

    def validate_comma_dr23(self, *args):
        try:
            if debit_entry23.get() == "":
                dr23.set(format(float(0), ',.2f'))
            else:
                comma23 = format(float(dr23.get()), ',.2f')
                dr23.set(comma23)
        except:
            debit_entry23.delete(0, END)
            dr23.set(format(float(0), ',.2f'))
        finally:
            self.updatetotal()
            
    def validate_comma_cr23(self, *args):
        try:
            if credit_entry23.get() == "":
                cr23.set(format(float(0), ',.2f'))
            else:
                comma23 = format(float(cr23.get()), ',.2f')
                cr23.set(comma23)
        except:
            credit_entry23.delete(0, END)
            cr23.set(format(float(0), ',.2f'))
        finally:
            self.updatetotal()
            
    def showdebit23(self, *args):
        global debit_entry23
        global dr23
        dr23 = StringVar()
        debit_entry23 = Entry(line_frame23, textvariable = dr23, font = fonts, relief = SUNKEN, bd = 2, width = 14, justify = "right")
        debit_entry23.grid(column = 2, row = 0)
        dr23.trace("w", lambda nm, idx, mode, var23 = dr23: self.validate_float23(var23))
        debit_entry23.bind("<Tab>", self.validate_comma_dr23)
                
    def showcredit23(self):
        global cr23
        cr23 = StringVar()
        global credit_entry23
        credit_entry23 = Entry(line_frame23, textvariable = cr23, font = fonts, relief = SUNKEN, bd = 2, width = 14, justify = "right")
        credit_entry23.grid(column = 3, row = 0)
        cr23.trace("w", lambda nm, idx, mode, var23 = cr23: self.validate_float23(var23)) 
        credit_entry23.bind("<Tab>", self.validate_comma_cr23)
        
    def showcomment23(self):  
        global comment23
        comment23 = StringVar()
        global comment_entry23
        comment_entry23 = Entry(line_frame23, textvariable = comment23, font = fonts, relief = SUNKEN, bd = 2, width = 20, justify = "left")
        comment_entry23.grid(column = 4, row = 0)

    def showrepeat23(self, *args):
        global repeat_button23
        repeat_button23 = Button(line_frame23, text = "clear", bg = buttonbg, font = clearbuttonfont, command = self.cleanentry23, bd =0)
        repeat_button23.grid(column = 5, row = 0, padx = 2)
        repeat_button23.bind("<Return>", self.cleanentry23)
 
    def cleanentry23(self, *args):
        code_entry23.delete(0, END)
        title_label23.config(text = "")
        debit_entry23.delete(0, END)
        credit_entry23.delete(0, END)
        comment_entry23.delete(0, END)

        debit_entry23.config(state = DISABLED)
        credit_entry23.config(state = DISABLED)
        comment_entry23.config(state = DISABLED)
        repeat_button23.config(state = DISABLED)

        self.updatetotal()

### line 24 ###
    def lineframe24(self, master):
        global line_frame24
        line_frame24 = Frame(master, bg = wc)
        line_frame24.grid(column = 0, row = 29, sticky = NW)

    def show_code24(self):
        global journalcode24
        journalcode24 = StringVar()
        global code_entry24
        code_entry24 = Entry(line_frame24, textvariable = journalcode24, font = fonts, relief = SUNKEN, bd = 2, width = 5, justify = "center")
        code_entry24.grid(column = 0, row = 0)
        code_entry24.bind("<Tab>", self.journalcode_val24)
        code_entry24.bind("<FocusOut>", self.journalcode_val24)
        code_entry24.bind("<Return>", self.journalcode_val24)
        
    def journalcode_val24(self, *args):
        global result24
        codeinput24 = journalcode24.get()
        if len(codeinput24) == 4 and codeinput24.isdigit():
            if company == "DBPSC":
                find_code24 = "SELECT title FROM chart WHERE code = ?"
            else:
                find_code24 = "SELECT title FROM dssichart WHERE code = ?"
            c.execute(find_code24, [codeinput24])
            result24 = c.fetchall()
            if result24:
                title_label24.config(text = result24[0][0])
                self.linestatenormal24()
                debit_entry24.focus()
                self.updatetotal()
            else:
                self.linestatedisabled24()
                code_entry24.delete(0, END)
                self.updatetotal()
        else:
            self.linestatedisabled24()
            code_entry24.delete(0, END)
            self.updatetotal()

    def linestatenormal24(self):
        debit_entry24.config(state = NORMAL)
        credit_entry24.config(state = NORMAL)
        comment_entry24.config(state = NORMAL)
        repeat_button24.config(state = NORMAL)

    def linestatedisabled24(self):
        title_label24.config(text = "")
        debit_entry24.delete(0, END)
        credit_entry24.delete(0, END)
        comment_entry24.delete(0, END)

        debit_entry24.config(state = DISABLED)
        credit_entry24.config(state = DISABLED)
        comment_entry24.config(state = DISABLED)
        repeat_button24.config(state = DISABLED)
      
    def showtitle24(self, *args):
        global title_label24
        title_label24 = Label(line_frame24, text = "", font = fonts, relief = SUNKEN, bd = 2, width = 33, anchor = W, bg = codetitlebg, fg = codetitlefg)
        title_label24.grid(column = 1, row = 0)

    def validate_float24(self, var24):
        global new_value24
        new_value24 = var24.get()
        try:
            new_value24 == "" or "," or float(new_value24)
            old_value24 = new_value24
        except:
            old_value24 = ""
            var24.set(old_value24)

    def validate_comma_dr24(self, *args):
        try:
            if debit_entry24.get() == "":
                dr24.set(format(float(0), ',.2f'))
            else:
                comma24 = format(float(dr24.get()), ',.2f')
                dr24.set(comma24)
        except:
            debit_entry24.delete(0, END)
            dr24.set(format(float(0), ',.2f'))
        finally:
            self.updatetotal()
            
    def validate_comma_cr24(self, *args):
        try:
            if credit_entry24.get() == "":
                cr24.set(format(float(0), ',.2f'))
            else:
                comma24 = format(float(cr24.get()), ',.2f')
                cr24.set(comma24)
        except:
            credit_entry24.delete(0, END)
            cr24.set(format(float(0), ',.2f'))
        finally:
            self.updatetotal()
            
    def showdebit24(self, *args):
        global debit_entry24
        global dr24
        dr24 = StringVar()
        debit_entry24 = Entry(line_frame24, textvariable = dr24, font = fonts, relief = SUNKEN, bd = 2, width = 14, justify = "right")
        debit_entry24.grid(column = 2, row = 0)
        dr24.trace("w", lambda nm, idx, mode, var24 = dr24: self.validate_float24(var24))
        debit_entry24.bind("<Tab>", self.validate_comma_dr24)
                
    def showcredit24(self):
        global cr24
        cr24 = StringVar()
        global credit_entry24
        credit_entry24 = Entry(line_frame24, textvariable = cr24, font = fonts, relief = SUNKEN, bd = 2, width = 14, justify = "right")
        credit_entry24.grid(column = 3, row = 0)
        cr24.trace("w", lambda nm, idx, mode, var24 = cr24: self.validate_float24(var24)) 
        credit_entry24.bind("<Tab>", self.validate_comma_cr24)
        
    def showcomment24(self):  
        global comment24
        comment24 = StringVar()
        global comment_entry24
        comment_entry24 = Entry(line_frame24, textvariable = comment24, font = fonts, relief = SUNKEN, bd = 2, width = 20, justify = "left")
        comment_entry24.grid(column = 4, row = 0)

    def showrepeat24(self, *args):
        global repeat_button24
        repeat_button24 = Button(line_frame24, text = "clear", bg = buttonbg, font = clearbuttonfont, command = self.cleanentry24, bd =0)
        repeat_button24.grid(column = 5, row = 0, padx = 2)
        repeat_button24.bind("<Return>", self.cleanentry24)
 
    def cleanentry24(self, *args):
        code_entry24.delete(0, END)
        title_label24.config(text = "")
        debit_entry24.delete(0, END)
        credit_entry24.delete(0, END)
        comment_entry24.delete(0, END)

        debit_entry24.config(state = DISABLED)
        credit_entry24.config(state = DISABLED)
        comment_entry24.config(state = DISABLED)
        repeat_button24.config(state = DISABLED)

        self.updatetotal()

### line 25 ###
    def lineframe25(self, master):
        global line_frame25
        line_frame25 = Frame(master, bg = wc)
        line_frame25.grid(column = 0, row = 30, sticky = NW)

    def show_code25(self):
        global journalcode25
        journalcode25 = StringVar()
        global code_entry25
        code_entry25 = Entry(line_frame25, textvariable = journalcode25, font = fonts, relief = SUNKEN, bd = 2, width = 5, justify = "center")
        code_entry25.grid(column = 0, row = 0)
        code_entry25.bind("<Tab>", self.journalcode_val25)
        code_entry25.bind("<FocusOut>", self.journalcode_val25)
        code_entry25.bind("<Return>", self.journalcode_val25)
        
    def journalcode_val25(self, *args):
        global result25
        codeinput25 = journalcode25.get()
        if len(codeinput25) == 4 and codeinput25.isdigit():
            if company == "DBPSC":
                find_code25 = "SELECT title FROM chart WHERE code = ?"
            else:
                find_code25 = "SELECT title FROM dssichart WHERE code = ?"
            c.execute(find_code25, [codeinput25])
            result25 = c.fetchall()
            if result25:
                title_label25.config(text = result25[0][0])
                self.linestatenormal25()
                debit_entry25.focus()
                self.updatetotal()
            else:
                self.linestatedisabled25()
                code_entry25.delete(0, END)
                self.updatetotal()
        else:
            self.linestatedisabled25()
            code_entry25.delete(0, END)
            self.updatetotal()

    def linestatenormal25(self):
        debit_entry25.config(state = NORMAL)
        credit_entry25.config(state = NORMAL)
        comment_entry25.config(state = NORMAL)
        repeat_button25.config(state = NORMAL)

    def linestatedisabled25(self):
        title_label25.config(text = "")
        debit_entry25.delete(0, END)
        credit_entry25.delete(0, END)
        comment_entry25.delete(0, END)

        debit_entry25.config(state = DISABLED)
        credit_entry25.config(state = DISABLED)
        comment_entry25.config(state = DISABLED)
        repeat_button25.config(state = DISABLED)
      
    def showtitle25(self, *args):
        global title_label25
        title_label25 = Label(line_frame25, text = "", font = fonts, relief = SUNKEN, bd = 2, width = 33, anchor = W, bg = codetitlebg, fg = codetitlefg)
        title_label25.grid(column = 1, row = 0)

    def validate_float25(self, var25):
        global new_value25
        new_value25 = var25.get()
        try:
            new_value25 == "" or "," or float(new_value25)
            old_value25 = new_value25
        except:
            old_value25 = ""
            var25.set(old_value25)

    def validate_comma_dr25(self, *args):
        try:
            if debit_entry25.get() == "":
                dr25.set(format(float(0), ',.2f'))
            else:
                comma25 = format(float(dr25.get()), ',.2f')
                dr25.set(comma25)
        except:
            debit_entry25.delete(0, END)
            dr25.set(format(float(0), ',.2f'))
        finally:
            self.updatetotal()
            
    def validate_comma_cr25(self, *args):
        try:
            if credit_entry25.get() == "":
                cr25.set(format(float(0), ',.2f'))
            else:
                comma25 = format(float(cr25.get()), ',.2f')
                cr25.set(comma25)
        except:
            credit_entry25.delete(0, END)
            cr25.set(format(float(0), ',.2f'))
        finally:
            self.updatetotal()
            
    def showdebit25(self, *args):
        global debit_entry25
        global dr25
        dr25 = StringVar()
        debit_entry25 = Entry(line_frame25, textvariable = dr25, font = fonts, relief = SUNKEN, bd = 2, width = 14, justify = "right")
        debit_entry25.grid(column = 2, row = 0)
        dr25.trace("w", lambda nm, idx, mode, var25 = dr25: self.validate_float25(var25))
        debit_entry25.bind("<Tab>", self.validate_comma_dr25)
                
    def showcredit25(self):
        global cr25
        cr25 = StringVar()
        global credit_entry25
        credit_entry25 = Entry(line_frame25, textvariable = cr25, font = fonts, relief = SUNKEN, bd = 2, width = 14, justify = "right")
        credit_entry25.grid(column = 3, row = 0)
        cr25.trace("w", lambda nm, idx, mode, var25 = cr25: self.validate_float25(var25)) 
        credit_entry25.bind("<Tab>", self.validate_comma_cr25)
        
    def showcomment25(self):  
        global comment25
        comment25 = StringVar()
        global comment_entry25
        comment_entry25 = Entry(line_frame25, textvariable = comment25, font = fonts, relief = SUNKEN, bd = 2, width = 20, justify = "left")
        comment_entry25.grid(column = 4, row = 0)

    def showrepeat25(self, *args):
        global repeat_button25
        repeat_button25 = Button(line_frame25, text = "clear", bg = buttonbg, font = clearbuttonfont, command = self.cleanentry25, bd =0)
        repeat_button25.grid(column = 5, row = 0, padx = 2)
        repeat_button25.bind("<Return>", self.cleanentry25)
 
    def cleanentry25(self, *args):
        code_entry25.delete(0, END)
        title_label25.config(text = "")
        debit_entry25.delete(0, END)
        credit_entry25.delete(0, END)
        comment_entry25.delete(0, END)

        debit_entry25.config(state = DISABLED)
        credit_entry25.config(state = DISABLED)
        comment_entry25.config(state = DISABLED)
        repeat_button25.config(state = DISABLED)

        self.updatetotal()

### line 26 ###
    def lineframe26(self, master):
        global line_frame26
        line_frame26 = Frame(master, bg = wc)
        line_frame26.grid(column = 0, row = 31, sticky = NW)

    def show_code26(self):
        global journalcode26
        journalcode26 = StringVar()
        global code_entry26
        code_entry26 = Entry(line_frame26, textvariable = journalcode26, font = fonts, relief = SUNKEN, bd = 2, width = 5, justify = "center")
        code_entry26.grid(column = 0, row = 0)
        code_entry26.bind("<Tab>", self.journalcode_val26)
        code_entry26.bind("<FocusOut>", self.journalcode_val26)
        code_entry26.bind("<Return>", self.journalcode_val26)
        
    def journalcode_val26(self, *args):
        global result26
        codeinput26 = journalcode26.get()
        if len(codeinput26) == 4 and codeinput26.isdigit():
            if company == "DBPSC":
                find_code26 = "SELECT title FROM chart WHERE code = ?"
            else:
                find_code26 = "SELECT title FROM dssichart WHERE code = ?"
            c.execute(find_code26, [codeinput26])
            result26 = c.fetchall()
            if result26:
                title_label26.config(text = result26[0][0])
                self.linestatenormal26()
                debit_entry26.focus()
                self.updatetotal()
            else:
                self.linestatedisabled26()
                code_entry26.delete(0, END)
                self.updatetotal()
        else:
            self.linestatedisabled26()
            code_entry26.delete(0, END)
            self.updatetotal()

    def linestatenormal26(self):
        debit_entry26.config(state = NORMAL)
        credit_entry26.config(state = NORMAL)
        comment_entry26.config(state = NORMAL)
        repeat_button26.config(state = NORMAL)

    def linestatedisabled26(self):
        title_label26.config(text = "")
        debit_entry26.delete(0, END)
        credit_entry26.delete(0, END)
        comment_entry26.delete(0, END)

        debit_entry26.config(state = DISABLED)
        credit_entry26.config(state = DISABLED)
        comment_entry26.config(state = DISABLED)
        repeat_button26.config(state = DISABLED)
      
    def showtitle26(self, *args):
        global title_label26
        title_label26 = Label(line_frame26, text = "", font = fonts, relief = SUNKEN, bd = 2, width = 33, anchor = W, bg = codetitlebg, fg = codetitlefg)
        title_label26.grid(column = 1, row = 0)

    def validate_float26(self, var26):
        global new_value26
        new_value26 = var26.get()
        try:
            new_value26 == "" or "," or float(new_value26)
            old_value26 = new_value26
        except:
            old_value26 = ""
            var26.set(old_value26)

    def validate_comma_dr26(self, *args):
        try:
            if debit_entry26.get() == "":
                dr26.set(format(float(0), ',.2f'))
            else:
                comma26 = format(float(dr26.get()), ',.2f')
                dr26.set(comma26)
        except:
            debit_entry26.delete(0, END)
            dr26.set(format(float(0), ',.2f'))
        finally:
            self.updatetotal()
            
    def validate_comma_cr26(self, *args):
        try:
            if credit_entry26.get() == "":
                cr26.set(format(float(0), ',.2f'))
            else:
                comma26 = format(float(cr26.get()), ',.2f')
                cr26.set(comma26)
        except:
            credit_entry26.delete(0, END)
            cr26.set(format(float(0), ',.2f'))
        finally:
            self.updatetotal()
            
    def showdebit26(self, *args):
        global debit_entry26
        global dr26
        dr26 = StringVar()
        debit_entry26 = Entry(line_frame26, textvariable = dr26, font = fonts, relief = SUNKEN, bd = 2, width = 14, justify = "right")
        debit_entry26.grid(column = 2, row = 0)
        dr26.trace("w", lambda nm, idx, mode, var26 = dr26: self.validate_float26(var26))
        debit_entry26.bind("<Tab>", self.validate_comma_dr26)
                
    def showcredit26(self):
        global cr26
        cr26 = StringVar()
        global credit_entry26
        credit_entry26 = Entry(line_frame26, textvariable = cr26, font = fonts, relief = SUNKEN, bd = 2, width = 14, justify = "right")
        credit_entry26.grid(column = 3, row = 0)
        cr26.trace("w", lambda nm, idx, mode, var26 = cr26: self.validate_float26(var26)) 
        credit_entry26.bind("<Tab>", self.validate_comma_cr26)
        
    def showcomment26(self):  
        global comment26
        comment26 = StringVar()
        global comment_entry26
        comment_entry26 = Entry(line_frame26, textvariable = comment26, font = fonts, relief = SUNKEN, bd = 2, width = 20, justify = "left")
        comment_entry26.grid(column = 4, row = 0)

    def showrepeat26(self, *args):
        global repeat_button26
        repeat_button26 = Button(line_frame26, text = "clear", bg = buttonbg, font = clearbuttonfont, command = self.cleanentry26, bd =0)
        repeat_button26.grid(column = 5, row = 0, padx = 2)
        repeat_button26.bind("<Return>", self.cleanentry26)
 
    def cleanentry26(self, *args):
        code_entry26.delete(0, END)
        title_label26.config(text = "")
        debit_entry26.delete(0, END)
        credit_entry26.delete(0, END)
        comment_entry26.delete(0, END)

        debit_entry26.config(state = DISABLED)
        credit_entry26.config(state = DISABLED)
        comment_entry26.config(state = DISABLED)
        repeat_button26.config(state = DISABLED)

        self.updatetotal()

### line 27 ###
    def lineframe27(self, master):
        global line_frame27
        line_frame27 = Frame(master, bg = wc)
        line_frame27.grid(column = 0, row = 32, sticky = NW)

    def show_code27(self):
        global journalcode27
        journalcode27 = StringVar()
        global code_entry27
        code_entry27 = Entry(line_frame27, textvariable = journalcode27, font = fonts, relief = SUNKEN, bd = 2, width = 5, justify = "center")
        code_entry27.grid(column = 0, row = 0)
        code_entry27.bind("<Tab>", self.journalcode_val27)
        code_entry27.bind("<FocusOut>", self.journalcode_val27)
        code_entry27.bind("<Return>", self.journalcode_val27)
        
    def journalcode_val27(self, *args):
        global result27
        codeinput27 = journalcode27.get()
        if len(codeinput27) == 4 and codeinput27.isdigit():
            if company == "DBPSC":
                find_code27 = "SELECT title FROM chart WHERE code = ?"
            else:
                find_code27 = "SELECT title FROM dssichart WHERE code = ?"
            c.execute(find_code27, [codeinput27])
            result27 = c.fetchall()
            if result27:
                title_label27.config(text = result27[0][0])
                self.linestatenormal27()
                debit_entry27.focus()
                self.updatetotal()
            else:
                self.linestatedisabled27()
                code_entry27.delete(0, END)
                self.updatetotal()
        else:
            self.linestatedisabled27()
            code_entry27.delete(0, END)
            self.updatetotal()

    def linestatenormal27(self):
        debit_entry27.config(state = NORMAL)
        credit_entry27.config(state = NORMAL)
        comment_entry27.config(state = NORMAL)
        repeat_button27.config(state = NORMAL)

    def linestatedisabled27(self):
        title_label27.config(text = "")
        debit_entry27.delete(0, END)
        credit_entry27.delete(0, END)
        comment_entry27.delete(0, END)

        debit_entry27.config(state = DISABLED)
        credit_entry27.config(state = DISABLED)
        comment_entry27.config(state = DISABLED)
        repeat_button27.config(state = DISABLED)
      
    def showtitle27(self, *args):
        global title_label27
        title_label27 = Label(line_frame27, text = "", font = fonts, relief = SUNKEN, bd = 2, width = 33, anchor = W, bg = codetitlebg, fg = codetitlefg)
        title_label27.grid(column = 1, row = 0)

    def validate_float27(self, var27):
        global new_value27
        new_value27 = var27.get()
        try:
            new_value27 == "" or "," or float(new_value27)
            old_value27 = new_value27
        except:
            old_value27 = ""
            var27.set(old_value27)

    def validate_comma_dr27(self, *args):
        try:
            if debit_entry27.get() == "":
                dr27.set(format(float(0), ',.2f'))
            else:
                comma27 = format(float(dr27.get()), ',.2f')
                dr27.set(comma27)
        except:
            debit_entry27.delete(0, END)
            dr27.set(format(float(0), ',.2f'))
        finally:
            self.updatetotal()
            
    def validate_comma_cr27(self, *args):
        try:
            if credit_entry27.get() == "":
                cr27.set(format(float(0), ',.2f'))
            else:
                comma27 = format(float(cr27.get()), ',.2f')
                cr27.set(comma27)
        except:
            credit_entry27.delete(0, END)
            cr27.set(format(float(0), ',.2f'))
        finally:
            self.updatetotal()
            
    def showdebit27(self, *args):
        global debit_entry27
        global dr27
        dr27 = StringVar()
        debit_entry27 = Entry(line_frame27, textvariable = dr27, font = fonts, relief = SUNKEN, bd = 2, width = 14, justify = "right")
        debit_entry27.grid(column = 2, row = 0)
        dr27.trace("w", lambda nm, idx, mode, var27 = dr27: self.validate_float27(var27))
        debit_entry27.bind("<Tab>", self.validate_comma_dr27)
                
    def showcredit27(self):
        global cr27
        cr27 = StringVar()
        global credit_entry27
        credit_entry27 = Entry(line_frame27, textvariable = cr27, font = fonts, relief = SUNKEN, bd = 2, width = 14, justify = "right")
        credit_entry27.grid(column = 3, row = 0)
        cr27.trace("w", lambda nm, idx, mode, var27 = cr27: self.validate_float27(var27)) 
        credit_entry27.bind("<Tab>", self.validate_comma_cr27)
        
    def showcomment27(self):  
        global comment27
        comment27 = StringVar()
        global comment_entry27
        comment_entry27 = Entry(line_frame27, textvariable = comment27, font = fonts, relief = SUNKEN, bd = 2, width = 20, justify = "left")
        comment_entry27.grid(column = 4, row = 0)

    def showrepeat27(self, *args):
        global repeat_button27
        repeat_button27 = Button(line_frame27, text = "clear", bg = buttonbg, font = clearbuttonfont, command = self.cleanentry27, bd =0)
        repeat_button27.grid(column = 5, row = 0, padx = 2)
        repeat_button27.bind("<Return>", self.cleanentry27)
 
    def cleanentry27(self, *args):
        code_entry27.delete(0, END)
        title_label27.config(text = "")
        debit_entry27.delete(0, END)
        credit_entry27.delete(0, END)
        comment_entry27.delete(0, END)

        debit_entry27.config(state = DISABLED)
        credit_entry27.config(state = DISABLED)
        comment_entry27.config(state = DISABLED)
        repeat_button27.config(state = DISABLED)

        self.updatetotal()

### line 28 ###
    def lineframe28(self, master):
        global line_frame28
        line_frame28 = Frame(master, bg = wc)
        line_frame28.grid(column = 0, row = 33, sticky = NW)

    def show_code28(self):
        global journalcode28
        journalcode28 = StringVar()
        global code_entry28
        code_entry28 = Entry(line_frame28, textvariable = journalcode28, font = fonts, relief = SUNKEN, bd = 2, width = 5, justify = "center")
        code_entry28.grid(column = 0, row = 0)
        code_entry28.bind("<Tab>", self.journalcode_val28)
        code_entry28.bind("<FocusOut>", self.journalcode_val28)
        code_entry28.bind("<Return>", self.journalcode_val28)
        
    def journalcode_val28(self, *args):
        global result28
        codeinput28 = journalcode28.get()
        if len(codeinput28) == 4 and codeinput28.isdigit():
            if company == "DBPSC":
                find_code28 = "SELECT title FROM chart WHERE code = ?"
            else:
                find_code28 = "SELECT title FROM dssichart WHERE code = ?"
            c.execute(find_code28, [codeinput28])
            result28 = c.fetchall()
            if result28:
                title_label28.config(text = result28[0][0])
                self.linestatenormal28()
                debit_entry28.focus()
                self.updatetotal()
            else:
                self.linestatedisabled28()
                code_entry28.delete(0, END)
                self.updatetotal()
        else:
            self.linestatedisabled28()
            code_entry28.delete(0, END)
            self.updatetotal()

    def linestatenormal28(self):
        debit_entry28.config(state = NORMAL)
        credit_entry28.config(state = NORMAL)
        comment_entry28.config(state = NORMAL)
        repeat_button28.config(state = NORMAL)

    def linestatedisabled28(self):
        title_label28.config(text = "")
        debit_entry28.delete(0, END)
        credit_entry28.delete(0, END)
        comment_entry28.delete(0, END)

        debit_entry28.config(state = DISABLED)
        credit_entry28.config(state = DISABLED)
        comment_entry28.config(state = DISABLED)
        repeat_button28.config(state = DISABLED)
      
    def showtitle28(self, *args):
        global title_label28
        title_label28 = Label(line_frame28, text = "", font = fonts, relief = SUNKEN, bd = 2, width = 33, anchor = W, bg = codetitlebg, fg = codetitlefg)
        title_label28.grid(column = 1, row = 0)

    def validate_float28(self, var28):
        global new_value28
        new_value28 = var28.get()
        try:
            new_value28 == "" or "," or float(new_value28)
            old_value28 = new_value28
        except:
            old_value28 = ""
            var28.set(old_value28)

    def validate_comma_dr28(self, *args):
        try:
            if debit_entry28.get() == "":
                dr28.set(format(float(0), ',.2f'))
            else:
                comma28 = format(float(dr28.get()), ',.2f')
                dr28.set(comma28)
        except:
            debit_entry28.delete(0, END)
            dr28.set(format(float(0), ',.2f'))
        finally:
            self.updatetotal()
            
    def validate_comma_cr28(self, *args):
        try:
            if credit_entry28.get() == "":
                cr28.set(format(float(0), ',.2f'))
            else:
                comma28 = format(float(cr28.get()), ',.2f')
                cr28.set(comma28)
        except:
            credit_entry28.delete(0, END)
            cr28.set(format(float(0), ',.2f'))
        finally:
            self.updatetotal()
            
    def showdebit28(self, *args):
        global debit_entry28
        global dr28
        dr28 = StringVar()
        debit_entry28 = Entry(line_frame28, textvariable = dr28, font = fonts, relief = SUNKEN, bd = 2, width = 14, justify = "right")
        debit_entry28.grid(column = 2, row = 0)
        dr28.trace("w", lambda nm, idx, mode, var28 = dr28: self.validate_float28(var28))
        debit_entry28.bind("<Tab>", self.validate_comma_dr28)
                
    def showcredit28(self):
        global cr28
        cr28 = StringVar()
        global credit_entry28
        credit_entry28 = Entry(line_frame28, textvariable = cr28, font = fonts, relief = SUNKEN, bd = 2, width = 14, justify = "right")
        credit_entry28.grid(column = 3, row = 0)
        cr28.trace("w", lambda nm, idx, mode, var28 = cr28: self.validate_float28(var28)) 
        credit_entry28.bind("<Tab>", self.validate_comma_cr28)
        
    def showcomment28(self):  
        global comment28
        comment28 = StringVar()
        global comment_entry28
        comment_entry28 = Entry(line_frame28, textvariable = comment28, font = fonts, relief = SUNKEN, bd = 2, width = 20, justify = "left")
        comment_entry28.grid(column = 4, row = 0)

    def showrepeat28(self, *args):
        global repeat_button28
        repeat_button28 = Button(line_frame28, text = "clear", bg = buttonbg, font = clearbuttonfont, command = self.cleanentry28, bd =0)
        repeat_button28.grid(column = 5, row = 0, padx = 2)
        repeat_button28.bind("<Return>", self.cleanentry28)
 
    def cleanentry28(self, *args):
        code_entry28.delete(0, END)
        title_label28.config(text = "")
        debit_entry28.delete(0, END)
        credit_entry28.delete(0, END)
        comment_entry28.delete(0, END)

        debit_entry28.config(state = DISABLED)
        credit_entry28.config(state = DISABLED)
        comment_entry28.config(state = DISABLED)
        repeat_button28.config(state = DISABLED)

        self.updatetotal()

### line 29 ###
    def lineframe29(self, master):
        global line_frame29
        line_frame29 = Frame(master, bg = wc)
        line_frame29.grid(column = 0, row = 34, sticky = NW)

    def show_code29(self):
        global journalcode29
        journalcode29 = StringVar()
        global code_entry29
        code_entry29 = Entry(line_frame29, textvariable = journalcode29, font = fonts, relief = SUNKEN, bd = 2, width = 5, justify = "center")
        code_entry29.grid(column = 0, row = 0)
        code_entry29.bind("<Tab>", self.journalcode_val29)
        code_entry29.bind("<FocusOut>", self.journalcode_val29)
        code_entry29.bind("<Return>", self.journalcode_val29)
        
    def journalcode_val29(self, *args):
        global result29
        codeinput29 = journalcode29.get()
        if len(codeinput29) == 4 and codeinput29.isdigit():
            if company == "DBPSC":
                find_code29 = "SELECT title FROM chart WHERE code = ?"
            else:
                find_code29 = "SELECT title FROM dssichart WHERE code = ?"
            c.execute(find_code29, [codeinput29])
            result29 = c.fetchall()
            if result29:
                title_label29.config(text = result29[0][0])
                self.linestatenormal29()
                debit_entry29.focus()
                self.updatetotal()
            else:
                self.linestatedisabled29()
                code_entry29.delete(0, END)
                self.updatetotal()
        else:
            self.linestatedisabled29()
            code_entry29.delete(0, END)
            self.updatetotal()

    def linestatenormal29(self):
        debit_entry29.config(state = NORMAL)
        credit_entry29.config(state = NORMAL)
        comment_entry29.config(state = NORMAL)
        repeat_button29.config(state = NORMAL)

    def linestatedisabled29(self):
        title_label29.config(text = "")
        debit_entry29.delete(0, END)
        credit_entry29.delete(0, END)
        comment_entry29.delete(0, END)

        debit_entry29.config(state = DISABLED)
        credit_entry29.config(state = DISABLED)
        comment_entry29.config(state = DISABLED)
        repeat_button29.config(state = DISABLED)
      
    def showtitle29(self, *args):
        global title_label29
        title_label29 = Label(line_frame29, text = "", font = fonts, relief = SUNKEN, bd = 2, width = 33, anchor = W, bg = codetitlebg, fg = codetitlefg)
        title_label29.grid(column = 1, row = 0)

    def validate_float29(self, var29):
        global new_value29
        new_value29 = var29.get()
        try:
            new_value29 == "" or "," or float(new_value29)
            old_value29 = new_value29
        except:
            old_value29 = ""
            var29.set(old_value29)

    def validate_comma_dr29(self, *args):
        try:
            if debit_entry29.get() == "":
                dr29.set(format(float(0), ',.2f'))
            else:
                comma29 = format(float(dr29.get()), ',.2f')
                dr29.set(comma29)
        except:
            debit_entry29.delete(0, END)
            dr29.set(format(float(0), ',.2f'))
        finally:
            self.updatetotal()
            
    def validate_comma_cr29(self, *args):
        try:
            if credit_entry29.get() == "":
                cr29.set(format(float(0), ',.2f'))
            else:
                comma29 = format(float(cr29.get()), ',.2f')
                cr29.set(comma29)
        except:
            credit_entry29.delete(0, END)
            cr29.set(format(float(0), ',.2f'))
        finally:
            self.updatetotal()
            
    def showdebit29(self, *args):
        global debit_entry29
        global dr29
        dr29 = StringVar()
        debit_entry29 = Entry(line_frame29, textvariable = dr29, font = fonts, relief = SUNKEN, bd = 2, width = 14, justify = "right")
        debit_entry29.grid(column = 2, row = 0)
        dr29.trace("w", lambda nm, idx, mode, var29 = dr29: self.validate_float29(var29))
        debit_entry29.bind("<Tab>", self.validate_comma_dr29)
                
    def showcredit29(self):
        global cr29
        cr29 = StringVar()
        global credit_entry29
        credit_entry29 = Entry(line_frame29, textvariable = cr29, font = fonts, relief = SUNKEN, bd = 2, width = 14, justify = "right")
        credit_entry29.grid(column = 3, row = 0)
        cr29.trace("w", lambda nm, idx, mode, var29 = cr29: self.validate_float29(var29)) 
        credit_entry29.bind("<Tab>", self.validate_comma_cr29)
        
    def showcomment29(self):  
        global comment29
        comment29 = StringVar()
        global comment_entry29
        comment_entry29 = Entry(line_frame29, textvariable = comment29, font = fonts, relief = SUNKEN, bd = 2, width = 20, justify = "left")
        comment_entry29.grid(column = 4, row = 0)

    def showrepeat29(self, *args):
        global repeat_button29
        repeat_button29 = Button(line_frame29, text = "clear", bg = buttonbg, font = clearbuttonfont, command = self.cleanentry29, bd =0)
        repeat_button29.grid(column = 5, row = 0, padx = 2)
        repeat_button29.bind("<Return>", self.cleanentry29)
 
    def cleanentry29(self, *args):
        code_entry29.delete(0, END)
        title_label29.config(text = "")
        debit_entry29.delete(0, END)
        credit_entry29.delete(0, END)
        comment_entry29.delete(0, END)

        debit_entry29.config(state = DISABLED)
        credit_entry29.config(state = DISABLED)
        comment_entry29.config(state = DISABLED)
        repeat_button29.config(state = DISABLED)

        self.updatetotal()

### line 30 ###
    def lineframe30(self, master):
        global line_frame30
        line_frame30 = Frame(master, bg = wc)
        line_frame30.grid(column = 0, row = 35, sticky = NW)

    def show_code30(self):
        global journalcode30
        journalcode30 = StringVar()
        global code_entry30
        code_entry30 = Entry(line_frame30, textvariable = journalcode30, font = fonts, relief = SUNKEN, bd = 2, width = 5, justify = "center")
        code_entry30.grid(column = 0, row = 0)
        code_entry30.bind("<Tab>", self.journalcode_val30)
        code_entry30.bind("<FocusOut>", self.journalcode_val30)
        code_entry30.bind("<Return>", self.journalcode_val30)
        
    def journalcode_val30(self, *args):
        global result30
        codeinput30 = journalcode30.get()
        if len(codeinput30) == 4 and codeinput30.isdigit():
            if company == "DBPSC":
                find_code30 = "SELECT title FROM chart WHERE code = ?"
            else:
                find_code30 = "SELECT title FROM dssichart WHERE code = ?"
            c.execute(find_code30, [codeinput30])
            result30 = c.fetchall()
            if result30:
                title_label30.config(text = result30[0][0])
                self.linestatenormal30()
                debit_entry30.focus()
                self.updatetotal()
            else:
                self.linestatedisabled30()
                code_entry30.delete(0, END)
                self.updatetotal()
        else:
            self.linestatedisabled30()
            code_entry30.delete(0, END)
            self.updatetotal()

    def linestatenormal30(self):
        debit_entry30.config(state = NORMAL)
        credit_entry30.config(state = NORMAL)
        comment_entry30.config(state = NORMAL)
        repeat_button30.config(state = NORMAL)

    def linestatedisabled30(self):
        title_label30.config(text = "")
        debit_entry30.delete(0, END)
        credit_entry30.delete(0, END)
        comment_entry30.delete(0, END)

        debit_entry30.config(state = DISABLED)
        credit_entry30.config(state = DISABLED)
        comment_entry30.config(state = DISABLED)
        repeat_button30.config(state = DISABLED)
      
    def showtitle30(self, *args):
        global title_label30
        title_label30 = Label(line_frame30, text = "", font = fonts, relief = SUNKEN, bd = 2, width = 33, anchor = W, bg = codetitlebg, fg = codetitlefg)
        title_label30.grid(column = 1, row = 0)

    def validate_float30(self, var30):
        global new_value30
        new_value30 = var30.get()
        try:
            new_value30 == "" or "," or float(new_value30)
            old_value30 = new_value30
        except:
            old_value30 = ""
            var30.set(old_value30)

    def validate_comma_dr30(self, *args):
        try:
            if debit_entry30.get() == "":
                dr30.set(format(float(0), ',.2f'))
            else:
                comma30 = format(float(dr30.get()), ',.2f')
                dr30.set(comma30)
        except:
            debit_entry30.delete(0, END)
            dr30.set(format(float(0), ',.2f'))
        finally:
            self.updatetotal()
            
    def validate_comma_cr30(self, *args):
        try:
            if credit_entry30.get() == "":
                cr30.set(format(float(0), ',.2f'))
            else:
                comma30 = format(float(cr30.get()), ',.2f')
                cr30.set(comma30)
        except:
            credit_entry30.delete(0, END)
            cr30.set(format(float(0), ',.2f'))
        finally:
            self.updatetotal()
            
    def showdebit30(self, *args):
        global debit_entry30
        global dr30
        dr30 = StringVar()
        debit_entry30 = Entry(line_frame30, textvariable = dr30, font = fonts, relief = SUNKEN, bd = 2, width = 14, justify = "right")
        debit_entry30.grid(column = 2, row = 0)
        dr30.trace("w", lambda nm, idx, mode, var30 = dr30: self.validate_float30(var30))
        debit_entry30.bind("<Tab>", self.validate_comma_dr30)
                
    def showcredit30(self):
        global cr30
        cr30 = StringVar()
        global credit_entry30
        credit_entry30 = Entry(line_frame30, textvariable = cr30, font = fonts, relief = SUNKEN, bd = 2, width = 14, justify = "right")
        credit_entry30.grid(column = 3, row = 0)
        cr30.trace("w", lambda nm, idx, mode, var30 = cr30: self.validate_float30(var30)) 
        credit_entry30.bind("<Tab>", self.validate_comma_cr30)
        
    def showcomment30(self):  
        global comment30
        comment30 = StringVar()
        global comment_entry30
        comment_entry30 = Entry(line_frame30, textvariable = comment30, font = fonts, relief = SUNKEN, bd = 2, width = 20, justify = "left")
        comment_entry30.grid(column = 4, row = 0)

    def showrepeat30(self, *args):
        global repeat_button30
        repeat_button30 = Button(line_frame30, text = "clear", bg = buttonbg, font = clearbuttonfont, command = self.cleanentry30, bd =0)
        repeat_button30.grid(column = 5, row = 0, padx = 2)
        repeat_button30.bind("<Return>", self.cleanentry30)
 
    def cleanentry30(self, *args):
        code_entry30.delete(0, END)
        title_label30.config(text = "")
        debit_entry30.delete(0, END)
        credit_entry30.delete(0, END)
        comment_entry30.delete(0, END)

        debit_entry30.config(state = DISABLED)
        credit_entry30.config(state = DISABLED)
        comment_entry30.config(state = DISABLED)
        repeat_button30.config(state = DISABLED)

        self.updatetotal()

### posting functions ###

    def debitcreditchecker(self, *args):
        self.allvars()
        errorlines = []
        for x in range(30):
            if len(codelist[x].get()) == 4:
                try:
                    if float(drs[x].get().replace(",","")) > 0 and float(crs[x].get().replace(",","")) > 0:
                        errorlines.append("Please check line " + str(x + 1) + " with code " + codelist[x].get())
                    if drs[x].get() == "0.00" and crs[x].get() == "0.00":
                        errorlines.append("Please check line " + str(x + 1) + " with code " + codelist[x].get())
                    if drs[x].get() == "" and crs[x].get() == "":
                        errorlines.append("Please check line " + str(x + 1) + " with code " + codelist[x].get())
                except:
                    errorlines.append("Please check line " + str(x + 1) + " with code " + codelist[x].get())
        if len(errorlines) == 0:
            self.balancer()
        else:
            messagebox.showerror("Posting Protocol", errorlines)

    def balancer(self):
        self.updatetotal()
        try:
            if round(sum(drvalues),2) == round(sum(crvalues),2) and (sum(drvalues) + sum(crvalues)) > 0:
                if len(date_entry.get()) == 10:
                    self.lockchecker()
                    if datelockresult == "open":
                        self.final()
                    else:
                        messagebox.showerror("Period Locker", "Date entered is locked!")
                else:
                    messagebox.showerror("Date Checker", "Date entered is invalid!")
            else:
                messagebox.showerror("Error in Totals", "Please check totals!")
        except Exception as e:
            messagebox.showerror("Error!", e)

    def final(self, *args):
        finalresult = messagebox.askyesno("Posting Protocol", "Are you sure?")
        if finalresult == True:
            self.postmaster()
        else:
            pass
        
    def disablelines(self):
        self.allvars()
        for x in range(30):
            codelist[x].config(state = DISABLED)
            titles[x].config(fg = "grey")
            drs[x].config(state = DISABLED)
            crs[x].config(state = DISABLED)
            comments[x].config(state = DISABLED)
            repeats[x].config(state = DISABLED)
        debit_total.config(fg = "grey")
        credit_total.config(fg = "grey")

    def constantvalues(self):
        global ind, num, typ, dat, ref, par, use, mod, cli, sta, com, cre, splitm, splitd, splity
        ind = plusone
        num = newrecord
        typ = newtype
        dat = date_entry.get()
        cli = client_entry.get()
        ref = reference_entry.get()
        par = particulars_entry.get()
        use = self.username.get()
        mod = today.strftime('%m-%d-%Y')
        sta = "posted" #posted or final
        com = company
        cre = today.strftime('%m-%d-%Y')
        splitm = splitdate[0]
        splitd = splitdate[1]
        splity = splitdate[2]

    def linemaster(self):
        lin = 1
        for i in range(30):
            if len(codelist[i].get()) == 4:
                self.codeclass(codelist[i])
                validlines.append([lin, ind, num, typ, dat, ref, par, use, classcode, codelist[i].get(), titles[i].cget("text"), float(drs[i].get().replace(",", "")), float(crs[i].get().replace(",", "")), comments[i].get(), mod, cli, sta, com, cre, splitm, splitd, splity, book])
                lin += 1
        
    def postmaster(self):
        self.disablelines()
        date_entry.config(state = DISABLED)
        client_entry.config(state = DISABLED)
        particulars_entry.config(state = DISABLED)
        reference_entry.config(state = DISABLED)
         
        global validlines
        validlines = []
        
        self.numbermaster()
        self.constantvalues()
        self.linemaster()

        for x in validlines:
            if company == "DBPSC":
                insert = """INSERT INTO DBPSC (
                                   line, indexing, number, type, date,
                                   reference, particulars, user, class, code,
                                   title, debit, credit, comment, modified,
                                   client, status, company, created, month,
                                   day, year, book
                                   ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)"""
            else:
                insert = """INSERT INTO DSSI (
                                   line, indexing, number, type, date,
                                   reference, particulars, user, class, code,
                                   title, debit, credit, comment, modified,
                                   client, status, company, created, month,
                                   day, year, book
                                   ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)"""
            c.execute(insert, x)
            conn.commit()

        post_button.config(state = DISABLED)
        messagebox.showinfo("Posting Status", "Posting Successful! A new journal record has been created!")
        print_button.config(state = NORMAL)
        
    def printer(self, *args):   
        comfind = "SELECT * FROM company WHERE short = ?"
        c.execute(comfind, [company])
        comres = c.fetchall()
        if use == "mbsiano":
            proxy = "jsluching"
        else:
            proxy = use
        wb = openpyxl.load_workbook(path + "jv.xlsx") #path
        sheet = wb.active
        sheet['A1'] = comres[0][0]
        sheet['A2'] = comres[0][2]
        sheet['A3'] = comres[0][1]
        sheet['D6'] = dat
        sheet['J6'] = num
        sheet['D8'] = cli
        sheet['D10'] = par
        sheet['D12'] = ref
        sheet['J71'] = cre
        sheet['B69'] = proxy.upper()
        journalentries = []
        for i in range(30):
            if codelist[i].get() != '' and len(codelist[i].get()) == 4:
                journalentries.append([codelist[i].get(),titles[i].cget('text'),float(drs[i].get().replace(",","")),float(crs[i].get().replace(",",""))])
            else:
                pass
        row = 16
        for n in range(len(journalentries)):
            sheet['E' + str(row)] = journalentries[n][0]
            sheet['B' + str(row)] = journalentries[n][1]
            sheet['H' + str(row)] = journalentries[n][2]
            sheet['J' + str(row)] = journalentries[n][3]
            row += 1
        self.filtermaster(sheet, 'E', 16, 60)
        wb.save(savepath + "newrecord.xlsx") #path
        os.startfile(savepath + "newrecord.xlsx", "print") #path

    def printer2(self, *args):
        comfind = "SELECT * FROM company WHERE short = ?"
        c.execute(comfind, [company])
        comres = c.fetchall()
        if self.username.get() == "mbsiano":
            proxy = "jsluching"
        else:
            proxy = self.username.get()
        wb = openpyxl.load_workbook(path + "jv.xlsx") #path
        sheet = wb.active
        sheet['A1'] = comres[0][0]
        sheet['A2'] = comres[0][2]
        sheet['A3'] = comres[0][1]
        sheet['D6'] = date_entry.get()
        sheet['J6'] = vouchernumber.get('1.0', 'end-1c')
        sheet['D8'] = client_entry.get()
        sheet['D10'] = particulars_entry.get()
        sheet['D12'] = reference_entry.get()
        sheet['J71'] = v[19]
        sheet['B69'] = proxy.upper()
        journalentries = []
        for i in range(30):
            if codelist[i].get() != '' and len(codelist[i].get()) == 4:
                journalentries.append([codelist[i].get(),titles[i].cget('text'),float(drs[i].get().replace(",","")),float(crs[i].get().replace(",",""))])
            else:
                pass
        row = 16
        for n in range(len(journalentries)):
            sheet['E' + str(row)] = journalentries[n][0]
            sheet['B' + str(row)] = journalentries[n][1]
            sheet['H' + str(row)] = journalentries[n][2]
            sheet['J' + str(row)] = journalentries[n][3]
            row += 1
        self.filtermaster(sheet, 'E', 16, 60)
        wb.save(savepath + "viewrecord.xlsx") #path 
        os.startfile(savepath + "viewrecord.xlsx", "print") #path

    def filtermaster(self, sheet, col, x, y):
        for i in range(x,y):
            a = sheet[col + str(i)].value
            if a == "-":
                sheet.row_dimensions[i].hidden = True

root = Tk()
root.title(maintitle)
root.iconbitmap(iconpath + "icon.ico") #path
add = ImageTk.PhotoImage(Image.open(iconpath + "add.png")) #path
less = ImageTk.PhotoImage(Image.open(iconpath + "less.png")) #path
homeicon = ImageTk.PhotoImage(Image.open(iconpath + "home.png")) #path
previousicon = ImageTk.PhotoImage(Image.open(iconpath + "previous.png")) #path
nexticon = ImageTk.PhotoImage(Image.open(iconpath + "next.png")) #path
viewicon = ImageTk.PhotoImage(Image.open(iconpath + "view.png")) #path
plusicon = ImageTk.PhotoImage(Image.open(iconpath + "plus.png")) #path
updateicon = ImageTk.PhotoImage(Image.open(iconpath + "update.png")) #path
importericon = ImageTk.PhotoImage(Image.open(iconpath + "importer.png")) #path
reporticon = ImageTk.PhotoImage(Image.open(iconpath + "report.png")) #path
charticon = ImageTk.PhotoImage(Image.open(iconpath + "chart.png")) #path
biricon = ImageTk.PhotoImage(Image.open(iconpath + "bir.png")) #path
lockicon = ImageTk.PhotoImage(Image.open(iconpath + "lock.png")) #path
disbursementicon = ImageTk.PhotoImage(Image.open(iconpath + "disbursement.png")) #path
compilericon = ImageTk.PhotoImage(Image.open(iconpath + "compiler.png")) #path
accountingicon = ImageTk.PhotoImage(Image.open(iconpath + "accounting.png")) #path
financeicon = ImageTk.PhotoImage(Image.open(iconpath + "finance.png")) #path
findicon = ImageTk.PhotoImage(Image.open(iconpath + "find.png")) #path
texticon = ImageTk.PhotoImage(Image.open(iconpath + "text.png")) #path
bcdicon = ImageTk.PhotoImage(Image.open(iconpath + "bcd.png")) #path
gsadicon = ImageTk.PhotoImage(Image.open(iconpath + "gsad.png")) #path
companyicon = ImageTk.PhotoImage(Image.open(iconpath + "company.png")) #path
logouticon = ImageTk.PhotoImage(Image.open(iconpath + "logout.png")) #path
exiticon = ImageTk.PhotoImage(Image.open(iconpath + "exit.png")) #path
loginicon = ImageTk.PhotoImage(Image.open(iconpath + "login.png")) #path
registericon = ImageTk.PhotoImage(Image.open(iconpath + "register.png")) #path
changeicon = ImageTk.PhotoImage(Image.open(iconpath + "change.png")) #path
sendicon = ImageTk.PhotoImage(Image.open(iconpath + "send.png")) #path
roundedicon = ImageTk.PhotoImage(Image.open(iconpath + "rounded.png")) #path
buttonicon = ImageTk.PhotoImage(Image.open(iconpath + "button.png")) #path
dbpscicon = ImageTk.PhotoImage(Image.open(iconpath + "dbpsc.png")) #path
dssiicon = ImageTk.PhotoImage(Image.open(iconpath + "dssi.png")) #path
root.geometry(geometry)
root.config(background = mainbg)
root.resizable(width = False, height = False)
try:
    conn = sqlite3.connect(path + "charlie.db") #path
    c = conn.cursor()
    main(root)
except:
    messagebox.showerror("Connection Error", "MJ's PC must be turned on!")
root.mainloop()
